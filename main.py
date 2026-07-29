import hashlib, zipfile
from openpyxl.worksheet.worksheet import Worksheet
import sys, os, shutil, json, subprocess, csv, re, html
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
try:
    from pypdf import PdfReader
except ModuleNotFoundError:
    PdfReader = None
from PySide6.QtCore import Qt, QPropertyAnimation, QEasingCurve, QTimer
from PySide6.QtGui import QBrush, QPen, QPixmap, QTextDocument, QKeySequence, QShortcut, QIcon
from PySide6.QtWidgets import (QAbstractItemView, QApplication, QCalendarWidget, QComboBox, QDialog, QFileDialog, QFormLayout, QFrame,
    QGraphicsOpacityEffect, QGraphicsScene, QGraphicsView, QGridLayout, QGroupBox, QHBoxLayout, QHeaderView, QInputDialog, QLabel, QLineEdit,
    QListWidget, QMainWindow, QMessageBox, QProgressBar, QPushButton, QScrollArea, QSizePolicy, QSplitter, QStackedWidget, QTabWidget, QTableWidget,
    QTableWidgetItem, QTextBrowser, QTextEdit, QVBoxLayout, QWidget,)
from PySide6.QtPrintSupport import QPrinter, QPrintDialog
from typing import Any
from datetime import date, datetime, time, timedelta


# Bereinigte Einzeldatei-Version: veraltete, nicht mehr erreichbare Übergangsbausteine wurden entfernt.
APP_NAME = "ImmoVerwaltung"
APP_VERSION = "11.0"

SCHEMA: dict[str, list[str]] = {
    "Objekte": ["Objektname", "Vermieter", "Einheiten", "HV Rechnungen (PDF)", "Baurechnungen (PDF)", "Status", "Notiz", "Vollmachten (PDF)", "Abrechnungen", "Kontoauszüge (PDF)", "Hauskonto (PDF)", "Objektunterlagen (PDF)", "Buchhaltung", "Objektordner", "Vermieter E-Mail-Adresse", "Baujahr", "Grundstücksfläche", "Grundrisse (PDF)", "HV-Verträge (PDF)"],
    "Betriebskosten": ["Anschlussnutzer", "Zählernummer", "Geschoss", "Unterlagen (PDF)", "BK Art", "Objektordner", "Wohnungsordner", "Kundennummer", "Vertragsnummer", "Versorger", "Zählernummer Zusatz", "Abschlag monatlich", "Jahresverbrauch", "Abrechnungsjahr", "Fällig am", "Vertragsbeginn", "Vertragsende", "SEPA", "Rechnung PDF", "Vertrag PDF", "Bemerkungen", "Objektordner"],
    "Wohnungen": ["Objekt", "Wohnung", "Etage", "Größe qm", "Zimmer", "Status", "Objektordner", "Wohnungsordner", "Objektordner"],
    "Mieter": ["Mieter", "Mieter-Status", "Ort", "Wohnfläche", "Telefonnummer", "Mietbeginn", "Mietende", "Miete", "Kaltmiete", "BK monatlich", "Heizkosten", "BK jährlich", "Ist-BK jährlich", "Nachzahlung", "Guthaben", "BK-Status", "BK (PDF)", "HK (PDF)", "Objektordner", "Wohnungsordner", "E-Mail-Adresse"],
    "Mietverträge": ["Mieter", "Wohnung", "Beginn", "Ende", "Kaltmiete", "Nebenkosten", "Kaution", "Status", "Mietvertrag (PDF)", "Objektordner"],
    "Zahlungen": ["Datum", "Name/Quelle", "Betrag", "Verwendungszweck", "IBAN", "Status", "Kostenart", "Rechnungsjahr", "Fälligkeit", "Rechnung (PDF)", "Objektordner" ],
    "Zahlungsprüfung": ["Monat", "Mieter", "Wohnung", "Soll", "Gezahlt", "Differenz", "Status", "Zuordnung", "Treffer", "Objektordner"],
    "Mahnwesen": ["Mieter", "Wohnung", "Offener Betrag", "Mahnstufe", "Status", "Letzte Aktion", "Objektordner"],
    "Mitarbeiter": ["Name", "Bereich", "E-Mail", "Telefon", "Status"],
    "Stundennachweise": ["Mitarbeiter", "Datum", "Objekt", "Tätigkeit", "Stunden", "Status"],
    "Termine": ["ID", "Titel", "Datum", "Beginn", "Ende", "Objekt", "Ort", "Notiz", "Status", "Objektordner"],
    "Dienstleister": ["Firma", "Gewerk", "Ansprechpartner", "Telefon", "E-Mail", "Status", "Objektordner", "Kundennummer"],
    "Rechnungen": ["Rechnungsnr.", "Dienstleister", "Objekt", "Datum", "Netto", "MwSt.", "Brutto", "Status", "Objektordner"],
    "Dokumente": ["Titel", "Typ", "Bezug", "Datum", "Status", "Dateipfad", "Objektordner"],
    "Aktenstruktur": ["Objekt / Adresse", "Eigentümer", "Ordnerfarbe", "Ordnerart", "Unterlagen / Register", "Status", "Ordnerpfad"],
    "Grundsteuer": ["Objekt", "Steuernummer", "Jahr", "Betrag", "Fälligkeit", "Bescheid (PDF)", "Status", "Notiz", "Objektordner"],
    "Versicherungen": ["Objekt", "Versicherung", "Versicherungsnummer", "Gesellschaft", "Beitrag", "Fälligkeit", "Status", "Vertrag (PDF)", "Notiz", "Objektordner"],
    "Brand- und Arbeitsschutz": ["Objekt", "Prüfung", "Termin", "Nächste Prüfung", "Firma", "Kosten", "Prüfprotokoll (PDF)", "Status", "Notiz", "Objektordner"],
    "E-Mail": ["Datum", "Bereich", "Betreff", "Empfänger", "Absender", "Status", "PDF / Anlage (PDF)", "Notiz", "Objektordner"],
    "Wichtige Verträge": ["Objekt", "Vertragstyp", "Vertragspartner", "Beginn", "Ende", "Kosten", "Status", "Vertrag (PDF)", "Notiz", "Objektordner"],
    "Vermieterauskunft": ["Objekt", "Vermieter", "Datum", "Bereich", "Auskunft", "Status", "Dokument (PDF)", "Notiz", "Objektordner"],

    "Wohnungsgeberauskunft": ["Mieter", "Objekt", "Wohnung", "Einzugsdatum", "Auszugsdatum", "Ausgestellt am", "Status", "Wohnungsgeberauskunft (PDF)", "Objektordner", "Wohnungsordner", "Notiz"],
    "Übergabeprotokolle": ["Mieter", "Objekt", "Wohnung", "Art", "Datum", "Zählerstände", "Schlüssel", "Status", "Übergabeprotokoll (PDF)", "Objektordner", "Wohnungsordner", "Notiz"],
    "HV-Rechnungen": ["Datum", "Rechnungsnr.", "Objekt", "Vermieter", "Leistung", "Betrag netto", "MwSt.", "Betrag brutto", "Status", "Fällig am", "Zahlung am", "Rechnung PDF", "Notiz"],
    "Versorger": ["Versorger", "Art", "Kundennummer", "Vertragsnummer", "Objekt", "Wohnung", "Ansprechpartner", "Telefon", "E-Mail", "Vertragsbeginn", "Vertragsende", "Kündigungsfrist", "Fällig am", "Status", "Vertrag PDF", "Rechnung PDF", "Notiz", "Objektordner"],
    "Fristen": ["Titel", "Bereich", "Objekt", "Wohnung", "Mieter", "Fällig am", "Priorität", "Status", "Erinnerung", "Dokument PDF", "Objektordner", "Wohnungsordner", "Notiz"],
    "Aufgaben": ["Aufgabe", "Bereich", "Objekt", "Wohnung", "Mieter", "Verantwortlich", "Priorität", "Fällig am", "Status", "Dokument PDF", "Objektordner", "Wohnungsordner", "Notiz"],
    "Schäden": ["Schaden", "Objekt", "Wohnung", "Mieter", "Datum", "Beschreibung", "Priorität", "Status", "Handwerker", "Versicherung", "Kosten", "Foto/PDF", "Rechnung PDF", "Objektordner", "Wohnungsordner", "Notiz"],
    "Schlüssel": ["Objekt", "Wohnung", "Schlüsseltyp", "Schlüsselnummer", "Empfänger", "Ausgegeben am", "Rückgabe am", "Anzahl", "Status", "Objektordner", "Wohnungsordner", "Notiz"],
    "Ereignisprotokoll": ["Datum", "Uhrzeit", "Bereich", "Aktion", "Objekt", "Wohnung", "Mieter", "Beschreibung", "Benutzer", "Status"],
    "Kontoauszug-Import": ["Datum", "Buchungstext", "Auftraggeber", "Verwendungszweck", "Betrag", "Erkannt als", "Mieter", "Objekt", "Wohnung", "Status", "Hinweis"],
    "Mietkonto-Abgleich": ["Monat", "Jahr", "Mieter", "Objekt", "Wohnung", "Sollmiete", "Gezahlt", "Differenz", "Status", "Hinweis"],
    "Belegscanner": ["PDF", "Lieferant", "Rechnungsnummer", "Rechnungsdatum", "Fällig am", "Kundennummer", "Netto", "MwSt.", "Brutto", "Objekt", "Wohnung", "Kostenart", "Betriebskostenrelevant", "Umlagefähig", "Status", "Hinweis"],
    "Mitarbeiter-Login": ["Name", "Benutzername", "Passwort", "Rolle", "Status", "Letzter Login", "Hinweis"],
}
DATA_FILES: dict[str, str] = {
    "Objekte": "objekte.xlsx", "Betriebskosten": "betriebskosten.xlsx", "Wohnungen": "wohnungen.xlsx",
    "Mieter": "mieter.xlsx", "Mietverträge": "mietvertraege.xlsx", "Zahlungen": "zahlungen.xlsx",
    "Zahlungsprüfung": "zahlungspruefung.xlsx", "Mahnwesen": "mahnwesen.xlsx", "Mitarbeiter": "mitarbeiter.xlsx",
    "Stundennachweise": "stundennachweise.xlsx", "Termine": "termine.xlsx", "Dienstleister": "dienstleister.xlsx", "Rechnungen": "rechnungen.xlsx",
    "Dokumente": "dokumente.xlsx", "Aktenstruktur": "aktenstruktur.xlsx",    "Grundsteuer": "grundsteuer.xlsx",
    "Versicherungen": "versicherungen.xlsx",    "Brand- und Arbeitsschutz": "brand_arbeitsschutz.xlsx",
    "E-Mail": "email.xlsx",
    "Wichtige Verträge": "wichtige_vertraege.xlsx",
    "Vermieterauskunft": "vermieterauskunft.xlsx",

    "Wohnungsgeberauskunft": "wohnungsgeberauskunft.xlsx",
    "Übergabeprotokolle": "uebergabeprotokolle.xlsx",
    "HV-Rechnungen": "hv_rechnungen.xlsx",
    "Versorger": "versorger.xlsx",
    "Fristen": "fristen.xlsx",
    "Aufgaben": "aufgaben.xlsx",
    "Schäden": "schaeden.xlsx",
    "Schlüssel": "schluessel.xlsx",
    "Ereignisprotokoll": "ereignisprotokoll.xlsx",
    "Kontoauszug-Import": "kontoauszug_import.xlsx",
    "Mietkonto-Abgleich": "mietkonto_abgleich.xlsx",
    "Belegscanner": "belegscanner.xlsx",
    "Mitarbeiter-Login": "mitarbeiter_login.xlsx",
}
DATA: dict[str, list[list[str]]] = {k: [] for k in SCHEMA}

if getattr(sys, "frozen", False):
    APP_DIR = Path(sys.executable).resolve().parent
    RESOURCE_DIR = Path(getattr(sys, "_MEIPASS", APP_DIR))
else:
    APP_DIR = Path(__file__).resolve().parent
    RESOURCE_DIR = APP_DIR

# Kompatibilitätsname für bestehende Daten- und Konfigurationspfade.
BASE_DIR = APP_DIR


def resource_path(relative_path: str | Path) -> Path:
    """Liefert einen Asset-Pfad im Quellcode und in einer PyInstaller-EXE."""
    return RESOURCE_DIR / Path(relative_path)


CONFIG_FILE = APP_DIR / "config.json"
LOGO_FILE = resource_path("assets/dbs_logo.png")
APP_ICON_FILE = resource_path("assets/dbs_logo.ico")
DEFAULT_CONFIG = {"modus":"lokal","basis_pfad":"","daten_pfad":"daten","dokumente_pfad":"dokumente/pdf","exports_pfad":"exports","akten_pfad":"akten","backup_pfad":"backups","feedback_pfad":"feedback","theme":"hell","start_vollbild":"ja","auto_backup_start":"nein","tabellen_zeilenhoehe":"34", "firma_name":"DBS Hausverwaltung", "firma_untertitel":"Kaufmännische und technische Verwaltung", "firma_leistung":"Rentabilität, Betriebskosten, Hausservice", "firma_slogan":"Mit uns erwirtschaften Sie Gewinn!", "firma_inhaber":"Inh. Dipl. Betriebsw. Ulrike Schreiner", "firma_strasse":"Kirchstraße 3", "firma_plz_ort":"04703 Leisnig", "firma_telefon":"034321/13432", "firma_email":"", "firma_web":"", "firma_bank":"", "firma_iban":"", "firma_bic":"", "firma_steuer":"", "firma_logo":"", "firma_gruss":"Mit freundlichen Grüßen\nDBS Hausverwaltung", "bk_hinweis":"Bitte prüfen Sie die Abrechnung. Ein Guthaben wird verrechnet bzw. ausgezahlt; eine Nachzahlung ist zum angegebenen Termin fällig."}

def lade_config():
    if not CONFIG_FILE.exists():
        CONFIG_FILE.write_text(json.dumps(DEFAULT_CONFIG, indent=4, ensure_ascii=False), encoding="utf-8")
        return DEFAULT_CONFIG.copy()
    try:
        cfg = DEFAULT_CONFIG.copy()
        cfg.update(json.loads(CONFIG_FILE.read_text(encoding="utf-8")))
        return cfg
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return DEFAULT_CONFIG.copy()


def stelle_config_defaults_sicher() -> None:
    changed = False
    for key, value in DEFAULT_CONFIG.items():
        if key not in CONFIG:
            CONFIG[key] = value
            changed = True
    if changed:
        CONFIG_FILE.write_text(json.dumps(CONFIG, indent=4, ensure_ascii=False), encoding="utf-8")

CONFIG = lade_config()
stelle_config_defaults_sicher()

def config_pfad(key):
    value = str(CONFIG.get(key, "")).strip()
    basis = str(CONFIG.get("basis_pfad", "")).strip()
    if value.startswith("\\\\") or Path(value).is_absolute():
        return Path(value)
    if basis:
        return Path(basis) / value
    return BASE_DIR / value

DATEN_DIR = config_pfad("daten_pfad")
DOKUMENTE_DIR = config_pfad("dokumente_pfad")
EXPORT_DIR = config_pfad("exports_pfad")
AKTEN_DIR = config_pfad("akten_pfad")
BACKUP_DIR = config_pfad("backup_pfad")
FEEDBACK_DIR = config_pfad("feedback_pfad")
for target_dir in [DATEN_DIR, DOKUMENTE_DIR, EXPORT_DIR, AKTEN_DIR, BACKUP_DIR, FEEDBACK_DIR]:
    target_dir.mkdir(parents=True, exist_ok=True)


def excel_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, timedelta):
        return str(value.total_seconds())
    if isinstance(value, (str, int, float, bool)):
        return str(value)
    text_value = getattr(value, "text", None)
    if isinstance(text_value, str):
        return text_value
    formula_value = getattr(value, "value", None)
    if isinstance(formula_value, (str, int, float, bool)):
        return str(formula_value)
    return value.__class__.__name__


def erstelle_schnellbackup() -> Path | None:

    try:
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        ziel = BACKUP_DIR / ("backup_" + datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + ".zip")

        with zipfile.ZipFile(ziel, "w", zipfile.ZIP_DEFLATED) as zf:
            for backup_source in [DATEN_DIR, DOKUMENTE_DIR, EXPORT_DIR, AKTEN_DIR]:
                if not backup_source.exists():
                    continue
                for file in backup_source.rglob("*"):
                    if file.is_file():
                        try:
                            zf.write(file, file.relative_to(BASE_DIR))
                        except ValueError:
                            zf.write(file, file.name)

            if CONFIG_FILE.exists():
                zf.write(CONFIG_FILE, "config.json")

        return ziel
    except OSError as exc:
        QMessageBox.warning(None, "Backup-Fehler", str(exc))
        return None

def xlsx_pfad(titel): return DATEN_DIR / DATA_FILES[titel]

def erstelle_xlsx_wenn_fehlend(titel: str) -> None:
    p = xlsx_pfad(titel)

    if p.exists():
        return

    wb = Workbook()
    ws = wb.active

    if ws is None:
        return

    ws.title = titel[:31]
    ws.append(SCHEMA[titel])

    wb.save(p)

def lade_tabelle(titel: str) -> list[list[str]]:
    erstelle_xlsx_wenn_fehlend(titel)

    wb = load_workbook(xlsx_pfad(titel), data_only=True)
    ws = wb.active

    if ws is None:
        wb.close()
        return []

    # Optional für den Typechecker
    assert isinstance(ws, Worksheet)

    daten: list[list[str]] = []

    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            werte = [
                excel_text(value)
                for value in row
            ]

            if not any(werte):
                continue

            while len(werte) < len(SCHEMA[titel]):
                werte.append("")

            daten.append(werte[:len(SCHEMA[titel])])

    finally:
        wb.close()

    return daten

def speichere_tabelle(titel: str) -> None:
    if titel not in SCHEMA:
        raise ValueError(f"Unbekannte Tabelle: {titel}")

    # Bestehende Spalten werden konsistent befüllt; es werden keine Spalten ergänzt.
    normalizer = globals().get("objektordner_zuordnungen_normalisieren")
    if callable(normalizer):
        normalizer(titel)

    wb = Workbook()
    ws = wb.active

    if not isinstance(ws, Worksheet):
        raise RuntimeError("Kein aktives Excel-Arbeitsblatt vorhanden.")

    spalten = SCHEMA[titel]
    ws.title = titel[:31]
    ws.append(spalten)

    for row in DATA.get(titel, []):
        vals = list(row)

        while len(vals) < len(spalten):
            vals.append("")

        ws.append(vals[:len(spalten)])

    for col_index, col in enumerate(ws.columns, start=1):
        first_cell = col[0]
        value = excel_text(first_cell.value)
        ws.column_dimensions[get_column_letter(col_index)].width = min(
            max(len(value) + 4, 18),
            38,
        )

    try:
        wb.save(xlsx_pfad(titel))
    except PermissionError:
        QMessageBox.warning(
            None,
            "Excel-Datei geöffnet",
            f"Bitte Excel schließen:\n{xlsx_pfad(titel)}",
        )
    except OSError as exc:
        QMessageBox.warning(None, "Speicherfehler", str(exc))

def lade_alle():
    for titel in SCHEMA:
        if titel != "Zahlungsprüfung":
            DATA[titel] = lade_tabelle(titel)
    pruefe_zahlungen()
    speichere_tabelle("Zahlungsprüfung")

def to_float(value: Any, default: float = 0.0) -> float:
    """Liest Zahlen robust, auch im deutschen Format wie 1.234,56 €."""
    try:
        if value is None:
            return default
        if isinstance(value, (int, float)):
            return float(value)

        text_value = str(value).strip()
        if not text_value:
            return default

        # Währung, Leerzeichen und sonstige Zusätze entfernen.
        text_value = re.sub(r"[^0-9,.+\\-]", "", text_value)

        if "," in text_value and "." in text_value:
            # Das zuletzt vorkommende Trennzeichen gilt als Dezimaltrennzeichen.
            if text_value.rfind(",") > text_value.rfind("."):
                text_value = text_value.replace(".", "").replace(",", ".")
            else:
                text_value = text_value.replace(",", "")
        elif "," in text_value:
            text_value = text_value.replace(".", "").replace(",", ".")
        elif text_value.count(".") > 1:
            # Mehrere Punkte ohne Komma sind üblicherweise Tausendertrennzeichen.
            parts = text_value.split(".")
            text_value = "".join(parts[:-1]) + "." + parts[-1]

        return float(text_value)
    except (ValueError, TypeError):
        return default

def pruefe_zahlungen() -> None:
    rows: list[list[str]] = []

    for m in DATA["Mieter"]:
        name = m[0] if len(m) > 0 else ""
        ort = m[1] if len(m) > 1 else ""
        soll = m[6] if len(m) > 6 else "0"

        soll_f = to_float(soll)
        gezahlt = 0.0
        treffer = 0

        for z in DATA["Zahlungen"]:
            zahlungstext = " ".join(str(value) for value in z).lower()

            if name and name.lower() in zahlungstext:
                betrag = z[2] if len(z) > 2 else "0"
                gezahlt += to_float(betrag)
                treffer = 80

        diff = soll_f - gezahlt

        if abs(diff) < 0.01 and soll_f > 0:
            status = "Bezahlt"
        elif gezahlt == 0:
            status = "Offen"
        else:
            status = "Teilzahlung"

        rows.append([
            "Aktuell",
            name,
            ort,
            f"{soll_f:.2f}",
            f"{gezahlt:.2f}",
            f"{diff:.2f}",
            status,
            "Automatisch" if treffer else "Keine Zuordnung",
            f"{treffer} %",
        ])

    DATA["Zahlungsprüfung"] = rows
def system_datei_oeffnen(path: str | Path) -> None:
    p = Path(path)

    if not p.is_absolute():
        p = BASE_DIR / p

    if not p.exists():
        QMessageBox.warning(None, "Nicht gefunden", f"Nicht gefunden:\n{p}")
        return

    if sys.platform.startswith("win"):
        os.startfile(str(p))
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(p)])
    else:
        subprocess.Popen(["xdg-open", str(p)])

def kopiere_pdf_ins_projekt(path: str) -> str:
    quelle = Path(path)
    ziel = DOKUMENTE_DIR / quelle.name
    zaehler = 2

    while ziel.exists():
        ziel = DOKUMENTE_DIR / f"{quelle.stem}_{zaehler}{quelle.suffix}"
        zaehler += 1

    shutil.copy2(quelle, ziel)

    if ziel.is_relative_to(BASE_DIR):
        return str(ziel.relative_to(BASE_DIR)).replace("\\", "/")

    return str(ziel)

def frage_ja_nein(titel: str, text: str) -> bool:
    box = QMessageBox()
    box.setWindowTitle(titel)
    box.setText(text)

    ja = box.addButton("Ja", QMessageBox.ButtonRole.YesRole)
    nein = box.addButton("Nein", QMessageBox.ButtonRole.NoRole)
    box.setDefaultButton(nein)
    box.exec()

    return box.clickedButton() == ja

def norm_key(v: Any) -> str:
    t = str(v or "").lower().strip()
    for a, b in [("ä", "ae"), ("ö", "oe"), ("ü", "ue"), ("ß", "ss")]:
        t = t.replace(a, b)
    for ch in " _-./\\:;":
        t = t.replace(ch, "")
    return t

def lese_xlsx_import(path: str, titel: str) -> tuple[int, int]:
    if titel not in SCHEMA:
        raise ValueError(f"Unbekannte Tabelle: {titel}")

    wb = load_workbook(path, data_only=True)

    try:
        ws = wb.active

        if not isinstance(ws, Worksheet):
            return 0, 0

        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return 0, 0

        ziel = SCHEMA[titel]
        headers = [excel_text(v) for v in rows[0]]

        mapping: dict[int, int] = {}

        for i, header in enumerate(headers):
            header_key = norm_key(header)

            for j, feld in enumerate(ziel):
                feld_key = norm_key(feld)

                if (
                    header_key == feld_key
                    or header_key in feld_key
                    or feld_key in header_key
                ):
                    mapping[j] = i

        data_rows = rows[1:] if mapping else rows
        existing = {"|".join(norm_key(value) for value in row) for row in DATA[titel]}

        imported = 0
        duplicates = 0

        for raw in data_rows:
            new_row = ["" for _ in ziel]

            if mapping:
                for target_index, source_index in mapping.items():
                    if source_index < len(raw):

                        value = raw[source_index]
                        new_row[target_index] = excel_text(value)
            else:
                for index in range(min(len(raw), len(ziel))):
                    value = raw[index]
                    new_row[index] = excel_text(value)

            if not any(new_row):
                continue

            key = "|".join(norm_key(value) for value in new_row)

            if key in existing:
                duplicates += 1
                continue

            DATA[titel].append(new_row)
            existing.add(key)
            imported += 1

        DATA[titel].sort(key=lambda row: "|".join(norm_key(value) for value in row))
        speichere_tabelle(titel)

        return imported, duplicates

    finally:
        wb.close()

def baue_stylesheet():
    dunkel = str(CONFIG.get("theme", "hell")).lower() == "dunkel"

    if dunkel:
        bg = "#09111f"
        panel = "#111d30"
        panel_hover = "#17263d"
        panel_alt = "#0d192a"
        text = "#f8fafc"
        muted = "#9fb0c8"
        border = "#263750"
        input_bg = "#0d192a"
        header_bg = "#0b1525"
        table_alt = "#132239"
    else:
        bg = "#eef3f9"
        panel = "#ffffff"
        panel_hover = "#f7faff"
        panel_alt = "#f5f8fc"
        text = "#0b1628"
        muted = "#64748b"
        border = "#dce5f0"
        input_bg = "#ffffff"
        header_bg = "#ffffff"
        table_alt = "#f7f9fc"

    return f"""
    QWidget {{
        font-family: "Segoe UI", Arial;
        font-size: 14px;
        color: {text};
    }}

    QMainWindow {{
        background: {bg};
    }}

    QToolTip {{
        background: {panel};
        color: {text};
        border: 1px solid {border};
        padding: 7px;
        border-radius: 7px;
    }}

    #sidebar {{
        background: #071a35;
        border-right: 1px solid #16345d;
    }}

    #logoPanel {{
        background: #071a35;
        border-bottom: 1px solid #193b68;
    }}

    #brandText {{
        color: #dbeafe;
        font-size: 11px;
        font-weight: 700;
        padding-left: 14px;
        padding-bottom: 8px;
    }}

    #sidebarTitle {{
        color: #ffffff;
        font-size: 16px;
        font-weight: 900;
        padding: 4px 16px 10px 16px;
    }}

    #navButton {{
        color: #f00000;
        background: transparent;
        border: none;
        text-align: left;
        padding: 10px 14px;
        font-weight: 750;
        min-height: 42px;
        min-width: 255px;
        border-radius: 11px;
        margin: 2px 10px;
    }}

    #navButton:hover {{
        background: rgba(56, 111, 196, 0.30);
        color: #ffffff;
    }}

    #navButton[active="true"] {{
        background: #2563eb;
        color: #ffffff;
        border-left: 4px solid #93c5fd;
    }}

    #sidebarToggle {{
        background: rgba(255,255,255,0.08);
        color: #ffffff;
        border: 1px solid rgba(255,255,255,0.14);
        border-radius: 10px;
        min-width: 38px;
        max-width: 38px;
        min-height: 36px;
        padding: 0;
    }}

    #sidebarToggle:hover {{
        background: #2563eb;
        border-color: #60a5fa;
    }}

    #content {{
        background: {bg};
    }}

    #topbar {{
        background: {header_bg};
        border-bottom: 1px solid {border};
    }}

    #header {{
        font-size: 21px;
        font-weight: 900;
        color: {text};
    }}

    #topSearch {{
        background: {input_bg};
        border: 1px solid {border};
        border-radius: 11px;
        padding: 7px 12px;
        min-width: 320px;
        max-width: 520px;
    }}

    #topAction {{
        background: {panel_alt};
        border: 1px solid {border};
        border-radius: 10px;
        min-width: 38px;
        max-width: 38px;
        min-height: 36px;
        padding: 0;
    }}

    #topAction:hover {{
        background: #eaf1ff;
        border-color: #2563eb;
    }}

    #notificationButton {{
        background: #fff7ed;
        color: #c2410c;
        border: 1px solid #fed7aa;
        border-radius: 10px;
        min-height: 36px;
        padding: 5px 12px;
        font-weight: 900;
    }}

    #notificationButton:hover {{
        background: #ffedd5;
        border-color: #fb923c;
    }}

    #userPill {{
        background: {panel_alt};
        border: 1px solid {border};
        border-radius: 13px;
        padding: 6px 12px;
        font-weight: 800;
    }}

    #pageTitle {{
        font-size: 29px;
        font-weight: 900;
        color: {text};
        padding: 5px 0 2px 0;
    }}

    #subTitle {{
        color: {muted};
        font-size: 15px;
        padding-bottom: 10px;
    }}

    #metricCard {{
        background: {panel};
        border: 1px solid {border};
        border-radius: 16px;
        padding: 17px;
        min-height: 98px;
    }}

    #metricCard:hover {{
        border: 1px solid #3b82f6;
        background: {panel_hover};
    }}

    #metricIcon {{
        color: #2563eb;
        font-size: 28px;
        font-weight: 900;
    }}

    #metricTitle {{
        color: {muted};
        font-weight: 750;
        font-size: 13px;
    }}

    #metricValue {{
        font-size: 28px;
        font-weight: 900;
        color: {text};
    }}

    #chartPanel {{
        background: {panel};
        border: 1px solid {border};
        border-radius: 16px;
        padding: 15px;
        min-height: 230px;
    }}

    QLabel {{
        color: {text};
    }}

    QGroupBox {{
        background: {panel};
        border: 1px solid {border};
        border-radius: 14px;
        margin-top: 14px;
        padding: 14px;
        font-weight: 900;
    }}

    QGroupBox::title {{
        subcontrol-origin: margin;
        left: 14px;
        padding: 0 8px;
        color: #2563eb;
    }}

    QPushButton {{
        padding: 8px 15px;
        border: 1px solid {border};
        border-radius: 10px;
        background: {panel};
        color: {text};
        min-height: 35px;
        min-width: 108px;
        font-size: 14px;
        font-weight: 750;
    }}

    QPushButton:hover {{
        background: #edf4ff;
        border-color: #3b82f6;
    }}

    QPushButton:pressed {{
        background: #2563eb;
        color: #ffffff;
    }}

    #primaryButton {{
        background: #2563eb;
        color: #ffffff;
        border: 1px solid #1d4ed8;
    }}

    #primaryButton:hover {{
        background: #1d4ed8;
        color: #ffffff;
    }}

    QLineEdit, QTextEdit, QPlainTextEdit, QComboBox, QDateEdit, QSpinBox, QDoubleSpinBox {{
        padding: 8px 10px;
        border: 1px solid {border};
        border-radius: 10px;
        background: {input_bg};
        color: {text};
        min-height: 30px;
        selection-background-color: #2563eb;
    }}

    QLineEdit:focus, QTextEdit:focus, QComboBox:focus, QDateEdit:focus {{
        border: 1px solid #3b82f6;
    }}

    QTableWidget {{
        background: {panel};
        color: {text};
        border: 1px solid {border};
        border-radius: 13px;
        gridline-color: {border};
        selection-background-color: #2563eb;
        selection-color: #ffffff;
        alternate-background-color: {table_alt};
    }}

    QHeaderView::section {{
        background: {panel_alt};
        color: {text};
        padding: 10px;
        font-weight: 900;
        border: none;
        border-right: 1px solid {border};
        border-bottom: 1px solid {border};
    }}

    QTabWidget::pane {{
        border: 1px solid {border};
        border-radius: 14px;
        background: {panel};
        top: -1px;
    }}

    QTabBar::tab {{
        padding: 10px 16px;
        margin: 3px;
        border-radius: 9px;
        background: {panel_alt};
        color: {muted};
        font-weight: 800;
    }}

    QTabBar::tab:selected {{
        background: #2563eb;
        color: #ffffff;
    }}

    QScrollBar:vertical {{
        width: 13px;
        background: transparent;
        margin: 3px;
    }}

    QScrollBar::handle:vertical {{
        background: #8eabd0;
        min-height: 60px;
        border-radius: 6px;
    }}

    QScrollBar::handle:vertical:hover {{
        background: #2563eb;
    }}

    QScrollBar:horizontal {{
        height: 13px;
        background: transparent;
        margin: 3px;
    }}

    QScrollBar::handle:horizontal {{
        background: #8eabd0;
        min-width: 60px;
        border-radius: 6px;
    }}

    QScrollBar::handle:horizontal:hover {{
        background: #2563eb;
    }}

    QSplitter::handle {{
        background: {border};
    }}

    QCalendarWidget QWidget {{
        alternate-background-color: {panel_alt};
    }}
    """



def dashboard_kennzahlen() -> dict[str, float]:
    """Berechnet Dashboardwerte über Spaltennamen statt feste Indexpositionen."""
    wohnungen = DATA.get("Wohnungen", [])
    mieter = DATA.get("Mieter", [])
    zahlungen = DATA.get("Zahlungen", [])
    rechnungen = DATA.get("Rechnungen", [])

    gesamt_wohnungen = len(wohnungen)
    vermietet = 0
    frei = 0

    for row in wohnungen:
        status = feldwert("Wohnungen", row, ["Status"]).lower().strip()
        if any(w in status for w in ["vermietet", "aktiv", "belegt"]):
            vermietet += 1
        elif any(w in status for w in ["frei", "leer", "unvermietet", "verfügbar"]):
            frei += 1

    monatsmiete = 0.0
    for row in mieter:
        # Bevorzugt Gesamtmiete, ersatzweise Kaltmiete.
        mietwert = feldwert("Mieter", row, ["Miete"])
        if not mietwert:
            mietwert = feldwert("Mieter", row, ["Kaltmiete"])
        monatsmiete += to_float(mietwert)

    offene_bk = 0.0
    offene_rechnungen = 0.0

    # Offene Zahlungsposten anhand der vorhandenen Spaltennamen auswerten.
    for row in zahlungen:
        betrag = to_float(feldwert("Zahlungen", row, ["Betrag"]))
        status = feldwert("Zahlungen", row, ["Status"]).lower()
        kostenart = feldwert("Zahlungen", row, ["Kostenart"]).lower()

        if "offen" not in status and "fällig" not in status and "faellig" not in status:
            continue

        if any(w in kostenart for w in ["betrieb", "bk", "abwasser", "nebenkosten", "heizkosten"]):
            offene_bk += abs(betrag)
        else:
            offene_rechnungen += abs(betrag)

    # Offene Datensätze aus der eigenen Rechnungstabelle ergänzen.
    for row in rechnungen:
        status = feldwert("Rechnungen", row, ["Status"]).lower()
        if "offen" not in status and "fällig" not in status and "faellig" not in status:
            continue

        betrag_text = feldwert(
            "Rechnungen",
            row,
            ["Brutto", "Betrag brutto", "Betrag", "Netto"],
        )
        offene_rechnungen += abs(to_float(betrag_text))

    quote = (vermietet / gesamt_wohnungen * 100) if gesamt_wohnungen else 0.0

    return {
        "objekte": float(len(DATA.get("Objekte", []))),
        "wohnungen": float(gesamt_wohnungen),
        "mieter": float(len(mieter)),
        "leerstand": float(frei),
        "vermietungsquote": quote,
        "monatsmiete": monatsmiete,
        "jahresmiete": monatsmiete * 12,
        "offene_bk": offene_bk,
        "offene_rechnungen": offene_rechnungen,
        "vermietet": float(vermietet),
        "frei": float(frei),
    }

def euro(value: float) -> str:
    return f"{value:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")

class PdfUploadFeld(QWidget):
    def __init__(self, wert=""):
        super().__init__(); lay=QHBoxLayout(self); lay.setContentsMargins(0,0,0,0)
        self.input=QLineEdit(str(wert)); choose=QPushButton("PDF auswählen"); btn_oeffnen=QPushButton("Öffnen")
        choose.clicked.connect(self.waehlen); btn_oeffnen.clicked.connect(self.oeffnen)
        lay.addWidget(self.input); lay.addWidget(choose); lay.addWidget(btn_oeffnen)
    def waehlen(self):
        p,_=QFileDialog.getOpenFileName(self,"PDF auswählen","","PDF-Dateien (*.pdf)")
        if p: self.input.setText(kopiere_pdf_ins_projekt(p))
    def oeffnen(self):
        if self.input.text().strip(): system_datei_oeffnen(self.input.text().strip())
    def text(self): return self.input.text()


class OrdnerAuswahlFeld(QWidget):
    def __init__(self, wert=""):
        super().__init__()
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        self.input = QLineEdit(str(wert))
        choose = QPushButton("Ordner auswählen")
        open_btn = QPushButton("Öffnen")
        choose.clicked.connect(self.waehlen)
        open_btn.clicked.connect(self.oeffnen)
        lay.addWidget(self.input)
        lay.addWidget(choose)
        lay.addWidget(open_btn)

    def waehlen(self):
        path = QFileDialog.getExistingDirectory(self, "Ordner auswählen")
        if path:
            self.input.setText(path)

    def oeffnen(self):
        path = self.input.text().strip()
        if path:
            system_datei_oeffnen(path)
        else:
            QMessageBox.information(self, "Ordner", "Es ist kein Ordner hinterlegt.")

    def text(self):
        return self.input.text()


def flex_wert(row: list[str], felder: list[str], namen: list[str]) -> str:
    for name in namen:
        for i, feld in enumerate(felder):
            if norm_key(feld) == norm_key(name) and i < len(row):
                wert = str(row[i] or "").strip()
                if wert:
                    return wert
    return ""



class EingabeDialog(QDialog):
    def __init__(self, titel, felder, werte=None, bereich=""):
        super().__init__()
        self.bereich = str(bereich or "").strip()
        self.setWindowTitle(titel)
        self.resize(960, 760)

        # Die Anzeige ist gruppiert, aber gespeichert wird exakt nach Original-Spaltenindex.
        self.inputs_by_index: dict[int, QWidget] = {}
        self.felder = list(felder)
        self.felder_count = len(felder)

        root = QVBoxLayout(self)
        root.setContentsMargins(18, 18, 18, 18)

        header = QLabel(titel)
        header.setObjectName("pageTitle")
        root.addWidget(header)

        info = QLabel("Eingabemaske mit Scrollbereich – alle Felder bleiben erreichbar.")
        info.setObjectName("subTitle")
        root.addWidget(info)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(4, 4, 12, 4)
        content_layout.setSpacing(14)

        werte = list(werte or [""] * len(felder))
        while len(werte) < len(felder):
            werte.append("")

        gruppen = self._gruppen_bilden(felder, werte)

        for gruppenname, eintraege in gruppen:
            box = QGroupBox(gruppenname)
            form = QFormLayout(box)
            form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
            form.setFormAlignment(Qt.AlignmentFlag.AlignTop)
            form.setHorizontalSpacing(18)
            form.setVerticalSpacing(10)

            for original_index, feld, wert in eintraege:
                widget = self._widget_fuer_feld(feld, wert)
                self.inputs_by_index[original_index] = widget
                form.addRow(feld + ":", widget)

            content_layout.addWidget(box)

        content_layout.addStretch()
        scroll.setWidget(content)
        root.addWidget(scroll, 1)

        buttons = QHBoxLayout()
        buttons.addStretch()

        save = QPushButton("Speichern")
        save.setObjectName("primaryButton")
        cancel = QPushButton("Abbrechen")

        save.clicked.connect(self._validate_and_accept)
        cancel.clicked.connect(self.reject)

        buttons.addWidget(save)
        buttons.addWidget(cancel)
        root.addLayout(buttons)

    @staticmethod
    def _gruppen_bilden(felder, werte):
        allgemein = []
        vertrag = []
        finanzen = []
        dokumente = []
        ordner = []
        notizen = []

        for index, (feld, wert) in enumerate(zip(felder, werte)):
            low = str(feld).lower()
            eintrag = (index, feld, wert)

            if "pdf" in low or "foto" in low or "dokument" in low or "unterlagen" in low:
                dokumente.append(eintrag)
            elif "ordner" in low or "pfad" in low:
                ordner.append(eintrag)
            elif any(x in low for x in ["kundennummer", "vertragsnummer", "vertrag", "versorger", "sepa", "beginn", "ende", "kündigungsfrist"]):
                vertrag.append(eintrag)
            elif any(x in low for x in ["betrag", "miete", "kosten", "netto", "mwst", "brutto", "abschlag", "verbrauch", "jahr", "fällig", "zahlung", "guthaben", "nachzahlung", "kaution"]):
                finanzen.append(eintrag)
            elif any(x in low for x in ["notiz", "bemerkung", "beschreibung", "tätigkeit", "letzte aktion", "auskunft"]):
                notizen.append(eintrag)
            else:
                allgemein.append(eintrag)

        result = []
        if allgemein:
            result.append(("Allgemein", allgemein))
        if vertrag:
            result.append(("Vertrag / Versorger", vertrag))
        if finanzen:
            result.append(("Finanzen / Fristen", finanzen))
        if dokumente:
            result.append(("Dokumente / PDF", dokumente))
        if ordner:
            result.append(("Ordner / Ablage", ordner))
        if notizen:
            result.append(("Notizen", notizen))

        return result

    @staticmethod
    def _widget_fuer_feld(feld, wert):
        low = str(feld).lower()

        if "ordner" in low:
            return OrdnerAuswahlFeld(wert)

        if "pdf" in low or "foto" in low:
            return PdfUploadFeld(wert)

        if any(x in low for x in ["notiz", "bemerkung", "beschreibung", "verwendungszweck", "tätigkeit", "letzte aktion", "unterlagen / register", "auskunft"]):
            widget = QTextEdit(str(wert))
            widget.setMinimumHeight(95)
            return widget

        if low in [
            "status", "mahnstufe", "bereich", "gewerk", "typ", "bk art",
            "ordnerfarbe", "ordnerart", "sepa", "versorger", "priorität",
            "art", "schlüsseltyp","buchhaltung","eigentumsart"
        ]:
            widget = QComboBox()
            widget.addItems([
                "",
                "Ja", "Nein", "Eigenbestand", "Fremdverwaltung", "Eigene Buchhaltung", "Eigentümerabrechnung",
                "Aktiv", "Offen", "Neu", "In Bearbeitung", "Wartet",
                "Geprüft", "Freigegeben", "Bezahlt", "Erledigt", "Abgeschlossen",
                "Vermietet", "Frei", "Unklar", "Aktiv", "Gekündigt", "Ausgezogen", "Interessent",
                "Niedrig", "Normal", "Hoch", "Kritisch", "Dringend",
                "Vorbereitet", "Importiert", "Abgelegt", "Entwurf", "Angelegt",
                "Stadtwerke", "Wasserverband", "Abwasserzweckverband", "Müllentsorgung",
                "Gasversorger", "Stromversorger", "Hausmeister", "Versicherung",
                "Grundsteuer", "Schornsteinfeger", "Rauchwarnmelder", "Heizungswartung",
                "Legionellenprüfung", "Energieausweis",
                "Wasser", "Heizung", "Strom", "Gas", "Müll", "Sonstiges",
                "Fällig", "In Prüfung", "Gekündigt", "Verlängert",
                "Erstellt", "Übergeben", "Einzug", "Auszug",
                "Miete", "Nebenkosten", "Betriebskosten", "Abwasser",
                "Brand- und Arbeitsschutz", "Reparatur", "Instandhaltung",
                "Haustür", "Wohnungstür", "Briefkasten", "Keller", "Garage", "Zählerraum",
                "1", "2", "3",
            ])
            widget.setCurrentText(str(wert))
            return widget

        return QLineEdit(str(wert))


    @staticmethod
    def _widget_wert(widget: QWidget | None) -> str:
        if widget is None:
            return ""
        if isinstance(widget, QTextEdit):
            return widget.toPlainText().strip()
        if isinstance(widget, QComboBox):
            return widget.currentText().strip()
        if hasattr(widget, "text"):
            return str(widget.text()).strip()
        return ""

    @staticmethod
    def _widget_setzen(widget: QWidget | None, value: str) -> None:
        if widget is None:
            return
        if isinstance(widget, QTextEdit):
            widget.setPlainText(value)
        elif isinstance(widget, QComboBox):
            widget.setCurrentText(value)
        elif hasattr(widget, "setText"):
            widget.setText(value)

    def _validate_and_accept(self) -> None:
        """Erzwingt für objektbezogene Masken einen gültigen Objektordner."""
        object_folder_indexes = [
            index for index, field in enumerate(self.felder)
            if norm_key(field) in {"objektordner", "objekt-ordner", "objektordner-id"}
        ]
        object_indexes = [
            index for index, field in enumerate(self.felder)
            if norm_key(field) in {"objekt", "objektname", "objekt / adresse"}
        ]

        # Masken ohne Objektbezug bleiben unverändert.
        if not object_folder_indexes and not object_indexes:
            self.accept()
            return

        # In den Objekt-Stammdaten darf ein neuer Objektordner entstehen.
        # Andere Bereiche verlangen weiterhin einen bereits bekannten Ordner.
        if self.bereich == "Objekte":
            objektname = ""
            for index in object_indexes:
                objektname = self._widget_wert(self.inputs_by_index.get(index))
                if objektname:
                    break

            objektordner = ""
            for index in object_folder_indexes:
                objektordner = self._widget_wert(self.inputs_by_index.get(index))
                if objektordner:
                    break

            neuer_bezug = objektordner or objektname
            if not objektordner_wert_plausibel(neuer_bezug):
                QMessageBox.warning(
                    self,
                    "Objektordner erforderlich",
                    "Bitte einen gültigen Objektnamen oder Objektordner eingeben.",
                )
                return

            # Ausschließlich bestehende Felder befüllen; keine Excel-Strukturänderung.
            for index in object_folder_indexes:
                if not self._widget_wert(self.inputs_by_index.get(index)):
                    self._widget_setzen(self.inputs_by_index.get(index), neuer_bezug)

            self.accept()
            return

        candidates: list[str] = []
        for index in object_folder_indexes + object_indexes:
            value = self._widget_wert(self.inputs_by_index.get(index))
            if value:
                candidates.append(value)

        valid = ""
        for candidate in candidates:
            valid = normalisiere_gueltigen_objektordner(candidate)
            if valid:
                break

        if not valid:
            QMessageBox.warning(
                self,
                "Objektordner erforderlich",
                "Dieser Datensatz muss einem vorhandenen, eindeutigen Objektordner "
                "zugeordnet werden. Bitte im Feld Objekt oder Objektordner einen "
                "gültigen Objektordner auswählen.",
            )
            return

        # Alle vorhandenen Objektordner-Spalten konsistent befüllen.
        for index in object_folder_indexes:
            self._widget_setzen(self.inputs_by_index.get(index), valid)

        # Ein leeres Objektfeld ebenfalls mit dem eindeutigen Bezug versehen.
        for index in object_indexes:
            widget = self.inputs_by_index.get(index)
            if not self._widget_wert(widget):
                self._widget_setzen(widget, valid)

        self.accept()

    def values(self):
        out = []

        for index in range(self.felder_count):
            widget = self.inputs_by_index.get(index)

            if widget is None:
                out.append("")
            elif isinstance(widget, QTextEdit):
                out.append(widget.toPlainText())
            elif isinstance(widget, QComboBox):
                out.append(widget.currentText())
            elif hasattr(widget, "text"):
                out.append(widget.text())
            else:
                out.append("")

        return out





class EinstellungenDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Einstellungen")
        self.resize(920, 640)

        layout = QVBoxLayout(self)
        title = QLabel("Einstellungen")
        title.setObjectName("pageTitle")
        layout.addWidget(title)

        tabs = QTabWidget()
        layout.addWidget(tabs, 1)

        tab_pfade = QWidget()
        form = QFormLayout(tab_pfade)

        self.basis = QLineEdit(str(CONFIG.get("basis_pfad", "")))
        self.daten = QLineEdit(str(CONFIG.get("daten_pfad", "daten")))
        self.dokumente = QLineEdit(str(CONFIG.get("dokumente_pfad", "dokumente/pdf")))
        self.exports = QLineEdit(str(CONFIG.get("exports_pfad", "exports")))
        self.akten = QLineEdit(str(CONFIG.get("akten_pfad", "akten")))
        self.backup = QLineEdit(str(CONFIG.get("backup_pfad", "backups")))
        self.feedback = QLineEdit(str(CONFIG.get("feedback_pfad", "feedback")))

        form.addRow("Basis-Speicherort:", self._ordnerzeile(self.basis))
        form.addRow("Datenordner:", self._ordnerzeile(self.daten))
        form.addRow("Dokumente:", self._ordnerzeile(self.dokumente))
        form.addRow("Exports:", self._ordnerzeile(self.exports))
        form.addRow("Akten:", self._ordnerzeile(self.akten))
        form.addRow("Backups:", self._ordnerzeile(self.backup))
        form.addRow("Feedback:", self._ordnerzeile(self.feedback))
        tabs.addTab(tab_pfade, "Speicherorte")

        tab_darstellung = QWidget()
        form2 = QFormLayout(tab_darstellung)

        self.theme = QComboBox()
        self.theme.addItems(["hell", "dunkel"])
        self.theme.setCurrentText(str(CONFIG.get("theme", "hell")))

        self.start_vollbild = QComboBox()
        self.start_vollbild.addItems(["ja", "nein"])
        self.start_vollbild.setCurrentText(str(CONFIG.get("start_vollbild", "ja")))

        self.zeilenhoehe = QComboBox()
        self.zeilenhoehe.addItems(["30", "34", "38", "42", "48"])
        self.zeilenhoehe.setCurrentText(str(CONFIG.get("tabellen_zeilenhoehe", "34")))

        form2.addRow("Darstellung:", self.theme)
        form2.addRow("Start im Vollbild:", self.start_vollbild)
        form2.addRow("Tabellen-Zeilenhöhe:", self.zeilenhoehe)
        tabs.addTab(tab_darstellung, "Darstellung")

        tab_sicherheit = QWidget()
        form3 = QFormLayout(tab_sicherheit)

        self.auto_backup = QComboBox()
        self.auto_backup.addItems(["nein", "ja"])
        self.auto_backup.setCurrentText(str(CONFIG.get("auto_backup_start", "nein")))

        backup_btn = QPushButton("Backup jetzt erstellen")
        backup_btn.setObjectName("primaryButton")
        backup_btn.clicked.connect(self.backup_jetzt)

        config_btn = QPushButton("Konfigurationsordner öffnen")
        config_btn.clicked.connect(lambda: system_datei_oeffnen(str(CONFIG_FILE.parent)))

        form3.addRow("Backup beim Programmstart:", self.auto_backup)
        form3.addRow("Manuelles Backup:", backup_btn)
        form3.addRow("Konfiguration:", config_btn)
        tabs.addTab(tab_sicherheit, "Sicherheit")

        tab_firma = QWidget()
        firma_form = QFormLayout(tab_firma)
        self.firma_felder: dict[str, QLineEdit | QTextEdit] = {}
        firma_definitionen = [
            ("firma_name", "Firmenname"),
            ("firma_untertitel", "Untertitel"),
            ("firma_leistung", "Leistungszeile"),
            ("firma_slogan", "Slogan"),
            ("firma_inhaber", "Inhaber / Geschäftsführung"),
            ("firma_strasse", "Straße"),
            ("firma_plz_ort", "PLZ / Ort"),
            ("firma_telefon", "Telefon"),
            ("firma_email", "E-Mail"),
            ("firma_web", "Webseite"),
            ("firma_bank", "Bank"),
            ("firma_iban", "IBAN"),
            ("firma_bic", "BIC"),
            ("firma_steuer", "Steuernummer / USt-ID"),
        ]
        for key, label in firma_definitionen:
            edit = QLineEdit(str(CONFIG.get(key, "")))
            self.firma_felder[key] = edit
            firma_form.addRow(label + ":", edit)

        self.firma_logo = QLineEdit(str(CONFIG.get("firma_logo", "")))
        logo_row = QWidget()
        logo_lay = QHBoxLayout(logo_row)
        logo_lay.setContentsMargins(0, 0, 0, 0)
        logo_btn = QPushButton("Logo wählen")
        logo_btn.clicked.connect(self.waehle_firmenlogo)
        logo_lay.addWidget(self.firma_logo, 1)
        logo_lay.addWidget(logo_btn)
        firma_form.addRow("Logo:", logo_row)

        self.firma_gruss = QTextEdit(str(CONFIG.get("firma_gruss", "")))
        self.firma_gruss.setMinimumHeight(75)
        firma_form.addRow("Grußformel:", self.firma_gruss)

        self.bk_hinweis = QTextEdit(str(CONFIG.get("bk_hinweis", "")))
        self.bk_hinweis.setMinimumHeight(90)
        firma_form.addRow("BK-Hinweistext:", self.bk_hinweis)
        tabs.addTab(tab_firma, "Firmenlayout")

        buttons = QHBoxLayout()
        buttons.addStretch()
        save = QPushButton("Speichern")
        save.setObjectName("primaryButton")
        cancel = QPushButton("Abbrechen")
        save.clicked.connect(self.speichern)
        cancel.clicked.connect(self.reject)
        buttons.addWidget(save)
        buttons.addWidget(cancel)
        layout.addLayout(buttons)

    def _ordnerzeile(self, line_edit: QLineEdit) -> QWidget:
        row = QWidget()
        layout = QHBoxLayout(row)
        layout.setContentsMargins(0, 0, 0, 0)

        choose = QPushButton("Wählen")
        choose.setMinimumWidth(90)
        choose.clicked.connect(lambda: self.waehle_ordner(line_edit))

        open_btn = QPushButton("Öffnen")
        open_btn.setMinimumWidth(90)
        open_btn.clicked.connect(lambda: self.oeffne_ordner(line_edit))

        layout.addWidget(line_edit, 1)
        layout.addWidget(choose)
        layout.addWidget(open_btn)
        return row

    def waehle_ordner(self, line_edit: QLineEdit) -> None:
        start = line_edit.text().strip()
        path = QFileDialog.getExistingDirectory(self, "Ordner wählen", start)
        if path:
            line_edit.setText(path)

    def oeffne_ordner(self, line_edit: QLineEdit) -> None:
        value = line_edit.text().strip()
        if not value:
            QMessageBox.information(self, "Ordner", "Es ist kein Ordner hinterlegt.")
            return

        path = Path(value)
        if not path.is_absolute():
            basis = self.basis.text().strip()
            path = Path(basis) / value if basis else BASE_DIR / value

        if not path.exists():
            QMessageBox.warning(self, "Ordner", "Ordner wurde nicht gefunden:\n" + str(path))
            return

        system_datei_oeffnen(str(path))

    def backup_jetzt(self) -> None:
        ziel = erstelle_schnellbackup()
        if ziel is not None:
            QMessageBox.information(self, "Backup", "Backup erstellt:\n" + str(ziel))

    def waehle_firmenlogo(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Firmenlogo auswählen",
            self.firma_logo.text().strip(),
            "Bilder (*.png *.jpg *.jpeg *.bmp)",
        )
        if path:
            self.firma_logo.setText(path)

    def speichern(self):
        CONFIG["basis_pfad"] = self.basis.text().strip()
        CONFIG["daten_pfad"] = self.daten.text().strip() or "daten"
        CONFIG["dokumente_pfad"] = self.dokumente.text().strip() or "dokumente/pdf"
        CONFIG["exports_pfad"] = self.exports.text().strip() or "exports"
        CONFIG["akten_pfad"] = self.akten.text().strip() or "akten"
        CONFIG["backup_pfad"] = self.backup.text().strip() or "backups"
        CONFIG["feedback_pfad"] = self.feedback.text().strip() or "feedback"
        CONFIG["theme"] = self.theme.currentText()
        CONFIG["start_vollbild"] = self.start_vollbild.currentText()
        CONFIG["auto_backup_start"] = self.auto_backup.currentText()
        CONFIG["tabellen_zeilenhoehe"] = self.zeilenhoehe.currentText()
        for key, widget in self.firma_felder.items():
            CONFIG[key] = widget.text().strip()
        CONFIG["firma_logo"] = self.firma_logo.text().strip()
        CONFIG["firma_gruss"] = self.firma_gruss.toPlainText().strip()
        CONFIG["bk_hinweis"] = self.bk_hinweis.toPlainText().strip()

        CONFIG_FILE.write_text(json.dumps(CONFIG, indent=4, ensure_ascii=False), encoding="utf-8")
        QMessageBox.information(self, "Gespeichert", "Einstellungen gespeichert. Bitte Programm neu starten.")
        self.accept()






def normalize_relation_value(value: Any) -> str:
    """Normalisiert Werte für das interne Beziehungsmodell."""
    return norm_key(value)


def beziehungsfelder() -> list[str]:
    """Ermittelt Tabellenköpfe, die in mehreren Tabellen vorkommen."""
    counts: dict[str, int] = {}
    original: dict[str, str] = {}

    for felder in SCHEMA.values():
        for feld in felder:
            key = normalize_relation_value(feld)
            if not key:
                continue
            counts[key] = counts.get(key, 0) + 1
            original.setdefault(key, feld)

    return sorted(original[key] for key, count in counts.items() if count > 1)


def finde_beziehungen(feldname: str, suchwert: str) -> list[dict[str, str]]:
    """Findet alle Datensätze mit gleichem Tabellenkopf und gleichem Wert."""
    feld_key = normalize_relation_value(feldname)
    wert_key = normalize_relation_value(suchwert)
    results: list[dict[str, str]] = []

    if not feld_key or not wert_key:
        return results

    for bereich, rows in DATA.items():
        felder = SCHEMA.get(bereich, [])

        passende_spalten = [
            index
            for index, feld in enumerate(felder)
            if (
                normalize_relation_value(feld) == feld_key
                or feld_key in normalize_relation_value(feld)
                or normalize_relation_value(feld) in feld_key
            )
        ]

        if not passende_spalten:
            continue

        for row_index, row in enumerate(rows, start=1):
            for col_index in passende_spalten:
                if col_index < len(row) and normalize_relation_value(row[col_index]) == wert_key:
                    results.append({
                        "bereich": bereich,
                        "zeile": str(row_index),
                        "feld": felder[col_index],
                        "wert": str(row[col_index]),
                        "inhalt": " | ".join(str(value) for value in row),
                    })
                    break

    return results


def finde_freie_suche_beziehungen(suchwert: str) -> list[dict[str, str]]:
    """Findet Treffer über alle Tabellen und Spalten, wenn kein Beziehungsfeld gewählt wurde."""
    wert_key = normalize_relation_value(suchwert)
    results: list[dict[str, str]] = []

    if not wert_key:
        return results

    for bereich, rows in DATA.items():
        felder = SCHEMA.get(bereich, [])

        for row_index, row in enumerate(rows, start=1):
            for col_index, value in enumerate(row):
                if wert_key in normalize_relation_value(value):
                    feld = felder[col_index] if col_index < len(felder) else f"Spalte {col_index + 1}"
                    results.append({
                        "bereich": bereich,
                        "zeile": str(row_index),
                        "feld": feld,
                        "wert": str(value),
                        "inhalt": " | ".join(str(v) for v in row),
                    })
                    break

    return results


class BeziehungsAnalyseDialog(QDialog):
    """Analyse-Center: verbindet Tabellen über gleiche Tabellenköpfe."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Analyse-Center")
        self.resize(1200, 760)

        layout = QVBoxLayout(self)

        title = QLabel("Analyse-Center")
        title.setObjectName("pageTitle")
        layout.addWidget(title)

        info = QLabel("Tabellen werden im Hintergrund über gleiche Spaltenköpfe verbunden. Die Excel-Dateien bleiben unverändert.")
        info.setObjectName("metricTitle")
        layout.addWidget(info)

        search_row = QHBoxLayout()

        self.feld = QComboBox()
        self.feld.addItem("Freie Suche")
        self.feld.addItems(beziehungsfelder())
        self.feld.setMinimumWidth(260)

        self.suche = QLineEdit()
        self.suche.setPlaceholderText("Suchwert eingeben, z. B. Mietername, Objekt, Wohnung ...")
        self.suche.setMinimumWidth(360)

        btn = QPushButton("Analyse starten")
        btn.setObjectName("primaryButton")
        btn.clicked.connect(self.analysieren)

        search_row.addWidget(QLabel("Beziehungsfeld:"))
        search_row.addWidget(self.feld)
        search_row.addWidget(self.suche, 1)
        search_row.addWidget(btn)

        layout.addLayout(search_row)

        self.zusammenfassung = QLabel("Noch keine Analyse ausgeführt.")
        self.zusammenfassung.setObjectName("metricTitle")
        layout.addWidget(self.zusammenfassung)

        self.tabs = QTabWidget()
        layout.addWidget(self.tabs, 1)

        self.table_alle = QTableWidget()
        self.table_gruppiert = QTableWidget()

        self.tabs.addTab(self.table_alle, "Alle Treffer")
        self.tabs.addTab(self.table_gruppiert, "Gruppiert nach Bereich")

        buttons = QHBoxLayout()
        buttons.addStretch()

        export_btn = QPushButton("Analyse nach Excel exportieren")
        export_btn.clicked.connect(self.export_excel)

        close_btn = QPushButton("Schließen")
        close_btn.clicked.connect(self.accept)

        buttons.addWidget(export_btn)
        buttons.addWidget(close_btn)
        layout.addLayout(buttons)

        self.letzte_treffer: list[dict[str, str]] = []

    def analysieren(self) -> None:
        feld = self.feld.currentText()
        wert = self.suche.text().strip()

        if not wert:
            QMessageBox.information(self, "Analyse", "Bitte einen Suchwert eingeben.")
            return

        if feld == "Freie Suche":
            treffer = finde_freie_suche_beziehungen(wert)
        else:
            treffer = finde_beziehungen(feld, wert)

        self.letzte_treffer = treffer
        self.zusammenfassung.setText(
            f"Suchwert: {wert} | Beziehungsfeld: {feld} | Treffer: {len(treffer)}"
        )

        self._fuellen_alle(treffer)
        self._fuellen_gruppiert(treffer)

    def _fuellen_alle(self, treffer: list[dict[str, str]]) -> None:
        table = self.table_alle
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Bereich", "Zeile", "Feld", "Wert", "Inhalt"])
        table.setRowCount(len(treffer))
        table.setAlternatingRowColors(True)
        table.setWordWrap(False)
        table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)

        for row_index, item in enumerate(treffer):
            table.setItem(row_index, 0, QTableWidgetItem(item["bereich"]))
            table.setItem(row_index, 1, QTableWidgetItem(item["zeile"]))
            table.setItem(row_index, 2, QTableWidgetItem(item["feld"]))
            table.setItem(row_index, 3, QTableWidgetItem(item["wert"]))
            table.setItem(row_index, 4, QTableWidgetItem(item["inhalt"]))

        table.setColumnWidth(0, 200)
        table.setColumnWidth(1, 70)
        table.setColumnWidth(2, 180)
        table.setColumnWidth(3, 240)
        table.setColumnWidth(4, 680)

    def _fuellen_gruppiert(self, treffer: list[dict[str, str]]) -> None:
        gruppen: dict[str, int] = {}

        for item in treffer:
            bereich = item["bereich"]
            gruppen[bereich] = gruppen.get(bereich, 0) + 1

        table = self.table_gruppiert
        table.setColumnCount(2)
        table.setHorizontalHeaderLabels(["Bereich", "Treffer"])
        table.setRowCount(len(gruppen))
        table.setAlternatingRowColors(True)

        for row_index, bereich in enumerate(sorted(gruppen.keys())):
            table.setItem(row_index, 0, QTableWidgetItem(bereich))
            table.setItem(row_index, 1, QTableWidgetItem(str(gruppen[bereich])))

        table.setColumnWidth(0, 320)
        table.setColumnWidth(1, 120)

    def export_excel(self) -> None:
        if not self.letzte_treffer:
            QMessageBox.information(self, "Export", "Es gibt keine Treffer zum Exportieren.")
            return


        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        default_name = EXPORT_DIR / f"analyse_center_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

        ziel_text, _ = QFileDialog.getSaveFileName(
            self,
            "Analyse exportieren",
            str(default_name),
            "Excel-Dateien (*.xlsx)",
        )

        if not ziel_text:
            return

        ziel = Path(ziel_text)
        if ziel.suffix.lower() != ".xlsx":
            ziel = ziel.with_suffix(".xlsx")

        wb = Workbook()
        ws = wb.active

        if not isinstance(ws, Worksheet):
            QMessageBox.warning(self, "Export", "Excel-Arbeitsblatt konnte nicht erstellt werden.")
            return

        ws.title = "Analyse"
        ws.append(["Bereich", "Zeile", "Feld", "Wert", "Inhalt"])

        for item in self.letzte_treffer:
            ws.append([
                item["bereich"],
                item["zeile"],
                item["feld"],
                item["wert"],
                item["inhalt"],
            ])

        for col_index, _col in enumerate(ws.columns, start=1):
            ws.column_dimensions[get_column_letter(col_index)].width = 30

        wb.save(ziel)

        QMessageBox.information(self, "Export", f"Analyse exportiert:\n{ziel}")
        system_datei_oeffnen(str(ziel.parent))



def zeile_offen(row: list[str], felder: list[str]) -> bool:
    status = ""
    for i, feld in enumerate(felder):
        if norm_key(feld) == "status" and i < len(row):
            status = str(row[i]).lower()
            break
    if not status:
        return True
    return not any(w in status for w in ["erledigt", "bezahlt", "abgeschlossen", "archiviert"])


def dashboard_offene_liste(titel: str, limit: int = 8) -> list[list[str]]:
    felder = SCHEMA.get(titel, [])
    result: list[list[str]] = []
    for row in DATA.get(titel, []):
        if zeile_offen(row, felder):
            result.append(["" if v is None else str(v) for v in row])
    return result[:limit]


def dashboard_feld(row: list[str], felder: list[str], namen: list[str]) -> str:
    for name in namen:
        for i, feld in enumerate(felder):
            if norm_key(feld) == norm_key(name) and i < len(row):
                value = str(row[i] or "").strip()
                if value:
                    return value
    return ""

class DashboardSucheDialog(QDialog):
    """Suche nach Mietern, Vermietern und Objekten vom Dashboard aus."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Dashboard-Suche")
        self.resize(1150, 720)

        layout = QVBoxLayout(self)

        title = QLabel("Dashboard-Suche")
        title.setObjectName("pageTitle")
        layout.addWidget(title)

        info = QLabel("Suche nach Mieter, Vermieter oder Objekt. Alle passenden Daten werden gesammelt angezeigt.")
        info.setObjectName("metricTitle")
        layout.addWidget(info)

        row = QHBoxLayout()

        self.typ = QComboBox()
        self.typ.addItems(["Alle", "Mieter", "Vermieter", "Objekt"])
        self.typ.currentTextChanged.connect(self.suchen)

        self.search = QLineEdit()
        self.search.setPlaceholderText("Suchbegriff eingeben...")
        self.search.textChanged.connect(self.suchen)

        row.addWidget(QLabel("Suchen nach:"))
        row.addWidget(self.typ)
        row.addWidget(self.search, 1)
        layout.addLayout(row)

        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Bereich", "Typ", "Treffer", "Zeile", "Inhalt"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(0, 190)
        self.table.setColumnWidth(1, 120)
        self.table.setColumnWidth(2, 250)
        self.table.setColumnWidth(3, 70)
        self.table.setColumnWidth(4, 650)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.cellDoubleClicked.connect(self.oeffne_dokument_oder_ordner)
        layout.addWidget(self.table, 1)

        buttons = QHBoxLayout()
        buttons.addStretch()

        self.btn_oeffnen = QPushButton("PDF/Ordner öffnen")
        self.btn_oeffnen.clicked.connect(self.oeffne_dokument_oder_ordner)

        close_btn = QPushButton("Schließen")
        close_btn.clicked.connect(self.accept)

        buttons.addWidget(self.btn_oeffnen)
        buttons.addWidget(close_btn)
        layout.addLayout(buttons)

        self.results: list[tuple[str, str, str, int, list[str]]] = []

    @staticmethod
    def _spalten_fuer_typ(bereich: str, typ: str) -> list[int]:
        felder = SCHEMA.get(bereich, [])
        result: list[int] = []

        for index, feld in enumerate(felder):
            f = str(feld).lower()

            if typ == "Mieter" and "mieter" in f:
                result.append(index)
            elif typ == "Vermieter" and "vermieter" in f:
                result.append(index)
            elif typ == "Objekt" and ("objekt" in f or "wohnung" in f):
                result.append(index)

        return result

    def suchen(self) -> None:
        query = self.search.text().strip().lower()
        typ = self.typ.currentText()

        self.results = []

        if not query:
            self.table.setRowCount(0)
            return

        for bereich, rows in DATA.items():
            felder = SCHEMA.get(bereich, [])

            such_spalten: list[int] = []
            if typ != "Alle":
                such_spalten = self._spalten_fuer_typ(bereich, typ)

            for row_index, row in enumerate(rows, start=1):
                treffer_typ = ""
                treffer_text = ""

                if typ == "Alle":
                    pruefwerte = list(enumerate(row))
                else:
                    pruefwerte = [
                        (idx, row[idx])
                        for idx in such_spalten
                        if idx < len(row)
                    ]

                for col_index, value in pruefwerte:
                    value_text = str(value)
                    if query in value_text.lower():
                        feldname = felder[col_index] if col_index < len(felder) else f"Spalte {col_index + 1}"
                        treffer_typ = typ if typ != "Alle" else self._rate_typ(str(feldname))
                        treffer_text = f"{feldname}: {value_text}"
                        break

                if treffer_text:
                    self.results.append((
                        bereich,
                        treffer_typ,
                        treffer_text,
                        row_index,
                        [str(v) for v in row],
                    ))

        self.table.setRowCount(len(self.results))

        for row_idx, (bereich, treffer_typ, treffer, excel_row, row_values) in enumerate(self.results):
            self.table.setItem(row_idx, 0, QTableWidgetItem(bereich))
            self.table.setItem(row_idx, 1, QTableWidgetItem(treffer_typ))
            self.table.setItem(row_idx, 2, QTableWidgetItem(treffer))
            self.table.setItem(row_idx, 3, QTableWidgetItem(str(excel_row)))
            self.table.setItem(row_idx, 4, QTableWidgetItem(" | ".join(row_values)))

    @staticmethod
    def _rate_typ(feldname: str) -> str:
        f = feldname.lower()
        if "mieter" in f:
            return "Mieter"
        if "vermieter" in f:
            return "Vermieter"
        if "objekt" in f or "wohnung" in f:
            return "Objekt"
        return "Treffer"

    def oeffne_dokument_oder_ordner(self, *_args) -> None:
        row_index = self.table.currentRow()

        if row_index < 0 or row_index >= len(self.results):
            QMessageBox.information(self, "Hinweis", "Bitte zuerst einen Treffer auswählen.")
            return

        bereich, _typ, _treffer, _excel_row, row_values = self.results[row_index]
        felder = SCHEMA.get(bereich, [])

        for col_index, feld in enumerate(felder):
            if col_index >= len(row_values):
                continue

            feldname = str(feld).lower()
            value = row_values[col_index].strip()

            if value and ("pdf" in feldname or "ordner" in feldname or "pfad" in feldname):
                system_datei_oeffnen(value)
                return

        QMessageBox.information(self, "Kein Dokument", "Für diesen Treffer wurde keine PDF oder kein Ordner gefunden.")



def feldwert(titel: str, row: list[Any], namen: list[str]) -> str:
    felder = SCHEMA.get(titel, [])
    for name in namen:
        key = norm_key(name)
        for index, feld in enumerate(felder):
            if norm_key(feld) == key and index < len(row):
                value = str(row[index] or "").strip()
                if value:
                    return value
    return ""



def objektordner_rohliste() -> list[str]:
    """
    Liefert ausschließlich gültige Objektordner-Schlüssel aus den stabilen Stammdaten.
    Es werden keine Messwerte, Einheiten oder beliebige Tabellenfelder übernommen.
    """
    values: set[str] = set()

    # Objekte: Objektname und die bereits vorhandene Objektordner-Spalte.
    for row in DATA.get("Objekte", []):
        objektname = feldwert("Objekte", row, ["Objektname", "Objekt"])
        objektordner = feldwert("Objekte", row, ["Objektordner"])
        if objektname:
            values.add(objektname.strip())
        if objektordner:
            values.add(objektordner.strip())

    # Wohnungen: vorhandener Objektordner oder Objekt.
    for row in DATA.get("Wohnungen", []):
        value = feldwert("Wohnungen", row, ["Objektordner"])
        if not value:
            value = feldwert("Wohnungen", row, ["Objekt"])
        if value:
            values.add(value.strip())

    # Mieter: vorhandener Objektordner; nur eindeutiger Ort-Prefix als Altbestand-Fallback.
    for row in DATA.get("Mieter", []):
        value = feldwert("Mieter", row, ["Objektordner"])
        if value:
            values.add(value.strip())
            continue

        ort = feldwert("Mieter", row, ["Ort"])
        if "/" in ort:
            prefix = ort.split("/")[0].strip()
            if prefix:
                values.add(prefix)

    return sorted(v for v in values if objektordner_wert_plausibel(v))


def objektordner_wert_plausibel(value: Any) -> bool:
    """
    Verhindert, dass Einheiten/Messwerte wie m3, m², kWh oder Zahlen
    als Objektordner verwendet werden.
    """
    text_value = str(value or "").strip()

    if not text_value:
        return False

    key = norm_key(text_value)

    verbotene_werte = {
        "m3", "m²", "m2", "qm", "kwh", "kw", "liter", "l",
        "wasser", "strom", "gas", "heizung", "verbrauch",
        "monatlich", "jaehrlich", "jährlich",
        "ja", "nein", "aktiv", "offen", "bezahlt", "frei",
    }

    if key in {norm_key(v) for v in verbotene_werte}:
        return False

    # Reine Zahlen oder typische Messwerte nicht zulassen.
    if re.fullmatch(r"[-+]?\d+(?:[.,]\d+)?", text_value):
        return False

    if re.fullmatch(
        r"[-+]?\d+(?:[.,]\d+)?\s*(?:m3|m³|m2|m²|qm|kwh|kw|l|liter)",
        text_value,
        re.IGNORECASE,
    ):
        return False

    if len(text_value) < 3:
        return False

    return True


def normalisiere_gueltigen_objektordner(value: Any) -> str:
    """
    Gibt nur einen bekannten, gültigen Objektordner zurück.
    Abweichende Schreibweisen werden über norm_key abgeglichen.
    """
    raw = str(value or "").strip()

    if not objektordner_wert_plausibel(raw):
        return ""

    raw_key = norm_key(raw)

    for gueltig in objektordner_rohliste():
        if norm_key(gueltig) == raw_key:
            return gueltig

    return ""




def objektordner_fuer_datensatz(titel: str, row: list[Any]) -> str:
    """
    Zentrale und strenge Objektordner-Ermittlung.

    Priorität:
    1. vorhandenes Feld Objektordner
    2. eindeutiges Objekt/Objektname
    3. Mieterbeziehung
    4. Wohnungsbeziehung
    5. eindeutige Vermieterbeziehung

    Niemals verwendet werden Messwerte, Einheiten oder beliebige Fallback-Zellen.
    """
    # 1. Direkter Objektordner.
    direkt = feldwert(titel, row, ["Objektordner", "Objekt-Ordner", "Objektordner-ID"])
    valid = normalisiere_gueltigen_objektordner(direkt)
    if valid:
        return valid

    # 2. Direktes Objektfeld.
    objekt = feldwert(titel, row, ["Objekt", "Objektname", "Objekt / Adresse"])
    valid = normalisiere_gueltigen_objektordner(objekt)
    if valid:
        return valid

    # 3. Mieterbeziehung.
    mieter = feldwert(titel, row, ["Mieter", "Name/Quelle", "Empfänger"])
    if mieter:
        mieter_key = norm_key(mieter)

        for mrow in DATA.get("Mieter", []):
            mname = feldwert("Mieter", mrow, ["Mieter"])

            if not mname:
                continue

            if (
                norm_key(mname) == mieter_key
                or norm_key(mname) in mieter_key
                or mieter_key in norm_key(mname)
            ):
                m_objekt = feldwert("Mieter", mrow, ["Objektordner"])
                valid = normalisiere_gueltigen_objektordner(m_objekt)
                if valid:
                    return valid

                ort = feldwert("Mieter", mrow, ["Ort"])
                if "/" in ort:
                    valid = normalisiere_gueltigen_objektordner(
                        ort.split("/")[0].strip()
                    )
                    if valid:
                        return valid

    # 4. Wohnungsbeziehung.
    wohnung = feldwert(titel, row, ["Wohnung", "Wohnungsordner"])
    if wohnung:
        wohnung_key = norm_key(wohnung)
        passende_objekte: set[str] = set()

        for wrow in DATA.get("Wohnungen", []):
            wname = feldwert("Wohnungen", wrow, ["Wohnung", "Wohnungsordner"])

            if wname and norm_key(wname) == wohnung_key:
                kandidat = (
                    feldwert("Wohnungen", wrow, ["Objektordner"])
                    or feldwert("Wohnungen", wrow, ["Objekt"])
                )
                valid = normalisiere_gueltigen_objektordner(kandidat)
                if valid:
                    passende_objekte.add(valid)

        if len(passende_objekte) == 1:
            return next(iter(passende_objekte))

    # 5. Eindeutige Vermieterbeziehung.
    vermieter = feldwert(titel, row, ["Vermieter", "Eigentümer"])
    if vermieter:
        vermieter_key = norm_key(vermieter)
        passende_objekte: set[str] = set()

        for orow in DATA.get("Objekte", []):
            ov = feldwert("Objekte", orow, ["Vermieter", "Eigentümer"])

            if ov and norm_key(ov) == vermieter_key:
                oname = feldwert("Objekte", orow, ["Objektname", "Objekt"])
                valid = normalisiere_gueltigen_objektordner(oname)
                if valid:
                    passende_objekte.add(valid)

        if len(passende_objekte) == 1:
            return next(iter(passende_objekte))

    return ""



def objektordner_zuordnungen_normalisieren(titel: str) -> list[int]:
    """
    Schreibt einen eindeutig ermittelten Objektordner in bereits vorhandene
    Objekt-/Objektordner-Felder. Die Excelstruktur wird nicht verändert.
    Rückgabe: Zeilennummern, deren Bezug nicht eindeutig ermittelt werden konnte.
    """
    fields = SCHEMA.get(titel, [])
    object_folder_indexes = [
        index for index, field in enumerate(fields)
        if norm_key(field) in {"objektordner", "objekt-ordner", "objektordner-id"}
    ]
    object_indexes = [
        index for index, field in enumerate(fields)
        if norm_key(field) in {"objekt", "objektname", "objekt / adresse"}
    ]

    if not object_folder_indexes and not object_indexes:
        return []

    unresolved: list[int] = []
    for row_number, row in enumerate(DATA.get(titel, []), start=2):
        while len(row) < len(fields):
            row.append("")

        relation = objektordner_fuer_datensatz(titel, row)
        valid = normalisiere_gueltigen_objektordner(relation)
        if not valid:
            unresolved.append(row_number)
            continue

        for index in object_folder_indexes:
            row[index] = valid
        for index in object_indexes:
            if not str(row[index]).strip():
                row[index] = valid

    return unresolved



def alle_objektordner() -> list[str]:
    """Nur validierte Objektordner für Filter und Auswahllisten."""
    return objektordner_rohliste()


def objektordner_pruefbericht() -> list[dict[str, str]]:
    """Zeigt fehlerhafte oder nicht zuordenbare Datensätze, ohne Excel zu verändern."""
    result: list[dict[str, str]] = []

    for titel, rows in DATA.items():
        for row_index, row in enumerate(rows, start=1):
            relation = objektordner_fuer_datensatz(titel, row)

            direkter_wert = feldwert(
                titel,
                row,
                ["Objektordner", "Objekt-Ordner", "Objektordner-ID"],
            )

            if direkter_wert and not normalisiere_gueltigen_objektordner(direkter_wert):
                result.append({
                    "bereich": titel,
                    "zeile": str(row_index),
                    "problem": f"Ungültiger Objektordner: {direkter_wert}",
                    "objektordner": "",
                    "inhalt": " | ".join(str(v) for v in row),
                })
            elif not relation:
                # Nur Tabellen melden, bei denen Objektbezug sinnvoll ist.
                if titel in {
                    "Wohnungen", "Mieter", "Mietverträge", "Zahlungen",
                    "Betriebskosten", "Rechnungen", "Dokumente",
                    "Aufgaben", "Fristen", "Schäden", "Versorger",
                    "Übergabeprotokolle", "HV-Rechnungen",
                }:
                    result.append({
                        "bereich": titel,
                        "zeile": str(row_index),
                        "problem": "Nicht eindeutig zugeordnet",
                        "objektordner": "",
                        "inhalt": " | ".join(str(v) for v in row),
                    })

    return result


class ObjektordnerPruefungSeite(QWidget):
    """Prüfseite für Objektordner-Zuordnungen, rein lesend."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav

        root = QVBoxLayout(self)
        root.setContentsMargins(22, 20, 22, 20)
        root.setSpacing(14)

        title = QLabel("Objektordner-Prüfung")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel(
            "Prüft alle Zuordnungen. Ungültige Werte wie m3 werden nicht mehr als Objektordner verwendet. "
            "Die Excel-Dateien werden dabei nicht verändert."
        )
        info.setObjectName("subTitle")
        info.setWordWrap(True)
        root.addWidget(info)

        top = QHBoxLayout()

        refresh = QPushButton("Neu prüfen")
        refresh.setObjectName("primaryButton")
        refresh.clicked.connect(self.laden)

        export_btn = QPushButton("Prüfbericht exportieren")
        export_btn.clicked.connect(self.exportieren)

        top.addWidget(refresh)
        top.addWidget(export_btn)
        top.addStretch()
        root.addLayout(top)

        self.summary = QLabel("")
        self.summary.setObjectName("metricTitle")
        root.addWidget(self.summary)

        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(
            ["Bereich", "Zeile", "Problem", "Objektordner", "Inhalt"]
        )
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.cellDoubleClicked.connect(self._open_area)
        root.addWidget(self.table, 1)

        self.rows: list[dict[str, str]] = []
        self.laden()

    def laden(self) -> None:
        self.rows = objektordner_pruefbericht()
        self.summary.setText(
            f"Gültige Objektordner: {len(alle_objektordner())} | "
            f"Zu prüfende Datensätze: {len(self.rows)}"
        )

        self.table.setRowCount(len(self.rows))

        for row_index, item in enumerate(self.rows):
            values = [
                item.get("bereich", ""),
                item.get("zeile", ""),
                item.get("problem", ""),
                item.get("objektordner", ""),
                item.get("inhalt", ""),
            ]

            for col_index, value in enumerate(values):
                self.table.setItem(
                    row_index,
                    col_index,
                    QTableWidgetItem(str(value)),
                )

        for col_index, width in enumerate([180, 70, 280, 220, 760]):
            self.table.setColumnWidth(col_index, width)

    def _open_area(self, row: int, _column: int) -> None:
        item = self.table.item(row, 0)
        if item is not None and item.text().strip():
            self.nav(item.text().strip())

    def exportieren(self) -> None:
        if not self.rows:
            QMessageBox.information(
                self,
                "Objektordner-Prüfung",
                "Es wurden keine fehlerhaften Zuordnungen gefunden.",
            )
            return


        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        default = EXPORT_DIR / (
            "objektordner_pruefung_"
            + datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            + ".xlsx"
        )

        target, _ = QFileDialog.getSaveFileName(
            self,
            "Prüfbericht exportieren",
            str(default),
            "Excel-Dateien (*.xlsx)",
        )

        if not target:
            return

        target_path = Path(target)

        if target_path.suffix.lower() != ".xlsx":
            target_path = target_path.with_suffix(".xlsx")

        workbook = Workbook()
        sheet = workbook.active

        if not isinstance(sheet, Worksheet):
            QMessageBox.warning(
                self,
                "Export",
                "Arbeitsblatt konnte nicht erstellt werden.",
            )
            return

        sheet.title = "Objektordner-Prüfung"
        sheet.append(["Bereich", "Zeile", "Problem", "Objektordner", "Inhalt"])

        for item in self.rows:
            sheet.append([
                item.get("bereich", ""),
                item.get("zeile", ""),
                item.get("problem", ""),
                item.get("objektordner", ""),
                item.get("inhalt", ""),
            ])

        widths = [22, 10, 38, 30, 90]

        for col_index, width in enumerate(widths, start=1):
            sheet.column_dimensions[
                sheet.cell(1, col_index).column_letter
            ].width = width

        workbook.save(target_path)

        QMessageBox.information(
            self,
            "Objektordner-Prüfung",
            f"Prüfbericht exportiert:\n{target_path}",
        )



def globale_dashboard_suche(suchtext: str, objektordner: str = "") -> list[dict[str, str]]:
    query = str(suchtext or "").strip().lower()
    objekt_key = norm_key(objektordner)
    treffer: list[dict[str, str]] = []

    if not query and not objekt_key:
        return treffer

    for titel, rows in DATA.items():
        felder = SCHEMA.get(titel, [])

        for row_index, row in enumerate(rows, start=1):
            relation = objektordner_fuer_datensatz(titel, row)

            if objekt_key and norm_key(relation) != objekt_key:
                continue

            values = ["" if value is None else str(value) for value in row]
            combined = " ".join(values).lower()

            if query and query not in combined:
                continue

            gefundene_felder = []
            if query:
                for col_index, value in enumerate(values):
                    if query in value.lower():
                        feld = felder[col_index] if col_index < len(felder) else f"Spalte {col_index + 1}"
                        gefundene_felder.append(str(feld))

            treffer.append({
                "bereich": titel,
                "zeile": str(row_index),
                "objektordner": relation,
                "felder": ", ".join(gefundene_felder) if gefundene_felder else "Objektordner",
                "inhalt": " | ".join(values),
            })

    return treffer


class Dashboard(QWidget):
    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        root = QVBoxLayout(self); root.setContentsMargins(0, 0, 0, 0)
        scroll = QScrollArea(); scroll.setWidgetResizable(True); scroll.setFrameShape(QFrame.Shape.NoFrame)
        inner = QWidget(); layout = QVBoxLayout(inner); layout.setContentsMargins(28, 24, 28, 24); layout.setSpacing(16)
        top = QHBoxLayout(); title_block = QVBoxLayout()
        h = QLabel("Willkommen in Ihrer Immobilienverwaltung"); h.setObjectName("pageTitle")
        sub = QLabel("Übersicht & Analysen auf einen Blick"); sub.setObjectName("subTitle")
        title_block.addWidget(h); title_block.addWidget(sub)
        analyse_btn = QPushButton("Analyse-Center"); analyse_btn.setObjectName("primaryButton"); analyse_btn.setMinimumWidth(180); analyse_btn.clicked.connect(lambda: self.nav("Analyse-Center"))
        top.addLayout(title_block, 1); top.addWidget(analyse_btn); layout.addLayout(top)

        search_panel = QFrame()
        search_panel.setObjectName("chartPanel")
        search_layout = QVBoxLayout(search_panel)

        search_title = QLabel("Globale Suche über alle Tabellen")
        search_title.setObjectName("metricTitle")
        search_layout.addWidget(search_title)

        search_row = QHBoxLayout()

        self.dashboard_objekt_filter = QComboBox()
        self.dashboard_objekt_filter.setMinimumWidth(260)
        self.dashboard_objekt_filter.addItem("Alle Objektordner")
        self.dashboard_objekt_filter.addItems(alle_objektordner())
        self.dashboard_objekt_filter.currentTextChanged.connect(self.dashboard_suchen)

        self.dashboard_suchfeld = QLineEdit()
        self.dashboard_suchfeld.setPlaceholderText("Schlagwort aus allen Tabellen suchen ...")
        self.dashboard_suchfeld.setMinimumWidth(420)
        self.dashboard_suchfeld.textChanged.connect(self.dashboard_suchen)
        self.dashboard_suchfeld.returnPressed.connect(self.dashboard_suchen)

        search_button = QPushButton("Suchen")
        search_button.setObjectName("primaryButton")
        search_button.clicked.connect(self.dashboard_suchen)

        search_row.addWidget(QLabel("Objektordner:"))
        search_row.addWidget(self.dashboard_objekt_filter)
        search_row.addWidget(self.dashboard_suchfeld, 1)
        search_row.addWidget(search_button)
        search_layout.addLayout(search_row)

        self.dashboard_suchergebnis = QTableWidget()
        self.dashboard_suchergebnis.setColumnCount(5)
        self.dashboard_suchergebnis.setHorizontalHeaderLabels(
            ["Bereich", "Zeile", "Objektordner", "Trefferfeld", "Inhalt"]
        )
        self.dashboard_suchergebnis.setAlternatingRowColors(True)
        self.dashboard_suchergebnis.setWordWrap(False)
        self.dashboard_suchergebnis.setMinimumHeight(210)
        self.dashboard_suchergebnis.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.dashboard_suchergebnis.cellDoubleClicked.connect(self.dashboard_treffer_oeffnen)
        search_layout.addWidget(self.dashboard_suchergebnis)

        layout.addWidget(search_panel)

        k = dashboard_kennzahlen(); grid = QGridLayout(); grid.setSpacing(14)
        cards = [("🏢", "Objekte", str(int(k["objekte"]))), ("🏠", "Wohnungen", str(int(k["wohnungen"]))), ("👥", "Mieter", str(int(k["mieter"]))), ("⌂", "Leerstand", str(int(k["leerstand"]))), ("%", "Vermietungsquote", f'{k["vermietungsquote"]:.0f} %'), ("€", "Monatsmiete", euro(k["monatsmiete"])), ("📅", "Jahresmiete", euro(k["jahresmiete"])), ("📄", "Offene Betriebskosten", euro(k["offene_bk"])), ("🧾", "Offene Rechnungen", euro(k["offene_rechnungen"]))]
        for index, (icon, title, value) in enumerate(cards): grid.addWidget(self._metric_card(icon, title, value), index // 5, index % 5)
        layout.addLayout(grid)
        tabs = QTabWidget()
        tabs.addTab(self._dashboard_tab(k), "Dashboard"); tabs.addTab(self._dashboard_tabelle_tab("Objektanalyse", "Objekte"), "Objektanalyse"); tabs.addTab(self._dashboard_tabelle_tab("Vermieteranalyse", "Objekte"), "Vermieteranalyse"); tabs.addTab(self._dashboard_tabelle_tab("Mieteranalyse", "Mieter"), "Mieteranalyse"); tabs.addTab(self._dashboard_tabelle_tab("Finanzanalyse", "Zahlungen"), "Finanzanalyse"); tabs.addTab(self._dashboard_diagramm_tab(), "Diagramme"); tabs.addTab(self._dashboard_berichte_tab(), "Berichte"); tabs.addTab(self._dashboard_export_tab(), "Export")
        tabs.addTab(self._offene_posten_tab(), "Offene Vorgänge")
        layout.addWidget(tabs, 1); scroll.setWidget(inner); root.addWidget(scroll)

    def dashboard_suchen(self) -> None:
        query = self.dashboard_suchfeld.text().strip()
        selected = self.dashboard_objekt_filter.currentText().strip()
        objekt = "" if selected == "Alle Objektordner" else selected

        treffer = globale_dashboard_suche(query, objekt)
        self.dashboard_treffer = treffer
        self.dashboard_suchergebnis.setRowCount(len(treffer))

        for row_index, item in enumerate(treffer):
            values = [
                item.get("bereich", ""),
                item.get("zeile", ""),
                item.get("objektordner", ""),
                item.get("felder", ""),
                item.get("inhalt", ""),
            ]
            for col_index, value in enumerate(values):
                self.dashboard_suchergebnis.setItem(
                    row_index, col_index, QTableWidgetItem(str(value))
                )

        for col, width in enumerate([180, 70, 240, 220, 760]):
            self.dashboard_suchergebnis.setColumnWidth(col, width)

    def dashboard_treffer_oeffnen(self, row: int, _column: int) -> None:
        item = self.dashboard_suchergebnis.item(row, 0)
        if item is not None and item.text().strip():
            self.nav(item.text().strip())

    @staticmethod
    def _metric_card(icon: str, title: str, value: str) -> QFrame:
        card = QFrame(); card.setObjectName("metricCard"); lay = QHBoxLayout(card); icon_label = QLabel(icon); icon_label.setObjectName("metricIcon"); icon_label.setFixedWidth(55); texts = QVBoxLayout(); title_label = QLabel(title); title_label.setObjectName("metricTitle"); value_label = QLabel(value); value_label.setObjectName("metricValue"); texts.addWidget(title_label); texts.addWidget(value_label); lay.addWidget(icon_label); lay.addLayout(texts, 1); return card
    def _dashboard_tab(self, k: dict[str, float]) -> QWidget:
        page = QWidget(); grid = QGridLayout(page); grid.setSpacing(12); grid.addWidget(self._vermietet_chart(k), 0, 0); grid.addWidget(self._panel("Mieteinnahmen", "Monatsentwicklung aus Zahlungsdaten"), 0, 1); grid.addWidget(self._panel("Einnahmen / Ausgaben", "Auswertung nach Kostenart und Status"), 0, 2); grid.addWidget(self._panel("Betriebskosten pro Objekt", "Summen aus Zahlungen/Kostenarten"), 1, 0); grid.addWidget(self._quote_panel(k), 1, 1); grid.addWidget(self._offene_posten(k), 1, 2); return page
    @staticmethod
    def _panel(title: str, body: str) -> QFrame:
        panel = QFrame(); panel.setObjectName("chartPanel"); lay = QVBoxLayout(panel); label = QLabel(title); label.setObjectName("metricTitle"); body_label = QLabel(body); body_label.setAlignment(Qt.AlignmentFlag.AlignCenter); lay.addWidget(label); lay.addStretch(); lay.addWidget(body_label); lay.addStretch(); return panel
    def _vermietet_chart(self, k: dict[str, float]) -> QFrame:
        total = max(k["vermietet"] + k["frei"], 1); quote = k["vermietet"] / total * 100; body = f'{quote:.0f} %' + chr(10) + chr(10) + f'Vermietet: {int(k["vermietet"])}    Frei: {int(k["frei"])}'; return self._panel("Vermietet / Frei", body)
    def _quote_panel(self, k: dict[str, float]) -> QFrame:
        wohnungen = max(k["wohnungen"], 1); quote = k["leerstand"] / wohnungen * 100; return self._panel("Leerstandsquote", f"{quote:.0f} %")
    def _offene_posten(self, k: dict[str, float]) -> QFrame:
        body = f'Offene Betriebskosten: {euro(k["offene_bk"])}' + chr(10) + f'Offene Rechnungen: {euro(k["offene_rechnungen"])}'; return self._panel("Offene Posten", body)

    def _offene_posten_tab(self) -> QWidget:
        page = QWidget()
        layout = QGridLayout(page)
        layout.setSpacing(14)
        layout.addWidget(self._offene_tabelle_panel("Fristen", ["Titel", "Objekt", "Wohnung", "Fällig am", "Priorität", "Status"]), 0, 0)
        layout.addWidget(self._offene_tabelle_panel("Aufgaben", ["Aufgabe", "Objekt", "Wohnung", "Fällig am", "Priorität", "Status"]), 0, 1)
        layout.addWidget(self._offene_tabelle_panel("Schäden", ["Schaden", "Objekt", "Wohnung", "Datum", "Priorität", "Status"]), 1, 0)
        layout.addWidget(self._offene_tabelle_panel("Rechnungen", ["Rechnungsnummer", "Dienstleister", "Objekt", "Datum", "Betrag brutto", "Status"]), 1, 1)
        return page

    def _offene_tabelle_panel(self, titel: str, spalten: list[str]) -> QFrame:
        panel = QFrame()
        panel.setObjectName("chartPanel")
        lay = QVBoxLayout(panel)
        label = QLabel(f"Offene {titel}")
        label.setObjectName("metricTitle")
        lay.addWidget(label)

        table = QTableWidget()
        table.setAlternatingRowColors(True)
        table.setWordWrap(False)
        table.setColumnCount(len(spalten))
        table.setHorizontalHeaderLabels(spalten)
        table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        felder = SCHEMA.get(titel, [])
        rows = dashboard_offene_liste(titel, 8)
        table.setRowCount(len(rows))

        for r, row in enumerate(rows):
            for c, feld in enumerate(spalten):
                wert = dashboard_feld(row, felder, [feld])
                if not wert and c < len(row):
                    wert = row[c]
                table.setItem(r, c, QTableWidgetItem(str(wert)))

        for c in range(len(spalten)):
            table.setColumnWidth(c, 140)

        lay.addWidget(table, 1)

        btn = QPushButton(f"{titel} öffnen")
        btn.setObjectName("primaryButton")
        btn.clicked.connect(lambda checked=False, t=titel: self.nav(t))
        lay.addWidget(btn)
        return panel


    def _dashboard_tabelle_tab(self, title: str, datenbereich: str) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)

        label = QLabel(title)
        label.setObjectName("metricTitle")
        layout.addWidget(label)

        table = QTableWidget()
        table.setAlternatingRowColors(True)
        table.setWordWrap(False)
        table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)

        headers = SCHEMA.get(datenbereich, [])
        rows = DATA.get(datenbereich, [])

        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setRowCount(len(rows))

        for r, row in enumerate(rows):
            for c in range(len(headers)):
                value = row[c] if c < len(row) else ""
                table.setItem(r, c, QTableWidgetItem(str(value)))

        for c in range(min(len(headers), 12)):
            table.setColumnWidth(c, 160)

        layout.addWidget(table, 1)

        btn = QPushButton(f"{datenbereich} öffnen")
        btn.setObjectName("primaryButton")
        btn.clicked.connect(lambda checked=False, t=datenbereich: self.nav(t))
        layout.addWidget(btn)

        return page

    @staticmethod
    def _dashboard_diagramm_tab() -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)

        view = QGraphicsView()
        scene = QGraphicsScene()
        view.setScene(scene)

        k = dashboard_kennzahlen()
        daten = [
            ("Objekte", k.get("objekte", 0)),
            ("Wohnungen", k.get("wohnungen", 0)),
            ("Mieter", k.get("mieter", 0)),
            ("Vermietet", k.get("vermietet", 0)),
            ("Frei", k.get("frei", 0)),
        ]

        scene.addText("Bestandsübersicht").setPos(30, 10)
        maxv = max([float(v) for _n, v in daten] + [1.0])
        y = 70

        for name, value in daten:
            width = int(float(value) / maxv * 650)
            scene.addText(name).setPos(30, y - 22)
            scene.addRect(30, y, max(width, 4), 32, QPen(), QBrush())
            scene.addText(str(int(value))).setPos(30 + max(width, 4) + 15, y + 4)
            y += 70

        scene.setSceneRect(0, 0, 900, 420)
        layout.addWidget(view)
        return page

    def _dashboard_berichte_tab(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)

        info = QLabel("Berichte sind aktiv: Nutze Analyse-Center, Zahlenanalyse, Buchhaltung, BK-Automatik und Objekt-Cockpit für druck-/exportfähige Auswertungen.")
        info.setWordWrap(True)
        layout.addWidget(info)

        grid = QGridLayout()
        buttons = [
            ("Analyse-Center öffnen", "Analyse-Center"),
            ("Zahlenanalyse öffnen", "Zahlenanalyse"),
            ("Buchhaltung öffnen", "Buchhaltung"),
            ("BK-Automatik öffnen", "BK-Automatik"),
            ("Objekt-Cockpit öffnen", "Objekt-Cockpit"),
        ]

        for i, (label, ziel) in enumerate(buttons):
            b = QPushButton(label)
            b.setObjectName("primaryButton" if i == 0 else "")
            b.clicked.connect(lambda checked=False, z=ziel: self.nav(z))
            grid.addWidget(b, i // 2, i % 2)

        layout.addLayout(grid)
        layout.addStretch()
        return page

    def _dashboard_export_tab(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)

        info = QLabel("Exportfunktionen sind aktiv in Globale Suche, Akten-Center, Buchhaltung, BK-Automatik und Zahlenanalyse.")
        info.setWordWrap(True)
        layout.addWidget(info)

        for label, ziel in [
            ("Globale Suche Export", "Globale Suche"),
            ("Akten-Center Export", "Akten-Center"),
            ("Buchhaltung Export", "Buchhaltung"),
            ("BK-Automatik Export", "BK-Automatik"),
            ("Zahlenanalyse Export", "Zahlenanalyse"),
        ]:
            b = QPushButton(label)
            b.clicked.connect(lambda checked=False, z=ziel: self.nav(z))
            layout.addWidget(b)

        layout.addStretch()
        return page


    def _placeholder_tab(self, title: str, text: str) -> QWidget:
        page = QWidget(); layout = QVBoxLayout(page); layout.addWidget(self._panel(title, text)); return page


def widget_enthaelt_text(widget: QWidget, suchbegriffe: list[str]) -> bool:
    try:
        texte = []
        if hasattr(widget, "text"):
            texte.append(str(widget.text()))
        if hasattr(widget, "placeholderText"):
            texte.append(str(widget.placeholderText()))
        if hasattr(widget, "objectName"):
            texte.append(str(widget.objectName()))
        joined = " ".join(texte).lower()
        return any(s.lower() in joined for s in suchbegriffe)
    except (OSError, ValueError, TypeError, AttributeError, RuntimeError, KeyError, IndexError):
        return False



class TabellenSeite(QWidget):
    def __init__(self,titel):
        super().__init__(); self.titel=titel; self.felder=SCHEMA[titel]; self.daten=DATA[titel]; self.anzeige: list[list[str]]=[]
        root=QVBoxLayout(self); root.setContentsMargins(14,14,14,14)
        top=QHBoxLayout(); title=QLabel(titel); title.setObjectName("pageTitle"); self.search=QLineEdit(); self.search.setMinimumWidth(260); self.search.setPlaceholderText("Suchen..."); self.search.textChanged.connect(self.laden)
        self.objekt_filter = QComboBox()
        self.objekt_filter.setMinimumWidth(260)
        self.objekt_filter.currentTextChanged.connect(self.laden)  # Filter wirkt nur auf die Anzeige/Suche
        top.addWidget(title); top.addStretch()
        top.addWidget(QLabel("Objektordner:"))
        top.addWidget(self.objekt_filter)
        top.addWidget(self.search)
        for text,fn in [("+ Neu",self.neu),("Bearbeiten",self.bearbeiten),("Löschen",self.loeschen),("Excel öffnen",self.excel_oeffnen),("XLSX Import",self.xlsx_import),("XLSX Export",self.xlsx_export),("PDF hochladen",self.pdf_upload),("PDF öffnen",self.pdf_oeffnen),("Ordner wählen",self.ordner_waehlen),("Ordner öffnen",self.ordner_oeffnen)]:
            b=QPushButton(text)
            b.setMinimumWidth(125)
            b.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Fixed)
            b.clicked.connect(fn)
            top.addWidget(b)

        toolbar_widget = QWidget()
        toolbar_widget.setLayout(top)
        toolbar_scroll = QScrollArea()
        toolbar_scroll.setWidgetResizable(True)
        toolbar_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        toolbar_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        toolbar_scroll.setFixedHeight(74)
        toolbar_scroll.setWidget(toolbar_widget)
        root.addWidget(toolbar_scroll)
        self.table=QTableWidget(); self.table.setAlternatingRowColors(True); self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows); self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection); self.table.verticalHeader().setDefaultSectionSize(int(str(CONFIG.get("tabellen_zeilenhoehe", "34")))); self.table.setWordWrap(False); self.table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel); self.table.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel); self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn); self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn); self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed); self.table.horizontalHeader().setStretchLastSection(False); self.table.verticalHeader().setVisible(True); self.table.cellDoubleClicked.connect(self.zelle_doppelklick)
        root.addWidget(self.table,1); self.laden()

    def spaltenbreiten(self):
        for c,feld in enumerate(self.felder):
            n = str(feld).lower()
            w = 360 if ("pdf" in n or "ordner" in n) else (330 if any(x in n for x in ["notiz", "unterlagen", "konto", "rechnung"]) else (260 if any(x in n for x in ["mieter", "objekt", "vermieter"]) else 180))
            self.table.setColumnWidth(c, w)


    def objektordner_index(self) -> int | None:
        for index, feld in enumerate(self.felder):
            if "objektordner" in str(feld).lower():
                return index
        return None

    def aktualisiere_objekt_filter(self) -> None:
        current = self.objekt_filter.currentText().strip()

        self.objekt_filter.blockSignals(True)
        self.objekt_filter.clear()
        self.objekt_filter.addItem("Alle Objektordner")

        for wert in alle_objektordner():
            self.objekt_filter.addItem(wert)

        if current:
            idx = self.objekt_filter.findText(current)
            if idx >= 0:
                self.objekt_filter.setCurrentIndex(idx)

        self.objekt_filter.blockSignals(False)

    def laden(self) -> None:
        suche = self.search.text().lower().strip()

        self.aktualisiere_objekt_filter()
        gewaehlt = self.objekt_filter.currentText().strip()

        rows = list(self.daten)

        if gewaehlt and gewaehlt != "Alle Objektordner":
            gewaehlt_key = norm_key(gewaehlt)
            rows = [
                row
                for row in rows
                if norm_key(objektordner_fuer_datensatz(self.titel, row)) == gewaehlt_key
            ]

        rows = [
            row
            for row in rows
            if not suche
            or suche in " ".join(str(value) for value in row).lower()
        ]

        self.anzeige = rows
        self.table.setRowCount(len(rows))
        self.table.setColumnCount(len(self.felder))
        self.table.setHorizontalHeaderLabels(self.felder)

        for row_index, row in enumerate(rows):
            for col_index in range(len(self.felder)):
                value = row[col_index] if col_index < len(row) else ""
                text_value = "" if value is None else str(value)
                item = QTableWidgetItem(text_value)

                if "pdf" in str(self.felder[col_index]).lower() and text_value:
                    item.setText(f"📎 {Path(text_value).name}")
                    item.setToolTip(text_value)

                self.table.setItem(row_index, col_index, item)

        self.spaltenbreiten()

    def idx(self):
        r=self.table.currentRow()
        if r<0: QMessageBox.information(self,"Hinweis","Bitte Zeile auswählen."); return None
        if r<len(self.anzeige):
            try: return self.daten.index(self.anzeige[r])
            except ValueError: return r
        return None
    def neu(self):
        # Beim Anlegen eines Dienstleisters den aktuell gewählten Objektordner übernehmen.
        # Bestehende Datensätze und Excel-Spalten bleiben dabei vollständig erhalten.
        startwerte = ["" for _ in self.felder]
        if self.titel == "Dienstleister":
            gewaehlt = self.objekt_filter.currentText().strip()
            if gewaehlt and gewaehlt != "Alle Objektordner":
                for index, feld in enumerate(self.felder):
                    if norm_key(feld) == norm_key("Objektordner"):
                        startwerte[index] = gewaehlt
                        break

        d=EingabeDialog(
            f"{self.titel} anlegen",
            self.felder,
            startwerte,
            bereich=self.titel,
        )
        if d.exec():
            self.daten.append(d.values())
            speichere_tabelle(self.titel)
            self.after_change()
    def bearbeiten(self):
        i=self.idx()
        if i is None: return
        d=EingabeDialog(f"{self.titel} bearbeiten", self.felder, self.daten[i], bereich=self.titel)
        if d.exec(): self.daten[i]=d.values(); speichere_tabelle(self.titel); self.after_change()
    def loeschen(self):
        i=self.idx()
        if i is not None and frage_ja_nein("Löschen","Eintrag wirklich löschen?"): del self.daten[i]; speichere_tabelle(self.titel); self.after_change()
    def after_change(self):
        if self.titel in ["Mieter","Zahlungen"]: pruefe_zahlungen(); speichere_tabelle("Zahlungsprüfung")
        self.laden()
    def excel_oeffnen(self): speichere_tabelle(self.titel); system_datei_oeffnen(xlsx_pfad(self.titel))
    def xlsx_export(self):
        speichere_tabelle(self.titel); ziel=EXPORT_DIR/DATA_FILES[self.titel].replace(".xlsx","_export.xlsx"); shutil.copy2(xlsx_pfad(self.titel),ziel); QMessageBox.information(self,"Export",f"Export erstellt:\n{ziel}"); system_datei_oeffnen(ziel)
    def xlsx_import(self):
        p,_=QFileDialog.getOpenFileName(self,"XLSX importieren","","Excel-Dateien (*.xlsx)")
        if p: imp,dup=lese_xlsx_import(p,self.titel); QMessageBox.information(self,"Import",f"Importiert: {imp}\nDubletten: {dup}"); self.laden()
    def pdf_spalten(self) -> list[int]:
        return [i for i, feld in enumerate(self.felder) if "pdf" in str(feld).lower()]
    def pdf_upload(self):
        i=self.idx()
        if i is None: return
        cols=self.pdf_spalten()
        if not cols: QMessageBox.information(self,"PDF","Diese Tabelle hat keine PDF-Spalte."); return
        choice,ok=QInputDialog.getItem(self,"PDF-Spalte","PDF speichern in:",[self.felder[c] for c in cols],0,False)
        if not ok: return
        col=self.felder.index(choice); p,_=QFileDialog.getOpenFileName(self,"PDF auswählen","","PDF-Dateien (*.pdf)")
        if p:
            while len(self.daten[i])<len(self.felder): self.daten[i].append("")
            self.daten[i][col]=kopiere_pdf_ins_projekt(p); speichere_tabelle(self.titel); self.laden()

    def pdf_oeffnen(self) -> None:
        i = self.idx()

        if i is None:
            return

        col = self.table.currentColumn()

        ist_pdf_spalte = False

        if 0 <= col < len(self.felder):
            feldname = str(self.felder[col]).lower()
            ist_pdf_spalte = "pdf" in feldname

        if not ist_pdf_spalte:
            cols = self.pdf_spalten()

            if not cols:
                return

            col = cols[0]

        path = ""

        if 0 <= i < len(self.daten):
            row = self.daten[i]

            if 0 <= col < len(row):
                path = str(row[col])

        if path:
            system_datei_oeffnen(path)


    def ordner_spalten(self) -> list[int]:
        return [i for i, feld in enumerate(self.felder) if "ordner" in str(feld).lower()]

    def ordner_waehlen(self) -> None:
        i = self.idx()
        if i is None:
            return

        cols = self.ordner_spalten()
        if not cols:
            QMessageBox.information(self, "Ordner", "Diese Tabelle hat keine Ordner-Spalte.")
            return

        choice, ok = QInputDialog.getItem(
            self,
            "Ordner-Spalte",
            "Ordner speichern in:",
            [self.felder[c] for c in cols],
            0,
            False,
        )
        if not ok:
            return

        col = self.felder.index(choice)
        path = QFileDialog.getExistingDirectory(self, "Ordner auswählen")
        if not path:
            return

        while len(self.daten[i]) < len(self.felder):
            self.daten[i].append("")

        self.daten[i][col] = path
        speichere_tabelle(self.titel)
        self.laden()

    def ordner_oeffnen(self) -> None:
        i = self.idx()
        if i is None:
            return

        col = self.table.currentColumn()
        ist_ordner_spalte = False

        if 0 <= col < len(self.felder):
            ist_ordner_spalte = "ordner" in str(self.felder[col]).lower()

        if not ist_ordner_spalte:
            cols = self.ordner_spalten()
            if not cols:
                QMessageBox.information(self, "Ordner", "Diese Tabelle hat keine Ordner-Spalte.")
                return
            col = cols[0]

        path = ""
        if 0 <= i < len(self.daten):
            row = self.daten[i]
            if 0 <= col < len(row):
                path = str(row[col]).strip()

        if path:
            system_datei_oeffnen(path)
        else:
            QMessageBox.information(self, "Ordner", "Es ist kein Ordner hinterlegt.")




    def zelle_doppelklick(self, _row: int, col: int) -> None:
        if 0 <= col < len(self.felder) and "pdf" in str(self.felder[col]).lower():
            self.pdf_oeffnen()





class BeziehungsAnalyseSeite(QWidget):
    """Eingebettetes Analyse-Center als normale Seite im Hauptfenster."""

    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        layout.setContentsMargins(18, 18, 18, 18)

        title = QLabel("Analyse-Center")
        title.setObjectName("pageTitle")
        layout.addWidget(title)

        info = QLabel("Tabellen werden über gleiche oder ähnliche Spaltenköpfe logisch verbunden. Excel-Dateien bleiben unverändert.")
        info.setObjectName("metricTitle")
        layout.addWidget(info)

        row = QHBoxLayout()

        self.feld = QComboBox()
        self.feld.addItem("Freie Suche")
        self.feld.addItems(beziehungsfelder())
        self.feld.setMinimumWidth(260)

        self.suche = QLineEdit()
        self.suche.setPlaceholderText("Suchwert eingeben, z. B. Mietername, Objekt, Wohnung ...")

        btn = QPushButton("Analyse starten")
        btn.setObjectName("primaryButton")
        btn.clicked.connect(self.analysieren)

        export_btn = QPushButton("Excel exportieren")
        export_btn.clicked.connect(self.export_excel)

        row.addWidget(QLabel("Beziehungsfeld:"))
        row.addWidget(self.feld)
        row.addWidget(self.suche, 1)
        row.addWidget(btn)
        row.addWidget(export_btn)
        layout.addLayout(row)

        self.zusammenfassung = QLabel("Noch keine Analyse ausgeführt.")
        self.zusammenfassung.setObjectName("metricTitle")
        layout.addWidget(self.zusammenfassung)

        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Bereich", "Zeile", "Feld", "Wert", "Inhalt"])
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        layout.addWidget(self.table, 1)

        self.letzte_treffer: list[dict[str, str]] = []

    def analysieren(self) -> None:
        feld = self.feld.currentText()
        wert = self.suche.text().strip()

        if not wert:
            QMessageBox.information(self, "Analyse", "Bitte einen Suchwert eingeben.")
            return

        if feld == "Freie Suche":
            treffer = finde_freie_suche_beziehungen(wert)
        else:
            treffer = finde_beziehungen(feld, wert)

        self.letzte_treffer = treffer
        self.zusammenfassung.setText(f"Suchwert: {wert} | Beziehungsfeld: {feld} | Treffer: {len(treffer)}")

        self.table.setRowCount(len(treffer))

        for row_index, item in enumerate(treffer):
            self.table.setItem(row_index, 0, QTableWidgetItem(item["bereich"]))
            self.table.setItem(row_index, 1, QTableWidgetItem(item["zeile"]))
            self.table.setItem(row_index, 2, QTableWidgetItem(item["feld"]))
            self.table.setItem(row_index, 3, QTableWidgetItem(item["wert"]))
            self.table.setItem(row_index, 4, QTableWidgetItem(item["inhalt"]))

        self.table.setColumnWidth(0, 190)
        self.table.setColumnWidth(1, 70)
        self.table.setColumnWidth(2, 170)
        self.table.setColumnWidth(3, 240)
        self.table.setColumnWidth(4, 700)

    def export_excel(self) -> None:
        dialog = BeziehungsAnalyseDialog(self)
        dialog.letzte_treffer = self.letzte_treffer
        dialog.export_excel()


def zahl_aus_text(value: Any) -> float:
    """Konvertiert deutsche Zahlenformate robust in float."""
    text = str(value or "").strip()
    if not text:
        return 0.0

    text = (
        text.replace("€", "")
        .replace("EUR", "")
        .replace("eur", "")
        .replace(" ", "")
        .replace(".", "")
        .replace(",", ".")
    )

    try:
        return float(text)
    except (ValueError, TypeError):
        return 0.0


def ist_zahlenfeld(feldname: str) -> bool:
    name = norm_key(feldname)
    keywords = [
        "betrag",
        "miete",
        "kaltmiete",
        "nebenkosten",
        "kaution",
        "kosten",
        "beitrag",
        "brutto",
        "netto",
        "mwst",
        "gezahlt",
        "soll",
        "differenz",
        "nachzahlung",
        "guthaben",
        "bkmonatlich",
        "bkjahrlich",
        "istbkjahrlich",
        "wohnflache",
        "grosse",
        "stunden",
        "verbrauch",
    ]
    return any(k in name for k in keywords)


def zahlenanalyse_gesamt() -> list[dict[str, Any]]:
    """Analysiert numerische Spalten über alle Tabellen, ohne Excel-Dateien zu verändern."""
    results: list[dict[str, Any]] = []

    for bereich, rows in DATA.items():
        felder = SCHEMA.get(bereich, [])

        for col_index, feld in enumerate(felder):
            if not ist_zahlenfeld(feld):
                continue

            werte: list[float] = []
            for row in rows:
                if col_index < len(row):
                    wert = zahl_aus_text(row[col_index])
                    if abs(wert) > 0.0001:
                        werte.append(wert)

            if not werte:
                continue

            total = sum(werte)
            count = len(werte)
            avg = total / count if count else 0.0

            results.append({
                "bereich": bereich,
                "feld": str(feld),
                "anzahl": count,
                "summe": total,
                "durchschnitt": avg,
                "minimum": min(werte),
                "maximum": max(werte),
            })

    results.sort(key=lambda item: abs(float(item["summe"])), reverse=True)
    return results


class ZahlenAnalyseSeite(QWidget):
    """Zahlenanalyse mit sichtbarem Diagramm."""

    def __init__(self):
        super().__init__()

        layout = QVBoxLayout(self)
        layout.setContentsMargins(18, 18, 18, 18)

        title = QLabel("Zahlenanalyse & Diagramme")
        title.setObjectName("pageTitle")
        layout.addWidget(title)

        info = QLabel("Auswertung aller erkannten Zahlenfelder. Die Excel-Dateien werden nur gelesen und nicht verändert.")
        info.setObjectName("metricTitle")
        layout.addWidget(info)

        button_row = QHBoxLayout()
        aktualisieren = QPushButton("Analyse aktualisieren")
        aktualisieren.setObjectName("primaryButton")
        aktualisieren.clicked.connect(self.laden)

        export = QPushButton("Excel exportieren")
        export.clicked.connect(self.export_excel)

        button_row.addWidget(aktualisieren)
        button_row.addWidget(export)
        button_row.addStretch()
        layout.addLayout(button_row)

        self.tabs = QTabWidget()
        layout.addWidget(self.tabs, 1)

        self.table = QTableWidget()
        self.chart = QGraphicsView()
        self.chart_scene = QGraphicsScene()
        self.chart.setScene(self.chart_scene)

        self.tabs.addTab(self.table, "Zahlenübersicht")
        self.tabs.addTab(self.chart, "Diagramm")

        self.daten: list[dict[str, Any]] = []
        self.laden()

    def laden(self) -> None:
        self.daten = zahlenanalyse_gesamt()
        self._tabelle_fuellen()
        self._diagramm_zeichnen()

    def _tabelle_fuellen(self) -> None:
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            "Bereich",
            "Zahlenfeld",
            "Anzahl",
            "Summe",
            "Durchschnitt",
            "Minimum",
            "Maximum",
        ])
        self.table.setRowCount(len(self.daten))
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)

        for row_index, item in enumerate(self.daten):
            row_values = [
                item["bereich"],
                item["feld"],
                str(item["anzahl"]),
                f'{item["summe"]:.2f}',
                f'{item["durchschnitt"]:.2f}',
                f'{item["minimum"]:.2f}',
                f'{item["maximum"]:.2f}',
            ]

            for col_index, value in enumerate(row_values):
                self.table.setItem(row_index, col_index, QTableWidgetItem(str(value)))

        self.table.setColumnWidth(0, 180)
        self.table.setColumnWidth(1, 220)
        for col in range(2, 7):
            self.table.setColumnWidth(col, 130)

    def _diagramm_zeichnen(self) -> None:
        self.chart_scene.clear()

        top = self.daten[:10]
        if not top:
            self.chart_scene.addText("Keine Zahlenwerte gefunden.")
            return

        max_value = max(abs(float(item["summe"])) for item in top) or 1.0
        x = 40
        y = 40
        max_width = 700
        bar_height = 28
        gap = 18

        title = self.chart_scene.addText("Top-Zahlenfelder nach Summe")
        title.setPos(x, 5)

        for index, item in enumerate(top):
            value = float(item["summe"])
            width = int((abs(value) / max_value) * max_width)
            yy = y + index * (bar_height + gap)

            label = f'{item["bereich"]} / {item["feld"]}'
            self.chart_scene.addText(label).setPos(x, yy - 18)

            self.chart_scene.addRect(
                x,
                yy,
                max(width, 4),
                bar_height,
                QPen(),
                QBrush()
            )

            value_text = self.chart_scene.addText(f'{value:.2f}')
            value_text.setPos(x + max(width, 4) + 12, yy + 2)

        self.chart_scene.setSceneRect(0, 0, 950, y + len(top) * (bar_height + gap) + 80)

    def export_excel(self) -> None:
        if not self.daten:
            QMessageBox.information(self, "Export", "Es gibt keine Zahlenanalyse zum Exportieren.")
            return


        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        default_name = EXPORT_DIR / f"zahlenanalyse_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

        ziel_text, _ = QFileDialog.getSaveFileName(
            self,
            "Zahlenanalyse exportieren",
            str(default_name),
            "Excel-Dateien (*.xlsx)",
        )

        if not ziel_text:
            return

        ziel = Path(ziel_text)
        if ziel.suffix.lower() != ".xlsx":
            ziel = ziel.with_suffix(".xlsx")

        wb = Workbook()
        ws = wb.active

        if not isinstance(ws, Worksheet):
            QMessageBox.warning(self, "Export", "Excel-Arbeitsblatt konnte nicht erstellt werden.")
            return

        ws.title = "Zahlenanalyse"
        ws.append(["Bereich", "Zahlenfeld", "Anzahl", "Summe", "Durchschnitt", "Minimum", "Maximum"])

        for item in self.daten:
            ws.append([
                item["bereich"],
                item["feld"],
                item["anzahl"],
                item["summe"],
                item["durchschnitt"],
                item["minimum"],
                item["maximum"],
            ])

        for col_index, _col in enumerate(ws.columns, start=1):
            ws.column_dimensions[get_column_letter(col_index)].width = 24

        wb.save(ziel)

        QMessageBox.information(self, "Export", f"Zahlenanalyse exportiert:\n{ziel}")
        system_datei_oeffnen(str(ziel.parent))


def buchhaltung_betrag(value: Any) -> float:
    text = str(value or "").strip()
    if not text:
        return 0.0
    text = text.replace("€", "").replace("EUR", "").replace("eur", "").replace(" ", "")
    text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except (ValueError, TypeError):
        return 0.0


def buchhaltung_jahr_aus_text(value: Any) -> str:
    text = str(value or "")
    match = re.search(r"(20\d{2}|19\d{2})", text)
    return match.group(1) if match else ""


def buchhaltung_monat_aus_datum(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    parts = re.split(r"[./-]", text)
    if len(parts) >= 2:
        # dd.mm.yyyy oder yyyy-mm-dd
        if len(parts[0]) == 4:
            return parts[1].zfill(2)
        return parts[1].zfill(2)
    return ""



def objekt_buchhaltungsart() -> dict[str, str]:
    mapping: dict[str, str] = {}
    felder = SCHEMA.get("Objekte", [])
    objekt_idx = 0
    buchhaltung_idx = None

    for i, feld in enumerate(felder):
        if norm_key(feld) in ["buchhaltung", "eigentumsart"]:
            buchhaltung_idx = i

    for row in DATA.get("Objekte", []):
        objekt = str(row[objekt_idx]).strip() if len(row) > objekt_idx else ""
        art = str(row[buchhaltung_idx]).strip() if buchhaltung_idx is not None and len(row) > buchhaltung_idx else ""
        if objekt:
            mapping[norm_key(objekt)] = art

    return mapping


def ist_hausverwaltung_objekt(objekt: Any) -> bool:
    """Ordnet sichtbare Fremdverwaltungsobjekte intern der HV-Buchhaltung zu."""
    mapping = objekt_buchhaltungsart()
    art = mapping.get(norm_key(objekt), "").lower().strip()
    return (
        "fremdverwaltung" in art
        or "fremd verwaltung" in art
        or "hausverwaltung" in art  # Kompatibilität zu bereits gespeicherten Werten
        or art == "hv"
    )


def ist_eigenbestand_objekt(objekt: Any) -> bool:
    mapping = objekt_buchhaltungsart()
    art = mapping.get(norm_key(objekt), "").lower().strip()
    if ist_hausverwaltung_objekt(objekt):
        return False
    return any(w in art for w in ["eigen", "eigene", "eigenbestand", "eigene buchhaltung"])


def ist_versorger_kostenart(value: Any) -> bool:
    text = str(value or "").lower()
    return any(w in text for w in [
        "versorger", "stadtwerke", "wasser", "abwasser", "strom", "gas",
        "müll", "muell", "schornstein", "grundsteuer", "versicherung",
        "heizung", "hausmeister", "wartung"
    ])

def buchhaltung_daten(jahr_filter: str = "", monat_filter: str = "") -> list[dict[str, Any]]:
    buchungen: list[dict[str, Any]] = []

    # Einnahmen: HV-Rechnungen
    for row in DATA.get("HV-Rechnungen", []):
        datum = row[0] if len(row) > 0 else ""
        nr = row[1] if len(row) > 1 else ""
        objekt = row[2] if len(row) > 2 else ""
        vermieter = row[3] if len(row) > 3 else ""
        leistung = row[4] if len(row) > 4 else "HV-Rechnung"
        netto = buchhaltung_betrag(row[5] if len(row) > 5 else "")
        brutto = buchhaltung_betrag(row[7] if len(row) > 7 else "")
        status = row[8] if len(row) > 8 else ""
        jahr = buchhaltung_jahr_aus_text(datum)
        monat = buchhaltung_monat_aus_datum(datum)
        betrag = brutto if brutto else netto

        if jahr_filter and str(jahr) != jahr_filter:
            continue
        if monat_filter and monat != monat_filter:
            continue

        if betrag:
            buchungen.append({
                "quelle": "HV-Rechnungen",
                "datum": str(datum),
                "jahr": str(jahr),
                "monat": str(monat),
                "typ": "Einnahme",
                "name": str(vermieter),
                "kostenart": "HV-Rechnung",
                "status": str(status),
                "betrag": betrag,
                "zweck": f"{nr} / {objekt} / {leistung}",
            })

    # Einnahmen/Sonstige echte Zahlungseingänge aus Zahlungen bleiben möglich,
    # aber Miet-/HV-Schätzungen werden nicht mehr automatisch erzeugt.
    for row in DATA.get("Zahlungen", []):
        datum = row[0] if len(row) > 0 else ""
        name = row[1] if len(row) > 1 else ""
        betrag = buchhaltung_betrag(row[2] if len(row) > 2 else "")
        zweck = row[3] if len(row) > 3 else ""
        status = row[5] if len(row) > 5 else ""
        kostenart = row[6] if len(row) > 6 else "Sonstiges"
        jahr = row[7] if len(row) > 7 else buchhaltung_jahr_aus_text(datum)
        monat = buchhaltung_monat_aus_datum(datum)

        text = f"{kostenart} {zweck} {name}".lower()
        ist_sonstige_einnahme = any(w in text for w in ["sonstige einnahme", "gutschrift", "zinsertrag"])

        if not ist_sonstige_einnahme:
            continue
        if jahr_filter and str(jahr) != jahr_filter:
            continue
        if monat_filter and monat != monat_filter:
            continue

        buchungen.append({
            "quelle": "Zahlungen",
            "datum": str(datum),
            "jahr": str(jahr),
            "monat": str(monat),
            "typ": "Einnahme",
            "name": str(name),
            "kostenart": str(kostenart),
            "status": str(status),
            "betrag": betrag,
            "zweck": str(zweck),
        })

    # Ausgaben: Rechnungen nur dann in eigene Buchhaltung,
    # wenn das Objekt Eigenbestand / eigene Buchhaltung ist.
    for row in DATA.get("Rechnungen", []):
        nr = row[0] if len(row) > 0 else ""
        dienstleister = row[1] if len(row) > 1 else ""
        objekt = row[2] if len(row) > 2 else ""
        datum = row[3] if len(row) > 3 else ""
        brutto = buchhaltung_betrag(row[6] if len(row) > 6 else "")
        status = row[7] if len(row) > 7 else ""
        jahr = buchhaltung_jahr_aus_text(datum)
        monat = buchhaltung_monat_aus_datum(datum)

        if not ist_eigenbestand_objekt(objekt):
            continue
        if jahr_filter and str(jahr) != jahr_filter:
            continue
        if monat_filter and monat != monat_filter:
            continue

        if brutto:
            buchungen.append({
                "quelle": "Rechnungen Eigenbestand",
                "datum": str(datum),
                "jahr": str(jahr),
                "monat": str(monat),
                "typ": "Ausgabe",
                "name": str(dienstleister),
                "kostenart": "Rechnung",
                "status": str(status),
                "betrag": brutto,
                "zweck": f"Rechnung {nr} / {objekt}",
            })

    # Ausgaben: Versorger/BK nur bei Eigenbestand
    for row in DATA.get("Betriebskosten", []):
        felder = SCHEMA.get("Betriebskosten", [])
        def val(field_name: str) -> str:
            for i, feld in enumerate(felder):
                if norm_key(feld) == norm_key(field_name) and i < len(row):
                    return str(row[i])
            return ""

        objekt = val("Objektordner") or val("Objekt")
        kostenart = val("BK Art")
        versorger = val("Versorger")
        abschlag = buchhaltung_betrag(val("Abschlag monatlich"))
        jahr = val("Abrechnungsjahr")
        faellig = val("Fällig am")
        monat = buchhaltung_monat_aus_datum(faellig)

        if not ist_eigenbestand_objekt(objekt):
            continue
        if not ist_versorger_kostenart(kostenart + " " + versorger):
            continue
        if jahr_filter and str(jahr) != jahr_filter:
            continue
        if monat_filter and monat != monat_filter:
            continue

        if abschlag:
            buchungen.append({
                "quelle": "Versorger/BK Eigenbestand",
                "datum": str(faellig),
                "jahr": str(jahr),
                "monat": str(monat),
                "typ": "Ausgabe",
                "name": str(versorger),
                "kostenart": str(kostenart),
                "status": "",
                "betrag": abschlag,
                "zweck": f"{objekt} / Kundennr. {val('Kundennummer')} / Vertrag {val('Vertragsnummer')}",
            })

    return buchungen


def buchhaltung_summe(buchungen: list[dict[str, Any]], typ: str) -> float:
    return sum(float(b["betrag"]) for b in buchungen if b["typ"] == typ)


LOHN_DATEI = APP_DIR / "lohnabrechnungen.json"


def lade_lohnabrechnungen() -> list[dict[str, Any]]:
    """Lädt Lohnabrechnungen aus einer separaten JSON-Datei; die Excel-Struktur bleibt unverändert."""
    if not LOHN_DATEI.exists():
        return []
    try:
        daten = json.loads(LOHN_DATEI.read_text(encoding="utf-8"))
        return daten if isinstance(daten, list) else []
    except (OSError, json.JSONDecodeError):
        return []


def speichere_lohnabrechnungen(daten: list[dict[str, Any]]) -> None:
    LOHN_DATEI.write_text(json.dumps(daten, indent=2, ensure_ascii=False), encoding="utf-8")


def lohn_berechnen(d: dict[str, Any]) -> dict[str, float]:
    brutto = max(0.0, to_float(d.get("brutto")))
    steuer = max(0.0, to_float(d.get("lohnsteuer")))
    soli = max(0.0, to_float(d.get("soli")))
    kirche = max(0.0, to_float(d.get("kirchensteuer")))
    kv_an = max(0.0, to_float(d.get("kv_an")))
    rv_an = max(0.0, to_float(d.get("rv_an")))
    av_an = max(0.0, to_float(d.get("av_an")))
    pv_an = max(0.0, to_float(d.get("pv_an")))
    sonstige = max(0.0, to_float(d.get("sonstige_abzuege")))
    kv_ag = max(0.0, to_float(d.get("kv_ag")))
    rv_ag = max(0.0, to_float(d.get("rv_ag")))
    av_ag = max(0.0, to_float(d.get("av_ag")))
    pv_ag = max(0.0, to_float(d.get("pv_ag")))
    umlagen = max(0.0, to_float(d.get("umlagen_ag")))
    netto = brutto - steuer - soli - kirche - kv_an - rv_an - av_an - pv_an - sonstige
    ag_anteile = kv_ag + rv_ag + av_ag + pv_ag + umlagen
    return {
        "brutto": brutto,
        "abzuege": steuer + soli + kirche + kv_an + rv_an + av_an + pv_an + sonstige,
        "netto": netto,
        "ag_anteile": ag_anteile,
        "arbeitgeberkosten": brutto + ag_anteile,
    }


class LohnDialog(QDialog):
    FELDER = [
        ("mitarbeiter", "Mitarbeiter"), ("personalnummer", "Personalnummer"),
        ("abrechnungsmonat", "Abrechnungsmonat (YYYY-MM)"), ("zahlungsdatum", "Zahlungsdatum"),
        ("brutto", "Bruttolohn"), ("lohnsteuer", "Lohnsteuer"), ("soli", "Solidaritätszuschlag"),
        ("kirchensteuer", "Kirchensteuer"), ("kv_an", "Krankenversicherung AN"),
        ("rv_an", "Rentenversicherung AN"), ("av_an", "Arbeitslosenversicherung AN"),
        ("pv_an", "Pflegeversicherung AN"), ("sonstige_abzuege", "Sonstige Abzüge"),
        ("kv_ag", "Krankenversicherung AG"), ("rv_ag", "Rentenversicherung AG"),
        ("av_ag", "Arbeitslosenversicherung AG"), ("pv_ag", "Pflegeversicherung AG"),
        ("umlagen_ag", "Umlagen / BG AG"), ("kostenstelle", "Kostenstelle / Objekt"),
        ("notiz", "Notiz"),
    ]

    def __init__(self, daten: dict[str, Any] | None = None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Lohnabrechnung")
        self.resize(720, 760)
        root = QVBoxLayout(self)
        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        content = QWidget(); form = QFormLayout(content)
        self.inputs: dict[str, QLineEdit] = {}
        daten = daten or {}
        for key, label in self.FELDER:
            edit = QLineEdit(str(daten.get(key, "")))
            self.inputs[key] = edit
            form.addRow(label + ":", edit)
        scroll.setWidget(content); root.addWidget(scroll, 1)
        buttons = QHBoxLayout(); buttons.addStretch()
        ok = QPushButton("Speichern"); ok.setObjectName("primaryButton"); ok.clicked.connect(self.accept)
        cancel = QPushButton("Abbrechen"); cancel.clicked.connect(self.reject)
        buttons.addWidget(ok); buttons.addWidget(cancel); root.addLayout(buttons)

    def values(self) -> dict[str, Any]:
        d = {key: edit.text().strip() for key, edit in self.inputs.items()}
        d.update(lohn_berechnen(d))
        return d


class LohnabrechnungWidget(QWidget):
    """Lohnabrechnungen mit manuellen gesetzlichen Abzügen, PDF- und DATEV-Export."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.daten = lade_lohnabrechnungen()
        root = QVBoxLayout(self)
        info = QLabel("Lohnabrechnungen: Brutto, Abzüge, Netto, Arbeitgeberanteile, PDF und DATEV-Buchungsstapel. Steuer- und SV-Werte werden aus der Lohnabrechnung übernommen und nicht steuerlich berechnet.")
        info.setWordWrap(True); info.setObjectName("subTitle"); root.addWidget(info)
        row = QHBoxLayout()
        for text, slot, primary in [
            ("Neu", self.neu, True), ("Bearbeiten", self.bearbeiten, False), ("Löschen", self.loeschen, False),
            ("Abrechnung PDF", self.pdf_export, False), ("DATEV Export", self.datev_export, False),
        ]:
            b=QPushButton(text); b.clicked.connect(slot)
            if primary: b.setObjectName("primaryButton")
            row.addWidget(b)
        row.addStretch(); root.addLayout(row)
        self.table=QTableWidget(); root.addWidget(self.table,1)
        self.laden()

    def laden(self):
        headers=["Monat","Mitarbeiter","Personalnr.","Brutto","Abzüge","Netto","AG-Anteile","Arbeitgeberkosten","Kostenstelle"]
        self.table.setColumnCount(len(headers)); self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(self.daten)); self.table.setAlternatingRowColors(True)
        for r,d in enumerate(self.daten):
            calc=lohn_berechnen(d)
            vals=[d.get("abrechnungsmonat",""),d.get("mitarbeiter",""),d.get("personalnummer",""),
                  f'{calc["brutto"]:.2f}',f'{calc["abzuege"]:.2f}',f'{calc["netto"]:.2f}',
                  f'{calc["ag_anteile"]:.2f}',f'{calc["arbeitgeberkosten"]:.2f}',d.get("kostenstelle","")]
            for c,v in enumerate(vals): self.table.setItem(r,c,QTableWidgetItem(str(v)))
        for c,w in enumerate([90,200,100,100,100,100,110,130,180]): self.table.setColumnWidth(c,w)

    def neu(self):
        dlg=LohnDialog(parent=self)
        if dlg.exec()==QDialog.DialogCode.Accepted:
            self.daten.append(dlg.values()); speichere_lohnabrechnungen(self.daten); self.laden()

    def _index(self) -> int:
        return self.table.currentRow()

    def bearbeiten(self):
        i=self._index()
        if i<0: QMessageBox.information(self,"Lohnabrechnung","Bitte eine Abrechnung auswählen."); return
        dlg=LohnDialog(self.daten[i],self)
        if dlg.exec()==QDialog.DialogCode.Accepted:
            self.daten[i]=dlg.values(); speichere_lohnabrechnungen(self.daten); self.laden()

    def loeschen(self):
        i=self._index()
        if i<0: return
        if frage_ja_nein("Lohnabrechnung löschen","Ausgewählte Lohnabrechnung wirklich löschen?"):
            self.daten.pop(i); speichere_lohnabrechnungen(self.daten); self.laden()

    def pdf_export(self):
        i=self._index()
        if i<0: QMessageBox.information(self,"PDF","Bitte eine Abrechnung auswählen."); return
        d=self.daten[i]; c=lohn_berechnen(d)
        default=EXPORT_DIR / f"lohnabrechnung_{d.get('abrechnungsmonat','')}_{norm_key(d.get('mitarbeiter',''))}.pdf"
        ziel,_=QFileDialog.getSaveFileName(self,"Lohnabrechnung PDF",str(default),"PDF-Dateien (*.pdf)")
        if not ziel: return
        if not ziel.lower().endswith('.pdf'): ziel += '.pdf'
        rows=''.join(f"<tr><td>{html.escape(label)}</td><td style='text-align:right'>{html.escape(str(d.get(key,'')))}</td></tr>" for key,label in LohnDialog.FELDER if key not in {'notiz'})
        rows += f"<tr><th>Netto</th><th style='text-align:right'>{euro(c['netto'])}</th></tr><tr><th>Arbeitgeberkosten</th><th style='text-align:right'>{euro(c['arbeitgeberkosten'])}</th></tr>"
        doc=QTextDocument(); doc.setHtml(f"<h1>Lohnabrechnung</h1><h2>{html.escape(str(d.get('mitarbeiter','')))} – {html.escape(str(d.get('abrechnungsmonat','')))}</h2><table width='100%' cellspacing='4'>{rows}</table><p>{html.escape(str(d.get('notiz','')))}</p>")
        printer=QPrinter(QPrinter.PrinterMode.HighResolution); printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat); printer.setOutputFileName(ziel); doc.print_(printer)
        QMessageBox.information(self,"PDF",f"Lohnabrechnung erstellt:\n{ziel}")

    def datev_export(self):
        if not self.daten: QMessageBox.information(self,"DATEV","Keine Lohnabrechnungen vorhanden."); return
        default=EXPORT_DIR / f"DATEV_Lohn_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.csv"
        ziel,_=QFileDialog.getSaveFileName(self,"DATEV-Lohnexport",str(default),"CSV-Dateien (*.csv)")
        if not ziel: return
        if not ziel.lower().endswith('.csv'): ziel += '.csv'
        headers=["Umsatz (ohne Soll/Haben-Kz)","Soll/Haben-Kennzeichen","WKZ Umsatz","Konto","Gegenkonto (ohne BU-Schlüssel)","Belegdatum","Belegfeld 1","Buchungstext","KOST1 - Kostenstelle"]
        with open(ziel,'w',newline='',encoding='cp1252',errors='replace') as fh:
            w=csv.writer(fh,delimiter=';',quotechar='"',quoting=csv.QUOTE_MINIMAL); w.writerow(headers)
            for d in self.daten:
                c=lohn_berechnen(d); beleg=str(d.get('zahlungsdatum') or d.get('abrechnungsmonat','')).replace('-','')[-4:]
                name=str(d.get('mitarbeiter','')); monat=str(d.get('abrechnungsmonat','')); kost=str(d.get('kostenstelle',''))
                buchungen=[
                    (c['brutto'],'S','4120','1740',f'Bruttolohn {name} {monat}'),
                    (c['ag_anteile'],'S','4130','1740',f'AG-Anteile {name} {monat}'),
                    (c['netto'],'H','1740','1200',f'Nettolohn {name} {monat}'),
                    (max(0.0,c['abzuege']),'H','1740','1740',f'Abzüge Lohn {name} {monat}'),
                ]
                for betrag,sh,konto,gk,text in buchungen:
                    if betrag: w.writerow([f"{betrag:.2f}".replace('.',','),sh,'EUR',konto,gk,beleg,d.get('personalnummer',''),text,kost])
        QMessageBox.information(self,"DATEV",f"DATEV-Lohnexport erstellt:\n{ziel}\n\nKonten 4120/4130/1740/1200 sind Standardvorgaben und müssen zum Kontenrahmen des Steuerbüros passen.")


def exportiere_datev_buchungen(buchungen: list[dict[str, Any]], parent=None) -> None:
    if not buchungen:
        QMessageBox.information(parent, "DATEV", "Keine Buchungen zum Exportieren.")
        return
    default = EXPORT_DIR / f"DATEV_Buchhaltung_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.csv"
    ziel, _ = QFileDialog.getSaveFileName(parent, "DATEV-Buchungsstapel", str(default), "CSV-Dateien (*.csv)")
    if not ziel:
        return
    if not ziel.lower().endswith(".csv"):
        ziel += ".csv"
    headers = ["Umsatz (ohne Soll/Haben-Kz)", "Soll/Haben-Kennzeichen", "WKZ Umsatz", "Konto", "Gegenkonto (ohne BU-Schlüssel)", "Belegdatum", "Belegfeld 1", "Buchungstext", "KOST1 - Kostenstelle"]
    with open(ziel, "w", newline="", encoding="cp1252", errors="replace") as fh:
        writer = csv.writer(fh, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        for b in buchungen:
            typ = str(b.get("typ", ""))
            konto = "1200" if typ == "Einnahme" else "4900"
            gegenkonto = "8400" if typ == "Einnahme" else "1200"
            sh = "S" if typ == "Ausgabe" else "H"
            datum = re.sub(r"[^0-9]", "", str(b.get("datum", "")))[-4:]
            writer.writerow([
                f"{float(b.get('betrag', 0)):.2f}".replace(".", ","), sh, "EUR", konto, gegenkonto,
                datum, str(b.get("quelle", ""))[:36], str(b.get("zweck", ""))[:60], str(b.get("name", ""))[:36]
            ])
    QMessageBox.information(parent, "DATEV", f"DATEV-Buchungsstapel erstellt:\n{ziel}\n\nKonten 1200/8400/4900 sind Standardvorgaben und müssen zum Kontenrahmen des Steuerbüros passen.")




STEUER_KATEGORIEN = {
    "Miete": "Einnahmen Vermietung",
    "Nebenkosten": "Nebenkostenvorauszahlungen",
    "Betriebskosten": "Nebenkostenvorauszahlungen",
    "Heizkosten": "Heizkostenvorauszahlungen",
    "HV-Rechnung": "Erlöse Hausverwaltung",
    "Rechnung": "Instandhaltung / Fremdleistungen",
    "Grundsteuer": "Grundsteuer",
    "Versicherung": "Versicherungen",
    "Wasser": "Wasser / Abwasser",
    "Abwasser": "Wasser / Abwasser",
    "Müll": "Müllentsorgung",
    "Strom": "Allgemeinstrom",
    "Gas": "Heizung / Energie",
    "Heizung": "Heizung / Energie",
    "Hausmeister": "Hausmeister",
    "Schornsteinfeger": "Schornsteinfeger",
    "Reparatur": "Instandhaltung / Reparatur",
    "Instandhaltung": "Instandhaltung / Reparatur",
    "Zinsen": "Schuldzinsen",
    "Tilgung": "Nicht abzugsfähige Tilgung",
    "Software": "Software / EDV",
    "Telefon": "Telefon / Internet",
    "Porto": "Porto / Versand",
    "Steuerberater": "Rechts- und Beratungskosten",
}


def steuer_kategorie(buchung: dict[str, Any]) -> str:
    text = " ".join(str(buchung.get(k, "")) for k in ("kostenart", "zweck", "quelle")).lower()
    for suchwort, kategorie in STEUER_KATEGORIEN.items():
        if suchwort.lower() in text:
            return kategorie
    return "Sonstige Einnahmen" if buchung.get("typ") == "Einnahme" else "Sonstige Ausgaben"


def steuer_bereich(buchung: dict[str, Any]) -> str:
    quelle = str(buchung.get("quelle", "")).lower()
    zweck = str(buchung.get("zweck", "")).lower()
    if "hv" in quelle or "hausverwaltung" in quelle or "hv-" in zweck:
        return "Hausverwaltung"
    return "Eigenbestand"


def steuer_jahresdaten(jahr: str) -> dict[str, Any]:
    buchungen = buchhaltung_daten(jahr, "")
    kategorien: dict[str, dict[str, float]] = {}
    bereiche: dict[str, dict[str, float]] = {}
    offene: list[dict[str, Any]] = []
    beleglos: list[dict[str, Any]] = []
    for b in buchungen:
        kat = steuer_kategorie(b)
        ber = steuer_bereich(b)
        kategorien.setdefault(kat, {"Einnahme": 0.0, "Ausgabe": 0.0})
        bereiche.setdefault(ber, {"Einnahme": 0.0, "Ausgabe": 0.0})
        typ = str(b.get("typ", ""))
        betrag = float(b.get("betrag", 0) or 0)
        if typ in ("Einnahme", "Ausgabe"):
            kategorien[kat][typ] += betrag
            bereiche[ber][typ] += betrag
        status = str(b.get("status", "")).lower()
        if "offen" in status or "fällig" in status or "faellig" in status:
            offene.append(b)
        if not str(b.get("quelle", "")).strip() or not str(b.get("datum", "")).strip():
            beleglos.append(b)
    lohn = []
    for d in lade_lohnabrechnungen():
        monat = str(d.get("abrechnungsmonat", ""))
        if jahr and not monat.startswith(jahr):
            continue
        c = lohn_berechnen(d)
        lohn.append({**d, **c})
    return {"jahr": jahr, "buchungen": buchungen, "kategorien": kategorien, "bereiche": bereiche,
            "offene": offene, "beleglos": beleglos, "lohn": lohn}


def _sheet_breite(ws: Worksheet) -> None:
    for idx, col in enumerate(ws.columns, start=1):
        maxlen = 0
        for cell in col:
            maxlen = max(maxlen, len(excel_text(cell.value)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(maxlen + 2, 12), 45)


def exportiere_steuerberater_paket(jahr: str, parent=None) -> None:
    if not jahr:
        QMessageBox.information(parent, "Jahresabschluss", "Bitte ein Jahr auswählen.")
        return
    daten = steuer_jahresdaten(jahr)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    ziel_text, _ = QFileDialog.getSaveFileName(parent, "Steuerberater-Paket speichern",
        str(EXPORT_DIR / f"Steuerberater_Paket_{jahr}.zip"), "ZIP-Dateien (*.zip)")
    if not ziel_text:
        return
    ziel = Path(ziel_text)
    if ziel.suffix.lower() != ".zip":
        ziel = ziel.with_suffix(".zip")
    tmp = EXPORT_DIR / f"steuerberater_{jahr}_temp"
    if tmp.exists():
        shutil.rmtree(tmp)
    tmp.mkdir(parents=True, exist_ok=True)

    wb = Workbook(); ws = wb.active
    assert isinstance(ws, Worksheet)
    ws.title = "Jahresübersicht"
    ws.append(["Bereich", "Einnahmen", "Ausgaben", "Ergebnis"])
    for ber, vals in sorted(daten["bereiche"].items()):
        ws.append([ber, vals["Einnahme"], vals["Ausgabe"], vals["Einnahme"]-vals["Ausgabe"]])
    ws.append([]); ws.append(["Steuerkategorie", "Einnahmen", "Ausgaben", "Saldo"])
    for kat, vals in sorted(daten["kategorien"].items()):
        ws.append([kat, vals["Einnahme"], vals["Ausgabe"], vals["Einnahme"]-vals["Ausgabe"]])
    _sheet_breite(ws)

    ws2 = wb.create_sheet("Buchungen")
    ws2.append(["Datum","Typ","Bereich","Steuerkategorie","Name","Kostenart","Status","Betrag","Quelle","Zweck"])
    for b in daten["buchungen"]:
        ws2.append([b.get("datum",""), b.get("typ",""), steuer_bereich(b), steuer_kategorie(b),
                    b.get("name",""), b.get("kostenart",""), b.get("status",""), float(b.get("betrag",0) or 0),
                    b.get("quelle",""), b.get("zweck","")])
    _sheet_breite(ws2)

    ws3 = wb.create_sheet("Offene Posten")
    ws3.append(["Datum","Typ","Name","Betrag","Status","Quelle","Zweck"])
    for b in daten["offene"]:
        ws3.append([b.get("datum",""),b.get("typ",""),b.get("name",""),float(b.get("betrag",0) or 0),b.get("status",""),b.get("quelle",""),b.get("zweck","")])
    _sheet_breite(ws3)

    ws4 = wb.create_sheet("Lohn")
    ws4.append(["Monat","Mitarbeiter","Personalnummer","Brutto","Abzüge","Netto","AG-Anteile","Arbeitgeberkosten","Kostenstelle"])
    for d in daten["lohn"]:
        ws4.append([d.get("abrechnungsmonat",""),d.get("mitarbeiter",""),d.get("personalnummer",""),d.get("brutto",0),d.get("abzuege",0),d.get("netto",0),d.get("ag_anteile",0),d.get("arbeitgeberkosten",0),d.get("kostenstelle","")])
    _sheet_breite(ws4)
    xlsx = tmp / f"Jahresabschluss_{jahr}.xlsx"; wb.save(xlsx)

    # DATEV-nahe Buchungsdatei ohne Dialog
    csv_path = tmp / f"DATEV_Buchungen_{jahr}.csv"
    headers=["Umsatz (ohne Soll/Haben-Kz)","Soll/Haben-Kennzeichen","WKZ Umsatz","Konto","Gegenkonto (ohne BU-Schlüssel)","Belegdatum","Belegfeld 1","Buchungstext","KOST1 - Kostenstelle"]
    with csv_path.open("w", newline="", encoding="cp1252", errors="replace") as fh:
        w=csv.writer(fh, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL); w.writerow(headers)
        for b in daten["buchungen"]:
            typ=str(b.get("typ","")); konto="1200" if typ=="Einnahme" else "4900"; gegen="8400" if typ=="Einnahme" else "1200"; sh="H" if typ=="Einnahme" else "S"
            datum=re.sub(r"[^0-9]", "", str(b.get("datum","")))[-4:]
            w.writerow([f"{float(b.get('betrag',0) or 0):.2f}".replace(".",","),sh,"EUR",konto,gegen,datum,str(b.get("quelle",""))[:36],str(b.get("zweck",""))[:60],steuer_bereich(b)])

    (tmp / "Hinweise.txt").write_text(
        "Steuerberater-Paket für " + jahr + "\n\n"
        "Enthalten: Jahresübersicht, Buchungen, offene Posten, Lohnübersicht und DATEV-nahe CSV.\n"
        "Die Konten 1200/8400/4900 sowie steuerliche Einordnungen sind Vorschläge und müssen vom Steuerbüro geprüft werden.\n"
        "Tilgungen, Abschreibungen, private Anteile und Umsatzsteuer-Sachverhalte sind anhand der Originalbelege endgültig zu beurteilen.\n",
        encoding="utf-8")
    (tmp / "Pruefbericht.json").write_text(json.dumps({
        "jahr":jahr, "buchungen":len(daten["buchungen"]), "offene_posten":len(daten["offene"]),
        "unvollstaendige_buchungen":len(daten["beleglos"]), "lohnabrechnungen":len(daten["lohn"])
    }, indent=2, ensure_ascii=False), encoding="utf-8")

    with zipfile.ZipFile(ziel, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in tmp.rglob("*"):
            if f.is_file(): zf.write(f, f.relative_to(tmp))
    shutil.rmtree(tmp, ignore_errors=True)
    QMessageBox.information(parent, "Jahresabschluss", f"Steuerberater-Paket erstellt:\n{ziel}")


class JahresabschlussWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        root=QVBoxLayout(self)
        info=QLabel("Jahresabschluss-Vorbereitung für Eigenbestand und Hausverwaltung: Steuerkategorien, offene Posten, Lohnkosten und Steuerberater-Paket.")
        info.setWordWrap(True); info.setObjectName("subTitle"); root.addWidget(info)
        row=QHBoxLayout(); self.jahr=QComboBox(); self.jahr.addItem("")
        jahre=set()
        for b in buchhaltung_daten("", ""):
            if str(b.get("jahr","")).strip(): jahre.add(str(b.get("jahr")).strip())
        for d in lade_lohnabrechnungen():
            m=str(d.get("abrechnungsmonat",""));
            if len(m)>=4: jahre.add(m[:4])
        for j in sorted(jahre): self.jahr.addItem(j)
        laden=QPushButton("Auswerten"); laden.setObjectName("primaryButton"); laden.clicked.connect(self.laden)
        paket=QPushButton("Steuerberater-Paket ZIP"); paket.clicked.connect(lambda: exportiere_steuerberater_paket(self.jahr.currentText().strip(), self))
        row.addWidget(QLabel("Jahr:")); row.addWidget(self.jahr); row.addWidget(laden); row.addWidget(paket); row.addStretch(); root.addLayout(row)
        self.summary=QLabel("Bitte Jahr auswählen."); self.summary.setObjectName("metricValue"); root.addWidget(self.summary)
        self.tabs=QTabWidget(); root.addWidget(self.tabs,1)
        self.kat=QTableWidget(); self.ber=QTableWidget(); self.pruef=QTableWidget()
        self.tabs.addTab(self.kat,"Steuerkategorien"); self.tabs.addTab(self.ber,"Bereiche"); self.tabs.addTab(self.pruef,"Prüfung")

    def laden(self):
        jahr=self.jahr.currentText().strip()
        if not jahr: QMessageBox.information(self,"Jahresabschluss","Bitte Jahr auswählen."); return
        d=steuer_jahresdaten(jahr)
        ein=sum(v["Einnahme"] for v in d["bereiche"].values()); aus=sum(v["Ausgabe"] for v in d["bereiche"].values())
        lohn=sum(float(x.get("arbeitgeberkosten",0) or 0) for x in d["lohn"])
        self.summary.setText(f"Einnahmen {euro(ein)} | Ausgaben {euro(aus)} | Ergebnis {euro(ein-aus)} | Arbeitgeberkosten {euro(lohn)}")
        rows=sorted(d["kategorien"].items()); self.kat.setColumnCount(4); self.kat.setHorizontalHeaderLabels(["Steuerkategorie","Einnahmen","Ausgaben","Saldo"]); self.kat.setRowCount(len(rows))
        for r,(k,v) in enumerate(rows):
            for c,x in enumerate([k,euro(v["Einnahme"]),euro(v["Ausgabe"]),euro(v["Einnahme"]-v["Ausgabe"])]): self.kat.setItem(r,c,QTableWidgetItem(str(x)))
        rows2=sorted(d["bereiche"].items()); self.ber.setColumnCount(4); self.ber.setHorizontalHeaderLabels(["Bereich","Einnahmen","Ausgaben","Ergebnis"]); self.ber.setRowCount(len(rows2))
        for r,(k,v) in enumerate(rows2):
            for c,x in enumerate([k,euro(v["Einnahme"]),euro(v["Ausgabe"]),euro(v["Einnahme"]-v["Ausgabe"])]): self.ber.setItem(r,c,QTableWidgetItem(str(x)))
        checks=[("Buchungen",len(d["buchungen"])),("Offene Posten",len(d["offene"])),("Unvollständige Buchungen",len(d["beleglos"])),("Lohnabrechnungen",len(d["lohn"]))]
        self.pruef.setColumnCount(2); self.pruef.setHorizontalHeaderLabels(["Prüfpunkt","Anzahl"]); self.pruef.setRowCount(len(checks))
        for r,(k,v) in enumerate(checks): self.pruef.setItem(r,0,QTableWidgetItem(k)); self.pruef.setItem(r,1,QTableWidgetItem(str(v)))



# ========================= VERSION 10.0: FINANZCENTER =========================
FINANZCENTER_DIR = APP_DIR / "finanzcenter"
FINANZCENTER_DIR.mkdir(parents=True, exist_ok=True)


def _json_laden(path: Path, standard: Any) -> Any:
    try:
        if not path.exists():
            return standard
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError, TypeError):
        return standard


def _json_speichern(path: Path, daten: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(daten, indent=2, ensure_ascii=False), encoding="utf-8")
    tmp.replace(path)


class FinanzRegisterWidget(QWidget):
    """Kompakter JSON-basierter Stammdateneditor ohne Änderung der Excel-Struktur."""
    def __init__(self, titel: str, dateiname: str, felder: list[str], parent=None):
        super().__init__(parent)
        self.titel, self.felder = titel, felder
        self.path = FINANZCENTER_DIR / dateiname
        self.daten: list[dict[str, str]] = _json_laden(self.path, [])
        lay = QVBoxLayout(self)
        info = QLabel(f"{titel}: steuerlich relevante Zusatzdaten. Speicherung separat, Excel-Struktur bleibt unverändert.")
        info.setObjectName("subTitle"); lay.addWidget(info)
        bar=QHBoxLayout()
        for label, fn, primary in [("Neu", self.neu, True),("Bearbeiten", self.bearbeiten, False),("Löschen", self.loeschen, False),("CSV Export", self.export_csv, False)]:
            b=QPushButton(label); b.clicked.connect(fn)
            if primary: b.setObjectName("primaryButton")
            bar.addWidget(b)
        bar.addStretch(); lay.addLayout(bar)
        self.table=QTableWidget(); self.table.setAlternatingRowColors(True); lay.addWidget(self.table,1)
        self.aktualisieren()

    def aktualisieren(self):
        self.table.setColumnCount(len(self.felder)); self.table.setHorizontalHeaderLabels(self.felder)
        self.table.setRowCount(len(self.daten))
        for r,d in enumerate(self.daten):
            for c,f in enumerate(self.felder): self.table.setItem(r,c,QTableWidgetItem(str(d.get(f,""))))
        self.table.resizeColumnsToContents()

    def _dialog(self, data=None):
        dlg=EingabeDialog(self.titel, self.felder, [str((data or {}).get(f,"")) for f in self.felder])
        if dlg.exec()!=QDialog.DialogCode.Accepted: return None
        vals=dlg.values(); return {f:(vals[i] if i<len(vals) else "") for i,f in enumerate(self.felder)}

    def neu(self):
        d=self._dialog()
        if d is not None: self.daten.append(d); _json_speichern(self.path,self.daten); self.aktualisieren()

    def bearbeiten(self):
        i=self.table.currentRow()
        if i<0: QMessageBox.information(self,self.titel,"Bitte einen Datensatz auswählen."); return
        d=self._dialog(self.daten[i])
        if d is not None: self.daten[i]=d; _json_speichern(self.path,self.daten); self.aktualisieren()

    def loeschen(self):
        i=self.table.currentRow()
        if i>=0 and frage_ja_nein(self.titel,"Ausgewählten Datensatz wirklich löschen?"):
            del self.daten[i]; _json_speichern(self.path,self.daten); self.aktualisieren()

    def export_csv(self):
        if not self.daten: QMessageBox.information(self,"Export","Keine Daten vorhanden."); return
        ziel,_=QFileDialog.getSaveFileName(self,"CSV exportieren",str(EXPORT_DIR/f"{norm_key(self.titel)}_{datetime.now():%Y-%m-%d}.csv"),"CSV (*.csv)")
        if not ziel: return
        with open(ziel,"w",newline="",encoding="utf-8-sig") as fh:
            w=csv.DictWriter(fh,fieldnames=self.felder,delimiter=";"); w.writeheader(); w.writerows(self.daten)
        QMessageBox.information(self,"Export",f"Export erstellt:\n{ziel}")


class AnlagenverzeichnisWidget(FinanzRegisterWidget):
    def __init__(self,parent=None):
        super().__init__("Anlagenverzeichnis","anlagen.json",["Anlage","Bereich","Objekt","Anschaffungsdatum","Anschaffungskosten","Nutzungsdauer Jahre","AfA-Methode","Restbuchwert","Beleg/PDF","Notiz"],parent)
        btn=QPushButton("AfA neu berechnen"); btn.clicked.connect(self.afa_berechnen); self.layout().insertWidget(2,btn)
    def afa_berechnen(self):
        jahr=datetime.now().year
        for d in self.daten:
            kosten=to_float(d.get("Anschaffungskosten")); dauer=max(1,to_float(d.get("Nutzungsdauer Jahre"),1))
            try: start=int(str(d.get("Anschaffungsdatum",jahr))[:4])
            except ValueError: start=jahr
            afa=kosten/dauer; vergangen=max(0,min(dauer,jahr-start+1)); d["Restbuchwert"]=f"{max(0,kosten-afa*vergangen):.2f}"
        _json_speichern(self.path,self.daten); self.aktualisieren(); QMessageBox.information(self,"AfA","Lineare AfA-Vorschau aktualisiert. Steuerliche Prüfung bleibt erforderlich.")


class UmsatzsteuerWidget(QWidget):
    def __init__(self,parent=None):
        super().__init__(parent); lay=QVBoxLayout(self)
        info=QLabel("Umsatzsteuer-Auswertung aus den vorhandenen Buchungen. Steuersätze werden aus MwSt.-Feldern bzw. Beträgen abgeleitet."); info.setObjectName("subTitle"); lay.addWidget(info)
        bar=QHBoxLayout(); self.jahr=QComboBox(); self.jahr.addItems([str(y) for y in range(datetime.now().year-8,datetime.now().year+2)]); self.jahr.setCurrentText(str(datetime.now().year))
        b=QPushButton("Berechnen"); b.setObjectName("primaryButton"); b.clicked.connect(self.laden); bar.addWidget(QLabel("Jahr:")); bar.addWidget(self.jahr); bar.addWidget(b); bar.addStretch(); lay.addLayout(bar)
        self.table=QTableWidget(); lay.addWidget(self.table,1); self.summe=QLabel(); self.summe.setObjectName("metricValue"); lay.addWidget(self.summe); self.laden()
    def laden(self):
        jahr=self.jahr.currentText(); gruppen={"0 %":[0,0],"7 %":[0,0],"19 %":[0,0]}
        for b in buchhaltung_daten(jahr,""):
            betrag=abs(to_float(b.get("betrag"))); art=str(b.get("art","")).lower(); text=" ".join(str(v) for v in b.values()).lower()
            satz="7 %" if "7 %" in text or "7%" in text else "0 %" if "steuerfrei" in text or "0 %" in text else "19 %"
            steuer=betrag*(7/107 if satz=="7 %" else 19/119 if satz=="19 %" else 0)
            idx=0 if "einnah" in art else 1; gruppen[satz][idx]+=steuer
        self.table.setColumnCount(4); self.table.setHorizontalHeaderLabels(["Steuersatz","Umsatzsteuer","Vorsteuer","Zahllast"]); self.table.setRowCount(3)
        ust=vor=0
        for r,(satz,(u,v)) in enumerate(gruppen.items()):
            ust+=u; vor+=v
            for c,x in enumerate([satz,euro(u),euro(v),euro(u-v)]): self.table.setItem(r,c,QTableWidgetItem(str(x)))
        self.summe.setText(f"Voraussichtliche Zahllast: {euro(ust-vor)}")


class BelegAssistentWidget(QWidget):
    def __init__(self,parent=None):
        super().__init__(parent); lay=QVBoxLayout(self)
        info=QLabel("Beleg-Assistent: liest PDF-Text lokal aus und schlägt Lieferant, Betrag, Rechnungsnummer, Kostenart und DATEV-Konto vor."); info.setObjectName("subTitle"); lay.addWidget(info)
        bar=QHBoxLayout(); b=QPushButton("PDF analysieren"); b.setObjectName("primaryButton"); b.clicked.connect(self.analysieren); bar.addWidget(b); bar.addStretch(); lay.addLayout(bar)
        self.result=QTableWidget(); lay.addWidget(self.result,1)
    def analysieren(self):
        if PdfReader is None: QMessageBox.warning(self,"PDF","pypdf ist nicht installiert."); return
        path,_=QFileDialog.getOpenFileName(self,"Beleg auswählen","","PDF (*.pdf)")
        if not path:return
        try:
            reader=PdfReader(path); text="\n".join((p.extract_text() or "") for p in reader.pages[:8])
        except Exception as exc: QMessageBox.warning(self,"PDF",str(exc)); return
        lower=text.lower(); nummer=re.search(r"(?:rechnungs(?:nummer|nr\.?|nr)|belegnummer)\s*[:#]?\s*([\w\-/]+)",text,re.I)
        amounts=re.findall(r"(?<!\d)(\d{1,3}(?:\.\d{3})*,\d{2}|\d+[.,]\d{2})\s*(?:€|eur)",text,re.I)
        betrag=max((to_float(a) for a in amounts),default=0.0)
        regeln=[("schornstein","Schornsteinfeger","4200"),("versicherung","Versicherung","4360"),("grundsteuer","Grundsteuer","4510"),("wasser","Wasser/Abwasser","4240"),("strom","Strom","4240"),("reparatur","Instandhaltung","4805"),("wartung","Wartung","4805"),("software","Software","4964")]
        kostenart,konto="Sonstige Kosten","4900"
        for key,art,kto in regeln:
            if key in lower: kostenart,konto=art,kto; break
        lines=[x.strip() for x in text.splitlines() if x.strip()]; lieferant=lines[0][:100] if lines else Path(path).stem
        daten=[("PDF",path),("Lieferant (Vorschlag)",lieferant),("Rechnungsnummer",nummer.group(1) if nummer else ""),("Bruttobetrag",euro(betrag)),("Kostenart",kostenart),("DATEV-Konto Vorschlag",konto),("Hinweis","Vorschläge vor Buchung fachlich prüfen")]
        self.result.setColumnCount(2); self.result.setHorizontalHeaderLabels(["Feld","Erkannt/Vorschlag"]); self.result.setRowCount(len(daten))
        for r,(k,v) in enumerate(daten): self.result.setItem(r,0,QTableWidgetItem(k)); self.result.setItem(r,1,QTableWidgetItem(str(v)))
        self.result.resizeColumnsToContents()


class SteuercenterWidget(QWidget):
    def __init__(self,parent=None):
        super().__init__(parent); lay=QVBoxLayout(self)
        info=QLabel("Steuercenter: EÜR-nahe Übersicht, Werbungskosten, Lohn, AfA, Darlehenszinsen und Jahresakte."); info.setObjectName("subTitle"); lay.addWidget(info)
        bar=QHBoxLayout(); self.jahr=QComboBox(); self.jahr.addItems([str(y) for y in range(datetime.now().year-8,datetime.now().year+2)]); self.jahr.setCurrentText(str(datetime.now().year))
        b=QPushButton("Auswertung"); b.setObjectName("primaryButton"); b.clicked.connect(self.laden); paket=QPushButton("Komplette Jahresakte ZIP"); paket.clicked.connect(self.paket)
        bar.addWidget(QLabel("Steuerjahr:")); bar.addWidget(self.jahr); bar.addWidget(b); bar.addWidget(paket); bar.addStretch(); lay.addLayout(bar)
        self.table=QTableWidget(); lay.addWidget(self.table,1); self.laden()
    def laden(self):
        jahr=self.jahr.currentText(); b=buchhaltung_daten(jahr,""); ein=sum(abs(to_float(x.get("betrag"))) for x in b if "einnah" in str(x.get("typ", x.get("art", ""))).lower()); aus=sum(abs(to_float(x.get("betrag"))) for x in b if "ausgab" in str(x.get("typ", x.get("art", ""))).lower())
        lohn=sum(to_float(x.get("arbeitgeberkosten")) for x in lade_lohnabrechnungen() if jahr in str(x.get("abrechnungsmonat","")))
        anlagen=_json_laden(FINANZCENTER_DIR/"anlagen.json",[]); afa=sum(to_float(x.get("Anschaffungskosten"))/max(1,to_float(x.get("Nutzungsdauer Jahre"),1)) for x in anlagen)
        darlehen=_json_laden(FINANZCENTER_DIR/"darlehen.json",[]); zinsen=sum(to_float(x.get("Zinsen Jahr")) for x in darlehen if not x.get("Jahr") or str(x.get("Jahr"))==jahr)
        rows=[("Einnahmen",ein),("Ausgaben laufend",aus),("Arbeitgeberkosten",lohn),("AfA-Vorschau",afa),("Darlehenszinsen",zinsen),("Vorläufiges Ergebnis",ein-aus-lohn-afa-zinsen)]
        self.table.setColumnCount(2); self.table.setHorizontalHeaderLabels(["Position","Betrag"]); self.table.setRowCount(len(rows))
        for r,(k,v) in enumerate(rows): self.table.setItem(r,0,QTableWidgetItem(k)); self.table.setItem(r,1,QTableWidgetItem(euro(v)))
    def paket(self):
        exportiere_steuerberater_paket(self.jahr.currentText(),self)



# -----------------------------------------------------------------------------
# Professionelles Übergabecenter für Lohnbüro und Steuerberater
# Speicherung ausschließlich in JSON unter finanzcenter; die Excel-Schemata
# SCHEMA und DATA_FILES bleiben unverändert.
# -----------------------------------------------------------------------------
PERSONALSTAMM_DATEI = FINANZCENTER_DIR / "personalstamm.json"
LOHNBEWEGUNGEN_DATEI = FINANZCENTER_DIR / "lohnbewegungen.json"
STAMMDATENAENDERUNGEN_DATEI = FINANZCENTER_DIR / "stammdatenaenderungen.json"


def _json_liste_laden(path: Path) -> list[dict[str, Any]]:
    daten = _json_laden(path, [])
    return daten if isinstance(daten, list) else []


def _json_liste_speichern(path: Path, daten: list[dict[str, Any]]) -> None:
    _json_speichern(path, daten)


class DatensatzDialog(QDialog):
    """Generischer, scrollbarer JSON-Datensatzdialog ohne Excel-Änderung."""
    def __init__(self, titel: str, felder: list[tuple[str, str]], daten: dict[str, Any] | None = None, parent=None):
        super().__init__(parent)
        self.setWindowTitle(titel)
        self.resize(760, 780)
        self.felder = felder
        self.inputs: dict[str, QWidget] = {}
        daten = daten or {}
        root = QVBoxLayout(self)
        title = QLabel(titel); title.setObjectName("pageTitle"); root.addWidget(title)
        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        content = QWidget(); form = QFormLayout(content)
        multiline = {"notiz", "aenderung", "hinweis", "beschreibung"}
        for key, label in felder:
            if key in multiline:
                widget = QTextEdit(str(daten.get(key, ""))); widget.setMinimumHeight(90)
            elif key in {"status", "beschaeftigungsart", "abrechnungsstatus"}:
                widget = QComboBox()
                werte = ["", "Aktiv", "Inaktiv", "Vorbereitet", "Geprüft", "Freigegeben", "Übergeben", "Abgerechnet",
                         "Vollzeit", "Teilzeit", "Minijob", "Kurzfristig", "Ausbildung", "Werkstudent"]
                widget.addItems(werte); widget.setCurrentText(str(daten.get(key, "")))
            else:
                widget = QLineEdit(str(daten.get(key, "")))
                if key in {"steuer_id", "sv_nummer", "iban"}: widget.setEchoMode(QLineEdit.EchoMode.PasswordEchoOnEdit)
            self.inputs[key] = widget
            form.addRow(label + ":", widget)
        scroll.setWidget(content); root.addWidget(scroll, 1)
        buttons = QHBoxLayout(); buttons.addStretch()
        ok = QPushButton("Speichern"); ok.setObjectName("primaryButton"); ok.clicked.connect(self.accept)
        cancel = QPushButton("Abbrechen"); cancel.clicked.connect(self.reject)
        buttons.addWidget(ok); buttons.addWidget(cancel); root.addLayout(buttons)

    def values(self) -> dict[str, Any]:
        result: dict[str, Any] = {}
        for key, _label in self.felder:
            widget = self.inputs[key]
            if isinstance(widget, QTextEdit): result[key] = widget.toPlainText().strip()
            elif isinstance(widget, QComboBox): result[key] = widget.currentText().strip()
            else: result[key] = widget.text().strip()
        return result


class JsonRegisterWidget(QWidget):
    def __init__(self, titel: str, path: Path, felder: list[tuple[str, str]], parent=None):
        super().__init__(parent)
        self.titel, self.path, self.felder = titel, path, felder
        self.daten = _json_liste_laden(path)
        root = QVBoxLayout(self)
        info = QLabel(f"{titel}: lokale Zusatzdaten für die strukturierte Übergabe. Keine Änderung der Excel-Tabellen.")
        info.setWordWrap(True); info.setObjectName("subTitle"); root.addWidget(info)
        row = QHBoxLayout()
        for text, slot, primary in [("Neu", self.neu, True), ("Bearbeiten", self.bearbeiten, False), ("Löschen", self.loeschen, False), ("Excel Export", self.excel_export, False)]:
            b = QPushButton(text); b.clicked.connect(slot)
            if primary: b.setObjectName("primaryButton")
            row.addWidget(b)
        row.addStretch(); root.addLayout(row)
        self.table = QTableWidget(); root.addWidget(self.table, 1)
        self.laden()

    def laden(self):
        labels = [label for _key, label in self.felder]
        self.table.setColumnCount(len(labels)); self.table.setHorizontalHeaderLabels(labels)
        self.table.setRowCount(len(self.daten)); self.table.setAlternatingRowColors(True)
        for r, d in enumerate(self.daten):
            for c, (key, _label) in enumerate(self.felder):
                wert = str(d.get(key, ""))
                if key in {"steuer_id", "sv_nummer", "iban"} and wert: wert = "••••" + wert[-4:]
                self.table.setItem(r, c, QTableWidgetItem(wert))
        self.table.resizeColumnsToContents()

    def _index(self): return self.table.currentRow()
    def neu(self):
        dlg = DatensatzDialog(self.titel, self.felder, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.daten.append(dlg.values()); _json_liste_speichern(self.path, self.daten); self.laden()
    def bearbeiten(self):
        i = self._index()
        if i < 0: QMessageBox.information(self, self.titel, "Bitte einen Datensatz auswählen."); return
        dlg = DatensatzDialog(self.titel, self.felder, self.daten[i], self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.daten[i] = dlg.values(); _json_liste_speichern(self.path, self.daten); self.laden()
    def loeschen(self):
        i = self._index()
        if i >= 0 and frage_ja_nein(self.titel, "Ausgewählten Datensatz wirklich löschen?"):
            self.daten.pop(i); _json_liste_speichern(self.path, self.daten); self.laden()
    def excel_export(self):
        if not self.daten: QMessageBox.information(self, "Export", "Keine Daten vorhanden."); return
        ziel, _ = QFileDialog.getSaveFileName(self, "Excel Export", str(EXPORT_DIR / (norm_key(self.titel) + ".xlsx")), "Excel (*.xlsx)")
        if not ziel: return
        if not ziel.lower().endswith('.xlsx'): ziel += '.xlsx'
        wb = Workbook(); ws = wb.active; ws.title = self.titel[:31]
        ws.append([label for _key, label in self.felder])
        for d in self.daten: ws.append([d.get(key, "") for key, _label in self.felder])
        for i, _ in enumerate(ws.columns, 1): ws.column_dimensions[get_column_letter(i)].width = 24
        wb.save(ziel); QMessageBox.information(self, "Export", f"Export erstellt:\n{ziel}")


PERSONAL_FELDER = [
    ("personalnummer", "Personalnummer"), ("name", "Name"), ("geburtsdatum", "Geburtsdatum"),
    ("adresse", "Adresse"), ("steuer_id", "Steuer-ID"), ("steuerklasse", "Steuerklasse"),
    ("kinderfreibetrag", "Kinderfreibetrag"), ("konfession", "Konfession"),
    ("sv_nummer", "Sozialversicherungsnummer"), ("krankenkasse", "Krankenkasse"),
    ("beschaeftigungsart", "Beschäftigungsart"), ("eintritt", "Eintritt"), ("austritt", "Austritt"),
    ("wochenstunden", "Wochenstunden"), ("stundenlohn", "Stundenlohn"), ("monatsgehalt", "Monatsgehalt"),
    ("iban", "IBAN"), ("kostenstelle", "Kostenstelle / Objekt"), ("status", "Status"), ("notiz", "Notiz")]

BEWEGUNGS_FELDER = [
    ("abrechnungsmonat", "Abrechnungsmonat YYYY-MM"), ("personalnummer", "Personalnummer"),
    ("mitarbeiter", "Mitarbeiter"), ("sollstunden", "Sollstunden"), ("iststunden", "Iststunden"),
    ("ueberstunden", "Überstunden"), ("urlaubstage", "Urlaubstage"), ("krankheitstage", "Krankheitstage"),
    ("zuschlaege", "Zuschläge EUR"), ("praemie", "Prämie EUR"), ("sachbezug", "Sachbezug EUR"),
    ("fahrtkosten", "Fahrtkostenzuschuss EUR"), ("einmalzahlung", "Einmalzahlung EUR"),
    ("stammdaten_geaendert", "Stammdaten geändert Ja/Nein"), ("abrechnungsstatus", "Abrechnungsstatus"),
    ("notiz", "Notiz")]

AENDERUNGS_FELDER = [
    ("gueltig_ab", "Gültig ab"), ("personalnummer", "Personalnummer"), ("mitarbeiter", "Mitarbeiter"),
    ("feld", "Geändertes Feld"), ("alt", "Alter Wert"), ("neu", "Neuer Wert"),
    ("gemeldet_am", "Gemeldet am"), ("status", "Status"), ("aenderung", "Beschreibung / Hinweis")]


class LohnbueroCenterWidget(QWidget):
    """Vorbereitung und strukturierte Übergabe an externe/interne Lohnbuchhaltung."""
    def __init__(self, parent=None):
        super().__init__(parent)
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)

        # Der komplette Center-Inhalt liegt in einem echten Scrollbereich. Dadurch
        # bleiben Kopfzeile, Aktionen und Register auch auf kleineren Bildschirmen
        # sowie bei hoher Windows-Anzeigeskalierung vollständig erreichbar.
        self.scroll_area = QScrollArea(self)
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setFrameShape(QFrame.Shape.NoFrame)
        self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        content = QWidget()
        content.setMinimumWidth(760)
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(18, 18, 18, 18)
        content_layout.setSpacing(12)

        title = QLabel("Lohnbüro-Center")
        title.setObjectName("pageTitle")
        content_layout.addWidget(title)

        info = QLabel("Personalstamm, Monatsbewegungen, Änderungen und vorhandene Lohnabrechnungen werden zu einem prüfbaren Übergabepaket zusammengeführt. Gesetzliche Lohnberechnungen erfolgen weiterhin im eingesetzten Lohnsystem.")
        info.setWordWrap(True)
        info.setObjectName("subTitle")
        content_layout.addWidget(info)

        bar = QHBoxLayout()
        self.monat = QLineEdit(datetime.now().strftime('%Y-%m'))
        self.monat.setMaximumWidth(120)
        paket = QPushButton("Lohnbüro-Paket erstellen")
        paket.setObjectName("primaryButton")
        paket.clicked.connect(self.export_paket)
        pruefen = QPushButton("Vollständigkeit prüfen")
        pruefen.clicked.connect(self.pruefen)
        bar.addWidget(QLabel("Monat:"))
        bar.addWidget(self.monat)
        bar.addWidget(pruefen)
        bar.addWidget(paket)
        bar.addStretch()
        content_layout.addLayout(bar)

        self.tabs = QTabWidget()
        self.tabs.setMinimumHeight(560)
        content_layout.addWidget(self.tabs, 1)

        self.personal = JsonRegisterWidget("Personalstamm", PERSONALSTAMM_DATEI, PERSONAL_FELDER, self)
        self.bewegungen = JsonRegisterWidget("Monatsbewegungen", LOHNBEWEGUNGEN_DATEI, BEWEGUNGS_FELDER, self)
        self.aenderungen = JsonRegisterWidget("Stammdatenänderungen", STAMMDATENAENDERUNGEN_DATEI, AENDERUNGS_FELDER, self)
        self.tabs.addTab(self.personal, "Personalstamm")
        self.tabs.addTab(self.bewegungen, "Monatsbewegungen")
        self.tabs.addTab(self.aenderungen, "Änderungen")

        self.pruef_table = QTableWidget()
        self.pruef_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.pruef_table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.tabs.addTab(self.pruef_table, "Prüfprotokoll")

        # Tabellen behalten zusätzlich ihre eigenen Scrollleisten für viele
        # Mitarbeiter und breite Datensätze.
        for register in (self.personal, self.bewegungen, self.aenderungen):
            register.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
            register.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        self.scroll_area.setWidget(content)
        root.addWidget(self.scroll_area)

    def _pruefungen(self, monat: str) -> list[tuple[str, str, str]]:
        personal = _json_liste_laden(PERSONALSTAMM_DATEI); beweg = _json_liste_laden(LOHNBEWEGUNGEN_DATEI)
        lohn = [x for x in lade_lohnabrechnungen() if str(x.get('abrechnungsmonat','')) == monat]
        result = []
        aktive = [p for p in personal if str(p.get('status','')).lower() not in {'inaktiv','ausgetreten'}]
        for p in aktive:
            pn = str(p.get('personalnummer','')).strip(); name = str(p.get('name','')).strip()
            fehlend = [label for key, label in PERSONAL_FELDER if key in {'personalnummer','name','steuer_id','sv_nummer','krankenkasse','eintritt','beschaeftigungsart'} and not str(p.get(key,'')).strip()]
            result.append((name or pn, "Personalstamm", "OK" if not fehlend else "Fehlt: " + ", ".join(fehlend)))
            hat_bewegung = any(str(x.get('abrechnungsmonat','')) == monat and (str(x.get('personalnummer','')) == pn or str(x.get('mitarbeiter','')) == name) for x in beweg)
            result.append((name or pn, "Monatsbewegung", "Vorhanden" if hat_bewegung else "Nicht erfasst"))
            hat_lohn = any(str(x.get('personalnummer','')) == pn or str(x.get('mitarbeiter','')) == name for x in lohn)
            result.append((name or pn, "Lohnabrechnung", "Vorhanden" if hat_lohn else "Noch nicht vorhanden"))
        if not aktive: result.append(("Gesamt", "Personalstamm", "Keine aktiven Mitarbeiter erfasst"))
        return result

    def pruefen(self):
        monat = self.monat.text().strip(); rows = self._pruefungen(monat)
        self.pruef_table.setColumnCount(3); self.pruef_table.setHorizontalHeaderLabels(["Mitarbeiter", "Prüfung", "Ergebnis"]); self.pruef_table.setRowCount(len(rows))
        for r, values in enumerate(rows):
            for c, value in enumerate(values): self.pruef_table.setItem(r,c,QTableWidgetItem(value))
        self.pruef_table.resizeColumnsToContents(); self.tabs.setCurrentWidget(self.pruef_table)

    def export_paket(self):
        monat = self.monat.text().strip()
        if not re.fullmatch(r"\d{4}-\d{2}", monat): QMessageBox.warning(self, "Monat", "Bitte Monat als YYYY-MM eingeben."); return
        ziel, _ = QFileDialog.getSaveFileName(self, "Lohnbüro-Paket", str(EXPORT_DIR / f"Lohnbuero_Paket_{monat}.zip"), "ZIP (*.zip)")
        if not ziel: return
        if not ziel.lower().endswith('.zip'): ziel += '.zip'
        import tempfile
        with tempfile.TemporaryDirectory() as td:
            p = Path(td); personal = _json_liste_laden(PERSONALSTAMM_DATEI)
            beweg = [x for x in _json_liste_laden(LOHNBEWEGUNGEN_DATEI) if str(x.get('abrechnungsmonat','')) == monat]
            aend = [x for x in _json_liste_laden(STAMMDATENAENDERUNGEN_DATEI) if monat[:4] in str(x.get('gueltig_ab',''))]
            lohn = [x for x in lade_lohnabrechnungen() if str(x.get('abrechnungsmonat','')) == monat]
            pruef = self._pruefungen(monat)
            for filename, data in [("personalstamm.json", personal),("monatsbewegungen.json", beweg),("stammdatenaenderungen.json", aend),("lohnabrechnungen.json", lohn)]:
                (p/filename).write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding='utf-8')
            wb=Workbook(); ws=wb.active; ws.title="Personalstamm"; ws.append([label for _k,label in PERSONAL_FELDER]); [ws.append([d.get(k,'') for k,_l in PERSONAL_FELDER]) for d in personal]
            ws2=wb.create_sheet("Monatsbewegungen"); ws2.append([label for _k,label in BEWEGUNGS_FELDER]); [ws2.append([d.get(k,'') for k,_l in BEWEGUNGS_FELDER]) for d in beweg]
            ws3=wb.create_sheet("Lohnabrechnungen"); headers=sorted({k for d in lohn for k in d.keys()}); ws3.append(headers); [ws3.append([d.get(k,'') for k in headers]) for d in lohn]
            ws4=wb.create_sheet("Prüfprotokoll"); ws4.append(["Mitarbeiter","Prüfung","Ergebnis"]); [ws4.append(list(r)) for r in pruef]
            wb.save(p/f"Lohnbuero_Uebergabe_{monat}.xlsx")
            with open(p/f"Monatsbewegungen_{monat}.csv",'w',newline='',encoding='utf-8-sig') as fh:
                w=csv.DictWriter(fh, fieldnames=[k for k,_l in BEWEGUNGS_FELDER], delimiter=';'); w.writeheader(); w.writerows(beweg)
            (p/"HINWEISE.txt").write_text("Übergabepaket für die Lohnbuchhaltung. Steuer- und Sozialversicherungsberechnung sowie Meldungen müssen im zugelassenen Lohnsystem geprüft und durchgeführt werden. Sensible personenbezogene Daten geschützt übertragen.",encoding='utf-8')
            with zipfile.ZipFile(ziel,'w',zipfile.ZIP_DEFLATED) as zf:
                for f in p.iterdir(): zf.write(f,f.name)
        QMessageBox.information(self,"Lohnbüro",f"Übergabepaket erstellt:\n{ziel}\n\nBitte wegen personenbezogener Daten nur verschlüsselt/sicher übertragen.")


class SteuerberaterUebergabeWidget(QWidget):
    """Erweiterte, prüfbare Übergabe für Steuerberater und Jahresabschluss."""
    def __init__(self,parent=None):
        super().__init__(parent); root=QVBoxLayout(self)
        title=QLabel("Steuerberater-Übergabecenter"); title.setObjectName("pageTitle"); root.addWidget(title)
        info=QLabel("Erstellt ein getrenntes Jahrespaket mit Buchungsjournal, Objekt-/Bereichsauswertung, Lohnsummen, Anlagen, Darlehen, Umsatzsteuer und Prüfhinweisen.")
        info.setWordWrap(True); info.setObjectName("subTitle"); root.addWidget(info)
        bar=QHBoxLayout(); self.jahr=QComboBox(); self.jahr.addItems([str(y) for y in range(datetime.now().year-8,datetime.now().year+2)]); self.jahr.setCurrentText(str(datetime.now().year))
        laden=QPushButton("Prüfen"); laden.clicked.connect(self.pruefen); paket=QPushButton("Steuerberater-Paket erstellen"); paket.setObjectName("primaryButton"); paket.clicked.connect(self.export)
        bar.addWidget(QLabel("Jahr:")); bar.addWidget(self.jahr); bar.addWidget(laden); bar.addWidget(paket); bar.addStretch(); root.addLayout(bar)
        self.table=QTableWidget(); root.addWidget(self.table,1); self.pruefen()
    def _daten(self):
        jahr=self.jahr.currentText(); buch=buchhaltung_daten(jahr,''); lohn=[x for x in lade_lohnabrechnungen() if jahr in str(x.get('abrechnungsmonat',''))]
        ein=buchhaltung_summe(buch,'Einnahme'); aus=buchhaltung_summe(buch,'Ausgabe'); ag=sum(lohn_berechnen(x)['arbeitgeberkosten'] for x in lohn)
        anlagen=_json_liste_laden(FINANZCENTER_DIR/'anlagen.json'); darlehen=_json_liste_laden(FINANZCENTER_DIR/'darlehen.json')
        return jahr,buch,lohn,anlagen,darlehen,ein,aus,ag
    def pruefen(self):
        jahr,buch,lohn,anlagen,darlehen,ein,aus,ag=self._daten()
        rows=[("Buchungen",len(buch),"OK" if buch else "Keine Buchungen"),("Einnahmen",euro(ein),""),("Ausgaben",euro(aus),""),("Lohnabrechnungen",len(lohn),""),("Arbeitgeberkosten",euro(ag),""),("Anlagen",len(anlagen),"AfA fachlich prüfen"),("Darlehen",len(darlehen),"Zinsbescheinigungen beifügen"),("Vorläufiges Ergebnis",euro(ein-aus-ag),"Keine Steuererklärung")]
        self.table.setColumnCount(3); self.table.setHorizontalHeaderLabels(["Bereich","Wert","Hinweis"]); self.table.setRowCount(len(rows))
        for r,row in enumerate(rows):
            for c,v in enumerate(row): self.table.setItem(r,c,QTableWidgetItem(str(v)))
        self.table.resizeColumnsToContents()
    def export(self):
        jahr,buch,lohn,anlagen,darlehen,ein,aus,ag=self._daten()
        ziel,_=QFileDialog.getSaveFileName(self,"Steuerberater-Paket",str(EXPORT_DIR/f"Steuerberater_Uebergabe_{jahr}.zip"),"ZIP (*.zip)")
        if not ziel:return
        if not ziel.lower().endswith('.zip'): ziel += '.zip'
        import tempfile
        with tempfile.TemporaryDirectory() as td:
            p=Path(td); wb=Workbook(); ws=wb.active; ws.title="Buchungsjournal"
            bh=["Quelle","Datum","Jahr","Monat","Typ","Name","Kostenart","Status","Betrag","Zweck"]; ws.append(bh)
            for x in buch: ws.append([x.get(k.lower(),'') for k in bh])
            ws2=wb.create_sheet("Lohnsummen"); ws2.append(["Monat","Mitarbeiter","Personalnummer","Brutto","Netto","AG-Anteile","Arbeitgeberkosten","Kostenstelle"])
            for d in lohn:
                c=lohn_berechnen(d); ws2.append([d.get('abrechnungsmonat',''),d.get('mitarbeiter',''),d.get('personalnummer',''),c['brutto'],c['netto'],c['ag_anteile'],c['arbeitgeberkosten'],d.get('kostenstelle','')])
            ws3=wb.create_sheet("Jahresübersicht"); ws3.append(["Position","Betrag"]); [ws3.append(r) for r in [("Einnahmen",ein),("Ausgaben",aus),("Arbeitgeberkosten",ag),("Vorläufiges Ergebnis",ein-aus-ag)]]
            for title,data in [("Anlagen",anlagen),("Darlehen",darlehen)]:
                s=wb.create_sheet(title); keys=sorted({k for d in data for k in d.keys()}); s.append(keys); [s.append([d.get(k,'') for k in keys]) for d in data]
            wb.save(p/f"Steuerberater_Uebergabe_{jahr}.xlsx")
            with open(p/f"Buchungsjournal_{jahr}.csv",'w',newline='',encoding='utf-8-sig') as fh:
                keys=['quelle','datum','jahr','monat','typ','name','kostenart','status','betrag','zweck']; w=csv.DictWriter(fh,fieldnames=keys,delimiter=';'); w.writeheader(); w.writerows([{k:x.get(k,'') for k in keys} for x in buch])
            (p/'PRUEF_HINWEISE.txt').write_text("Vorläufige Arbeitsunterlagen für den Steuerberater. Kontenzuordnung, Steuerschlüssel, Umsatzsteuerpflicht, AfA, Privatanteile und Abgrenzungen müssen fachlich geprüft werden. Dieses Paket ersetzt weder Buchführung noch Steuererklärung durch befugte Personen.",encoding='utf-8')
            with zipfile.ZipFile(ziel,'w',zipfile.ZIP_DEFLATED) as zf:
                for f in p.iterdir(): zf.write(f,f.name)
        QMessageBox.information(self,"Steuerberater",f"Übergabepaket erstellt:\n{ziel}")

class BuchhaltungSeite(QWidget):
    """Eigenes Unterprogramm für Buchhaltung, gelistet und berechnet aus bestehenden Tabellen."""

    def __init__(self):
        super().__init__()
        self.buchungen: list[dict[str, Any]] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 22)
        root.setSpacing(14)

        title = QLabel("Buchhaltung")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel("Finanzcenter Professional: Buchhaltung, Lohn, DATEV, Anlagen/AfA, Darlehen, Umsatzsteuer, Bank, Beleganalyse und Steuerjahresabschluss.")
        info.setObjectName("subTitle")
        root.addWidget(info)

        filter_row = QHBoxLayout()
        self.jahr = QComboBox()
        self.monat = QComboBox()
        self.jahr.addItem("")
        self.monat.addItem("")
        self.monat.addItems([str(i).zfill(2) for i in range(1, 13)])

        refresh = QPushButton("Aktualisieren")
        refresh.setObjectName("primaryButton")
        refresh.clicked.connect(self.laden)

        export = QPushButton("Excel Export")
        export.clicked.connect(self.export_excel)
        datev = QPushButton("DATEV Export")
        datev.clicked.connect(lambda: exportiere_datev_buchungen(self.buchungen, self))

        filter_row.addWidget(QLabel("Jahr:"))
        filter_row.addWidget(self.jahr)
        filter_row.addWidget(QLabel("Monat:"))
        filter_row.addWidget(self.monat)
        filter_row.addWidget(refresh)
        filter_row.addWidget(export)
        filter_row.addWidget(datev)
        filter_row.addStretch()
        root.addLayout(filter_row)

        self.kpi_grid = QGridLayout()
        self.kpi_grid.setSpacing(14)
        root.addLayout(self.kpi_grid)

        self.tabs = QTabWidget()
        root.addWidget(self.tabs, 1)

        self.table = QTableWidget()
        self.table_kostenarten = QTableWidget()
        self.table_offen = QTableWidget()
        self.chart = QGraphicsView()
        self.chart_scene = QGraphicsScene()
        self.chart.setScene(self.chart_scene)

        self.tabs.addTab(self.table, "Buchungen")
        self.tabs.addTab(self.table_kostenarten, "Kostenarten")
        self.tabs.addTab(self.table_offen, "Offene Posten")
        self.tabs.addTab(self.chart, "Diagramm")
        self.lohn_widget = LohnabrechnungWidget(self)
        self.tabs.addTab(self.lohn_widget, "Lohnabrechnungen")
        self.lohnbuero_widget = LohnbueroCenterWidget(self)
        self.tabs.addTab(self.lohnbuero_widget, "Lohnbüro-Center")
        self.steuerberater_uebergabe_widget = SteuerberaterUebergabeWidget(self)
        self.tabs.addTab(self.steuerberater_uebergabe_widget, "Steuerberater-Übergabe")
        self.jahresabschluss_widget = JahresabschlussWidget(self)
        self.tabs.addTab(self.jahresabschluss_widget, "Jahresabschluss / Steuerberater")
        self.konten_widget = FinanzRegisterWidget("Kontenplan", "kontenplan.json", ["Kontonummer", "Kontoname", "Kontotyp", "Bereich", "DATEV-Konto", "Steuerschlüssel", "Aktiv", "Notiz"], self)
        self.tabs.addTab(self.konten_widget, "Kontenplan")
        self.anlagen_widget = AnlagenverzeichnisWidget(self)
        self.tabs.addTab(self.anlagen_widget, "Anlagen / AfA")
        self.darlehen_widget = FinanzRegisterWidget("Darlehensverwaltung", "darlehen.json", ["Jahr", "Bank", "Darlehensnummer", "Objekt", "Ursprungsbetrag", "Zinssatz", "Tilgung Jahr", "Zinsen Jahr", "Restschuld", "Sondertilgung", "Zinsbescheinigung PDF", "Notiz"], self)
        self.tabs.addTab(self.darlehen_widget, "Darlehen")
        self.ust_widget = UmsatzsteuerWidget(self)
        self.tabs.addTab(self.ust_widget, "Umsatzsteuer")
        self.bank_widget = FinanzRegisterWidget("Bankkonten / Salden", "bankkonten.json", ["Kontoname", "IBAN", "Bank", "Bereich", "Objekt", "Anfangssaldo", "Aktueller Saldo", "Stand", "Notiz"], self)
        self.tabs.addTab(self.bank_widget, "Bank / Konten")
        self.beleg_widget = BelegAssistentWidget(self)
        self.tabs.addTab(self.beleg_widget, "Beleg-Assistent")
        self.steuer_widget = SteuercenterWidget(self)
        self.tabs.addTab(self.steuer_widget, "Steuercenter")

        self._jahre_fuellen()
        self.laden()

    def _jahre_fuellen(self) -> None:
        jahre = set()
        for row in DATA.get("Zahlungen", []):
            if len(row) > 7 and str(row[7]).strip():
                jahre.add(str(row[7]).strip())
            elif len(row) > 0:
                jahr = buchhaltung_jahr_aus_text(row[0])
                if jahr:
                    jahre.add(jahr)

        for row in DATA.get("Rechnungen", []):
            if len(row) > 3:
                jahr = buchhaltung_jahr_aus_text(row[3])
                if jahr:
                    jahre.add(jahr)

        for row in DATA.get("HV-Rechnungen", []):
            if len(row) > 0:
                jahr = buchhaltung_jahr_aus_text(row[0])
                if jahr:
                    jahre.add(jahr)

        for jahr in sorted(jahre):
            self.jahr.addItem(jahr)

    def laden(self) -> None:
        jahr = self.jahr.currentText().strip()
        monat = self.monat.currentText().strip()
        self.buchungen = buchhaltung_daten(jahr, monat)

        self._kpis()
        self._buchungen_tabelle()
        self._kostenarten_tabelle()
        self._offene_tabelle()
        self._diagramm()

    @staticmethod
    def _karte(titel: str, wert: str, icon: str) -> QFrame:
        card = QFrame()
        card.setObjectName("metricCard")
        lay = QVBoxLayout(card)
        t = QLabel(f"{icon}  {titel}")
        t.setObjectName("metricTitle")
        v = QLabel(wert)
        v.setObjectName("metricValue")
        lay.addWidget(t)
        lay.addWidget(v)
        return card

    def _kpis(self) -> None:
        while self.kpi_grid.count():
            item = self.kpi_grid.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        einnahmen = buchhaltung_summe(self.buchungen, "Einnahme")
        ausgaben = buchhaltung_summe(self.buchungen, "Ausgabe")
        saldo = einnahmen - ausgaben
        offen = sum(float(b["betrag"]) for b in self.buchungen if "offen" in str(b["status"]).lower())

        karten = [
            ("Einnahmen", euro(einnahmen), "€"),
            ("Ausgaben", euro(ausgaben), "↘"),
            ("Saldo", euro(saldo), "Σ"),
            ("Offene Posten", euro(offen), "⚠"),
            ("Buchungen", str(len(self.buchungen)), "▤"),
        ]

        for i, (titel, wert, icon) in enumerate(karten):
            self.kpi_grid.addWidget(self._karte(titel, wert, icon), 0, i)

    def _buchungen_tabelle(self) -> None:
        headers = ["Quelle", "Datum", "Jahr", "Monat", "Typ", "Name", "Kostenart", "Status", "Betrag", "Zweck"]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(self.buchungen))
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)

        for r, b in enumerate(self.buchungen):
            values = [
                b["quelle"], b["datum"], b["jahr"], b["monat"], b["typ"], b["name"],
                b["kostenart"], b["status"], f'{float(b["betrag"]):.2f}', b["zweck"]
            ]
            for c, value in enumerate(values):
                self.table.setItem(r, c, QTableWidgetItem(str(value)))

        widths = [120, 100, 80, 80, 110, 220, 180, 120, 120, 420]
        for c, w in enumerate(widths):
            self.table.setColumnWidth(c, w)

    def _kostenarten_tabelle(self) -> None:
        sums: dict[str, dict[str, float]] = {}
        for b in self.buchungen:
            key = str(b["kostenart"] or "Sonstiges")
            sums.setdefault(key, {"Einnahme": 0.0, "Ausgabe": 0.0})
            sums[key][b["typ"]] += float(b["betrag"])

        rows = sorted(sums.items())
        self.table_kostenarten.setColumnCount(4)
        self.table_kostenarten.setHorizontalHeaderLabels(["Kostenart", "Einnahmen", "Ausgaben", "Saldo"])
        self.table_kostenarten.setRowCount(len(rows))
        self.table_kostenarten.setAlternatingRowColors(True)

        for r, (key, vals) in enumerate(rows):
            ein = vals["Einnahme"]
            aus = vals["Ausgabe"]
            saldo = ein - aus
            for c, value in enumerate([key, f"{ein:.2f}", f"{aus:.2f}", f"{saldo:.2f}"]):
                self.table_kostenarten.setItem(r, c, QTableWidgetItem(str(value)))

        self.table_kostenarten.setColumnWidth(0, 260)
        for c in range(1, 4):
            self.table_kostenarten.setColumnWidth(c, 140)

    def _offene_tabelle(self) -> None:
        offen = [b for b in self.buchungen if "offen" in str(b["status"]).lower()]
        headers = ["Quelle", "Datum", "Typ", "Name", "Kostenart", "Betrag", "Zweck"]
        self.table_offen.setColumnCount(len(headers))
        self.table_offen.setHorizontalHeaderLabels(headers)
        self.table_offen.setRowCount(len(offen))
        self.table_offen.setAlternatingRowColors(True)

        for r, b in enumerate(offen):
            values = [b["quelle"], b["datum"], b["typ"], b["name"], b["kostenart"], f'{float(b["betrag"]):.2f}', b["zweck"]]
            for c, value in enumerate(values):
                self.table_offen.setItem(r, c, QTableWidgetItem(str(value)))

        widths = [120, 100, 110, 220, 180, 120, 420]
        for c, w in enumerate(widths):
            self.table_offen.setColumnWidth(c, w)

    def _diagramm(self) -> None:
        self.chart_scene.clear()
        ein = buchhaltung_summe(self.buchungen, "Einnahme")
        aus = buchhaltung_summe(self.buchungen, "Ausgabe")
        max_value = max(ein, aus, 1.0)

        self.chart_scene.addText("Buchhaltung: Einnahmen / Ausgaben").setPos(30, 10)

        daten = [("Einnahmen", ein), ("Ausgaben", aus), ("Saldo", ein - aus)]
        y = 70
        for label, value in daten:
            width = int(abs(value) / max_value * 650)
            self.chart_scene.addText(label).setPos(30, y - 22)
            self.chart_scene.addRect(30, y, max(width, 4), 32, QPen(), QBrush())
            self.chart_scene.addText(euro(value)).setPos(30 + max(width, 4) + 15, y + 4)
            y += 70

        self.chart_scene.setSceneRect(0, 0, 900, 320)

    def export_excel(self) -> None:
        if not self.buchungen:
            QMessageBox.information(self, "Export", "Keine Buchungen zum Exportieren.")
            return


        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        ziel_default = EXPORT_DIR / f"buchhaltung_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

        ziel_text, _ = QFileDialog.getSaveFileName(
            self,
            "Buchhaltung exportieren",
            str(ziel_default),
            "Excel-Dateien (*.xlsx)",
        )

        if not ziel_text:
            return

        ziel = Path(ziel_text)
        if ziel.suffix.lower() != ".xlsx":
            ziel = ziel.with_suffix(".xlsx")

        wb = Workbook()
        ws = wb.active

        if not isinstance(ws, Worksheet):
            QMessageBox.warning(self, "Export", "Excel-Arbeitsblatt konnte nicht erstellt werden.")
            return

        ws.title = "Buchhaltung"
        ws.append(["Quelle", "Datum", "Jahr", "Monat", "Typ", "Name", "Kostenart", "Status", "Betrag", "Zweck"])

        for b in self.buchungen:
            ws.append([
                b["quelle"], b["datum"], b["jahr"], b["monat"], b["typ"], b["name"],
                b["kostenart"], b["status"], float(b["betrag"]), b["zweck"]
            ])

        for col_index, _col in enumerate(ws.columns, start=1):
            ws.column_dimensions[get_column_letter(col_index)].width = 24

        wb.save(ziel)
        QMessageBox.information(self, "Export", f"Buchhaltung exportiert:\n{ziel}")
        system_datei_oeffnen(str(ziel.parent))


def feld_index(titel: str, feldname: str) -> int | None:
    for index, feld in enumerate(SCHEMA.get(titel, [])):
        if norm_key(feld) == norm_key(feldname):
            return index
    return None


def zeilenwert(titel: str, row: list[str], feldname: str) -> str:
    index = feld_index(titel, feldname)
    if index is None or index >= len(row):
        return ""
    return str(row[index])


def betrag_de(value: Any) -> float:
    text = str(value or "").strip()
    if not text:
        return 0.0
    text = text.replace("€", "").replace("EUR", "").replace("eur", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except (ValueError, TypeError):
        return 0.0


def waehrung(value: float) -> str:
    return f"{value:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")


def jahr_aus_datum_oder_text(value: Any, fallback: str = "") -> str:
    text = str(value or "")
    match = re.search(r"(20\d{2}|19\d{2})", text)
    return match.group(1) if match else fallback


def mieter_objekt_wohnung(row: list[str]) -> tuple[str, str]:
    ort = zeilenwert("Mieter", row, "Ort")
    objektordner = zeilenwert("Mieter", row, "Objektordner")
    wohnungsordner = zeilenwert("Mieter", row, "Wohnungsordner")

    objekt = objektordner.strip()
    wohnung = wohnungsordner.strip()

    if not objekt and "/" in ort:
        objekt = ort.split("/")[0].strip()

    if not wohnung and "/" in ort:
        wohnung = ort.split("/")[-1].strip()

    return objekt, wohnung







BK_JAHRESKOSTEN_FILE = APP_DIR / "bk_jahreskosten.json"
BK_UMLAGESCHLUESSEL = [
    "Wohnfläche (m²)",
    "Wohnungen (gleichmäßig)",
    "Mieter / Personen (gleichmäßig)",
    "BK-Vorauszahlung",
]
BK_STATUSWERTE = ["Offen", "Erfasst", "Geprüft"]
BK_JAHRESSTATUS = ["Entwurf", "In Bearbeitung", "Fertig"]
STANDARD_BK_KOSTENARTEN = [
    "Grundsteuer", "Wasser", "Abwasser", "Müllabfuhr", "Allgemeinstrom",
    "Schornsteinfeger", "Gebäudeversicherung", "Haftpflichtversicherung",
    "Hausmeister", "Hausreinigung", "Gartenpflege", "Winterdienst",
    "Straßenreinigung", "Heizungswartung", "Rauchwarnmelder", "Aufzug",
    "Legionellenprüfung", "Dachrinnenreinigung", "Sonstige Betriebskosten",
]


def bk_leere_position(art: str = "") -> dict[str, Any]:
    return {
        "art": str(art or "").strip(),
        "betrag": 0.0,
        "umlage": "Wohnfläche (m²)",
        "lieferant": "",
        "rechnungsnummer": "",
        "rechnungsdatum": "",
        "pdf": "",
        "status": "Offen",
        "bemerkung": "",
    }


def bk_datenbank_laden() -> dict[str, Any]:
    try:
        if not BK_JAHRESKOSTEN_FILE.exists():
            return {}
        data = json.loads(BK_JAHRESKOSTEN_FILE.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except (OSError, json.JSONDecodeError, TypeError):
        return {}


def bk_datenbank_speichern(daten: dict[str, Any]) -> None:
    BK_JAHRESKOSTEN_FILE.parent.mkdir(parents=True, exist_ok=True)
    tmp = BK_JAHRESKOSTEN_FILE.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(daten, indent=2, ensure_ascii=False), encoding="utf-8")
    tmp.replace(BK_JAHRESKOSTEN_FILE)


def bk_jahresdatensatz_laden(objekt: str, jahr: str) -> dict[str, Any]:
    daten = bk_datenbank_laden()
    raw = daten.get(norm_key(objekt), {}).get(str(jahr), {})
    # Rückwärtskompatibilität zu Version 7.9: Dort war der Jahreswert direkt eine Liste.
    if isinstance(raw, list):
        kosten = raw
        meta = {}
    elif isinstance(raw, dict):
        kosten = raw.get("kosten", [])
        meta = raw
    else:
        kosten, meta = [], {}

    result_kosten: list[dict[str, Any]] = []
    for item in kosten if isinstance(kosten, list) else []:
        if not isinstance(item, dict):
            continue
        pos = bk_leere_position(item.get("art", ""))
        pos.update({
            "betrag": float(item.get("betrag", 0) or 0),
            "umlage": str(item.get("umlage", "Wohnfläche (m²)") or "Wohnfläche (m²)"),
            "lieferant": str(item.get("lieferant", "") or ""),
            "rechnungsnummer": str(item.get("rechnungsnummer", "") or ""),
            "rechnungsdatum": str(item.get("rechnungsdatum", "") or ""),
            "pdf": str(item.get("pdf", "") or ""),
            "status": str(item.get("status", "Erfasst" if float(item.get("betrag", 0) or 0) > 0 else "Offen") or "Offen"),
            "bemerkung": str(item.get("bemerkung", "") or ""),
        })
        result_kosten.append(pos)

    return {
        "objekt": objekt,
        "jahr": str(jahr),
        "status": str(meta.get("status", "Entwurf") or "Entwurf"),
        "notiz": str(meta.get("notiz", "") or ""),
        "geaendert": str(meta.get("geaendert", "") or ""),
        "fertiggestellt": str(meta.get("fertiggestellt", "") or ""),
        "kosten": result_kosten,
    }


def bk_jahreskosten_laden(objekt: str, jahr: str) -> list[dict[str, Any]]:
    return bk_jahresdatensatz_laden(objekt, jahr)["kosten"]


def bk_jahresdatensatz_speichern(
    objekt: str,
    jahr: str,
    kosten: list[dict[str, Any]],
    status: str = "Entwurf",
    notiz: str = "",
) -> None:
    daten = bk_datenbank_laden()
    key = norm_key(objekt)
    daten.setdefault(key, {})
    vorher = daten[key].get(str(jahr), {})
    fertiggestellt = ""
    if isinstance(vorher, dict):
        fertiggestellt = str(vorher.get("fertiggestellt", "") or "")
    if status == "Fertig" and not fertiggestellt:
        fertiggestellt = datetime.now().isoformat(timespec="seconds")
    if status != "Fertig":
        fertiggestellt = ""

    clean: list[dict[str, Any]] = []
    for item in kosten:
        art = str(item.get("art", "") or "").strip()
        if not art:
            continue
        clean.append({
            "art": art,
            "betrag": round(float(item.get("betrag", 0) or 0), 2),
            "umlage": str(item.get("umlage", "Wohnfläche (m²)") or "Wohnfläche (m²)"),
            "lieferant": str(item.get("lieferant", "") or "").strip(),
            "rechnungsnummer": str(item.get("rechnungsnummer", "") or "").strip(),
            "rechnungsdatum": str(item.get("rechnungsdatum", "") or "").strip(),
            "pdf": str(item.get("pdf", "") or "").strip(),
            "status": str(item.get("status", "Offen") or "Offen"),
            "bemerkung": str(item.get("bemerkung", "") or "").strip(),
        })
    daten[key][str(jahr)] = {
        "objekt": objekt,
        "jahr": str(jahr),
        "status": status if status in BK_JAHRESSTATUS else "Entwurf",
        "notiz": str(notiz or ""),
        "geaendert": datetime.now().isoformat(timespec="seconds"),
        "fertiggestellt": fertiggestellt,
        "kosten": clean,
    }
    bk_datenbank_speichern(daten)


def bk_jahreskosten_speichern(objekt: str, jahr: str, kosten: list[dict[str, Any]]) -> None:
    bestand = bk_jahresdatensatz_laden(objekt, jahr)
    bk_jahresdatensatz_speichern(
        objekt, jahr, kosten,
        str(bestand.get("status", "Entwurf")),
        str(bestand.get("notiz", "")),
    )


def bk_rechnung_kopieren(path: str, objekt: str, jahr: str, kostenart: str) -> str:
    quelle = Path(path)
    if not quelle.exists():
        return ""
    ziel_dir = DOKUMENTE_DIR / "bk_belege" / safe_filename(jahr) / safe_filename(objekt)
    ziel_dir.mkdir(parents=True, exist_ok=True)
    basisname = safe_filename(kostenart) + quelle.suffix.lower()
    ziel = ziel_dir / basisname
    nummer = 2
    while ziel.exists() and ziel.resolve() != quelle.resolve():
        ziel = ziel_dir / f"{safe_filename(kostenart)}_{nummer}{quelle.suffix.lower()}"
        nummer += 1
    if ziel.resolve() != quelle.resolve():
        shutil.copy2(quelle, ziel)
    try:
        return str(ziel.relative_to(BASE_DIR)).replace("\\", "/")
    except ValueError:
        return str(ziel)


def bk_objekte() -> list[str]:
    """Liefert alle bekannten Objektbezüge ohne Änderung der Excel-Struktur."""
    werte: dict[str, str] = {}
    for titel in ["Objekte", "Wohnungen", "Mieter"]:
        for row in DATA.get(titel, []):
            kandidaten = [
                feldwert(titel, row, ["Objektname", "Objekt", "Objekt / Adresse"]),
                feldwert(titel, row, ["Objektordner"]),
            ]
            if titel == "Mieter":
                objekt, _wohnung = mieter_objekt_wohnung(row)
                kandidaten.append(objekt)
            for wert in kandidaten:
                text = str(wert or "").strip()
                if text:
                    werte.setdefault(norm_key(text), text)
    return sorted(werte.values(), key=lambda value: value.lower())


def bk_mieter_fuer_objekt(objekt: str) -> list[tuple[int, list[str]]]:
    key = norm_key(objekt)
    return [
        (index, row)
        for index, row in enumerate(DATA.get("Mieter", []))
        if key and norm_key(mieter_objekt_wohnung(row)[0]) == key
    ]


def bk_wohnungen_fuer_objekt(objekt: str) -> list[list[str]]:
    key = norm_key(objekt)
    return [
        row for row in DATA.get("Wohnungen", [])
        if key and norm_key(feldwert("Wohnungen", row, ["Objekt", "Objektordner"])) == key
    ]


def bk_wohnflaeche(row: list[str], wohnung: str, objekt: str) -> float:
    wert = betrag_de(zeilenwert("Mieter", row, "Wohnfläche"))
    if wert > 0:
        return wert
    for wrow in bk_wohnungen_fuer_objekt(objekt):
        wname = feldwert("Wohnungen", wrow, ["Wohnung", "Wohnungsordner"])
        if wohnung and norm_key(wname) == norm_key(wohnung):
            return betrag_de(feldwert("Wohnungen", wrow, ["Größe qm", "Wohnfläche", "Fläche"]))
    return 0.0


def bk_datum(value: Any) -> date | None:
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    match = re.search(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})", text)
    if match:
        try:
            return date(int(match.group(3)), int(match.group(2)), int(match.group(1)))
        except ValueError:
            return None
    return None


def bk_nutzungsdaten(row: list[str], jahr: str) -> tuple[date, date, int, int, float]:
    try:
        year = int(jahr)
    except ValueError:
        year = datetime.now().year
    year_start, year_end = date(year, 1, 1), date(year, 12, 31)
    start = bk_datum(zeilenwert("Mieter", row, "Mietbeginn")) or year_start
    end = bk_datum(zeilenwert("Mieter", row, "Mietende")) or year_end
    start = max(start, year_start)
    end = min(end, year_end)
    total_days = (year_end - year_start).days + 1
    occupied_days = max(0, (end - start).days + 1)
    factor = occupied_days / total_days if total_days else 0.0
    return start, end, occupied_days, total_days, factor


def bk_mietzeitraum(row: list[str], jahr: str) -> str:
    start, end, occupied, _total, _factor = bk_nutzungsdaten(row, jahr)
    if occupied <= 0:
        return "Kein Nutzungszeitraum im Abrechnungsjahr"
    return f"{start.strftime('%d.%m.%Y')} – {end.strftime('%d.%m.%Y')}"


def bk_vorauszahlung(row: list[str], jahr: str) -> float:
    _start, _end, _days, _total, factor = bk_nutzungsdaten(row, jahr)
    annual = betrag_de(zeilenwert("Mieter", row, "BK jährlich"))
    if annual > 0:
        return annual * factor
    monthly = betrag_de(zeilenwert("Mieter", row, "BK monatlich"))
    return monthly * 12.0 * factor


def bk_gewicht(schluessel: str, row: list[str], wohnung: str, objekt: str, jahr: str = "") -> float:
    _start, _end, _days, _total, time_factor = bk_nutzungsdaten(row, jahr or str(datetime.now().year))
    if time_factor <= 0:
        return 0.0
    if schluessel == "Wohnfläche (m²)":
        return bk_wohnflaeche(row, wohnung, objekt) * time_factor
    if schluessel == "BK-Vorauszahlung":
        return bk_vorauszahlung(row, jahr or str(datetime.now().year))
    return time_factor


def bk_abrechnung_berechnen(
    jahr: str,
    objekt: str = "",
    haus_kosten: float | None = None,
    schluessel: str = "Wohnfläche (m²)",
    kostenpositionen: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    """Erstellt detaillierte Abrechnungen unter Berücksichtigung des Mietzeitraums."""
    if not objekt:
        result: list[dict[str, Any]] = []
        for obj in bk_objekte():
            result.extend(bk_abrechnung_berechnen(jahr, obj, haus_kosten, schluessel, kostenpositionen))
        return result

    mieter_rows = bk_mieter_fuer_objekt(objekt)
    if not mieter_rows:
        return []

    positionen = list(kostenpositionen or bk_jahreskosten_laden(objekt, jahr))
    if not positionen and float(haus_kosten or 0) > 0:
        positionen = [{"art": "Betriebskosten Haus gesamt", "betrag": float(haus_kosten or 0), "umlage": schluessel}]
    positionen = [
        p for p in positionen
        if str(p.get("art", "")).strip() and float(p.get("betrag", 0) or 0) > 0
    ]
    if not positionen:
        return []

    basis: list[dict[str, Any]] = []
    for mieter_index, row in mieter_rows:
        _objekt, wohnung = mieter_objekt_wohnung(row)
        start, end, days, total_days, factor = bk_nutzungsdaten(row, jahr)
        if days <= 0:
            continue
        basis.append({
            "mieter_index": mieter_index,
            "row": row,
            "wohnung": wohnung,
            "vorauszahlung": bk_vorauszahlung(row, jahr),
            "nutzungstage": days,
            "jahrestage": total_days,
            "zeitfaktor": factor,
            "nutzungsbeginn": start,
            "nutzungsende": end,
            "details": [],
            "ist": 0.0,
        })
    if not basis:
        return []

    for position in positionen:
        art = str(position.get("art", "")).strip()
        betrag = float(position.get("betrag", 0) or 0)
        umlage = str(position.get("umlage", schluessel) or schluessel)
        gewichte = [
            max(0.0, bk_gewicht(umlage, b["row"], b["wohnung"], objekt, jahr))
            for b in basis
        ]
        gesamtgewicht = sum(gewichte)
        effektive_umlage = umlage
        if gesamtgewicht <= 0:
            gewichte = [float(b["zeitfaktor"]) for b in basis]
            gesamtgewicht = sum(gewichte)
            effektive_umlage = "Mieter / Personen (gleichmäßig)"
        for index, b in enumerate(basis):
            anteil = betrag * gewichte[index] / gesamtgewicht if gesamtgewicht else 0.0
            prozent = gewichte[index] / gesamtgewicht * 100.0 if gesamtgewicht else 0.0
            b["ist"] += anteil
            b["details"].append({
                "art": art,
                "lieferant": str(position.get("lieferant", "") or ""),
                "rechnungsnummer": str(position.get("rechnungsnummer", "") or ""),
                "rechnungsdatum": str(position.get("rechnungsdatum", "") or ""),
                "pdf": str(position.get("pdf", "") or ""),
                "umlage": effektive_umlage,
                "betrag": betrag,
                "gewicht": gewichte[index],
                "gesamtgewicht": gesamtgewicht,
                "anteil_prozent": prozent,
                "anteil": anteil,
            })

    haus_kosten_gesamt = sum(float(p.get("betrag", 0) or 0) for p in positionen)
    abrechnungen: list[dict[str, Any]] = []
    for b in basis:
        row = b["row"]
        ist = float(b["ist"])
        vorauszahlung = float(b["vorauszahlung"])
        differenz = ist - vorauszahlung
        abrechnungen.append({
            "jahr": jahr,
            "mieter_index": b["mieter_index"],
            "mieter": zeilenwert("Mieter", row, "Mieter"),
            "empfaenger_anschrift": zeilenwert("Mieter", row, "Ort"),
            "objekt": objekt,
            "wohnung": b["wohnung"],
            "wohnungsadresse": b["wohnung"] or zeilenwert("Mieter", row, "Ort") or objekt,
            "wohnflaeche": bk_wohnflaeche(row, b["wohnung"], objekt),
            "mietzeitraum": bk_mietzeitraum(row, jahr),
            "abrechnungszeitraum": f"01.01.{jahr} – 31.12.{jahr}" if jahr else "",
            "nutzungstage": b["nutzungstage"],
            "jahrestage": b["jahrestage"],
            "hauskosten": haus_kosten_gesamt,
            "umlage_schluessel": "je Kostenart",
            "anteil_prozent": (ist / haus_kosten_gesamt * 100.0) if haus_kosten_gesamt else 0.0,
            "vorauszahlung": vorauszahlung,
            "ist": ist,
            "nachzahlung": differenz if differenz > 0 else 0.0,
            "guthaben": abs(differenz) if differenz < 0 else 0.0,
            "status": "Nachzahlung" if differenz > 0.009 else ("Guthaben" if differenz < -0.009 else "Ausgeglichen"),
            "details": b["details"],
        })
    return abrechnungen


def bk_logo_uri() -> str:
    logo = str(CONFIG.get("firma_logo", "")).strip()
    if not logo:
        return ""
    path = Path(logo)
    if not path.is_absolute():
        path = BASE_DIR / path
    return path.resolve().as_uri() if path.exists() else ""


def bk_abrechnung_html(eintrag: dict[str, Any]) -> str:
    esc = lambda value: html.escape(str(value or ""))
    logo_uri = bk_logo_uri()
    logo_html = f"<img src='{esc(logo_uri)}' style='max-height:85px;max-width:220px'>" if logo_uri else ""
    heute = datetime.now().strftime("%d.%m.%Y")
    bankzeile = " · ".join(filter(None, [
        str(CONFIG.get("firma_bank", "")).strip(),
        ("IBAN " + str(CONFIG.get("firma_iban", "")).strip()) if CONFIG.get("firma_iban") else "",
        ("BIC " + str(CONFIG.get("firma_bic", "")).strip()) if CONFIG.get("firma_bic") else "",
    ]))
    nachzahlung = float(eintrag.get("nachzahlung", 0))
    guthaben = float(eintrag.get("guthaben", 0))
    ergebnis_label = "Nachzahlung" if nachzahlung > 0 else ("Guthaben" if guthaben > 0 else "Ergebnis")
    ergebnis_wert = nachzahlung or guthaben
    gruss = esc(CONFIG.get("firma_gruss", "")).replace("\n", "<br>")
    hinweis = esc(CONFIG.get("bk_hinweis", ""))
    detail_rows = "".join(
        "<tr><td>{}</td><td>{}</td><td>{}: {:.2f} %</td><td class='rechts'>{}</td><td class='rechts'>{}</td></tr>".format(
            esc(detail.get("art")),
            esc(detail.get("lieferant")),
            esc(detail.get("umlage")),
            float(detail.get("anteil_prozent", 0)),
            waehrung(float(detail.get("betrag", 0))),
            waehrung(float(detail.get("anteil", 0))),
        )
        for detail in eintrag.get("details", [])
    )
    return f"""
    <html><head><meta charset='utf-8'><style>
      @page {{ size: A4; margin: 15mm 14mm 15mm 18mm; }}
      body {{ font-family: Arial, Helvetica, sans-serif; font-size: 9.5pt; color:#111; line-height:1.3; }}
      .kopf {{ width:100%; border:0; margin-bottom:14px; }} .kopf td {{ border:0; vertical-align:top; }}
      .firma {{ text-align:right; font-size:9pt; }} .firma .name {{ font-size:16pt; font-weight:bold; color:#16365c; }}
      .slogan {{ font-weight:bold; color:#16365c; }} .absender {{ font-size:7.5pt; text-decoration:underline; margin-bottom:7px; }}
      .empfaenger {{ min-height:82px; width:52%; white-space:pre-line; }} .datum {{ text-align:right; margin:8px 0 15px; }}
      h1 {{ font-size:15pt; margin:0 0 14px; }} .stamm,.kosten {{ width:100%; border-collapse:collapse; }}
      .stamm {{ margin-bottom:15px; }} .stamm td {{ padding:3px 5px; border-bottom:1px solid #ddd; }}
      .stamm td:first-child {{ width:34%; font-weight:bold; }} .kosten th {{ background:#16365c; color:white; padding:6px; border:1px solid #16365c; }}
      .kosten td {{ padding:5px; border:1px solid #b8c2cc; }} .rechts {{ text-align:right; }}
      .ergebnis {{ margin-top:15px; border:2px solid #16365c; padding:10px; font-size:12pt; font-weight:bold; }}
      .hinweis {{ margin-top:15px; font-size:9pt; }} .fuss {{ margin-top:28px; font-size:8pt; border-top:1px solid #999; padding-top:6px; text-align:center; }}
    </style></head><body>
      <table class='kopf'><tr><td>{logo_html}</td><td class='firma'>
        <div class='name'>{esc(CONFIG.get('firma_name'))}</div><div>{esc(CONFIG.get('firma_untertitel'))}</div>
        <div>{esc(CONFIG.get('firma_leistung'))}</div><div class='slogan'>{esc(CONFIG.get('firma_slogan'))}</div>
        <div>{esc(CONFIG.get('firma_inhaber'))}</div><div>{esc(CONFIG.get('firma_strasse'))}, {esc(CONFIG.get('firma_plz_ort'))}</div>
        <div>Tel.: {esc(CONFIG.get('firma_telefon'))}</div><div>{esc(CONFIG.get('firma_email'))}</div>
      </td></tr></table>
      <div class='absender'>{esc(CONFIG.get('firma_name'))} · {esc(CONFIG.get('firma_strasse'))} · {esc(CONFIG.get('firma_plz_ort'))}</div>
      <div class='empfaenger'>Frau/Herrn<br><b>{esc(eintrag.get('mieter'))}</b><br>{esc(eintrag.get('empfaenger_anschrift'))}</div>
      <div class='datum'>{esc(CONFIG.get('firma_plz_ort')).split(' ')[-1] if CONFIG.get('firma_plz_ort') else ''}, {heute}</div>
      <h1>BETRIEBSKOSTENABRECHNUNG</h1>
      <table class='stamm'>
        <tr><td>Wohnung</td><td>{esc(eintrag.get('wohnung'))}, {esc(eintrag.get('objekt'))}</td></tr>
        <tr><td>Mietzeitraum</td><td>{esc(eintrag.get('mietzeitraum'))} ({esc(eintrag.get('nutzungstage'))} Tage)</td></tr>
        <tr><td>Abrechnungszeitraum</td><td>{esc(eintrag.get('abrechnungszeitraum'))}</td></tr>
        <tr><td>Umlageschlüssel</td><td>je Kostenart gemäß Aufstellung</td></tr>
      </table>
      <p>Sehr geehrte Damen und Herren,</p>
      <p>für den genannten Zeitraum ergibt sich folgende Betriebskostenabrechnung:</p>
      <table class='kosten'>
        <tr><th>Position</th><th>Versorger</th><th>Berechnungsgrundlage</th><th>Gesamt Haus</th><th>Ihr Anteil</th></tr>
        {detail_rows}
        <tr><td colspan='3'><b>Summe Betriebskosten</b></td><td class='rechts'><b>{waehrung(float(eintrag.get('hauskosten',0)))}</b></td><td class='rechts'><b>{waehrung(float(eintrag.get('ist',0)))}</b></td></tr>
        <tr><td colspan='4'><b>Geleistete BK-Vorauszahlungen</b></td><td class='rechts'><b>- {waehrung(float(eintrag.get('vorauszahlung',0)))}</b></td></tr>
      </table>
      <div class='ergebnis'>{ergebnis_label}: <span style='float:right'>{waehrung(ergebnis_wert)}</span></div>
      <p class='hinweis'>{hinweis}</p><p>{gruss}</p>
      <div class='fuss'>{esc(CONFIG.get('firma_name'))} · {esc(CONFIG.get('firma_strasse'))} · {esc(CONFIG.get('firma_plz_ort'))}{(' · ' + esc(bankzeile)) if bankzeile else ''}{(' · ' + esc(CONFIG.get('firma_steuer'))) if CONFIG.get('firma_steuer') else ''}</div>
    </body></html>
    """


def bk_pdf_speichern(eintrag: dict[str, Any]) -> Path:
    jahr = safe_filename(eintrag.get("jahr", ""))
    objekt = safe_filename(eintrag.get("objekt", "Objekt"))
    mieter = safe_filename(eintrag.get("mieter", "Mieter"))
    ziel_dir = DOKUMENTE_DIR / "bk_abrechnungen" / jahr / objekt
    ziel_dir.mkdir(parents=True, exist_ok=True)
    ziel = ziel_dir / f"BK_Abrechnung_{jahr}_{mieter}.pdf"
    doc = QTextDocument()
    doc.setHtml(bk_abrechnung_html(eintrag))
    printer = QPrinter(QPrinter.PrinterMode.HighResolution)
    printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
    printer.setOutputFileName(str(ziel))
    doc.print_(printer)
    return ziel


def bk_mieterfelder_aktualisieren(eintrag: dict[str, Any], pdf_path: Path | None = None) -> None:
    index = int(eintrag.get("mieter_index", -1))
    rows = DATA.get("Mieter", [])
    if index < 0 or index >= len(rows):
        return
    row = rows[index]
    updates = {
        "Ist-BK jährlich": f"{float(eintrag.get('ist',0)):.2f}",
        "Nachzahlung": f"{float(eintrag.get('nachzahlung',0)):.2f}",
        "Guthaben": f"{float(eintrag.get('guthaben',0)):.2f}",
        "BK-Status": str(eintrag.get("status", "")),
    }
    if pdf_path is not None:
        try:
            updates["BK (PDF)"] = str(pdf_path.relative_to(BASE_DIR)).replace("\\", "/")
        except ValueError:
            updates["BK (PDF)"] = str(pdf_path)
    for feld, wert in updates.items():
        idx = feld_index("Mieter", feld)
        if idx is not None:
            while len(row) <= idx:
                row.append("")
            row[idx] = wert


class BKAutomatikSeite(QWidget):
    """BK-Verwaltungscenter Professional: Erfassung, Belege, Prüfung, Abrechnung und Vergleich."""

    COL_ART, COL_BETRAG, COL_UMLAGE, COL_LIEFERANT, COL_NR, COL_DATUM, COL_PDF, COL_STATUS, COL_NOTIZ = range(9)

    def __init__(self):
        super().__init__()
        self.abrechnungen: list[dict[str, Any]] = []
        self._ladevorgang = False
        self._dirty = False
        self.autosave_timer = QTimer(self)
        self.autosave_timer.setSingleShot(True)
        self.autosave_timer.setInterval(800)
        self.autosave_timer.timeout.connect(self._autosave)

        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(10)
        title = QLabel("BK-Verwaltungscenter Professional")
        title.setObjectName("pageTitle")
        root.addWidget(title)
        info = QLabel("Rechnungen nach und nach erfassen, Belege hinterlegen, prüfen und anschließend alle Mieterabrechnungen erstellen.")
        info.setObjectName("subTitle")
        root.addWidget(info)

        kopf = QHBoxLayout()
        self.objekt = QComboBox()
        self.objekt.setMinimumWidth(280)
        self.objekt.addItems(bk_objekte())
        self.jahr = QComboBox()
        self.jahr.setEditable(True)
        self.jahr.addItems(self._jahre())
        self.jahr.setCurrentText(str(datetime.now().year - 1))
        self.jahresstatus = QComboBox()
        self.jahresstatus.addItems(BK_JAHRESSTATUS)
        self.jahresstatus.setMinimumWidth(150)
        kopf.addWidget(QLabel("Objekt:")); kopf.addWidget(self.objekt)
        kopf.addWidget(QLabel("Abrechnungsjahr:")); kopf.addWidget(self.jahr)
        kopf.addWidget(QLabel("Jahresstatus:")); kopf.addWidget(self.jahresstatus)
        kopf.addStretch()
        root.addLayout(kopf)

        statuszeile = QHBoxLayout()
        self.fortschritt = QProgressBar()
        self.fortschritt.setRange(0, 100)
        self.fortschritt.setMinimumWidth(320)
        self.fortschritt_label = QLabel("0 von 0 Positionen vollständig")
        self.offen_label = QLabel("Offen: 0")
        self.summe_label = QLabel("Summe: 0,00 €")
        self.summe_label.setObjectName("metricValue")
        statuszeile.addWidget(self.fortschritt)
        statuszeile.addWidget(self.fortschritt_label)
        statuszeile.addWidget(self.offen_label)
        statuszeile.addStretch()
        statuszeile.addWidget(self.summe_label)
        root.addLayout(statuszeile)

        splitter = QSplitter(Qt.Orientation.Vertical)
        root.addWidget(splitter, 1)
        oben = QWidget()
        oben_lay = QVBoxLayout(oben)
        oben_lay.setContentsMargins(0, 0, 0, 0)

        self.kostentabelle = QTableWidget(0, 9)
        self.kostentabelle.setHorizontalHeaderLabels([
            "Kostenart", "Jahresbetrag Haus", "Umlageschlüssel", "Lieferant/Versorger",
            "Rechnungsnummer", "Rechnungsdatum", "Rechnung PDF", "Status", "Bemerkung",
        ])
        widths = [230, 145, 210, 190, 145, 120, 210, 110, 260]
        for column, width in enumerate(widths):
            self.kostentabelle.setColumnWidth(column, width)
        self.kostentabelle.setAlternatingRowColors(True)
        self.kostentabelle.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        oben_lay.addWidget(self.kostentabelle, 1)

        kosten_btns = QHBoxLayout()
        buttons = [
            ("Kostenart hinzufügen", self.kostenart_hinzufuegen),
            ("Standard-Kostenarten", self.standard_kostenarten),
            ("Zeile entfernen", self.kostenart_entfernen),
            ("PDF zuordnen", self.pdf_zuordnen),
            ("PDF öffnen", self.pdf_oeffnen),
            ("Jetzt speichern", self.kosten_speichern),
        ]
        for label, slot in buttons:
            btn = QPushButton(label)
            btn.clicked.connect(slot)
            if label == "Jetzt speichern":
                btn.setObjectName("primaryButton")
            kosten_btns.addWidget(btn)
        kosten_btns.addStretch()
        oben_lay.addLayout(kosten_btns)
        splitter.addWidget(oben)

        unten = QWidget()
        unten_lay = QVBoxLayout(unten)
        unten_lay.setContentsMargins(0, 0, 0, 0)
        action = QHBoxLayout()
        self.berechnen_btn = QPushButton("Mieterabrechnungen erstellen")
        self.berechnen_btn.setObjectName("primaryButton")
        self.berechnen_btn.clicked.connect(self.laden)
        for btn in [
            self.berechnen_btn,
            self._button("Ausgewählte drucken", self.drucken),
            self._button("Ausgewählte als PDF", self.pdf_einzeln),
            self._button("Alle PDFs erstellen", self.pdf_alle),
            self._button("Jahr fertigstellen", self.jahr_fertigstellen),
            self._button("Firmenlayout bearbeiten", self.firmenlayout),
        ]:
            action.addWidget(btn)
        action.addStretch()
        unten_lay.addLayout(action)

        self.tabs = QTabWidget()
        self.table = QTableWidget()
        self.table_pruefung = QTableWidget()
        self.table_vergleich = QTableWidget()
        self.preview = QTextEdit()
        self.preview.setReadOnly(True)
        self.tabs.addTab(self.table, "Mieterabrechnungen")
        self.tabs.addTab(self.table_pruefung, "Plausibilitätsprüfung")
        self.tabs.addTab(self.table_vergleich, "Vorjahresvergleich")
        self.tabs.addTab(self.preview, "Firmenvordruck")
        unten_lay.addWidget(self.tabs, 1)
        splitter.addWidget(unten)
        splitter.setSizes([430, 390])

        self.table.itemSelectionChanged.connect(self.vorschau_aktualisieren)
        self.objekt.currentTextChanged.connect(self.kosten_laden)
        self.jahr.currentTextChanged.connect(self.kosten_laden)
        self.jahresstatus.currentTextChanged.connect(self._geaendert)
        self.kostentabelle.itemChanged.connect(self._geaendert)
        self.kosten_laden()

    @staticmethod
    def _button(text: str, slot) -> QPushButton:
        btn = QPushButton(text)
        btn.clicked.connect(slot)
        return btn

    @staticmethod
    def _jahre() -> list[str]:
        years = {str(datetime.now().year), str(datetime.now().year - 1)}
        for titel in ["Zahlungen", "Betriebskosten", "Mieter"]:
            for row in DATA.get(titel, []):
                for value in row:
                    year = jahr_aus_datum_oder_text(value)
                    if year:
                        years.add(year)
        return sorted(years, reverse=True)

    def kostenart_hinzufuegen(self, art: str = "", **values: Any) -> None:
        row = self.kostentabelle.rowCount()
        self.kostentabelle.insertRow(row)
        pos = bk_leere_position(art)
        pos.update(values)
        for column, key in [
            (self.COL_ART, "art"), (self.COL_BETRAG, "betrag"), (self.COL_LIEFERANT, "lieferant"),
            (self.COL_NR, "rechnungsnummer"), (self.COL_DATUM, "rechnungsdatum"),
            (self.COL_PDF, "pdf"), (self.COL_NOTIZ, "bemerkung"),
        ]:
            value = pos.get(key, "")
            if key == "betrag":
                value = "" if float(value or 0) <= 0 else f"{float(value):.2f}"
            self.kostentabelle.setItem(row, column, QTableWidgetItem(str(value)))
        umlage = QComboBox()
        umlage.addItems(BK_UMLAGESCHLUESSEL)
        umlage.setCurrentText(str(pos.get("umlage", BK_UMLAGESCHLUESSEL[0])))
        umlage.currentTextChanged.connect(self._geaendert)
        self.kostentabelle.setCellWidget(row, self.COL_UMLAGE, umlage)
        status = QComboBox()
        status.addItems(BK_STATUSWERTE)
        status.setCurrentText(str(pos.get("status", "Offen")))
        status.currentTextChanged.connect(self._geaendert)
        self.kostentabelle.setCellWidget(row, self.COL_STATUS, status)

    def standard_kostenarten(self) -> None:
        vorhandene = {
            norm_key(self.kostentabelle.item(row, self.COL_ART).text())
            for row in range(self.kostentabelle.rowCount())
            if self.kostentabelle.item(row, self.COL_ART)
        }
        for art in STANDARD_BK_KOSTENARTEN:
            if norm_key(art) not in vorhandene:
                self.kostenart_hinzufuegen(art)
        self._geaendert()

    def kostenart_entfernen(self) -> None:
        rows = sorted({item.row() for item in self.kostentabelle.selectedIndexes()}, reverse=True)
        for row in rows:
            self.kostentabelle.removeRow(row)
        self._geaendert()

    def kostenpositionen(self, auch_leere: bool = True) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []
        for row in range(self.kostentabelle.rowCount()):
            def text_at(column: int) -> str:
                item = self.kostentabelle.item(row, column)
                return item.text().strip() if item else ""
            umlage = self.kostentabelle.cellWidget(row, self.COL_UMLAGE)
            status = self.kostentabelle.cellWidget(row, self.COL_STATUS)
            pos = {
                "art": text_at(self.COL_ART),
                "betrag": betrag_de(text_at(self.COL_BETRAG)),
                "umlage": umlage.currentText() if isinstance(umlage, QComboBox) else BK_UMLAGESCHLUESSEL[0],
                "lieferant": text_at(self.COL_LIEFERANT),
                "rechnungsnummer": text_at(self.COL_NR),
                "rechnungsdatum": text_at(self.COL_DATUM),
                "pdf": text_at(self.COL_PDF),
                "status": status.currentText() if isinstance(status, QComboBox) else "Offen",
                "bemerkung": text_at(self.COL_NOTIZ),
            }
            if pos["art"] and (auch_leere or pos["betrag"] > 0):
                result.append(pos)
        return result

    def _geaendert(self, *_args) -> None:
        if self._ladevorgang:
            return
        self._dirty = True
        self.status_aktualisieren()
        self.autosave_timer.start()

    def _autosave(self) -> None:
        if not self._dirty:
            return
        self.kosten_speichern(leise=True)

    def kosten_laden(self, *_args) -> None:
        self.autosave_timer.stop()
        self._ladevorgang = True
        self.kostentabelle.blockSignals(True)
        self.kostentabelle.setRowCount(0)
        record = bk_jahresdatensatz_laden(self.objekt.currentText().strip(), self.jahr.currentText().strip())
        self.jahresstatus.setCurrentText(str(record.get("status", "Entwurf")))
        for pos in record.get("kosten", []):
            values = dict(pos)
            art = str(values.pop("art", ""))
            self.kostenart_hinzufuegen(art, **values)
        self.kostentabelle.blockSignals(False)
        self._ladevorgang = False
        self._dirty = False
        self.abrechnungen = []
        self._tabelle_fuellen()
        self.status_aktualisieren()
        self.pruefung_aktualisieren()
        self.vergleich_aktualisieren()
        self.vorschau_aktualisieren()

    def kosten_speichern(self, _checked: bool = False, leise: bool = False) -> bool:
        objekt, jahr = self.objekt.currentText().strip(), self.jahr.currentText().strip()
        if not objekt or not jahr:
            if not leise:
                QMessageBox.information(self, "BK-Verwaltungscenter", "Bitte Objekt und Jahr auswählen.")
            return False
        try:
            bk_jahresdatensatz_speichern(
                objekt, jahr, self.kostenpositionen(True), self.jahresstatus.currentText()
            )
        except OSError as exc:
            QMessageBox.warning(self, "Speicherfehler", str(exc))
            return False
        self._dirty = False
        if not leise:
            QMessageBox.information(self, "Gespeichert", "Der aktuelle Zwischenstand wurde gespeichert.")
        self.vergleich_aktualisieren()
        return True

    def status_aktualisieren(self) -> None:
        positionen = self.kostenpositionen(True)
        total = len(positionen)
        complete = sum(
            1 for pos in positionen
            if pos["betrag"] > 0 and pos["umlage"] and pos["status"] in {"Erfasst", "Geprüft"}
        )
        open_count = sum(1 for pos in positionen if pos["status"] == "Offen" or pos["betrag"] <= 0)
        percent = round(complete / total * 100) if total else 0
        self.fortschritt.setValue(percent)
        self.fortschritt_label.setText(f"{complete} von {total} Positionen vollständig ({percent} %)")
        self.offen_label.setText(f"Offen: {open_count}")
        self.summe_label.setText("Summe: " + waehrung(sum(float(pos["betrag"]) for pos in positionen)))
        self.berechnen_btn.setEnabled(any(pos["betrag"] > 0 for pos in positionen))

    def pdf_zuordnen(self) -> None:
        row = self.kostentabelle.currentRow()
        if row < 0:
            QMessageBox.information(self, "Beleg", "Bitte zuerst eine Kostenart markieren.")
            return
        art_item = self.kostentabelle.item(row, self.COL_ART)
        art = art_item.text().strip() if art_item else "Betriebskosten"
        path, _ = QFileDialog.getOpenFileName(self, "Rechnung auswählen", "", "PDF-Dateien (*.pdf)")
        if path:
            gespeichert = bk_rechnung_kopieren(path, self.objekt.currentText(), self.jahr.currentText(), art)
            self.kostentabelle.setItem(row, self.COL_PDF, QTableWidgetItem(gespeichert))
            status = self.kostentabelle.cellWidget(row, self.COL_STATUS)
            if isinstance(status, QComboBox) and status.currentText() == "Offen":
                status.setCurrentText("Erfasst")
            self._geaendert()

    def pdf_oeffnen(self) -> None:
        row = self.kostentabelle.currentRow()
        item = self.kostentabelle.item(row, self.COL_PDF) if row >= 0 else None
        path = item.text().strip() if item else ""
        if not path:
            QMessageBox.information(self, "Beleg", "Für diese Kostenart ist keine Rechnung hinterlegt.")
            return
        system_datei_oeffnen(path)

    def pruefungsliste(self) -> list[list[str]]:
        issues: list[list[str]] = []
        kosten = self.kostenpositionen(True)
        if not kosten:
            issues.append(["Kritisch", "Jahresabrechnung", "Keine Kostenarten vorhanden."])
        for pos in kosten:
            art = pos["art"] or "Unbenannte Kostenart"
            if pos["betrag"] <= 0:
                issues.append(["Offen", art, "Jahresbetrag fehlt."])
            if not pos["umlage"]:
                issues.append(["Kritisch", art, "Umlageschlüssel fehlt."])
            if pos["status"] == "Offen":
                issues.append(["Offen", art, "Position ist noch als offen markiert."])
            if pos["betrag"] > 0 and not pos["pdf"]:
                issues.append(["Hinweis", art, "Keine Originalrechnung als PDF hinterlegt."])
        for _index, row in bk_mieter_fuer_objekt(self.objekt.currentText().strip()):
            name = zeilenwert("Mieter", row, "Mieter")
            objekt, wohnung = mieter_objekt_wohnung(row)
            if not wohnung:
                issues.append(["Kritisch", name, "Wohnung konnte nicht zugeordnet werden."])
            if any(pos["umlage"] == "Wohnfläche (m²)" and pos["betrag"] > 0 for pos in kosten):
                if bk_wohnflaeche(row, wohnung, objekt) <= 0:
                    issues.append(["Kritisch", name, "Wohnfläche fehlt für Umlage nach m²."])
            if bk_vorauszahlung(row, self.jahr.currentText()) <= 0:
                issues.append(["Hinweis", name, "Keine BK-Vorauszahlung im Abrechnungszeitraum ermittelt."])
        return issues

    def pruefung_aktualisieren(self) -> None:
        issues = self.pruefungsliste()
        self.table_pruefung.setColumnCount(3)
        self.table_pruefung.setHorizontalHeaderLabels(["Priorität", "Bezug", "Beschreibung"])
        self.table_pruefung.setRowCount(len(issues))
        for row, values in enumerate(issues):
            for column, value in enumerate(values):
                self.table_pruefung.setItem(row, column, QTableWidgetItem(value))
        self.table_pruefung.setColumnWidth(0, 110)
        self.table_pruefung.setColumnWidth(1, 250)
        self.table_pruefung.setColumnWidth(2, 720)

    def laden(self) -> None:
        objekt, jahr = self.objekt.currentText().strip(), self.jahr.currentText().strip()
        kosten = [pos for pos in self.kostenpositionen(True) if pos["betrag"] > 0]
        if not objekt or not kosten:
            QMessageBox.information(self, "BK-Center", "Bitte Objekt und mindestens eine Kostenposition mit Betrag erfassen.")
            return
        self.kosten_speichern(leise=True)
        critical = [issue for issue in self.pruefungsliste() if issue[0] == "Kritisch"]
        if critical:
            text = "\n".join(f"• {issue[1]}: {issue[2]}" for issue in critical[:12])
            QMessageBox.warning(self, "Abrechnung nicht möglich", "Bitte zuerst folgende kritische Punkte beheben:\n\n" + text)
            self.pruefung_aktualisieren()
            self.tabs.setCurrentWidget(self.table_pruefung)
            return
        self.abrechnungen = bk_abrechnung_berechnen(jahr, objekt, kostenpositionen=kosten)
        if not self.abrechnungen:
            QMessageBox.warning(self, "BK-Center", "Für dieses Objekt wurden keine Mieter mit Nutzungszeitraum gefunden.")
        self._tabelle_fuellen()
        self.pruefung_aktualisieren()
        self.vorschau_aktualisieren()

    def _tabelle_fuellen(self) -> None:
        headers = ["Mieter", "Objekt", "Wohnung", "Nutzungstage", "Wohnfläche", "Vorauszahlung", "Ist-BK", "Nachzahlung", "Guthaben", "Status"]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(self.abrechnungen))
        self.table.setAlternatingRowColors(True)
        for row, item in enumerate(self.abrechnungen):
            values = [
                item["mieter"], item["objekt"], item["wohnung"], item["nutzungstage"],
                f'{item["wohnflaeche"]:.2f}', f'{item["vorauszahlung"]:.2f}',
                f'{item["ist"]:.2f}', f'{item["nachzahlung"]:.2f}',
                f'{item["guthaben"]:.2f}', item["status"],
            ]
            for column, value in enumerate(values):
                self.table.setItem(row, column, QTableWidgetItem(str(value)))
        widths = [220, 210, 170, 105, 110, 130, 120, 130, 120, 130]
        for column, width in enumerate(widths):
            self.table.setColumnWidth(column, width)
        if self.abrechnungen:
            self.table.selectRow(0)

    def vergleich_aktualisieren(self) -> None:
        objekt = self.objekt.currentText().strip()
        try:
            jahr = int(self.jahr.currentText().strip())
        except ValueError:
            jahr = datetime.now().year
        current = {norm_key(pos["art"]): pos for pos in bk_jahreskosten_laden(objekt, str(jahr))}
        previous = {norm_key(pos["art"]): pos for pos in bk_jahreskosten_laden(objekt, str(jahr - 1))}
        keys = sorted(set(current) | set(previous), key=lambda key: (current.get(key) or previous.get(key) or {}).get("art", "").lower())
        self.table_vergleich.setColumnCount(6)
        self.table_vergleich.setHorizontalHeaderLabels(["Kostenart", str(jahr - 1), str(jahr), "Differenz", "Änderung %", "Bewertung"])
        self.table_vergleich.setRowCount(len(keys))
        for row, key in enumerate(keys):
            old = float(previous.get(key, {}).get("betrag", 0) or 0)
            new = float(current.get(key, {}).get("betrag", 0) or 0)
            diff = new - old
            percent = (diff / old * 100) if old else (100.0 if new else 0.0)
            rating = "stark gestiegen" if percent > 15 else ("gestiegen" if percent > 5 else ("gesunken" if percent < -5 else "unauffällig"))
            art = str((current.get(key) or previous.get(key) or {}).get("art", ""))
            values = [art, waehrung(old), waehrung(new), waehrung(diff), f"{percent:.1f} %", rating]
            for column, value in enumerate(values):
                self.table_vergleich.setItem(row, column, QTableWidgetItem(value))
        for column, width in enumerate([260, 140, 140, 140, 120, 160]):
            self.table_vergleich.setColumnWidth(column, width)

    def jahr_fertigstellen(self) -> None:
        issues = self.pruefungsliste()
        blocking = [issue for issue in issues if issue[0] in {"Kritisch", "Offen"}]
        if blocking:
            QMessageBox.warning(self, "Noch nicht vollständig", "Die Jahresabrechnung enthält noch offene oder kritische Punkte. Details stehen in der Plausibilitätsprüfung.")
            self.tabs.setCurrentWidget(self.table_pruefung)
            return
        self.jahresstatus.setCurrentText("Fertig")
        self.kosten_speichern(leise=True)
        QMessageBox.information(self, "Fertiggestellt", "Die Jahresabrechnung wurde als fertig markiert. PDFs können nun endgültig erstellt werden.")

    def aktueller_eintrag(self) -> dict[str, Any] | None:
        row = self.table.currentRow()
        return self.abrechnungen[row] if 0 <= row < len(self.abrechnungen) else None

    def vorschau_aktualisieren(self) -> None:
        eintrag = self.aktueller_eintrag()
        self.preview.setHtml(bk_abrechnung_html(eintrag) if eintrag else "<p>Noch keine Abrechnung berechnet.</p>")

    def drucken(self) -> None:
        eintrag = self.aktueller_eintrag()
        if not eintrag:
            QMessageBox.information(self, "Drucken", "Bitte zuerst eine Abrechnung auswählen.")
            return
        doc = QTextDocument()
        doc.setHtml(bk_abrechnung_html(eintrag))
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        dialog = QPrintDialog(printer, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            doc.print_(printer)

    def pdf_einzeln(self) -> None:
        eintrag = self.aktueller_eintrag()
        if not eintrag:
            QMessageBox.information(self, "PDF", "Bitte zuerst eine Abrechnung auswählen.")
            return
        ziel = bk_pdf_speichern(eintrag)
        bk_mieterfelder_aktualisieren(eintrag, ziel)
        speichere_tabelle("Mieter")
        QMessageBox.information(self, "PDF", f"Abrechnung gespeichert und beim Mieter hinterlegt:\n{ziel}")
        system_datei_oeffnen(str(ziel))

    def pdf_alle(self) -> None:
        if not self.abrechnungen:
            QMessageBox.information(self, "PDF", "Bitte zuerst die Mieterabrechnungen erstellen.")
            return
        zielordner = None
        for eintrag in self.abrechnungen:
            ziel = bk_pdf_speichern(eintrag)
            zielordner = ziel.parent
            bk_mieterfelder_aktualisieren(eintrag, ziel)
        speichere_tabelle("Mieter")
        QMessageBox.information(self, "PDF", f"{len(self.abrechnungen)} Abrechnungen wurden erstellt und bei den Mietern hinterlegt.\n{zielordner}")
        if zielordner:
            system_datei_oeffnen(str(zielordner))

    def firmenlayout(self) -> None:
        dialog = EinstellungenDialog()
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.vorschau_aktualisieren()

    def closeEvent(self, event) -> None:
        if self._dirty:
            self.kosten_speichern(leise=True)
        super().closeEvent(event)

def safe_filename(text: str) -> str:
    text = str(text or "").strip()
    text = re.sub(r'[<>:"/\\\\|?*]', "_", text)
    text = re.sub(r"\s+", " ", text)
    text = text.rstrip(". ")
    return text or "akte"

def datei_ist_pdf(value: Any) -> bool:
    return str(value or "").lower().endswith(".pdf")


def dokumenten_treffer() -> list[dict[str, str]]:
    treffer: list[dict[str, str]] = []

    for bereich, rows in DATA.items():
        felder = SCHEMA.get(bereich, [])

        for row_index, row in enumerate(rows, start=1):
            suchtext = " ".join(str(v) for v in row)

            objekt = ""
            wohnung = ""
            mieter = ""

            for i, feld in enumerate(felder):
                value = str(row[i]) if i < len(row) else ""
                key = norm_key(feld)

                if key in ["objekt", "objektname", "objektordner"] and not objekt:
                    objekt = value
                elif key in ["wohnung", "wohnungsordner"] and not wohnung:
                    wohnung = value
                elif key in ["mieter", "name"] and not mieter:
                    mieter = value

            for col_index, value in enumerate(row):
                feld = felder[col_index] if col_index < len(felder) else f"Spalte {col_index+1}"
                value_text = str(value or "")

                if "pdf" in str(feld).lower() or datei_ist_pdf(value_text):
                    if value_text:
                        treffer.append({
                            "bereich": bereich,
                            "zeile": str(row_index),
                            "feld": str(feld),
                            "pfad": value_text,
                            "datei": Path(value_text).name,
                            "objekt": objekt,
                            "wohnung": wohnung,
                            "mieter": mieter,
                            "inhalt": suchtext,
                        })

    return treffer



def dokumenten_index_rows() -> list[dict[str, str]]:
    """
    Erstellt einen einheitlichen Dokumentenindex aus allen vorhandenen Tabellen.

    Die Funktion verändert keine Exceldatei. Relative Dateipfade werden gegen
    den Programm-, Dokumenten- und Datenordner geprüft.
    """
    result: list[dict[str, str]] = []
    seen: set[tuple[str, str, str]] = set()

    file_extensions = {
        ".pdf", ".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif",
        ".doc", ".docx", ".xls", ".xlsx", ".csv", ".txt", ".md",
        ".json", ".xml", ".zip",
    }

    file_field_tokens = {
        "pdf", "datei", "dokument", "pfad", "anlage", "anhang",
        "foto", "bild", "scan", "beleg", "rechnung",
    }

    for bereich, rows in DATA.items():
        felder = SCHEMA.get(bereich, [])

        for row_index, row in enumerate(rows, start=1):
            objekt = objektordner_fuer_datensatz(bereich, row)
            wohnung = feldwert(bereich, row, ["Wohnung", "Wohnungsordner"])
            mieter = feldwert(bereich, row, ["Mieter", "Name"])
            row_text = " | ".join(str(value or "") for value in row)

            for column_index, raw_value in enumerate(row):
                value = str(raw_value or "").strip()
                if not value:
                    continue

                feld = (
                    str(felder[column_index])
                    if column_index < len(felder)
                    else f"Spalte {column_index + 1}"
                )
                field_key = norm_key(feld)

                suffix = Path(value).suffix.lower()
                is_file_field = any(
                    token in field_key for token in file_field_tokens
                )
                looks_like_file = suffix in file_extensions

                if not is_file_field and not looks_like_file:
                    continue

                raw_path = Path(value)
                candidates: list[Path] = []

                if raw_path.is_absolute():
                    candidates.append(raw_path)
                else:
                    candidates.extend([
                        APP_DIR / raw_path,
                        DOKUMENTE_DIR / raw_path,
                        DATEN_DIR / raw_path,
                        raw_path,
                    ])

                existing_path = next(
                    (candidate for candidate in candidates if candidate.exists()),
                    None,
                )
                resolved_path = existing_path or candidates[0]

                title = Path(value).name or feld
                status = "Vorhanden" if existing_path is not None else "Fehlt"

                unique_key = (
                    bereich,
                    str(row_index),
                    str(resolved_path),
                )
                if unique_key in seen:
                    continue
                seen.add(unique_key)

                result.append({
                    "bereich": bereich,
                    "zeile": str(row_index),
                    "feld": feld,
                    "titel": title,
                    "datei": Path(value).name,
                    "pfad": str(resolved_path),
                    "objekt": objekt,
                    "wohnung": wohnung,
                    "mieter": mieter,
                    "status": status,
                    "dateityp": suffix.lstrip(".").upper() if suffix else "DATEI",
                    "inhalt": row_text,
                })

    result.sort(
        key=lambda item: (
            item.get("status", "") != "Fehlt",
            item.get("bereich", "").lower(),
            item.get("titel", "").lower(),
        )
    )
    return result

def dokument_kategorie(feld: str, datei: str) -> str:
    text = f"{feld} {datei}".lower()

    if "mietvertrag" in text:
        return "Mietvertrag"
    if "wohnungsgeber" in text:
        return "Wohnungsgeberauskunft"
    if "übergabe" in text or "uebergabe" in text:
        return "Übergabeprotokoll"
    if "betrieb" in text or "bk" in text:
        return "Betriebskosten"
    if "rechnung" in text:
        return "Rechnung"
    if "versicherung" in text:
        return "Versicherung"
    if "grundsteuer" in text:
        return "Grundsteuer"
    if "energie" in text:
        return "Energieausweis"
    if "foto" in text or "bild" in text:
        return "Foto"
    return "Sonstige"


def dokumenten_status_liste(entity_type: str, name: str) -> list[tuple[str, bool]]:
    docs = dokumenten_treffer()
    name_key = norm_key(name)

    def passt(doc: dict[str, str]) -> bool:
        if entity_type == "Objekt":
            return name_key and name_key in norm_key(doc.get("objekt", "") + " " + doc.get("inhalt", ""))
        if entity_type == "Wohnung":
            return name_key and name_key in norm_key(doc.get("wohnung", "") + " " + doc.get("inhalt", ""))
        if entity_type == "Mieter":
            return name_key and name_key in norm_key(doc.get("mieter", "") + " " + doc.get("inhalt", ""))
        return name_key in norm_key(doc.get("inhalt", ""))

    relevante = [d for d in docs if passt(d)]
    kategorien = {dokument_kategorie(d.get("feld", ""), d.get("datei", "")) for d in relevante}

    pflicht = ["Mietvertrag", "Wohnungsgeberauskunft", "Übergabeprotokoll", "Betriebskosten", "Rechnung"]
    return [(p, p in kategorien) for p in pflicht]


class AktenCenterSeite(QWidget):
    """Version 3.0: Digitale Akten- und Dokumentenübersicht."""

    def __init__(self):
        super().__init__()
        self.docs: list[dict[str, str]] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 22)
        root.setSpacing(14)

        title = QLabel("Akten- & Dokumenten-Center")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel("Zentrale digitale Objekt-, Wohnungs- und Mieterakte. Excel-Struktur bleibt unverändert.")
        info.setObjectName("subTitle")
        root.addWidget(info)

        filter_row = QHBoxLayout()

        self.entity_type = QComboBox()
        self.entity_type.addItems(["Alle", "Objekt", "Wohnung", "Mieter"])

        self.entity_name = QComboBox()
        self.entity_name.setEditable(True)

        self.search = QLineEdit()
        self.search.setPlaceholderText("Dokument, Kategorie, Objekt, Wohnung, Mieter oder Dateiname suchen ...")

        aktualisieren = QPushButton("Aktualisieren")
        aktualisieren.setObjectName("primaryButton")
        aktualisieren.clicked.connect(self.laden)

        oeffnen = QPushButton("PDF öffnen")
        oeffnen.clicked.connect(self.pdf_oeffnen)

        akte = QPushButton("Akte öffnen")
        akte.clicked.connect(self.akte_oeffnen)

        export = QPushButton("Liste exportieren")
        export.clicked.connect(self.export_excel)

        self.entity_type.currentTextChanged.connect(self.entity_liste_fuellen)

        filter_row.addWidget(QLabel("Akte:"))
        filter_row.addWidget(self.entity_type)
        filter_row.addWidget(self.entity_name)
        filter_row.addWidget(self.search, 1)
        filter_row.addWidget(aktualisieren)
        filter_row.addWidget(oeffnen)
        filter_row.addWidget(akte)
        filter_row.addWidget(export)

        root.addLayout(filter_row)

        self.tabs = QTabWidget()
        root.addWidget(self.tabs, 1)

        self.table = QTableWidget()
        self.status_table = QTableWidget()
        self.kategorie_table = QTableWidget()

        self.tabs.addTab(self.table, "Dokumente")
        self.tabs.addTab(self.status_table, "Dokumentenstatus")
        self.tabs.addTab(self.kategorie_table, "Kategorien")

        self.search.textChanged.connect(self.laden)
        self.entity_name.currentTextChanged.connect(self.laden)

        self.entity_liste_fuellen()
        self.laden()

    def entity_liste_fuellen(self) -> None:
        typ = self.entity_type.currentText()
        self.entity_name.blockSignals(True)
        self.entity_name.clear()
        self.entity_name.addItem("")

        werte: set[str] = set()

        if typ == "Objekt":
            for row in DATA.get("Objekte", []):
                if row:
                    werte.add(str(row[0]))
            for row in DATA.get("Wohnungen", []):
                if row:
                    werte.add(str(row[0]))
        elif typ == "Wohnung":
            for row in DATA.get("Wohnungen", []):
                if len(row) > 1:
                    werte.add(str(row[1]))
        elif typ == "Mieter":
            for row in DATA.get("Mieter", []):
                if row:
                    werte.add(str(row[0]))

        for wert in sorted(w for w in werte if w.strip()):
            self.entity_name.addItem(wert)

        self.entity_name.blockSignals(False)
        self.laden()

    def _gefilterte_docs(self) -> list[dict[str, str]]:
        docs = dokumenten_treffer()
        typ = self.entity_type.currentText()
        name = self.entity_name.currentText().strip()
        suche = self.search.text().strip().lower()

        result = []

        for doc in docs:
            combined = " ".join(str(v) for v in doc.values()).lower()

            if suche and suche not in combined:
                continue

            if typ == "Objekt" and name:
                if norm_key(name) not in norm_key(doc.get("objekt", "") + " " + doc.get("inhalt", "")):
                    continue
            elif typ == "Wohnung" and name:
                if norm_key(name) not in norm_key(doc.get("wohnung", "") + " " + doc.get("inhalt", "")):
                    continue
            elif typ == "Mieter" and name:
                if norm_key(name) not in norm_key(doc.get("mieter", "") + " " + doc.get("inhalt", "")):
                    continue

            result.append(doc)

        return result

    def laden(self) -> None:
        self.docs = self._gefilterte_docs()
        self._dokumente_tabelle()
        self._status_tabelle()
        self._kategorien_tabelle()

    def _dokumente_tabelle(self) -> None:
        headers = ["Kategorie", "Bereich", "Zeile", "Feld", "Datei", "Objekt", "Wohnung", "Mieter", "Pfad"]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(self.docs))
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)

        for r, doc in enumerate(self.docs):
            values = [
                dokument_kategorie(doc.get("feld", ""), doc.get("datei", "")),
                doc.get("bereich", ""),
                doc.get("zeile", ""),
                doc.get("feld", ""),
                doc.get("datei", ""),
                doc.get("objekt", ""),
                doc.get("wohnung", ""),
                doc.get("mieter", ""),
                doc.get("pfad", ""),
            ]
            for c, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                if c == 8:
                    item.setToolTip(str(value))
                self.table.setItem(r, c, item)

        widths = [150, 160, 70, 180, 260, 200, 160, 200, 420]
        for c, w in enumerate(widths):
            self.table.setColumnWidth(c, w)

    def _status_tabelle(self) -> None:
        typ = self.entity_type.currentText()
        name = self.entity_name.currentText().strip()

        if typ == "Alle" or not name:
            self.status_table.setColumnCount(2)
            self.status_table.setHorizontalHeaderLabels(["Hinweis", "Wert"])
            self.status_table.setRowCount(1)
            self.status_table.setItem(0, 0, QTableWidgetItem("Bitte Akte wählen"))
            self.status_table.setItem(0, 1, QTableWidgetItem("Objekt, Wohnung oder Mieter auswählen"))
            return

        status = dokumenten_status_liste(typ, name)
        self.status_table.setColumnCount(3)
        self.status_table.setHorizontalHeaderLabels(["Dokument", "Status", "Bewertung"])
        self.status_table.setRowCount(len(status))
        self.status_table.setAlternatingRowColors(True)

        for r, (dok, vorhanden) in enumerate(status):
            self.status_table.setItem(r, 0, QTableWidgetItem(dok))
            self.status_table.setItem(r, 1, QTableWidgetItem("Vorhanden" if vorhanden else "Fehlt"))
            self.status_table.setItem(r, 2, QTableWidgetItem("🟢 vollständig" if vorhanden else "🔴 fehlt"))

        self.status_table.setColumnWidth(0, 260)
        self.status_table.setColumnWidth(1, 160)
        self.status_table.setColumnWidth(2, 180)

    def _kategorien_tabelle(self) -> None:
        counts: dict[str, int] = {}

        for doc in self.docs:
            kat = dokument_kategorie(doc.get("feld", ""), doc.get("datei", ""))
            counts[kat] = counts.get(kat, 0) + 1

        rows = sorted(counts.items())
        self.kategorie_table.setColumnCount(2)
        self.kategorie_table.setHorizontalHeaderLabels(["Kategorie", "Anzahl"])
        self.kategorie_table.setRowCount(len(rows))
        self.kategorie_table.setAlternatingRowColors(True)

        for r, (kat, count) in enumerate(rows):
            self.kategorie_table.setItem(r, 0, QTableWidgetItem(kat))
            self.kategorie_table.setItem(r, 1, QTableWidgetItem(str(count)))

        self.kategorie_table.setColumnWidth(0, 280)
        self.kategorie_table.setColumnWidth(1, 120)

    def aktueller_pfad(self) -> str:
        row = self.table.currentRow()
        if row < 0:
            return ""
        item = self.table.item(row, 8)
        return item.text() if item else ""

    def pdf_oeffnen(self) -> None:
        pfad = self.aktueller_pfad()
        if not pfad:
            QMessageBox.information(self, "PDF öffnen", "Bitte zuerst ein Dokument auswählen.")
            return
        system_datei_oeffnen(pfad)

    def akte_oeffnen(self) -> None:
        typ = self.entity_type.currentText()
        name = self.entity_name.currentText().strip()

        if typ == "Alle" or not name:
            QMessageBox.information(self, "Akte öffnen", "Bitte Objekt, Wohnung oder Mieter auswählen.")
            return

        ordner = DOKUMENTE_DIR / typ.lower() / safe_filename(name)
        ordner.mkdir(parents=True, exist_ok=True)
        system_datei_oeffnen(str(ordner))

    def export_excel(self) -> None:
        if not self.docs:
            QMessageBox.information(self, "Export", "Keine Dokumente zum Exportieren.")
            return


        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        ziel_default = EXPORT_DIR / f"dokumentenliste_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

        ziel_text, _ = QFileDialog.getSaveFileName(
            self,
            "Dokumentenliste exportieren",
            str(ziel_default),
            "Excel-Dateien (*.xlsx)",
        )
        if not ziel_text:
            return

        ziel = Path(ziel_text)
        if ziel.suffix.lower() != ".xlsx":
            ziel = ziel.with_suffix(".xlsx")

        wb = Workbook()
        ws = wb.active

        if not isinstance(ws, Worksheet):
            QMessageBox.warning(self, "Export", "Excel-Arbeitsblatt konnte nicht erstellt werden.")
            return

        ws.title = "Dokumentenliste"
        ws.append(["Kategorie", "Bereich", "Zeile", "Feld", "Datei", "Objekt", "Wohnung", "Mieter", "Pfad"])

        for doc in self.docs:
            ws.append([
                dokument_kategorie(doc.get("feld", ""), doc.get("datei", "")),
                doc.get("bereich", ""),
                doc.get("zeile", ""),
                doc.get("feld", ""),
                doc.get("datei", ""),
                doc.get("objekt", ""),
                doc.get("wohnung", ""),
                doc.get("mieter", ""),
                doc.get("pfad", ""),
            ])

        for col_index, _col in enumerate(ws.columns, start=1):
            ws.column_dimensions[get_column_letter(col_index)].width = 24

        wb.save(ziel)
        QMessageBox.information(self, "Export", f"Dokumentenliste exportiert:\n{ziel}")
        system_datei_oeffnen(str(ziel.parent))


def globale_suche_treffer(suchtext: str) -> list[dict[str, str]]:
    suche = str(suchtext or "").strip().lower()
    if not suche:
        return []

    treffer: list[dict[str, str]] = []

    for bereich, rows in DATA.items():
        felder = SCHEMA.get(bereich, [])

        for row_index, row in enumerate(rows, start=1):
            zellen = ["" if v is None else str(v) for v in row]
            combined = " ".join(zellen).lower()

            if suche not in combined:
                continue

            passende_felder = []
            for i, value in enumerate(zellen):
                if suche in value.lower():
                    feld = felder[i] if i < len(felder) else f"Spalte {i+1}"
                    passende_felder.append(str(feld))

            hauptwert = ""
            for kandidat in ["Mieter", "Objekt", "Objektname", "Wohnung", "Versorger", "Kundennummer", "Vertragsnummer", "Rechnungsnr.", "Rechnungsnummer", "Titel", "Aufgabe", "Schaden"]:
                idx = feld_index(bereich, kandidat) if "feld_index" in globals() else None
                if idx is not None and idx < len(zellen) and zellen[idx]:
                    hauptwert = zellen[idx]
                    break

            if not hauptwert and zellen:
                hauptwert = zellen[0]

            treffer.append({
                "bereich": bereich,
                "zeile": str(row_index),
                "hauptwert": hauptwert,
                "felder": ", ".join(passende_felder) if passende_felder else "Volltext",
                "inhalt": " | ".join(zellen[:12]),
            })

    return treffer


class GlobaleSucheSeite(QWidget):
    """Zentrale intelligente Suche über alle Tabellen."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.treffer: list[dict[str, str]] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 22)
        root.setSpacing(14)

        title = QLabel("Globale Suche")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel("Eine Suche über Mieter, Objekte, Wohnungen, Rechnungen, Kundennummern, Vertragsnummern, Dokumente und Akten.")
        info.setObjectName("subTitle")
        root.addWidget(info)

        row = QHBoxLayout()
        self.search = QLineEdit()
        self.search.setPlaceholderText("Suchbegriff eingeben, z. B. Mieter, Objekt, Kundennummer, Vertragsnummer, PDF ...")
        self.search.returnPressed.connect(self.suchen)

        btn = QPushButton("Suchen")
        btn.setObjectName("primaryButton")
        btn.clicked.connect(self.suchen)

        open_btn = QPushButton("Bereich öffnen")
        open_btn.clicked.connect(self.bereich_oeffnen)

        export_btn = QPushButton("Treffer exportieren")
        export_btn.clicked.connect(self.export_excel)

        row.addWidget(self.search, 1)
        row.addWidget(btn)
        row.addWidget(open_btn)
        row.addWidget(export_btn)
        root.addLayout(row)

        self.summary = QLabel("Noch keine Suche ausgeführt.")
        self.summary.setObjectName("metricTitle")
        root.addWidget(self.summary)

        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Bereich", "Zeile", "Hauptwert", "Trefferfeld", "Inhalt"])
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.cellDoubleClicked.connect(lambda _r, _c: self.bereich_oeffnen())

        root.addWidget(self.table, 1)

        quick = QHBoxLayout()
        for label in ["Offen", "Miete", "Nachzahlung", "Guthaben"]:
            b = QPushButton(label)
            b.clicked.connect(lambda checked=False, text=label: self.quick_search(text))
            quick.addWidget(b)
        quick.addStretch()
        root.addLayout(quick)

    def quick_search(self, text: str) -> None:
        self.search.setText(text)
        self.suchen()

    def suchen(self) -> None:
        suche = self.search.text().strip()
        self.treffer = globale_suche_treffer(suche)
        self.summary.setText(f"Suchbegriff: {suche or '-'} | Treffer: {len(self.treffer)}")

        self.table.setRowCount(len(self.treffer))

        for r, item in enumerate(self.treffer):
            values = [
                item.get("bereich", ""),
                item.get("zeile", ""),
                item.get("hauptwert", ""),
                item.get("felder", ""),
                item.get("inhalt", ""),
            ]
            for c, value in enumerate(values):
                self.table.setItem(r, c, QTableWidgetItem(str(value)))

        widths = [180, 70, 260, 260, 760]
        for c, w in enumerate(widths):
            self.table.setColumnWidth(c, w)

    def aktueller_bereich(self) -> str:
        row = self.table.currentRow()
        if row < 0:
            return ""
        item = self.table.item(row, 0)
        return item.text() if item else ""

    def bereich_oeffnen(self) -> None:
        bereich = self.aktueller_bereich()
        if not bereich:
            QMessageBox.information(self, "Bereich öffnen", "Bitte zuerst einen Treffer auswählen.")
            return
        self.nav(bereich)

    def export_excel(self) -> None:
        if not self.treffer:
            QMessageBox.information(self, "Export", "Keine Treffer zum Exportieren.")
            return


        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        ziel_default = EXPORT_DIR / f"globale_suche_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

        ziel_text, _ = QFileDialog.getSaveFileName(
            self,
            "Suchtreffer exportieren",
            str(ziel_default),
            "Excel-Dateien (*.xlsx)",
        )
        if not ziel_text:
            return

        ziel = Path(ziel_text)
        if ziel.suffix.lower() != ".xlsx":
            ziel = ziel.with_suffix(".xlsx")

        wb = Workbook()
        ws = wb.active

        if not isinstance(ws, Worksheet):
            QMessageBox.warning(self, "Export", "Excel-Arbeitsblatt konnte nicht erstellt werden.")
            return

        ws.title = "Globale Suche"
        ws.append(["Bereich", "Zeile", "Hauptwert", "Trefferfeld", "Inhalt"])

        for item in self.treffer:
            ws.append([
                item.get("bereich", ""),
                item.get("zeile", ""),
                item.get("hauptwert", ""),
                item.get("felder", ""),
                item.get("inhalt", ""),
            ])

        for col_index, _col in enumerate(ws.columns, start=1):
            ws.column_dimensions[get_column_letter(col_index)].width = 24

        wb.save(ziel)
        QMessageBox.information(self, "Export", f"Suchtreffer exportiert:\n{ziel}")
        system_datei_oeffnen(str(ziel.parent))


def core_count(titel: str) -> int:
    return len(DATA.get(titel, []))


def core_text(row: list[Any]) -> str:
    return " ".join("" if v is None else str(v) for v in row).lower()


def core_status_count(titel: str, words: list[str]) -> int:
    return sum(1 for row in DATA.get(titel, []) if any(w.lower() in core_text(row) for w in words))


def core_sum(titel: str, feldnamen: list[str]) -> float:
    felder = SCHEMA.get(titel, [])
    indizes = []
    for name in feldnamen:
        for i, feld in enumerate(felder):
            if norm_key(feld) == norm_key(name):
                indizes.append(i)
    total = 0.0
    for row in DATA.get(titel, []):
        for i in indizes:
            if i < len(row):
                try:
                    total += to_float(row[i])
                except (OSError, ValueError, TypeError, AttributeError, RuntimeError, KeyError, IndexError):
                    pass
    return total


def cockpit_werte() -> dict[str, float]:
    einnahmen = core_sum("HV-Rechnungen", ["Betrag brutto", "Betrag netto"])
    ausgaben = core_sum("Rechnungen", ["Betrag brutto", "Betrag netto"])
    return {
        "objekte": float(core_count("Objekte")),
        "wohnungen": float(core_count("Wohnungen")),
        "mieter": float(core_count("Mieter")),
        "vermietet": float(core_status_count("Wohnungen", ["vermietet", "belegt"])),
        "frei": float(core_status_count("Wohnungen", ["frei", "leer", "unvermietet"])),
        "sollmiete": core_sum("Mieter", ["Miete", "Warmmiete"]),
        "kaltmiete": core_sum("Mieter", ["Kaltmiete"]),
        "bk_voraus": core_sum("Mieter", ["BK monatlich"]) * 12,
        "bk_nachzahlung": core_sum("Mieter", ["Nachzahlung"]),
        "bk_guthaben": core_sum("Mieter", ["Guthaben"]),
        "einnahmen": einnahmen,
        "ausgaben": ausgaben,
        "saldo": einnahmen - ausgaben,
        "offene_rechnungen": float(core_status_count("Rechnungen", ["offen"])),
        "offene_schaeden": float(core_status_count("Schäden", ["offen", "in bearbeitung"])),
        "offene_aufgaben": float(core_status_count("Aufgaben", ["offen", "neu", "in bearbeitung"])),
    }





class VerwaltungsCockpitSeite(QWidget):
    def __init__(self, nav):
        super().__init__()
        self.nav = nav

        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 22)
        root.setSpacing(14)

        title = QLabel("Verwaltungs-Cockpit PRO")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel("Live-Kennzahlen, Benachrichtigungen und Finanzübersicht.")
        info.setObjectName("subTitle")
        root.addWidget(info)

        top = QHBoxLayout()
        refresh = QPushButton("Aktualisieren")
        refresh.setObjectName("primaryButton")
        refresh.clicked.connect(self.laden)
        event_btn = QPushButton("Ereignisprotokoll")
        event_btn.clicked.connect(lambda: self.nav("Ereignisprotokoll"))
        top.addWidget(refresh)
        top.addWidget(event_btn)
        top.addStretch()
        root.addLayout(top)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        root.addWidget(scroll, 1)

        inner = QWidget()
        self.inner_layout = QVBoxLayout(inner)
        self.inner_layout.setContentsMargins(4, 4, 14, 4)
        self.inner_layout.setSpacing(18)
        scroll.setWidget(inner)

        self.grid = QGridLayout()
        self.grid.setSpacing(14)
        self.inner_layout.addLayout(self.grid)

        self.tabs = QTabWidget()
        self.table = QTableWidget()
        self.notifications = QTableWidget()
        self.chart = QGraphicsView()
        self.scene = QGraphicsScene()
        self.chart.setScene(self.scene)

        self.tabs.addTab(self.table, "Kennzahlen")
        self.tabs.addTab(self.notifications, "Benachrichtigungen")
        self.tabs.addTab(self.chart, "Diagramm")
        self.inner_layout.addWidget(self.tabs, 1)

        self.laden()

    def _card(self, titel: str, wert: str, icon: str, ziel: str = "") -> QFrame:
        card = QFrame()
        card.setObjectName("metricCard")
        card.setMinimumHeight(118)
        card.setMinimumWidth(210)
        lay = QVBoxLayout(card)
        lay.setContentsMargins(14, 12, 14, 12)
        t = QLabel(f"{icon}  {titel}")
        t.setObjectName("metricTitle")
        t.setWordWrap(True)
        v = QLabel(wert)
        v.setObjectName("metricValue")
        v.setWordWrap(True)
        lay.addWidget(t)
        lay.addWidget(v)
        if ziel:
            b = QPushButton("Öffnen")
            b.clicked.connect(lambda checked=False, z=ziel: self.nav(z))
            lay.addWidget(b)
        return card

    def laden(self) -> None:
        k = cockpit_werte()

        while self.grid.count():
            item = self.grid.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        cards = [
            ("Objekte", str(int(k["objekte"])), "🏢", "Objekte"),
            ("Wohnungen", str(int(k["wohnungen"])), "🏠", "Wohnungen"),
            ("Mieter", str(int(k["mieter"])), "👤", "Mieter"),
            ("Vermietet", str(int(k["vermietet"])), "✅", "Wohnungen"),
            ("Frei", str(int(k["frei"])), "🚪", "Wohnungen"),
            ("Sollmiete", euro(k["sollmiete"]), "💰", "Mieter"),
            ("BK Voraus", euro(k["bk_voraus"]), "📄", "Betriebskosten"),
            ("BK Nachzahlung", euro(k["bk_nachzahlung"]), "↗", "BK-Automatik"),
            ("BK Guthaben", euro(k["bk_guthaben"]), "↘", "BK-Automatik"),
            ("Einnahmen", euro(k["einnahmen"]), "💶", "Buchhaltung"),
            ("Ausgaben", euro(k["ausgaben"]), "💸", "Buchhaltung"),
            ("Saldo", euro(k["saldo"]), "Σ", "Buchhaltung"),
            ("Offene Rechnungen", str(int(k["offene_rechnungen"])), "🧾", "Rechnungen"),
            ("Offene Schäden", str(int(k["offene_schaeden"])), "⚠", "Schäden"),
            ("Offene Aufgaben", str(int(k["offene_aufgaben"])), "✓", "Aufgaben"),
        ]

        for i, data in enumerate(cards):
            self.grid.addWidget(self._card(*data), i // 5, i % 5)

        self._table(k)
        self._notifications(k)
        self._chart(k)

    def _table(self, k: dict[str, float]) -> None:
        rows = list(k.items())
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["Kennzahl", "Wert"])
        self.table.setRowCount(len(rows))
        self.table.setAlternatingRowColors(True)
        geld = {"sollmiete", "kaltmiete", "bk_voraus", "bk_nachzahlung", "bk_guthaben", "einnahmen", "ausgaben", "saldo"}
        for r, (name, value) in enumerate(rows):
            self.table.setItem(r, 0, QTableWidgetItem(str(name)))
            self.table.setItem(r, 1, QTableWidgetItem(euro(float(value)) if name in geld else str(int(value))))
        self.table.setColumnWidth(0, 260)
        self.table.setColumnWidth(1, 180)

    def _notifications(self, k: dict[str, float]) -> None:
        rows = []
        if k.get("offene_rechnungen", 0) > 0:
            rows.append(["🔴", "Offene Rechnungen", f'{int(k["offene_rechnungen"])} offen', "Rechnungen"])
        if k.get("offene_schaeden", 0) > 0:
            rows.append(["🟡", "Offene Schäden", f'{int(k["offene_schaeden"])} offen', "Schäden"])
        if k.get("offene_aufgaben", 0) > 0:
            rows.append(["🟡", "Offene Aufgaben", f'{int(k["offene_aufgaben"])} offen', "Aufgaben"])
        if k.get("frei", 0) > 0:
            rows.append(["🔵", "Freie Wohnungen", f'{int(k["frei"])} frei', "Wohnungen"])
        if k.get("bk_nachzahlung", 0) > 0:
            rows.append(["⚠", "BK Nachzahlungen", euro(k["bk_nachzahlung"]), "BK-Automatik"])
        if not rows:
            rows.append(["🟢", "Alles ruhig", "Keine kritischen Vorgänge.", "Dashboard"])

        self.notifications.setColumnCount(4)
        self.notifications.setHorizontalHeaderLabels(["Status", "Bereich", "Hinweis", "Ziel"])
        self.notifications.setRowCount(len(rows))
        self.notifications.setAlternatingRowColors(True)
        for r, row in enumerate(rows):
            for c, value in enumerate(row):
                self.notifications.setItem(r, c, QTableWidgetItem(str(value)))
        self.notifications.setColumnWidth(0, 80)
        self.notifications.setColumnWidth(1, 200)
        self.notifications.setColumnWidth(2, 420)
        self.notifications.setColumnWidth(3, 180)
        if not getattr(self, "_notifications_signal_connected", False):
            self.notifications.cellDoubleClicked.connect(self._notification_oeffnen)
            self._notifications_signal_connected = True

    def _notification_oeffnen(self, row: int, _column: int) -> None:
        item = self.notifications.item(row, 3)
        if item is not None and item.text().strip():
            self.nav(item.text().strip())

    def _chart(self, k: dict[str, float]) -> None:
        self.scene.clear()
        self.scene.addText("Finanzübersicht").setPos(30, 10)
        daten = [("Einnahmen", k["einnahmen"]), ("Ausgaben", k["ausgaben"]), ("Saldo", k["saldo"]), ("Sollmiete", k["sollmiete"])]
        maxv = max([abs(v) for _, v in daten] + [1.0])
        y = 70
        for label, value in daten:
            width = int(abs(value) / maxv * 650)
            self.scene.addText(label).setPos(30, y - 22)
            self.scene.addRect(30, y, max(width, 4), 32, QPen(), QBrush())
            self.scene.addText(euro(value)).setPos(30 + max(width, 4) + 15, y + 4)
            y += 70
        self.scene.setSceneRect(0, 0, 900, 360)


def objekt_liste() -> list[str]:
    werte = set()
    for row in DATA.get("Objekte", []):
        if row and str(row[0]).strip():
            werte.add(str(row[0]).strip())
    for row in DATA.get("Wohnungen", []):
        if row and str(row[0]).strip():
            werte.add(str(row[0]).strip())
    return sorted(werte)


def objekt_rows(titel: str, objekt: str) -> list[list[Any]]:
    key = norm_key(objekt)
    if not key:
        return DATA.get(titel, [])
    rows = []
    for row in DATA.get(titel, []):
        text = norm_key(" ".join(str(v) for v in row))
        if key in text:
            rows.append(row)
    return rows


def objekt_sum(titel: str, objekt: str, feldnamen: list[str]) -> float:
    felder = SCHEMA.get(titel, [])
    idxs = []
    for name in feldnamen:
        for i, feld in enumerate(felder):
            if norm_key(feld) == norm_key(name):
                idxs.append(i)
    total = 0.0
    for row in objekt_rows(titel, objekt):
        for i in idxs:
            if i < len(row):
                try:
                    total += to_float(row[i])
                except (OSError, ValueError, TypeError, AttributeError, RuntimeError, KeyError, IndexError):
                    pass
    return total


class ObjektCockpitSeite(QWidget):
    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 22)
        root.setSpacing(14)

        title = QLabel("Objekt-Cockpit")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel("Alle wichtigen Daten eines Objekts auf einer Seite.")
        info.setObjectName("subTitle")
        root.addWidget(info)

        top = QHBoxLayout()
        self.objekt = QComboBox()
        self.objekt.setEditable(True)
        self.objekt.addItems(objekt_liste())
        self.objekt.currentTextChanged.connect(self.laden)

        refresh = QPushButton("Aktualisieren")
        refresh.setObjectName("primaryButton")
        refresh.clicked.connect(self.laden)

        top.addWidget(QLabel("Objekt:"))
        top.addWidget(self.objekt, 1)
        top.addWidget(refresh)
        root.addLayout(top)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        root.addWidget(scroll, 1)

        inner = QWidget()
        self.layout = QVBoxLayout(inner)
        self.layout.setContentsMargins(4, 4, 14, 4)
        self.layout.setSpacing(16)
        scroll.setWidget(inner)

        self.cards = QGridLayout()
        self.cards.setSpacing(14)
        self.layout.addLayout(self.cards)

        self.tabs = QTabWidget()
        self.layout.addWidget(self.tabs, 1)

        self.tbl_wohnungen = QTableWidget()
        self.tbl_mieter = QTableWidget()
        self.tbl_rechnungen = QTableWidget()
        self.tbl_schaeden = QTableWidget()
        self.tbl_dokumente = QTableWidget()
        self.tbl_historie = QTableWidget()

        self.tabs.addTab(self.tbl_wohnungen, "Wohnungen")
        self.tabs.addTab(self.tbl_mieter, "Mieter")
        self.tabs.addTab(self.tbl_rechnungen, "Rechnungen")
        self.tabs.addTab(self.tbl_schaeden, "Schäden")
        self.tabs.addTab(self.tbl_dokumente, "Dokumente")
        self.tabs.addTab(self.tbl_historie, "Historie")

        self.laden()

    def _card(self, titel: str, wert: str, icon: str, ziel: str = "") -> QFrame:
        card = QFrame()
        card.setObjectName("metricCard")
        card.setMinimumHeight(105)
        card.setMinimumWidth(210)
        lay = QVBoxLayout(card)
        t = QLabel(f"{icon}  {titel}")
        t.setObjectName("metricTitle")
        v = QLabel(wert)
        v.setObjectName("metricValue")
        v.setWordWrap(True)
        lay.addWidget(t)
        lay.addWidget(v)
        if ziel:
            b = QPushButton("Öffnen")
            b.clicked.connect(lambda checked=False, z=ziel: self.nav(z))
            lay.addWidget(b)
        return card

    def laden(self) -> None:
        objekt = self.objekt.currentText().strip()
        while self.cards.count():
            item = self.cards.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        wohnungen = objekt_rows("Wohnungen", objekt)
        mieter = objekt_rows("Mieter", objekt)
        rechnungen = objekt_rows("Rechnungen", objekt)
        schaeden = objekt_rows("Schäden", objekt)
        aufgaben = objekt_rows("Aufgaben", objekt)

        miete = objekt_sum("Mieter", objekt, ["Miete", "Warmmiete"])
        bk = objekt_sum("Mieter", objekt, ["BK monatlich"]) * 12
        re_summe = objekt_sum("Rechnungen", objekt, ["Betrag brutto", "Betrag netto"])

        data = [
            ("Wohnungen", str(len(wohnungen)), "🏠", "Wohnungen"),
            ("Mieter", str(len(mieter)), "👤", "Mieter"),
            ("Miete Monat", euro(miete), "💰", "Mieter"),
            ("BK Voraus Jahr", euro(bk), "📄", "Betriebskosten"),
            ("Rechnungen", str(len(rechnungen)), "🧾", "Rechnungen"),
            ("Rechnungssumme", euro(re_summe), "💸", "Rechnungen"),
            ("Schäden", str(len(schaeden)), "⚠", "Schäden"),
            ("Aufgaben", str(len(aufgaben)), "✓", "Aufgaben"),
        ]
        for i, d in enumerate(data):
            self.cards.addWidget(self._card(*d), i // 4, i % 4)

        self._fill(self.tbl_wohnungen, "Wohnungen", wohnungen)
        self._fill(self.tbl_mieter, "Mieter", mieter)
        self._fill(self.tbl_rechnungen, "Rechnungen", rechnungen)
        self._fill(self.tbl_schaeden, "Schäden", schaeden)
        self._dokumente(objekt)
        self._historie(objekt)

    @staticmethod
    def _fill(table: QTableWidget, titel: str, rows: list[list[Any]]) -> None:
        headers = SCHEMA.get(titel, [])
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setRowCount(len(rows))
        table.setAlternatingRowColors(True)
        table.setWordWrap(False)
        table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        for r, row in enumerate(rows):
            for c in range(len(headers)):
                table.setItem(r, c, QTableWidgetItem(str(row[c] if c < len(row) else "")))
        for c in range(min(len(headers), 12)):
            table.setColumnWidth(c, 160)


    def _dokumente(self, objekt: str) -> None:
        docs = []
        try:
            for d in dokumenten_treffer():
                if norm_key(objekt) in norm_key(d.get("objekt", "") + " " + d.get("inhalt", "")):
                    docs.append(d)
        except (OSError, ValueError, TypeError, AttributeError, RuntimeError, KeyError, IndexError):
            docs = []

        headers = ["Kategorie", "Bereich", "Datei", "Feld", "Pfad"]
        self.tbl_dokumente.setColumnCount(len(headers))
        self.tbl_dokumente.setHorizontalHeaderLabels(headers)
        self.tbl_dokumente.setRowCount(len(docs))
        self.tbl_dokumente.setAlternatingRowColors(True)
        self.tbl_dokumente.setWordWrap(False)
        self.tbl_dokumente.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.tbl_dokumente.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)

        for r, d in enumerate(docs):
            values = [
                dokument_kategorie(d.get("feld", ""), d.get("datei", "")) if "dokument_kategorie" in globals() else "Dokument",
                d.get("bereich", ""),
                d.get("datei", ""),
                d.get("feld", ""),
                d.get("pfad", ""),
            ]
            for c, value in enumerate(values):
                self.tbl_dokumente.setItem(r, c, QTableWidgetItem(str(value)))

        widths = [170, 160, 260, 180, 420]
        for c, w in enumerate(widths):
            self.tbl_dokumente.setColumnWidth(c, w)


    def _historie(self, objekt: str) -> None:
        rows = []
        for titel in ["Rechnungen", "Schäden", "Fristen", "Aufgaben", "Betriebskosten", "Zahlungen"]:
            for row in objekt_rows(titel, objekt):
                datum = ""
                for i, feld in enumerate(SCHEMA.get(titel, [])):
                    if i < len(row) and ("datum" in norm_key(feld) or "faellig" in norm_key(feld) or "fällig" in norm_key(feld)):
                        datum = str(row[i])
                        break
                rows.append([datum, titel, " | ".join(str(v) for v in row[:8])])
        self.tbl_historie.setColumnCount(3)
        self.tbl_historie.setHorizontalHeaderLabels(["Datum", "Bereich", "Ereignis"])
        self.tbl_historie.setRowCount(len(rows))
        self.tbl_historie.setAlternatingRowColors(True)
        for r, row in enumerate(rows):
            for c, v in enumerate(row):
                self.tbl_historie.setItem(r, c, QTableWidgetItem(str(v)))
        self.tbl_historie.setColumnWidth(0, 120)
        self.tbl_historie.setColumnWidth(1, 160)
        self.tbl_historie.setColumnWidth(2, 760)


def konto_betrag(value: Any) -> float:
    text = str(value or "").strip()
    text = text.replace("€", "").replace("EUR", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except (OSError, ValueError, TypeError, AttributeError, RuntimeError, KeyError, IndexError):
        return 0.0


def konto_csv_lesen(pfad: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    encodings = ["utf-8-sig", "cp1252", "latin1"]
    content = None

    for enc in encodings:
        try:
            content = Path(pfad).read_text(encoding=enc)
            break
        except (OSError, ValueError, TypeError, AttributeError, RuntimeError, KeyError, IndexError):
            continue

    if content is None:
        return rows

    delimiter = ";"
    if content.count(",") > content.count(";"):
        delimiter = ","

    reader = csv.DictReader(content.splitlines(), delimiter=delimiter)
    for row in reader:
        norm = {str(k or "").strip().lower(): str(v or "").strip() for k, v in row.items()}
        rows.append(norm)

    return rows


def konto_find(row: dict[str, str], names: list[str]) -> str:
    for name in names:
        n = name.lower()
        for key, value in row.items():
            if n in key:
                return value
    return ""





def zahlungsabgleich_score(text: str, mietername: str, objekt: str, wohnung: str, soll: float, betrag: float) -> tuple[int, list[str]]:
    key = norm_key(text)
    score = 0
    gr = []

    if mietername and norm_key(mietername) in key:
        score += 60
        gr.append("Name")
    else:
        # Teilnamensuche
        parts = [p for p in norm_key(mietername).split() if len(p) >= 4]
        hits = sum(1 for p in parts if p in key)
        if hits:
            score += 20 * hits
            gr.append("Teilname")

    if objekt and norm_key(objekt) in key:
        score += 20
        gr.append("Objekt")
    if wohnung and norm_key(wohnung) in key:
        score += 15
        gr.append("Wohnung")

    if soll > 0:
        diff = abs(abs(betrag) - soll)
        if diff <= 2:
            score += 35
            gr.append("Betrag exakt")
        elif diff <= max(25, soll * 0.15):
            score += 15
            gr.append("Betrag ähnlich")

    if any(w in key for w in ["miete", "mietzahlung", "warmmiete", "kaltmiete", "wohnung"]):
        score += 20
        gr.append("Mietbegriff")

    return score, gr


def zahlungsabgleich_mieter_profi(text: str, betrag: float) -> dict[str, str]:
    bester = {
        "mieter": "",
        "objekt": "",
        "wohnung": "",
        "status": "Unklar",
        "hinweis": "Kein eindeutiger Mieter erkannt",
        "score": "0",
        "soll": "0",
    }

    best_score = 0
    for row in DATA.get("Mieter", []):
        felder = SCHEMA.get("Mieter", [])

        def val(field_name: str) -> str:
            for i, feld in enumerate(felder):
                if norm_key(feld) == norm_key(field_name) and i < len(row):
                    return str(row[i] or "")
            return ""

        mieter = val("Mieter") or (str(row[0]) if row else "")
        ort = val("Ort")
        objekt = val("Objektordner")
        wohnung = val("Wohnungsordner")

        if not objekt and "/" in ort:
            objekt = ort.split("/")[0].strip()
        if not wohnung and "/" in ort:
            wohnung = ort.split("/")[-1].strip()

        soll = konto_betrag(val("Miete") or val("Warmmiete"))
        score, gruende = zahlungsabgleich_score(text, mieter, objekt, wohnung, soll, betrag)

        if score > best_score:
            best_score = score
            status = "Erkannt" if score >= 70 else ("Prüfen" if score >= 45 else "Unklar")
            bester = {
                "mieter": mieter,
                "objekt": objekt,
                "wohnung": wohnung,
                "status": status,
                "hinweis": f"Score {score}: {', '.join(gruende) if gruende else 'keine starken Treffer'}",
                "score": str(score),
                "soll": f"{soll:.2f}",
            }

    return bester


def zahlungsabgleich_status(soll: float, betrag: float) -> str:
    if soll <= 0:
        return "Prüfen"
    diff = abs(betrag) - soll
    if abs(diff) <= 2:
        return "Bezahlt"
    if diff < 0:
        return "Teilzahlung"
    return "Überzahlung"


def konto_import_klassifizieren(row: dict[str, str]) -> list[str]:
    datum = konto_find(row, ["datum", "buchungstag", "valuta"])
    text = konto_find(row, ["buchungstext", "text", "umsatzart"])
    auftraggeber = konto_find(row, ["auftraggeber", "name", "beguenstigter", "begünstigter"])
    zweck = konto_find(row, ["verwendungszweck", "zweck", "beschreibung"])
    betrag_text = konto_find(row, ["betrag", "umsatz", "wert"])
    betrag = konto_betrag(betrag_text)

    combined = f"{text} {auftraggeber} {zweck}"
    lower = combined.lower()

    match = zahlungsabgleich_mieter_profi(combined, betrag)
    erkannt = "Mieteinnahme" if betrag > 0 and (match["status"] in ["Erkannt", "Prüfen"] or any(w in lower for w in ["miete", "mietzahlung", "warmmiete", "kaltmiete"])) else "Sonstiges"

    if erkannt == "Mieteinnahme":
        soll = konto_betrag(match.get("soll", "0"))
        zahlstatus = zahlungsabgleich_status(soll, betrag)
        hinweis = f'{match.get("hinweis","")} | Zahlungsstatus: {zahlstatus}'
    else:
        match = {"mieter": "", "objekt": "", "wohnung": "", "status": "Ignoriert", "hinweis": "Keine Mieteinnahme"}
        hinweis = match["hinweis"]

    return [
        datum,
        text,
        auftraggeber,
        zweck,
        str(betrag),
        erkannt,
        match.get("mieter", ""),
        match.get("objekt", ""),
        match.get("wohnung", ""),
        match.get("status", ""),
        hinweis,
    ]




def konto_pdf_text_lesen(pfad: str) -> str:
    if PdfReader is None:
        return ""
    try:
        reader = PdfReader(pfad)
        return "\n".join((page.extract_text() or "") for page in reader.pages)
    except (OSError, ValueError, TypeError, AttributeError, RuntimeError, KeyError, IndexError):
        return ""


def konto_pdf_buchungen_erkennen(text: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    lines = [re.sub(r"\s+", " ", line).strip() for line in str(text or "").splitlines() if line.strip()]
    datum_pattern = re.compile(r"(\d{2}\.\d{2}\.\d{4}|\d{2}\.\d{2}\.\d{2})")
    betrag_pattern = re.compile(r"(-?\d{1,3}(?:\.\d{3})*,\d{2}|-?\d+,\d{2})")
    current_date = ""
    buffer = ""

    for line in lines:
        date_match = datum_pattern.search(line)
        amounts = betrag_pattern.findall(line)

        if date_match and amounts:
            datum = date_match.group(1)
            betrag = amounts[-1]
            beschreibung = line.replace(datum, " ", 1).replace(betrag, " ", 1).strip()
            rows.append({
                "datum": datum,
                "buchungstext": beschreibung,
                "auftraggeber": "",
                "verwendungszweck": beschreibung,
                "betrag": betrag,
            })
            current_date = ""
            buffer = ""
        elif date_match:
            current_date = date_match.group(1)
            buffer = line.replace(current_date, " ", 1).strip()
        elif current_date:
            buffer = (buffer + " " + line).strip()
            amounts = betrag_pattern.findall(buffer)
            if amounts:
                betrag = amounts[-1]
                beschreibung = buffer.replace(betrag, " ", 1).strip()
                rows.append({
                    "datum": current_date,
                    "buchungstext": beschreibung,
                    "auftraggeber": "",
                    "verwendungszweck": beschreibung,
                    "betrag": betrag,
                })
                current_date = ""
                buffer = ""

    return rows


def mj_aus_datum(value: Any) -> tuple[str, str]:
    parts = re.split(r"[./-]", str(value or "").strip())
    if len(parts) < 3:
        return "", ""
    if len(parts[0]) == 4:
        return parts[1].zfill(2), parts[0]
    jahr = parts[2] if len(parts[2]) == 4 else "20" + parts[2]
    return parts[1].zfill(2), jahr


def mieterwert(row: list[Any], feldname: str) -> str:
    for i, feld in enumerate(SCHEMA.get("Mieter", [])):
        if norm_key(feld) == norm_key(feldname) and i < len(row):
            return str(row[i] or "")
    return ""


def mietkonto_berechnen(monat: str, jahr: str) -> list[list[str]]:
    result: list[list[str]] = []
    for mrow in DATA.get("Mieter", []):
        name = mieterwert(mrow, "Mieter") or (str(mrow[0]) if mrow else "")
        objekt = mieterwert(mrow, "Objektordner")
        wohnung = mieterwert(mrow, "Wohnungsordner")
        ort = mieterwert(mrow, "Ort")
        if not objekt and "/" in ort:
            objekt = ort.split("/")[0].strip()
        if not wohnung and "/" in ort:
            wohnung = ort.split("/")[-1].strip()

        soll = konto_betrag(mieterwert(mrow, "Miete") or mieterwert(mrow, "Warmmiete"))
        gezahlt = 0.0

        for zahlung in DATA.get("Zahlungen", []):
            zm, zj = mj_aus_datum(zahlung[0] if zahlung else "")
            ztext = " ".join(str(v) for v in zahlung)
            if monat and zm != monat:
                continue
            if jahr and zj != jahr:
                continue
            if norm_key(name) in norm_key(ztext) and "miete" in norm_key(ztext):
                gezahlt += konto_betrag(zahlung[2] if len(zahlung) > 2 else "")

        for imp in DATA.get("Kontoauszug-Import", []):
            if len(imp) < 10:
                continue
            im, ij = mj_aus_datum(imp[0])
            if monat and im != monat:
                continue
            if jahr and ij != jahr:
                continue
            if (
                str(imp[5]) == "Mieteinnahme"
                and str(imp[9]) in ["Erkannt", "Prüfen"]
                and norm_key(imp[6]) == norm_key(name)
            ):
                gezahlt += konto_betrag(imp[4])

        differenz = gezahlt - soll
        if abs(differenz) < 0.01:
            status = "Bezahlt"
            hinweis = "Sollmiete vollständig bezahlt"
        elif differenz < 0:
            status = "Offen"
            hinweis = f"Es fehlen {euro(abs(differenz))}"
        else:
            status = "Überzahlung"
            hinweis = f"Überzahlung {euro(differenz)}"

        result.append([
            str(monat), str(jahr), name, objekt, wohnung,
            f"{soll:.2f}", f"{gezahlt:.2f}", f"{differenz:.2f}",
            status, hinweis,
        ])
    return result


class KontoauszugImportSeite(QWidget):
    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.rows: list[list[str]] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 22)
        root.setSpacing(14)

        title = QLabel("Kontoauszug-Import")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel("CSV-Kontoauszüge einlesen, Mieteinnahmen automatisch erkennen und Mietern zuordnen.")
        info.setObjectName("subTitle")
        root.addWidget(info)

        top = QHBoxLayout()
        pdf_btn = QPushButton("PDF einlesen")
        pdf_btn.setObjectName("primaryButton")
        pdf_btn.clicked.connect(self.pdf_import)

        import_btn = QPushButton("CSV einlesen")
        import_btn.clicked.connect(self.csv_import)

        save_btn = QPushButton("Ergebnisse speichern")
        save_btn.clicked.connect(self.speichern)

        zahlungen_btn = QPushButton("Erkannte Mieten in Zahlungen übernehmen")
        zahlungen_btn.clicked.connect(self.in_zahlungen)

        top.addWidget(pdf_btn)
        top.addWidget(import_btn)
        top.addWidget(save_btn)
        top.addWidget(zahlungen_btn)
        top.addStretch()
        root.addLayout(top)

        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        root.addWidget(self.table, 1)

        self.laden()

    def laden(self) -> None:
        self.rows = DATA.get("Kontoauszug-Import", [])
        self._table()

    def _table(self) -> None:
        headers = SCHEMA.get("Kontoauszug-Import", [])
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(self.rows))

        for r, row in enumerate(self.rows):
            for c in range(len(headers)):
                value = row[c] if c < len(row) else ""
                self.table.setItem(r, c, QTableWidgetItem(str(value)))

        widths = [100, 180, 200, 360, 100, 140, 200, 180, 160, 120, 280]
        for c, w in enumerate(widths):
            self.table.setColumnWidth(c, w)


    def pdf_import(self) -> None:
        pfad, _ = QFileDialog.getOpenFileName(self, "Kontoauszug PDF wählen", "", "PDF-Dateien (*.pdf);;Alle Dateien (*.*)")
        if not pfad:
            return

        text = konto_pdf_text_lesen(pfad)
        if not text.strip():
            QMessageBox.warning(
                self,
                "PDF Import",
                "Aus dieser PDF konnte kein Text gelesen werden. Bitte prüfen, ob es ein echter Text-PDF-Kontoauszug ist und kein Scan."
            )
            return

        raw = konto_pdf_buchungen_erkennen(text)
        neue = [konto_import_klassifizieren(row) for row in raw]
        self.rows = neue
        DATA["Kontoauszug-Import"] = neue
        self._table()
        QMessageBox.information(self, "PDF Import", f"{len(neue)} Buchungen aus PDF erkannt.")


    def csv_import(self) -> None:
        pfad, _ = QFileDialog.getOpenFileName(self, "Kontoauszug CSV wählen", "", "CSV-Dateien (*.csv);;Alle Dateien (*.*)")
        if not pfad:
            return

        raw = konto_csv_lesen(pfad)
        neue = [konto_import_klassifizieren(row) for row in raw]
        self.rows = neue
        DATA["Kontoauszug-Import"] = neue
        self._table()
        QMessageBox.information(self, "Import", f"{len(neue)} Buchungen eingelesen.")

    def speichern(self) -> None:
        DATA["Kontoauszug-Import"] = self.rows
        speichere_tabelle("Kontoauszug-Import")
        QMessageBox.information(self, "Speichern", "Import-Ergebnisse gespeichert.")

    def in_zahlungen(self) -> None:
        zahlungen = DATA.setdefault("Zahlungen", [])
        count = 0
        duplicates = 0

        existing = set()
        for z in zahlungen:
            key = "|".join(str(v) for v in z[:4])
            existing.add(norm_key(key))

        for row in self.rows:
            if len(row) < 10:
                continue
            if row[5] != "Mieteinnahme" or row[9] not in ["Erkannt", "Prüfen"]:
                continue

            datum, _text, auftraggeber, zweck, betrag, _erkannt, mieter, objekt, wohnung, status, hinweis = row[:11]
            new_row = [
                datum,
                mieter or auftraggeber,
                betrag,
                zweck,
                "",
                "Bezahlt" if "Bezahlt" in hinweis else ("Teilzahlung" if "Teilzahlung" in hinweis else "Prüfen"),
                "Miete",
                datum[-4:] if len(datum) >= 4 else "",
                objekt,
                wohnung,
                hinweis,
            ]
            key = norm_key("|".join(str(v) for v in new_row[:4]))
            if key in existing:
                duplicates += 1
                continue

            zahlungen.append(new_row)
            existing.add(key)
            count += 1

        speichere_tabelle("Zahlungen")
        lade_tabelle("Zahlungen")
        QMessageBox.information(
            self,
            "Übernahme",
            f"{count} Mieteinnahmen übernommen. {duplicates} mögliche Doppelbuchungen übersprungen."
        )


class MietkontoAbgleichSeite(QWidget):
    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.rows = DATA.get("Mietkonto-Abgleich", [])
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 22)
        root.setSpacing(14)
        title = QLabel("Mietkonto-Abgleich")
        title.setObjectName("pageTitle")
        root.addWidget(title)
        info = QLabel("Sollmiete gegen Kontoauszug-Import und Zahlungstabelle prüfen.")
        info.setObjectName("subTitle")
        root.addWidget(info)
        top = QHBoxLayout()
        self.monat = QComboBox()
        self.monat.addItems([str(i).zfill(2) for i in range(1,13)])
        self.monat.setCurrentText("07")
        self.jahr = QComboBox()
        self.jahr.setEditable(True)
        self.jahr.addItems(["2026","2025","2024"])
        b_calc = QPushButton("Abgleich berechnen")
        b_calc.setObjectName("primaryButton")
        b_calc.clicked.connect(self.berechnen)
        b_save = QPushButton("Speichern")
        b_save.clicked.connect(self.speichern)
        b_open = QPushButton("Offene filtern")
        b_open.clicked.connect(self.offene)
        top.addWidget(QLabel("Monat:")); top.addWidget(self.monat)
        top.addWidget(QLabel("Jahr:")); top.addWidget(self.jahr)
        top.addWidget(b_calc); top.addWidget(b_save); top.addWidget(b_open); top.addStretch()
        root.addLayout(top)
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        root.addWidget(self.table, 1)
        self._table()

    def berechnen(self):
        self.rows = mietkonto_berechnen(self.monat.currentText(), self.jahr.currentText())
        DATA["Mietkonto-Abgleich"] = self.rows
        self._table()

    def speichern(self):
        DATA["Mietkonto-Abgleich"] = self.rows
        speichere_tabelle("Mietkonto-Abgleich")
        QMessageBox.information(self, "Mietkonto", "Mietkonto-Abgleich gespeichert.")

    def offene(self):
        self.rows = [r for r in self.rows if len(r) > 8 and r[8] != "Bezahlt"]
        self._table()

    def _table(self):
        headers = SCHEMA.get("Mietkonto-Abgleich", [])
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(self.rows))
        for r, row in enumerate(self.rows):
            for c in range(len(headers)):
                self.table.setItem(r, c, QTableWidgetItem(str(row[c] if c < len(row) else "")))
        for c, w in enumerate([80,80,220,200,160,120,120,120,130,320]):
            self.table.setColumnWidth(c, w)


class ZahlungsabgleichProSeite(QWidget):
    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 22)
        root.setSpacing(14)

        title = QLabel("Zahlungsabgleich PRO")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel("Ampelprüfung für erkannte Mieteinnahmen: bezahlt, Teilzahlung, Überzahlung, prüfen.")
        info.setObjectName("subTitle")
        root.addWidget(info)

        top = QHBoxLayout()
        refresh = QPushButton("Aktualisieren")
        refresh.setObjectName("primaryButton")
        refresh.clicked.connect(self.laden)
        konto = QPushButton("Kontoauszug-Import öffnen")
        konto.clicked.connect(lambda: self.nav("Kontoauszug-Import"))
        mietkonto = QPushButton("Mietkonto-Abgleich öffnen")
        mietkonto.clicked.connect(lambda: self.nav("Mietkonto-Abgleich"))
        top.addWidget(refresh)
        top.addWidget(konto)
        top.addWidget(mietkonto)
        top.addStretch()
        root.addLayout(top)

        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        root.addWidget(self.table, 1)
        self.laden()

    def laden(self) -> None:
        rows = []
        for r in DATA.get("Kontoauszug-Import", []):
            if len(r) < 11:
                continue
            if r[5] != "Mieteinnahme":
                continue
            betrag = konto_betrag(r[4])
            match = zahlungsabgleich_mieter_profi(" ".join(str(v) for v in r[:4]), betrag)
            soll = konto_betrag(match.get("soll", "0"))
            zahlstatus = zahlungsabgleich_status(soll, betrag)
            ampel = "🟢" if zahlstatus == "Bezahlt" else ("🟡" if zahlstatus in ["Teilzahlung", "Überzahlung"] else "🔴")
            differenz = abs(betrag) - soll
            rows.append([ampel, r[0], r[6], r[7], r[8], f"{soll:.2f}", f"{betrag:.2f}", f"{differenz:.2f}", zahlstatus, r[10]])

        headers = ["Ampel", "Datum", "Mieter", "Objekt", "Wohnung", "Soll", "Gezahlt", "Differenz", "Status", "Hinweis"]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            for c, v in enumerate(row):
                self.table.setItem(i, c, QTableWidgetItem(str(v)))
        for c, w in enumerate([70,100,220,200,160,110,110,110,130,420]):
            self.table.setColumnWidth(c, w)


def beleg_pdf_text(pfad: str) -> str:
    if PdfReader is None:
        return ""
    try:
        reader = PdfReader(pfad)
        parts = []
        for page in reader.pages:
            try:
                parts.append(page.extract_text() or "")
            except (OSError, ValueError, TypeError, AttributeError, RuntimeError, KeyError, IndexError):
                pass
        return "\n".join(parts)
    except (OSError, ValueError, TypeError, AttributeError, RuntimeError, KeyError, IndexError):
        return ""


def beleg_find(patterns: list[str], text: str) -> str:
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if m:
            return (m.group(1) if m.groups() else m.group(0)).strip()
    return ""


def beleg_betrag(text: str) -> str:
    return text.replace(".", "").replace(",", ".").replace("€", "").strip()


def beleg_kostenart(text: str) -> str:
    low = text.lower()
    mapping = [
        ("Abwasser", ["abwasser", "kanal"]),
        ("Wasser", ["wasser"]),
        ("Strom", ["strom", "energie", "allgemeinstrom"]),
        ("Gas", ["gas"]),
        ("Schornsteinfeger", ["schornstein"]),
        ("Versicherung", ["versicherung"]),
        ("Grundsteuer", ["grundsteuer"]),
        ("Hausmeister", ["hausmeister"]),
        ("Wartung", ["wartung", "heizung", "rauchmelder"]),
        ("Reparatur", ["reparatur", "instandsetzung", "handwerker"]),
        ("Müll", ["müll", "muell", "abfall"]),
    ]
    for kat, words in mapping:
        if any(w in low for w in words):
            return kat
    return "Sonstiges"


def beleg_objekt_match(text: str) -> tuple[str, str]:
    key = norm_key(text)
    objekt = ""
    wohnung = ""
    for row in DATA.get("Wohnungen", []):
        obj = str(row[0]) if len(row) > 0 else ""
        whg = str(row[1]) if len(row) > 1 else ""
        if obj and norm_key(obj) in key:
            objekt = obj
        if whg and norm_key(whg) in key:
            wohnung = whg
        if objekt:
            return objekt, wohnung
    for row in DATA.get("Objekte", []):
        obj = str(row[0]) if row else ""
        if obj and norm_key(obj) in key:
            return obj, ""
    return "", ""


def beleg_analyse(pfad: str) -> list[str]:
    text = beleg_pdf_text(pfad)
    if not text.strip():
        return [pfad, "", "", "", "", "", "", "", "", "", "", "", "", "", "Fehler", "PDF enthält keinen lesbaren Text"]

    clean = re.sub(r"\s+", " ", text)

    lieferant = ""
    for line in text.splitlines():
        if line.strip() and not any(w in line.lower() for w in ["rechnung", "kundennummer", "betrag"]):
            lieferant = line.strip()
            break

    re_nr = beleg_find([
        r"Rechnungs(?:nummer|nr\.?|-Nr\.?)\s*[:#]?\s*([A-Z0-9\-\/]+)",
        r"Rechnung\s*Nr\.?\s*[:#]?\s*([A-Z0-9\-\/]+)"
    ], clean)
    datum = beleg_find([r"Rechnungsdatum\s*[:#]?\s*(\d{2}\.\d{2}\.\d{4})", r"Datum\s*[:#]?\s*(\d{2}\.\d{2}\.\d{4})"], clean)
    faellig = beleg_find([r"Fällig(?:\sam)?\s*[:#]?\s*(\d{2}\.\d{2}\.\d{4})", r"Zahlbar bis\s*[:#]?\s*(\d{2}\.\d{2}\.\d{4})"], clean)
    kd = beleg_find([r"Kundennummer\s*[:#]?\s*([A-Z0-9\-\/]+)", r"Kunden-Nr\.?\s*[:#]?\s*([A-Z0-9\-\/]+)"], clean)

    brutto = beleg_find([r"(?:Brutto|Gesamtbetrag|Rechnungsbetrag|Zu zahlen)\s*[:#]?\s*([0-9\.\,]+)\s*€?", r"([0-9\.\,]+)\s*€\s*(?:Brutto|Gesamt)"], clean)
    netto = beleg_find([r"Netto\s*[:#]?\s*([0-9\.\,]+)\s*€?"], clean)
    mwst = beleg_find([r"(?:MwSt\.?|USt\.?|Mehrwertsteuer)\s*[:#]?\s*([0-9\.\,]+)\s*€?"], clean)

    if brutto:
        brutto = beleg_betrag(brutto)
    if netto:
        netto = beleg_betrag(netto)
    if mwst:
        mwst = beleg_betrag(mwst)

    objekt, wohnung = beleg_objekt_match(clean)
    kostenart = beleg_kostenart(clean)
    bk = "Ja" if kostenart in ["Abwasser", "Wasser", "Strom", "Gas", "Schornsteinfeger", "Versicherung", "Grundsteuer", "Hausmeister", "Wartung", "Müll"] else "Nein"
    umlage = "Ja" if bk == "Ja" else "Nein"

    hinweise = []
    if not re_nr:
        hinweise.append("Rechnungsnummer fehlt")
    if not brutto:
        hinweise.append("Bruttobetrag fehlt")
    if not objekt:
        hinweise.append("Objekt nicht erkannt")
    status = "Geprüft" if not hinweise else "Prüfen"

    return [pfad, lieferant, re_nr, datum, faellig, kd, netto, mwst, brutto, objekt, wohnung, kostenart, bk, umlage, status, "; ".join(hinweise)]


class BelegscannerSeite(QWidget):
    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.rows = DATA.get("Belegscanner", [])

        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 22)
        root.setSpacing(14)

        title = QLabel("Belegscanner PRO")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel("Rechnungs-PDFs einlesen, Daten erkennen und in Rechnungen/Betriebskosten übernehmen.")
        info.setObjectName("subTitle")
        root.addWidget(info)

        top = QHBoxLayout()
        pdf_btn = QPushButton("PDF-Rechnung einlesen")
        pdf_btn.setObjectName("primaryButton")
        pdf_btn.clicked.connect(self.pdf_import)
        save_btn = QPushButton("Speichern")
        save_btn.clicked.connect(self.speichern)
        rec_btn = QPushButton("In Rechnungen übernehmen")
        rec_btn.clicked.connect(self.in_rechnungen)
        bk_btn = QPushButton("In Betriebskosten übernehmen")
        bk_btn.clicked.connect(self.in_bk)
        top.addWidget(pdf_btn); top.addWidget(save_btn); top.addWidget(rec_btn); top.addWidget(bk_btn); top.addStretch()
        root.addLayout(top)

        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        root.addWidget(self.table, 1)
        self._table()

    def pdf_import(self):
        pfad, _ = QFileDialog.getOpenFileName(self, "Rechnungs-PDF wählen", "", "PDF-Dateien (*.pdf);;Alle Dateien (*.*)")
        if not pfad:
            return
        row = beleg_analyse(pfad)
        self.rows.append(row)
        DATA["Belegscanner"] = self.rows
        self._table()
        QMessageBox.information(self, "Belegscanner", "PDF analysiert.")

    def speichern(self):
        DATA["Belegscanner"] = self.rows
        speichere_tabelle("Belegscanner")
        QMessageBox.information(self, "Belegscanner", "Ergebnisse gespeichert.")

    def in_rechnungen(self):
        recs = DATA.setdefault("Rechnungen", [])
        count = 0
        existing = norm_key(" ".join(" ".join(str(v) for v in row) for row in recs))
        for r in self.rows:
            if len(r) < 16 or r[14] == "Fehler":
                continue
            if r[2] and norm_key(r[2]) in existing:
                continue
            recs.append([r[2], r[1], r[9], r[3], r[6], r[7], r[8], "Offen", r[4], r[0], r[15]])
            count += 1
        speichere_tabelle("Rechnungen")
        lade_tabelle("Rechnungen")
        QMessageBox.information(self, "Übernahme", f"{count} Belege in Rechnungen übernommen.")

    def in_bk(self):
        bk_rows = DATA.setdefault("Betriebskosten", [])
        count = 0
        for r in self.rows:
            if len(r) < 16 or r[12] != "Ja":
                continue
            bk_rows.append([r[9], r[10], r[11], r[8], "", r[3][-4:] if r[3] else "", r[4], r[5], r[2], r[1], r[0], r[15]])
            count += 1
        speichere_tabelle("Betriebskosten")
        lade_tabelle("Betriebskosten")
        QMessageBox.information(self, "Übernahme", f"{count} Belege in Betriebskosten übernommen.")

    def _table(self):
        headers = SCHEMA.get("Belegscanner", [])
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(self.rows))
        for r, row in enumerate(self.rows):
            for c in range(len(headers)):
                self.table.setItem(r, c, QTableWidgetItem(str(row[c] if c < len(row) else "")))
        widths = [260,180,150,120,120,150,100,100,100,180,150,160,130,100,110,340]
        for c, w in enumerate(widths):
            self.table.setColumnWidth(c, w)


CURRENT_USER: dict[str, str] = {"name": "", "rolle": "", "benutzername": ""}


def login_standard_benutzer() -> list[list[str]]:
    return [
        ["Julia", "julia", "julia123", "Admin", "Aktiv", "", "Administratorin"],
        ["Franzi", "franzi", "franzi123", "Admin", "Aktiv", "", "Administratorin"],
        ["Robert", "robert", "robert123", "Mitarbeiter", "Aktiv", "", "Standard-Testlogin"],
        ["Ralf", "ralf", "ralf123", "Mitarbeiter", "Aktiv", "", "Standard-Testlogin"],
        ["Admin", "admin", "admin123", "Admin", "Aktiv", "", "Standard-Testlogin"],
    ]


def login_benutzer_laden() -> list[list[str]]:
    rows = DATA.get("Mitarbeiter-Login", [])
    if not rows:
        DATA["Mitarbeiter-Login"] = login_standard_benutzer()
        rows = DATA.get("Mitarbeiter-Login", [])

    changed = False
    for row in rows:
        while len(row) < len(SCHEMA.get("Mitarbeiter-Login", [])):
            row.append("")
        name = norm_key(row[0] if row else "")
        if name in {"julia", "franzi"}:
            if len(row) > 3 and norm_key(row[3]) != "admin":
                row[3] = "Admin"
                changed = True
            if len(row) > 6 and not str(row[6]).strip():
                row[6] = "Administratorin"
                changed = True

    if changed or not xlsx_pfad("Mitarbeiter-Login").exists():
        try:
            speichere_tabelle("Mitarbeiter-Login")
        except (OSError, ValueError, TypeError, AttributeError, RuntimeError, KeyError, IndexError):
            pass

    return rows


def login_pruefen(benutzername: str, passwort: str) -> dict[str, str] | None:
    benutzername = str(benutzername or "").strip().lower()
    passwort = str(passwort or "").strip()

    for row in login_benutzer_laden():
        employee_name = str(row[0] if len(row) > 0 else "").strip()
        user = str(row[1] if len(row) > 1 else "").strip().lower()
        pw = str(row[2] if len(row) > 2 else "").strip()
        rolle = str(row[3] if len(row) > 3 else "").strip()
        status = str(row[4] if len(row) > 4 else "").strip().lower()

        if user == benutzername and pw == passwort and status == "aktiv":
            if norm_key(employee_name) in {"julia", "franzi"}:
                rolle = "Admin"
            return {"name": employee_name, "rolle": rolle, "benutzername": user}

    return None


def login_letzten_login_speichern(benutzername: str) -> None:
    try:
        rows = DATA.get("Mitarbeiter-Login", [])
        for row in rows:
            if len(row) > 1 and str(row[1]).strip().lower() == str(benutzername).strip().lower():
                while len(row) < len(SCHEMA["Mitarbeiter-Login"]):
                    row.append("")
                row[5] = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
                break
        speichere_tabelle("Mitarbeiter-Login")
    except (OSError, ValueError, TypeError, AttributeError, RuntimeError, KeyError, IndexError):
        pass


class LoginDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("ImmoVerwaltung – Mitarbeiter Login")
        self.setMinimumWidth(430)
        self.user: dict[str, str] | None = None

        layout = QVBoxLayout(self)
        layout.setContentsMargins(26, 24, 26, 24)
        layout.setSpacing(12)

        title = QLabel("Mitarbeiter Login")
        title.setObjectName("pageTitle")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)

        info = QLabel("Bitte mit Mitarbeiter-Zugang anmelden.")
        info.setAlignment(Qt.AlignmentFlag.AlignCenter)
        info.setObjectName("subTitle")
        layout.addWidget(info)

        self.user_combo = QComboBox()
        self.user_combo.setEditable(True)

        for row in login_benutzer_laden():
            if len(row) > 1 and str(row[4] if len(row) > 4 else "").lower() == "aktiv":
                self.user_combo.addItem(str(row[1]))

        self.password = QLineEdit()
        self.password.setEchoMode(QLineEdit.EchoMode.Password)
        self.password.setPlaceholderText("Passwort")
        self.password.returnPressed.connect(self.login)

        layout.addWidget(QLabel("Benutzername:"))
        layout.addWidget(self.user_combo)
        layout.addWidget(QLabel("Passwort:"))
        layout.addWidget(self.password)

        btn_row = QHBoxLayout()
        login_btn = QPushButton("Anmelden")
        login_btn.setObjectName("primaryButton")
        login_btn.clicked.connect(self.login)
        cancel_btn = QPushButton("Abbrechen")
        cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(login_btn)
        btn_row.addWidget(cancel_btn)
        layout.addLayout(btn_row)

        hint = QLabel("Standard-Zugänge: julia, franzi, robert, ralf, admin")
        hint.setWordWrap(True)
        hint.setStyleSheet("color:#64748b; font-size:10px;")
        layout.addWidget(hint)

    def login(self) -> None:
        user = self.user_combo.currentText().strip()
        pw = self.password.text().strip()
        result = login_pruefen(user, pw)

        if not result:
            QMessageBox.warning(self, "Login", "Benutzername oder Passwort falsch oder Benutzer nicht aktiv.")
            self.password.clear()
            self.password.setFocus()
            return

        self.user = result
        CURRENT_USER.update(result)
        login_letzten_login_speichern(result.get("benutzername", ""))
        self.accept()


def aufgaben_feld(row: list[Any], feldname: str) -> str:
    for index, feld in enumerate(SCHEMA.get("Aufgaben", [])):
        if norm_key(feld) == norm_key(feldname) and index < len(row):
            return str(row[index] or "").strip()
    return ""


def aufgaben_status_ist_offen(status: str) -> bool:
    return norm_key(status) not in {"erledigt", "abgeschlossen", "archiviert", "bezahlt"}


def datum_sort_key(value: str) -> tuple[int, int, int]:
    parts = re.split(r"[./-]", str(value or "").strip())
    if len(parts) >= 3:
        try:
            if len(parts[0]) == 4:
                return int(parts[0]), int(parts[1]), int(parts[2])
            jahr = int(parts[2]) if len(parts[2]) == 4 else 2000 + int(parts[2])
            return jahr, int(parts[1]), int(parts[0])
        except (ValueError, TypeError):
            pass
    return (9999, 12, 31)


def offene_aufgaben_rows() -> list[list[Any]]:
    return [row for row in DATA.get("Aufgaben", []) if aufgaben_status_ist_offen(aufgaben_feld(row, "Status"))]


def ueberfaellige_aufgaben_rows() -> list[list[Any]]:
    heute = date.today()
    result = []
    for row in offene_aufgaben_rows():
        faellig = aufgaben_feld(row, "Fällig am")
        if not faellig:
            continue
        parsed = None
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y"):
            try:
                parsed = datetime.strptime(faellig, fmt).date()
                break
            except ValueError:
                continue
        if parsed is not None and parsed < heute:
            result.append(row)
    return result


def objektchronik_rows(objektordner: str) -> list[list[str]]:
    objekt_key = norm_key(objektordner)
    events: list[list[str]] = []
    if not objekt_key:
        return events
    for titel in ["Mieter", "Mietverträge", "Zahlungen", "Rechnungen", "Betriebskosten", "Schäden", "Fristen", "Aufgaben", "Dokumente", "Ereignisprotokoll", "HV-Rechnungen", "Versorger", "Übergabeprotokolle"]:
        for row in DATA.get(titel, []):
            relation = objektordner_fuer_datensatz(titel, row)
            if norm_key(relation) != objekt_key:
                continue
            datum = ""
            for name in ["Datum", "Fällig am", "Mietbeginn", "Beginn", "Rechnungsdatum", "Zahlung am", "Termin", "Einzugsdatum", "Auszugsdatum"]:
                datum = feldwert(titel, row, [name])
                if datum:
                    break
            beschreibung = ""
            for name in ["Aufgabe", "Schaden", "Titel", "Leistung", "Verwendungszweck", "Beschreibung", "Notiz", "Bemerkungen", "Aktion"]:
                beschreibung = feldwert(titel, row, [name])
                if beschreibung:
                    break
            if not beschreibung:
                beschreibung = " | ".join(str(v) for v in row[:8] if str(v).strip())
            status = feldwert(titel, row, ["Status"])
            events.append([datum, titel, beschreibung, status, relation])
    events.sort(key=lambda item: datum_sort_key(item[0]))
    return events


class ArbeitsorganisationSeite(QWidget):
    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.filtered_rows: list[list[Any]] = []
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 22)
        root.setSpacing(14)
        title = QLabel("Arbeitsorganisation PRO")
        title.setObjectName("pageTitle")
        root.addWidget(title)
        info = QLabel("Aufgaben, Verantwortlichkeiten, Objektordner und Fälligkeiten zentral steuern.")
        info.setObjectName("subTitle")
        root.addWidget(info)
        filter_row = QHBoxLayout()
        self.objekt_filter = QComboBox()
        self.objekt_filter.addItem("Alle Objektordner")
        self.objekt_filter.addItems(alle_objektordner())
        self.objekt_filter.currentTextChanged.connect(self.laden)
        self.mitarbeiter_filter = QComboBox()
        self.mitarbeiter_filter.addItems(["Alle Mitarbeiter", "Julia", "Franzi", "Robert", "Ralf", "Admin"])
        self.mitarbeiter_filter.currentTextChanged.connect(self.laden)
        self.status_filter = QComboBox()
        self.status_filter.addItems(["Alle Status", "Offen", "Neu", "In Bearbeitung", "Wartet", "Erledigt"])
        self.status_filter.currentTextChanged.connect(self.laden)
        neu = QPushButton("Neue Aufgabe")
        neu.setObjectName("primaryButton")
        neu.clicked.connect(self.neue_aufgabe)
        bearbeiten = QPushButton("Bearbeiten")
        bearbeiten.clicked.connect(self.bearbeiten)
        erledigt = QPushButton("Als erledigt markieren")
        erledigt.clicked.connect(self.als_erledigt)
        handwerker = QPushButton("Handwerkerauftrag")
        handwerker.clicked.connect(self.handwerkerauftrag)
        filter_row.addWidget(QLabel("Objektordner:")); filter_row.addWidget(self.objekt_filter)
        filter_row.addWidget(QLabel("Mitarbeiter:")); filter_row.addWidget(self.mitarbeiter_filter)
        filter_row.addWidget(QLabel("Status:")); filter_row.addWidget(self.status_filter)
        filter_row.addWidget(neu); filter_row.addWidget(bearbeiten); filter_row.addWidget(erledigt); filter_row.addWidget(handwerker)
        root.addLayout(filter_row)
        self.metrics = QGridLayout(); root.addLayout(self.metrics)
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        root.addWidget(self.table, 1)
        self.laden()

    @staticmethod
    def _metric_card(title: str, value: str, icon: str) -> QFrame:
        card = QFrame(); card.setObjectName("metricCard")
        layout = QVBoxLayout(card)
        label = QLabel(f"{icon}  {title}"); label.setObjectName("metricTitle")
        amount = QLabel(value); amount.setObjectName("metricValue")
        layout.addWidget(label); layout.addWidget(amount)
        return card

    def laden(self) -> None:
        rows = list(DATA.get("Aufgaben", []))
        selected_objekt = self.objekt_filter.currentText().strip()
        if selected_objekt != "Alle Objektordner":
            key = norm_key(selected_objekt)
            rows = [row for row in rows if norm_key(objektordner_fuer_datensatz("Aufgaben", row)) == key]
        selected_mitarbeiter = self.mitarbeiter_filter.currentText().strip()
        if selected_mitarbeiter != "Alle Mitarbeiter":
            rows = [row for row in rows if norm_key(aufgaben_feld(row, "Verantwortlich")) == norm_key(selected_mitarbeiter)]
        selected_status = self.status_filter.currentText().strip()
        if selected_status != "Alle Status":
            rows = [row for row in rows if norm_key(aufgaben_feld(row, "Status")) == norm_key(selected_status)]
        self.filtered_rows = rows
        while self.metrics.count():
            item = self.metrics.takeAt(0)
            if item is not None and item.widget() is not None:
                item.widget().deleteLater()
        offen = sum(1 for row in rows if aufgaben_status_ist_offen(aufgaben_feld(row, "Status")))
        erledigt_count = sum(1 for row in rows if norm_key(aufgaben_feld(row, "Status")) == "erledigt")
        kritisch = sum(1 for row in rows if norm_key(aufgaben_feld(row, "Priorität")) in {"hoch", "kritisch", "dringend"})
        ueberfaellig = sum(1 for row in rows if row in ueberfaellige_aufgaben_rows())
        for index, data in enumerate([("Offen", str(offen), "🔵"), ("Überfällig", str(ueberfaellig), "🔴"), ("Kritisch", str(kritisch), "🟡"), ("Erledigt", str(erledigt_count), "🟢")]):
            self.metrics.addWidget(self._metric_card(*data), 0, index)
        headers = SCHEMA.get("Aufgaben", [])
        self.table.setColumnCount(len(headers)); self.table.setHorizontalHeaderLabels(headers); self.table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c in range(len(headers)):
                self.table.setItem(r, c, QTableWidgetItem(str(row[c] if c < len(row) else "")))
        for c in range(min(len(headers), 13)): self.table.setColumnWidth(c, 160)

    def _selected_original_index(self) -> int | None:
        row_index = self.table.currentRow()
        if row_index < 0 or row_index >= len(self.filtered_rows): return None
        selected_row = self.filtered_rows[row_index]
        for index, row in enumerate(DATA.get("Aufgaben", [])):
            if row is selected_row or row == selected_row: return index
        return None

    def neue_aufgabe(self) -> None:
        dialog = EingabeDialog("Neue Aufgabe", SCHEMA["Aufgaben"])
        if dialog.exec() == QDialog.DialogCode.Accepted:
            DATA["Aufgaben"].append(dialog.values()); speichere_tabelle("Aufgaben"); self.laden()

    def bearbeiten(self) -> None:
        index = self._selected_original_index()
        if index is None:
            QMessageBox.information(self, "Aufgaben", "Bitte eine Aufgabe auswählen."); return
        dialog = EingabeDialog("Aufgabe bearbeiten", SCHEMA["Aufgaben"], DATA["Aufgaben"][index])
        if dialog.exec() == QDialog.DialogCode.Accepted:
            DATA["Aufgaben"][index] = dialog.values(); speichere_tabelle("Aufgaben"); self.laden()

    def als_erledigt(self) -> None:
        index = self._selected_original_index()
        if index is None:
            QMessageBox.information(self, "Aufgaben", "Bitte eine Aufgabe auswählen."); return
        status_index = next((i for i, f in enumerate(SCHEMA["Aufgaben"]) if norm_key(f) == "status"), None)
        if status_index is None: return
        while len(DATA["Aufgaben"][index]) < len(SCHEMA["Aufgaben"]): DATA["Aufgaben"][index].append("")
        DATA["Aufgaben"][index][status_index] = "Erledigt"
        speichere_tabelle("Aufgaben"); self.laden()

    def handwerkerauftrag(self) -> None:
        values = ["" for _ in SCHEMA["Aufgaben"]]
        defaults = {"Aufgabe": "Handwerkerauftrag", "Bereich": "Technik", "Priorität": "Normal", "Status": "Neu", "Verantwortlich": "Robert"}
        selected_objekt = self.objekt_filter.currentText().strip()
        if selected_objekt != "Alle Objektordner":
            defaults["Objekt"] = selected_objekt; defaults["Objektordner"] = selected_objekt
        for i, feld in enumerate(SCHEMA["Aufgaben"]):
            if feld in defaults: values[i] = defaults[feld]
        dialog = EingabeDialog("Handwerkerauftrag anlegen", SCHEMA["Aufgaben"], values)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            DATA["Aufgaben"].append(dialog.values()); speichere_tabelle("Aufgaben"); self.laden()


class ObjektchronikSeite(QWidget):
    def __init__(self, nav):
        super().__init__(); self.nav = nav; self.rows: list[list[str]] = []
        root = QVBoxLayout(self); root.setContentsMargins(24, 22, 24, 22); root.setSpacing(14)
        title = QLabel("Digitale Objektchronik"); title.setObjectName("pageTitle"); root.addWidget(title)
        info = QLabel("Automatisch erzeugte Chronik aus Rechnungen, Zahlungen, Schäden, Aufgaben, Dokumenten und Verträgen."); info.setObjectName("subTitle"); root.addWidget(info)
        top = QHBoxLayout(); self.objekt = QComboBox(); self.objekt.setEditable(True); self.objekt.addItems(alle_objektordner()); self.objekt.currentTextChanged.connect(self.laden)
        refresh = QPushButton("Aktualisieren"); refresh.setObjectName("primaryButton"); refresh.clicked.connect(self.laden)
        export_btn = QPushButton("Chronik exportieren"); export_btn.clicked.connect(self.exportieren)
        top.addWidget(QLabel("Objektordner:")); top.addWidget(self.objekt, 1); top.addWidget(refresh); top.addWidget(export_btn); root.addLayout(top)
        self.table = QTableWidget(); self.table.setAlternatingRowColors(True); self.table.setWordWrap(False); self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Datum", "Bereich", "Ereignis", "Status", "Objektordner"])
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn); root.addWidget(self.table, 1)
        self.laden()

    def laden(self) -> None:
        self.rows = objektchronik_rows(self.objekt.currentText().strip()); self.table.setRowCount(len(self.rows))
        for r, row in enumerate(self.rows):
            for c, value in enumerate(row): self.table.setItem(r, c, QTableWidgetItem(str(value)))
        for c, width in enumerate([120, 180, 720, 140, 240]): self.table.setColumnWidth(c, width)

    def exportieren(self) -> None:
        if not self.rows:
            QMessageBox.information(self, "Chronik", "Keine Chronikeinträge vorhanden."); return
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        default = EXPORT_DIR / ("objektchronik_" + datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + ".xlsx")
        target, _ = QFileDialog.getSaveFileName(self, "Objektchronik exportieren", str(default), "Excel-Dateien (*.xlsx)")
        if not target: return
        target_path = Path(target)
        if target_path.suffix.lower() != ".xlsx": target_path = target_path.with_suffix(".xlsx")
        workbook = Workbook(); sheet = workbook.active
        if not isinstance(sheet, Worksheet): return
        sheet.title = "Objektchronik"; sheet.append(["Datum", "Bereich", "Ereignis", "Status", "Objektordner"])
        for row in self.rows: sheet.append(row)
        workbook.save(target_path)
        QMessageBox.information(self, "Export", f"Objektchronik exportiert:\n{target_path}")



def smart_objekt_status(objektordner: str) -> dict[str, Any]:
    """Ermittelt Ampel, Kennzahlen und Hinweise rein aus vorhandenen Tabellen."""
    objekt_key = norm_key(objektordner)

    def passend(titel: str, row: list[Any]) -> bool:
        return norm_key(objektordner_fuer_datensatz(titel, row)) == objekt_key

    offene_aufgaben = [
        row for row in DATA.get("Aufgaben", [])
        if passend("Aufgaben", row)
        and aufgaben_status_ist_offen(aufgaben_feld(row, "Status"))
    ]
    ueberfaellige_aufgaben = [
        row for row in ueberfaellige_aufgaben_rows()
        if passend("Aufgaben", row)
    ]

    offene_rechnungen = []
    for row in DATA.get("Rechnungen", []):
        if not passend("Rechnungen", row):
            continue
        status = feldwert("Rechnungen", row, ["Status"])
        if norm_key(status) not in {"bezahlt", "erledigt", "abgeschlossen"}:
            offene_rechnungen.append(row)

    offene_schaeden = []
    for row in DATA.get("Schäden", []):
        if not passend("Schäden", row):
            continue
        status = feldwert("Schäden", row, ["Status"])
        if norm_key(status) not in {"erledigt", "abgeschlossen", "behoben"}:
            offene_schaeden.append(row)

    freie_wohnungen = []
    for row in DATA.get("Wohnungen", []):
        if not passend("Wohnungen", row):
            continue
        status = feldwert("Wohnungen", row, ["Status"])
        if any(word in norm_key(status) for word in ["frei", "leer", "unvermietet"]):
            freie_wohnungen.append(row)

    mieter = [
        row for row in DATA.get("Mieter", [])
        if passend("Mieter", row)
    ]
    wohnungen = [
        row for row in DATA.get("Wohnungen", [])
        if passend("Wohnungen", row)
    ]
    dokumente = [
        row for row in DATA.get("Dokumente", [])
        if passend("Dokumente", row)
    ]

    monatsmiete = 0.0
    for row in mieter:
        value = feldwert("Mieter", row, ["Miete", "Warmmiete", "Kaltmiete"])
        monatsmiete += to_float(value)

    ausgaben = 0.0
    for row in DATA.get("Rechnungen", []):
        if passend("Rechnungen", row):
            ausgaben += to_float(
                feldwert("Rechnungen", row, ["Brutto", "Betrag brutto", "Betrag"])
            )

    severity = 0
    hinweise: list[str] = []

    if ueberfaellige_aufgaben:
        severity += 3
        hinweise.append(f"{len(ueberfaellige_aufgaben)} überfällige Aufgabe(n)")
    if offene_schaeden:
        severity += 3
        hinweise.append(f"{len(offene_schaeden)} offene Schadenmeldung(en)")
    if offene_rechnungen:
        severity += 2
        hinweise.append(f"{len(offene_rechnungen)} offene Rechnung(en)")
    if freie_wohnungen:
        severity += 2
        hinweise.append(f"{len(freie_wohnungen)} freie Wohnung(en)")
    if offene_aufgaben and not ueberfaellige_aufgaben:
        severity += 1
        hinweise.append(f"{len(offene_aufgaben)} offene Aufgabe(n)")

    if severity >= 5:
        ampel = "🔴"
        status_text = "Sofort handeln"
    elif severity >= 2:
        ampel = "🟡"
        status_text = "Aufmerksamkeit erforderlich"
    else:
        ampel = "🟢"
        status_text = "Alles in Ordnung"

    return {
        "ampel": ampel,
        "status": status_text,
        "hinweise": hinweise,
        "wohnungen": len(wohnungen),
        "mieter": len(mieter),
        "freie_wohnungen": len(freie_wohnungen),
        "offene_aufgaben": len(offene_aufgaben),
        "ueberfaellige_aufgaben": len(ueberfaellige_aufgaben),
        "offene_rechnungen": len(offene_rechnungen),
        "offene_schaeden": len(offene_schaeden),
        "dokumente": len(dokumente),
        "monatsmiete": monatsmiete,
        "jahresmiete": monatsmiete * 12,
        "ausgaben": ausgaben,
        "saldo": monatsmiete * 12 - ausgaben,
    }


def fristen_pro_status() -> list[dict[str, str]]:
    """Berechnet Fristenstatus anhand bestehender Tabelle Fristen."""

    heute = date.today()
    result: list[dict[str, str]] = []

    for row in DATA.get("Fristen", []):
        datum_text = feldwert("Fristen", row, ["Fällig am", "Termin"])
        parsed = None

        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y"):
            try:
                parsed = datetime.strptime(datum_text, fmt).date()
                break
            except ValueError:
                continue

        if parsed is None:
            tage = ""
            ampel = "⚪"
            bewertung = "Datum prüfen"
        else:
            delta = (parsed - heute).days
            tage = str(delta)

            if delta < 0:
                ampel = "🔴"
                bewertung = "Überfällig"
            elif delta <= 14:
                ampel = "🟠"
                bewertung = "Dringend"
            elif delta <= 60:
                ampel = "🟡"
                bewertung = "Bald fällig"
            else:
                ampel = "🟢"
                bewertung = "Im Plan"

        result.append({
            "ampel": ampel,
            "titel": feldwert("Fristen", row, ["Titel"]),
            "bereich": feldwert("Fristen", row, ["Bereich"]),
            "objekt": objektordner_fuer_datensatz("Fristen", row),
            "mieter": feldwert("Fristen", row, ["Mieter"]),
            "faellig": datum_text,
            "tage": tage,
            "prioritaet": feldwert("Fristen", row, ["Priorität"]),
            "status": feldwert("Fristen", row, ["Status"]),
            "bewertung": bewertung,
        })

    result.sort(
        key=lambda item: (
            {"🔴": 0, "🟠": 1, "🟡": 2, "🟢": 3, "⚪": 4}.get(item["ampel"], 5),
            datum_sort_key(item["faellig"]),
        )
    )
    return result



class SmartObjektakteSeite(QWidget):
    """Intelligente Objektakte mit Ampelstatus und lesbarem Scroll-Layout."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.current_objekt = ""

        root = QVBoxLayout(self)
        root.setContentsMargins(18, 16, 18, 16)
        root.setSpacing(12)

        title = QLabel("Smart Objektakte")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel(
            "Intelligente Objektübersicht mit Ampelstatus, Finanzen, Aufgaben, Schäden und Dokumenten."
        )
        info.setObjectName("subTitle")
        info.setWordWrap(True)
        root.addWidget(info)

        top = QHBoxLayout()
        self.objekt = QComboBox()
        self.objekt.setEditable(True)
        self.objekt.setMinimumWidth(280)
        self.objekt.addItems(alle_objektordner())
        self.objekt.currentTextChanged.connect(self.laden)

        refresh = QPushButton("Aktualisieren")
        refresh.setObjectName("primaryButton")
        refresh.clicked.connect(self.laden)

        chronik = QPushButton("Objektchronik öffnen")
        chronik.clicked.connect(lambda: self.nav("Objektchronik"))

        aufgaben = QPushButton("Arbeitsorganisation öffnen")
        aufgaben.clicked.connect(lambda: self.nav("Arbeitsorganisation PRO"))

        top.addWidget(QLabel("Objektordner:"))
        top.addWidget(self.objekt, 1)
        top.addWidget(refresh)
        top.addWidget(chronik)
        top.addWidget(aufgaben)
        root.addLayout(top)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        root.addWidget(scroll, 1)

        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(4, 4, 14, 4)
        content_layout.setSpacing(14)
        scroll.setWidget(content)

        self.status_panel = QFrame()
        self.status_panel.setObjectName("chartPanel")
        self.status_panel.setMinimumHeight(110)
        status_layout = QHBoxLayout(self.status_panel)
        status_layout.setContentsMargins(18, 14, 18, 14)

        self.ampel_label = QLabel("⚪")
        self.ampel_label.setStyleSheet("font-size:46px;")
        self.ampel_label.setMinimumWidth(70)

        self.status_label = QLabel("Kein Objekt ausgewählt")
        self.status_label.setObjectName("metricValue")
        self.status_label.setWordWrap(True)

        self.hinweis_label = QLabel("")
        self.hinweis_label.setWordWrap(True)
        self.hinweis_label.setMinimumHeight(42)

        status_layout.addWidget(self.ampel_label, 0, Qt.AlignmentFlag.AlignTop)
        text_layout = QVBoxLayout()
        text_layout.addWidget(self.status_label)
        text_layout.addWidget(self.hinweis_label)
        status_layout.addLayout(text_layout, 1)
        content_layout.addWidget(self.status_panel)

        self.cards_widget = QWidget()
        self.cards = QGridLayout(self.cards_widget)
        self.cards.setContentsMargins(0, 0, 0, 0)
        self.cards.setHorizontalSpacing(12)
        self.cards.setVerticalSpacing(12)
        content_layout.addWidget(self.cards_widget)

        self.tabs = QTabWidget()
        self.tabs.setMinimumHeight(420)
        content_layout.addWidget(self.tabs, 1)

        self.tbl_wohnungen = QTableWidget()
        self.tbl_mieter = QTableWidget()
        self.tbl_aufgaben = QTableWidget()
        self.tbl_rechnungen = QTableWidget()
        self.tbl_schaeden = QTableWidget()
        self.tbl_dokumente = QTableWidget()

        self.tabs.addTab(self.tbl_wohnungen, "Wohnungen")
        self.tabs.addTab(self.tbl_mieter, "Mieter")
        self.tabs.addTab(self.tbl_aufgaben, "Aufgaben")
        self.tabs.addTab(self.tbl_rechnungen, "Rechnungen")
        self.tabs.addTab(self.tbl_schaeden, "Schäden")
        self.tabs.addTab(self.tbl_dokumente, "Dokumente")

        self.laden()

    @staticmethod
    def _card(title: str, value: str, icon: str) -> QFrame:
        card = QFrame()
        card.setObjectName("metricCard")
        card.setMinimumWidth(190)
        card.setMaximumWidth(320)
        card.setMinimumHeight(105)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 12, 14, 12)

        label = QLabel(f"{icon}  {title}")
        label.setObjectName("metricTitle")
        label.setWordWrap(True)

        amount = QLabel(value)
        amount.setObjectName("metricValue")
        amount.setWordWrap(True)

        layout.addWidget(label)
        layout.addWidget(amount)
        return card

    def laden(self) -> None:
        objekt = self.objekt.currentText().strip()
        self.current_objekt = objekt

        if not objekt:
            self.ampel_label.setText("⚪")
            self.status_label.setText("Kein Objekt ausgewählt")
            self.hinweis_label.setText("")
            return

        status = smart_objekt_status(objekt)
        self.ampel_label.setText(str(status["ampel"]))
        self.status_label.setText(str(status["status"]))
        self.hinweis_label.setText(
            " · ".join(status["hinweise"]) if status["hinweise"] else "Keine kritischen Hinweise."
        )

        while self.cards.count():
            item = self.cards.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        cards = [
            ("Wohnungen", str(status["wohnungen"]), "🏠"),
            ("Mieter", str(status["mieter"]), "👥"),
            ("Freie Wohnungen", str(status["freie_wohnungen"]), "🚪"),
            ("Offene Aufgaben", str(status["offene_aufgaben"]), "🗂"),
            ("Überfällige Aufgaben", str(status["ueberfaellige_aufgaben"]), "⏰"),
            ("Offene Rechnungen", str(status["offene_rechnungen"]), "🧾"),
            ("Offene Schäden", str(status["offene_schaeden"]), "⚠"),
            ("Dokumente", str(status["dokumente"]), "📄"),
            ("Monatsmiete", euro(float(status["monatsmiete"])), "💶"),
            ("Jahresmiete", euro(float(status["jahresmiete"])), "📈"),
            ("Ausgaben", euro(float(status["ausgaben"])), "💸"),
            ("Objektsaldo", euro(float(status["saldo"])), "Σ"),
        ]

        # 3 Spalten statt 4: lesbarer auf kleineren Bildschirmen.
        for index, values in enumerate(cards):
            self.cards.addWidget(self._card(*values), index // 3, index % 3)

        self._fill(self.tbl_wohnungen, "Wohnungen", objekt)
        self._fill(self.tbl_mieter, "Mieter", objekt)
        self._fill(self.tbl_aufgaben, "Aufgaben", objekt)
        self._fill(self.tbl_rechnungen, "Rechnungen", objekt)
        self._fill(self.tbl_schaeden, "Schäden", objekt)
        self._fill(self.tbl_dokumente, "Dokumente", objekt)

    @staticmethod
    def _fill(table: QTableWidget, titel: str, objekt: str) -> None:
        rows = [
            row for row in DATA.get(titel, [])
            if norm_key(objektordner_fuer_datensatz(titel, row)) == norm_key(objekt)
        ]
        headers = SCHEMA.get(titel, [])

        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setRowCount(len(rows))
        table.setAlternatingRowColors(True)
        table.setWordWrap(False)
        table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        table.setMinimumHeight(360)

        for row_index, row in enumerate(rows):
            for col_index in range(len(headers)):
                value = row[col_index] if col_index < len(row) else ""
                table.setItem(row_index, col_index, QTableWidgetItem(str(value)))

        header = table.horizontalHeader()
        header.setStretchLastSection(False)

        for col_index in range(len(headers)):
            table.setColumnWidth(col_index, 150)

        if headers:
            table.setColumnWidth(0, 200)

class FristenmanagerProSeite(QWidget):
    """Fristenmanager mit Ampelbewertung und Objektfilter."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.rows: list[dict[str, str]] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 22)
        root.setSpacing(14)

        title = QLabel("Fristenmanager PRO")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel(
            "Überwacht bestehende Fristen automatisch und bewertet Dringlichkeit per Ampelsystem."
        )
        info.setObjectName("subTitle")
        root.addWidget(info)

        top = QHBoxLayout()
        self.objekt_filter = QComboBox()
        self.objekt_filter.addItem("Alle Objektordner")
        self.objekt_filter.addItems(alle_objektordner())
        self.objekt_filter.currentTextChanged.connect(self.laden)

        self.ampel_filter = QComboBox()
        self.ampel_filter.addItems(
            ["Alle Fristen", "Überfällig", "Dringend", "Bald fällig", "Im Plan", "Datum prüfen"]
        )
        self.ampel_filter.currentTextChanged.connect(self.laden)

        refresh = QPushButton("Aktualisieren")
        refresh.setObjectName("primaryButton")
        refresh.clicked.connect(self.laden)

        neue_frist = QPushButton("Neue Frist")
        neue_frist.clicked.connect(self.neue_frist)

        top.addWidget(QLabel("Objektordner:"))
        top.addWidget(self.objekt_filter)
        top.addWidget(QLabel("Bewertung:"))
        top.addWidget(self.ampel_filter)
        top.addWidget(refresh)
        top.addWidget(neue_frist)
        top.addStretch()
        root.addLayout(top)

        self.cards = QGridLayout()
        root.addLayout(self.cards)

        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "Ampel", "Titel", "Bereich", "Objektordner", "Mieter",
            "Fällig am", "Tage", "Priorität", "Status", "Bewertung"
        ])
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.cellDoubleClicked.connect(self._open_fristen)
        root.addWidget(self.table, 1)

        self.laden()

    @staticmethod
    def _card(title: str, value: str, icon: str) -> QFrame:
        card = QFrame()
        card.setObjectName("metricCard")
        layout = QVBoxLayout(card)
        label = QLabel(f"{icon}  {title}")
        label.setObjectName("metricTitle")
        amount = QLabel(value)
        amount.setObjectName("metricValue")
        layout.addWidget(label)
        layout.addWidget(amount)
        return card

    def laden(self) -> None:
        rows = fristen_pro_status()

        selected_objekt = self.objekt_filter.currentText().strip()
        if selected_objekt != "Alle Objektordner":
            rows = [
                row for row in rows
                if norm_key(row["objekt"]) == norm_key(selected_objekt)
            ]

        selected_status = self.ampel_filter.currentText().strip()
        if selected_status != "Alle Fristen":
            rows = [row for row in rows if row["bewertung"] == selected_status]

        self.rows = rows

        while self.cards.count():
            item = self.cards.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        counts = {
            "Überfällig": 0,
            "Dringend": 0,
            "Bald fällig": 0,
            "Im Plan": 0,
        }
        for item in fristen_pro_status():
            if item["bewertung"] in counts:
                counts[item["bewertung"]] += 1

        cards = [
            ("Überfällig", str(counts["Überfällig"]), "🔴"),
            ("Dringend", str(counts["Dringend"]), "🟠"),
            ("Bald fällig", str(counts["Bald fällig"]), "🟡"),
            ("Im Plan", str(counts["Im Plan"]), "🟢"),
        ]

        for index, card in enumerate(cards):
            self.cards.addWidget(self._card(*card), 0, index)

        self.table.setRowCount(len(rows))

        for row_index, item in enumerate(rows):
            values = [
                item["ampel"],
                item["titel"],
                item["bereich"],
                item["objekt"],
                item["mieter"],
                item["faellig"],
                item["tage"],
                item["prioritaet"],
                item["status"],
                item["bewertung"],
            ]
            for col_index, value in enumerate(values):
                self.table.setItem(row_index, col_index, QTableWidgetItem(str(value)))

        for col_index, width in enumerate([70, 240, 160, 240, 180, 110, 80, 120, 130, 140]):
            self.table.setColumnWidth(col_index, width)

    def neue_frist(self) -> None:
        dialog = EingabeDialog("Neue Frist", SCHEMA["Fristen"])
        if dialog.exec() == QDialog.DialogCode.Accepted:
            DATA["Fristen"].append(dialog.values())
            speichere_tabelle("Fristen")
            self.laden()

    def _open_fristen(self, _row: int, _column: int) -> None:
        self.nav("Fristen")



def objekt_gesundheitsindex(objektordner: str) -> dict[str, Any]:
    """Berechnet einen transparenten Objekt-Gesundheitsindex aus bestehenden Daten."""
    status = smart_objekt_status(objektordner)
    punkte = 100
    gruende: list[str] = []

    if status["ueberfaellige_aufgaben"]:
        abzug = min(25, int(status["ueberfaellige_aufgaben"]) * 5)
        punkte -= abzug
        gruende.append(f"-{abzug}: überfällige Aufgaben")

    if status["offene_schaeden"]:
        abzug = min(25, int(status["offene_schaeden"]) * 8)
        punkte -= abzug
        gruende.append(f"-{abzug}: offene Schäden")

    if status["offene_rechnungen"]:
        abzug = min(15, int(status["offene_rechnungen"]) * 3)
        punkte -= abzug
        gruende.append(f"-{abzug}: offene Rechnungen")

    if status["freie_wohnungen"]:
        abzug = min(20, int(status["freie_wohnungen"]) * 5)
        punkte -= abzug
        gruende.append(f"-{abzug}: Leerstand")

    if status["offene_aufgaben"] and not status["ueberfaellige_aufgaben"]:
        abzug = min(10, int(status["offene_aufgaben"]) * 2)
        punkte -= abzug
        gruende.append(f"-{abzug}: offene Aufgaben")

    punkte = max(0, min(100, punkte))

    if punkte >= 85:
        ampel = "🟢"
        bewertung = "Gesund"
    elif punkte >= 65:
        ampel = "🟡"
        bewertung = "Beobachten"
    else:
        ampel = "🔴"
        bewertung = "Kritisch"

    return {
        "punkte": punkte,
        "ampel": ampel,
        "bewertung": bewertung,
        "gruende": gruende,
    }


def plausibilitaetspruefungen(objektordner: str = "") -> list[dict[str, str]]:
    """Prüft typische Widersprüche, ohne Daten zu verändern."""
    result: list[dict[str, str]] = []
    objekt_key = norm_key(objektordner)

    def passend(titel: str, row: list[Any]) -> bool:
        if not objekt_key:
            return True
        return norm_key(objektordner_fuer_datensatz(titel, row)) == objekt_key

    # Freie Wohnung mit Zahlungseingang.
    for wrow in DATA.get("Wohnungen", []):
        if not passend("Wohnungen", wrow):
            continue
        status = feldwert("Wohnungen", wrow, ["Status"])
        if not any(word in norm_key(status) for word in ["frei", "leer", "unvermietet"]):
            continue

        wohnung = feldwert("Wohnungen", wrow, ["Wohnung", "Wohnungsordner"])
        objekt = objektordner_fuer_datensatz("Wohnungen", wrow)

        for zrow in DATA.get("Zahlungen", []):
            ztext = norm_key(" ".join(str(v) for v in zrow))
            if norm_key(wohnung) and norm_key(wohnung) in ztext and "miete" in ztext:
                result.append({
                    "ampel": "🔴",
                    "bereich": "Wohnungen/Zahlungen",
                    "objekt": objekt,
                    "hinweis": f"Freie Wohnung '{wohnung}' hat einen Mieteingang.",
                    "aktion": "Status und Zahlung prüfen",
                })
                break

    # Ausgezogener/gekündigter Mieter mit aktivem Vertrag.
    for mrow in DATA.get("Mieter", []):
        if not passend("Mieter", mrow):
            continue
        mieter = feldwert("Mieter", mrow, ["Mieter"])
        mstatus = feldwert("Mieter", mrow, ["Mieter-Status", "Status"])
        if norm_key(mstatus) not in {"ausgezogen", "gekuendigt", "gekündigt"}:
            continue

        for vrow in DATA.get("Mietverträge", []):
            vmieter = feldwert("Mietverträge", vrow, ["Mieter"])
            vstatus = feldwert("Mietverträge", vrow, ["Status"])
            if norm_key(vmieter) == norm_key(mieter) and norm_key(vstatus) in {"aktiv", "laufend"}:
                result.append({
                    "ampel": "🔴",
                    "bereich": "Mieter/Mietverträge",
                    "objekt": objektordner_fuer_datensatz("Mieter", mrow),
                    "hinweis": f"Mieter '{mieter}' ist {mstatus}, Vertrag aber aktiv.",
                    "aktion": "Mietvertrag prüfen",
                })

    # Doppelte Rechnungsnummer.
    gesehen: dict[str, str] = {}
    for rrow in DATA.get("Rechnungen", []):
        if not passend("Rechnungen", rrow):
            continue
        nummer = feldwert("Rechnungen", rrow, ["Rechnungsnr.", "Rechnungsnummer"])
        if not nummer:
            continue
        key = norm_key(nummer)
        if key in gesehen:
            result.append({
                "ampel": "🟠",
                "bereich": "Rechnungen",
                "objekt": objektordner_fuer_datensatz("Rechnungen", rrow),
                "hinweis": f"Doppelte Rechnungsnummer: {nummer}",
                "aktion": "Dublettenprüfung durchführen",
            })
        else:
            gesehen[key] = nummer

    # Fristen ohne Datum.
    for frow in DATA.get("Fristen", []):
        if not passend("Fristen", frow):
            continue
        titel = feldwert("Fristen", frow, ["Titel"])
        faellig = feldwert("Fristen", frow, ["Fällig am", "Termin"])
        if not faellig:
            result.append({
                "ampel": "🟡",
                "bereich": "Fristen",
                "objekt": objektordner_fuer_datensatz("Fristen", frow),
                "hinweis": f"Frist '{titel}' hat kein Fälligkeitsdatum.",
                "aktion": "Datum ergänzen",
            })

    # Betriebskosten ohne Kundennummer.
    for brow in DATA.get("Betriebskosten", []):
        if not passend("Betriebskosten", brow):
            continue
        versorger = feldwert("Betriebskosten", brow, ["Versorger"])
        kundennummer = feldwert("Betriebskosten", brow, ["Kundennummer"])
        if versorger and not kundennummer:
            result.append({
                "ampel": "🟡",
                "bereich": "Betriebskosten",
                "objekt": objektordner_fuer_datensatz("Betriebskosten", brow),
                "hinweis": f"Bei '{versorger}' fehlt die Kundennummer.",
                "aktion": "Kundennummer ergänzen",
            })

    return result


def jahrespruefung_objekt(objektordner: str) -> list[dict[str, str]]:
    """Erstellt einen automatischen Objekt-Prüfbericht."""
    pruefungen: list[dict[str, str]] = []
    status = smart_objekt_status(objektordner)
    health = objekt_gesundheitsindex(objektordner)

    pruefungen.append({
        "status": health["ampel"],
        "pruefpunkt": "Objekt-Gesundheitsindex",
        "ergebnis": f'{health["punkte"]} % – {health["bewertung"]}',
        "aktion": "; ".join(health["gruende"]) if health["gruende"] else "Keine Auffälligkeiten",
    })

    checks = [
        ("Wohnungen vorhanden", int(status["wohnungen"]) > 0, f'{status["wohnungen"]} Wohnung(en)'),
        ("Mieter vorhanden", int(status["mieter"]) > 0, f'{status["mieter"]} Mieter'),
        ("Keine überfälligen Aufgaben", int(status["ueberfaellige_aufgaben"]) == 0, f'{status["ueberfaellige_aufgaben"]} überfällig'),
        ("Keine offenen Schäden", int(status["offene_schaeden"]) == 0, f'{status["offene_schaeden"]} offen'),
        ("Keine offenen Rechnungen", int(status["offene_rechnungen"]) == 0, f'{status["offene_rechnungen"]} offen'),
        ("Kein Leerstand", int(status["freie_wohnungen"]) == 0, f'{status["freie_wohnungen"]} frei'),
        ("Dokumente vorhanden", int(status["dokumente"]) > 0, f'{status["dokumente"]} Dokument(e)'),
    ]

    for name, ok, detail in checks:
        pruefungen.append({
            "status": "🟢" if ok else "🔴",
            "pruefpunkt": name,
            "ergebnis": detail,
            "aktion": "Keine Aktion" if ok else "Prüfung erforderlich",
        })

    for item in plausibilitaetspruefungen(objektordner):
        pruefungen.append({
            "status": item["ampel"],
            "pruefpunkt": item["bereich"],
            "ergebnis": item["hinweis"],
            "aktion": item["aktion"],
        })

    return pruefungen


class WorkflowCenterSeite(QWidget):
    """Zentrale Vorschläge, Warnungen und Objekt-Gesundheitswerte."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav

        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 22)
        root.setSpacing(14)

        title = QLabel("Workflow-Center PRO")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel(
            "Automatische Vorschläge, Plausibilitätsprüfungen und Objekt-Gesundheitswerte."
        )
        info.setObjectName("subTitle")
        root.addWidget(info)

        top = QHBoxLayout()
        self.objekt = QComboBox()
        self.objekt.addItem("Alle Objektordner")
        self.objekt.addItems(alle_objektordner())
        self.objekt.currentTextChanged.connect(self.laden)

        refresh = QPushButton("Neu prüfen")
        refresh.setObjectName("primaryButton")
        refresh.clicked.connect(self.laden)

        jahrespruefung = QPushButton("Jahresprüfung öffnen")
        jahrespruefung.clicked.connect(lambda: self.nav("Jahresprüfung PRO"))

        top.addWidget(QLabel("Objektordner:"))
        top.addWidget(self.objekt)
        top.addWidget(refresh)
        top.addWidget(jahrespruefung)
        top.addStretch()
        root.addLayout(top)

        self.cards = QGridLayout()
        root.addLayout(self.cards)

        self.tabs = QTabWidget()
        root.addWidget(self.tabs, 1)

        self.tbl_warnungen = QTableWidget()
        self.tbl_gesundheit = QTableWidget()

        self.tabs.addTab(self.tbl_warnungen, "Warnungen & Vorschläge")
        self.tabs.addTab(self.tbl_gesundheit, "Objekt-Gesundheit")

        self.tbl_warnungen.cellDoubleClicked.connect(self._open_area)
        self.laden()

    @staticmethod
    def _card(title: str, value: str, icon: str) -> QFrame:
        card = QFrame()
        card.setObjectName("metricCard")
        layout = QVBoxLayout(card)
        label = QLabel(f"{icon}  {title}")
        label.setObjectName("metricTitle")
        amount = QLabel(value)
        amount.setObjectName("metricValue")
        layout.addWidget(label)
        layout.addWidget(amount)
        return card

    def laden(self) -> None:
        selected = self.objekt.currentText().strip()
        objekt = "" if selected == "Alle Objektordner" else selected

        warnungen = plausibilitaetspruefungen(objekt)

        while self.cards.count():
            item = self.cards.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        rot = sum(1 for item in warnungen if item["ampel"] == "🔴")
        orange = sum(1 for item in warnungen if item["ampel"] == "🟠")
        gelb = sum(1 for item in warnungen if item["ampel"] == "🟡")
        gesamt = len(warnungen)

        for index, card in enumerate([
            ("Kritisch", str(rot), "🔴"),
            ("Prüfen", str(orange), "🟠"),
            ("Hinweise", str(gelb), "🟡"),
            ("Gesamt", str(gesamt), "📋"),
        ]):
            self.cards.addWidget(self._card(*card), 0, index)

        self.tbl_warnungen.setColumnCount(5)
        self.tbl_warnungen.setHorizontalHeaderLabels(
            ["Ampel", "Bereich", "Objektordner", "Hinweis", "Empfohlene Aktion"]
        )
        self.tbl_warnungen.setRowCount(len(warnungen))
        self.tbl_warnungen.setAlternatingRowColors(True)
        self.tbl_warnungen.setWordWrap(False)
        self.tbl_warnungen.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)

        for row_index, item in enumerate(warnungen):
            values = [
                item["ampel"], item["bereich"], item["objekt"],
                item["hinweis"], item["aktion"],
            ]
            for col_index, value in enumerate(values):
                self.tbl_warnungen.setItem(
                    row_index, col_index, QTableWidgetItem(str(value))
                )

        for col_index, width in enumerate([70, 180, 240, 520, 320]):
            self.tbl_warnungen.setColumnWidth(col_index, width)

        objekte = [objekt] if objekt else alle_objektordner()
        self.tbl_gesundheit.setColumnCount(5)
        self.tbl_gesundheit.setHorizontalHeaderLabels(
            ["Ampel", "Objektordner", "Index", "Bewertung", "Begründung"]
        )
        self.tbl_gesundheit.setRowCount(len(objekte))
        self.tbl_gesundheit.setAlternatingRowColors(True)

        for row_index, obj in enumerate(objekte):
            health = objekt_gesundheitsindex(obj)
            values = [
                health["ampel"],
                obj,
                f'{health["punkte"]} %',
                health["bewertung"],
                "; ".join(health["gruende"]) if health["gruende"] else "Keine Auffälligkeiten",
            ]
            for col_index, value in enumerate(values):
                self.tbl_gesundheit.setItem(
                    row_index, col_index, QTableWidgetItem(str(value))
                )

        for col_index, width in enumerate([70, 260, 100, 150, 600]):
            self.tbl_gesundheit.setColumnWidth(col_index, width)

    def _open_area(self, row: int, _column: int) -> None:
        item = self.tbl_warnungen.item(row, 1)
        if item is None:
            return
        target = item.text().split("/")[0].strip()
        if target:
            self.nav(target)


class JahrespruefungProSeite(QWidget):
    """Automatischer Prüfbericht für einen Objektordner."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.rows: list[dict[str, str]] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 22)
        root.setSpacing(14)

        title = QLabel("Jahresprüfung PRO")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel(
            "Prüft Aufgaben, Schäden, Rechnungen, Leerstände, Dokumente und Plausibilität."
        )
        info.setObjectName("subTitle")
        root.addWidget(info)

        top = QHBoxLayout()
        self.objekt = QComboBox()
        self.objekt.setEditable(True)
        self.objekt.addItems(alle_objektordner())
        self.objekt.currentTextChanged.connect(self.laden)

        refresh = QPushButton("Objekt prüfen")
        refresh.setObjectName("primaryButton")
        refresh.clicked.connect(self.laden)

        export_btn = QPushButton("Prüfbericht exportieren")
        export_btn.clicked.connect(self.exportieren)

        top.addWidget(QLabel("Objektordner:"))
        top.addWidget(self.objekt, 1)
        top.addWidget(refresh)
        top.addWidget(export_btn)
        root.addLayout(top)

        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(
            ["Status", "Prüfpunkt", "Ergebnis", "Empfohlene Aktion"]
        )
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        root.addWidget(self.table, 1)

        self.laden()

    def laden(self) -> None:
        objekt = self.objekt.currentText().strip()
        self.rows = jahrespruefung_objekt(objekt) if objekt else []
        self.table.setRowCount(len(self.rows))

        for row_index, item in enumerate(self.rows):
            values = [
                item["status"], item["pruefpunkt"],
                item["ergebnis"], item["aktion"],
            ]
            for col_index, value in enumerate(values):
                self.table.setItem(
                    row_index, col_index, QTableWidgetItem(str(value))
                )

        for col_index, width in enumerate([80, 280, 520, 420]):
            self.table.setColumnWidth(col_index, width)

    def exportieren(self) -> None:
        if not self.rows:
            QMessageBox.information(self, "Jahresprüfung", "Kein Prüfbericht vorhanden.")
            return

        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        default = EXPORT_DIR / (
            "jahrespruefung_"
            + datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            + ".xlsx"
        )

        target, _ = QFileDialog.getSaveFileName(
            self,
            "Prüfbericht exportieren",
            str(default),
            "Excel-Dateien (*.xlsx)",
        )
        if not target:
            return

        target_path = Path(target)
        if target_path.suffix.lower() != ".xlsx":
            target_path = target_path.with_suffix(".xlsx")

        workbook = Workbook()
        sheet = workbook.active
        if not isinstance(sheet, Worksheet):
            QMessageBox.warning(self, "Export", "Arbeitsblatt konnte nicht erstellt werden.")
            return

        sheet.title = "Jahresprüfung"
        sheet.append(["Status", "Prüfpunkt", "Ergebnis", "Empfohlene Aktion"])

        for item in self.rows:
            sheet.append([
                item["status"],
                item["pruefpunkt"],
                item["ergebnis"],
                item["aktion"],
            ])

        for col_index, width in enumerate([12, 36, 70, 55], start=1):
            sheet.column_dimensions[sheet.cell(1, col_index).column_letter].width = width

        workbook.save(target_path)
        QMessageBox.information(
            self,
            "Jahresprüfung",
            f"Prüfbericht exportiert:\n{target_path}",
        )



def aktuelle_mitarbeiterrolle() -> str:
    return str(CURRENT_USER.get("rolle", "") or "").strip()


def aktueller_mitarbeitername() -> str:
    return str(CURRENT_USER.get("name", "") or "").strip()


def mitarbeiter_aufgaben(name: str, nur_offen: bool = True) -> list[list[Any]]:
    result: list[list[Any]] = []
    name_key = norm_key(name)

    for row in DATA.get("Aufgaben", []):
        verantwortlich = aufgaben_feld(row, "Verantwortlich")
        if name_key and norm_key(verantwortlich) != name_key:
            continue

        status = aufgaben_feld(row, "Status")
        if nur_offen and not aufgaben_status_ist_offen(status):
            continue

        result.append(row)

    result.sort(key=lambda row: datum_sort_key(aufgaben_feld(row, "Fällig am")))
    return result


def heutige_mitarbeiter_fristen(name: str) -> list[dict[str, str]]:

    heute = date.today()
    name_key = norm_key(name)
    result: list[dict[str, str]] = []

    for item in fristen_pro_status():
        mieter = norm_key(item.get("mieter", ""))
        bereich = norm_key(item.get("bereich", ""))
        titel = norm_key(item.get("titel", ""))

        mitarbeiter_bezug = (
            not name_key
            or name_key in mieter
            or name_key in bereich
            or name_key in titel
        )

        if not mitarbeiter_bezug and aktuelle_mitarbeiterrolle().lower() != "admin":
            continue

        parsed = None
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y"):
            try:
                parsed = datetime.strptime(item.get("faellig", ""), fmt).date()
                break
            except ValueError:
                continue

        if parsed == heute or item.get("bewertung") in {"Überfällig", "Dringend"}:
            result.append(item)

    return result


class MitarbeiterStartcenterSeite(QWidget):
    """Persönliches schwarzes Brett für den eingeloggten Mitarbeiter."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.name = aktueller_mitarbeitername()
        self.rolle = aktuelle_mitarbeiterrolle()
        self.aufgaben_rows: list[list[Any]] = []
        self.fristen_rows: list[dict[str, str]] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(20, 18, 20, 18)
        root.setSpacing(14)

        title = QLabel(f"Mein Arbeitstag – {self.name or 'Mitarbeiter'}")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel(
            "Persönliche Aufgaben, dringende Fristen und Schnellzugriffe für den aktuellen Arbeitstag."
        )
        info.setObjectName("subTitle")
        info.setWordWrap(True)
        root.addWidget(info)

        quick = QHBoxLayout()

        refresh = QPushButton("Aktualisieren")
        refresh.setObjectName("primaryButton")
        refresh.clicked.connect(self.laden)

        neue_aufgabe = QPushButton("Neue Aufgabe")
        neue_aufgabe.clicked.connect(self.neue_aufgabe)

        aufgaben = QPushButton("Arbeitsorganisation öffnen")
        aufgaben.clicked.connect(lambda: self.nav("Arbeitsorganisation PRO"))

        fristen = QPushButton("Fristenmanager öffnen")
        fristen.clicked.connect(lambda: self.nav("Fristenmanager PRO"))

        workflow = QPushButton("Workflow-Center öffnen")
        workflow.clicked.connect(lambda: self.nav("Workflow-Center PRO"))

        quick.addWidget(refresh)
        quick.addWidget(neue_aufgabe)
        quick.addWidget(aufgaben)
        quick.addWidget(fristen)
        quick.addWidget(workflow)
        quick.addStretch()
        root.addLayout(quick)

        self.cards = QGridLayout()
        self.cards.setSpacing(12)
        root.addLayout(self.cards)

        self.tabs = QTabWidget()
        root.addWidget(self.tabs, 1)

        self.tbl_aufgaben = QTableWidget()
        self.tbl_fristen = QTableWidget()
        self.tbl_uebersicht = QTableWidget()

        self.tabs.addTab(self.tbl_aufgaben, "Meine Aufgaben")
        self.tabs.addTab(self.tbl_fristen, "Dringende Fristen")
        self.tabs.addTab(self.tbl_uebersicht, "Tagesübersicht")

        self.tbl_aufgaben.setAlternatingRowColors(True)
        self.tbl_aufgaben.setWordWrap(False)
        self.tbl_aufgaben.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.tbl_aufgaben.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.tbl_aufgaben.cellDoubleClicked.connect(self._open_aufgaben)

        self.tbl_fristen.setAlternatingRowColors(True)
        self.tbl_fristen.setWordWrap(False)
        self.tbl_fristen.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.tbl_fristen.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.tbl_fristen.cellDoubleClicked.connect(self._open_fristen)

        self.tbl_uebersicht.setAlternatingRowColors(True)
        self.tbl_uebersicht.setWordWrap(False)

        self.laden()

    @staticmethod
    def _card(title: str, value: str, icon: str) -> QFrame:
        card = QFrame()
        card.setObjectName("metricCard")
        card.setMinimumHeight(100)

        layout = QVBoxLayout(card)
        label = QLabel(f"{icon}  {title}")
        label.setObjectName("metricTitle")
        label.setWordWrap(True)

        amount = QLabel(value)
        amount.setObjectName("metricValue")
        amount.setWordWrap(True)

        layout.addWidget(label)
        layout.addWidget(amount)
        return card

    def laden(self) -> None:
        self.name = aktueller_mitarbeitername()
        self.rolle = aktuelle_mitarbeiterrolle()

        self.aufgaben_rows = mitarbeiter_aufgaben(self.name, True)
        self.fristen_rows = heutige_mitarbeiter_fristen(self.name)

        while self.cards.count():
            item = self.cards.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        offen = len(self.aufgaben_rows)
        ueberfaellig = sum(
            1 for row in self.aufgaben_rows
            if row in ueberfaellige_aufgaben_rows()
        )
        dringend = sum(
            1 for row in self.aufgaben_rows
            if norm_key(aufgaben_feld(row, "Priorität")) in {"hoch", "kritisch", "dringend"}
        )
        fristen = len(self.fristen_rows)

        cards = [
            ("Offene Aufgaben", str(offen), "🗂"),
            ("Überfällig", str(ueberfaellig), "🔴"),
            ("Dringend", str(dringend), "🟠"),
            ("Fristen", str(fristen), "⏰"),
        ]

        for index, values in enumerate(cards):
            self.cards.addWidget(self._card(*values), 0, index)

        self._fill_aufgaben()
        self._fill_fristen()
        self._fill_uebersicht()

    def _fill_aufgaben(self) -> None:
        headers = SCHEMA.get("Aufgaben", [])
        self.tbl_aufgaben.setColumnCount(len(headers))
        self.tbl_aufgaben.setHorizontalHeaderLabels(headers)
        self.tbl_aufgaben.setRowCount(len(self.aufgaben_rows))

        for row_index, row in enumerate(self.aufgaben_rows):
            for col_index in range(len(headers)):
                value = row[col_index] if col_index < len(row) else ""
                self.tbl_aufgaben.setItem(
                    row_index, col_index, QTableWidgetItem(str(value))
                )

        for col_index in range(min(len(headers), 13)):
            self.tbl_aufgaben.setColumnWidth(col_index, 160)

    def _fill_fristen(self) -> None:
        headers = [
            "Ampel", "Titel", "Bereich", "Objektordner",
            "Fällig am", "Tage", "Priorität", "Bewertung"
        ]
        self.tbl_fristen.setColumnCount(len(headers))
        self.tbl_fristen.setHorizontalHeaderLabels(headers)
        self.tbl_fristen.setRowCount(len(self.fristen_rows))

        for row_index, item in enumerate(self.fristen_rows):
            values = [
                item.get("ampel", ""),
                item.get("titel", ""),
                item.get("bereich", ""),
                item.get("objekt", ""),
                item.get("faellig", ""),
                item.get("tage", ""),
                item.get("prioritaet", ""),
                item.get("bewertung", ""),
            ]
            for col_index, value in enumerate(values):
                self.tbl_fristen.setItem(
                    row_index, col_index, QTableWidgetItem(str(value))
                )

        for col_index, width in enumerate([70, 240, 170, 240, 110, 80, 120, 150]):
            self.tbl_fristen.setColumnWidth(col_index, width)

    def _fill_uebersicht(self) -> None:
        rows = [
            ["Mitarbeiter", self.name],
            ["Rolle", self.rolle],
            ["Offene Aufgaben", str(len(self.aufgaben_rows))],
            ["Dringende Fristen", str(len(self.fristen_rows))],
            ["Arbeitsstatus", "Handlungsbedarf" if self.aufgaben_rows or self.fristen_rows else "Alles erledigt"],
        ]

        self.tbl_uebersicht.setColumnCount(2)
        self.tbl_uebersicht.setHorizontalHeaderLabels(["Kennzahl", "Wert"])
        self.tbl_uebersicht.setRowCount(len(rows))

        for row_index, row in enumerate(rows):
            for col_index, value in enumerate(row):
                self.tbl_uebersicht.setItem(
                    row_index, col_index, QTableWidgetItem(str(value))
                )

        self.tbl_uebersicht.setColumnWidth(0, 260)
        self.tbl_uebersicht.setColumnWidth(1, 500)

    def neue_aufgabe(self) -> None:
        values = ["" for _ in SCHEMA["Aufgaben"]]

        defaults = {
            "Verantwortlich": self.name,
            "Status": "Neu",
            "Priorität": "Normal",
        }

        for index, feld in enumerate(SCHEMA["Aufgaben"]):
            if feld in defaults:
                values[index] = defaults[feld]

        dialog = EingabeDialog(
            "Neue persönliche Aufgabe",
            SCHEMA["Aufgaben"],
            values,
        )

        if dialog.exec() == QDialog.DialogCode.Accepted:
            DATA["Aufgaben"].append(dialog.values())
            speichere_tabelle("Aufgaben")
            self.laden()

    def _open_aufgaben(self, _row: int, _column: int) -> None:
        self.nav("Arbeitsorganisation PRO")

    def _open_fristen(self, _row: int, _column: int) -> None:
        self.nav("Fristenmanager PRO")



def kalender_datum(value: Any) -> date | None:
    """Parst die in den bestehenden Tabellen verwendeten Datumsformate."""
    text_value = str(value or "").strip()
    if not text_value:
        return None

    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y"):
        try:
            return datetime.strptime(text_value, fmt).date()
        except ValueError:
            continue

    return None


def kalender_eintraege() -> list[dict[str, str]]:
    """
    Erzeugt Kalendertermine aus bestehenden Aufgaben und Fristen.
    Es werden keine neuen Excelspalten oder Tabellen benötigt.
    """
    eintraege: list[dict[str, str]] = []

    for row in DATA.get("Aufgaben", []):
        datum_text = aufgaben_feld(row, "Fällig am")
        parsed = kalender_datum(datum_text)
        if parsed is None:
            continue

        status = aufgaben_feld(row, "Status")
        if not aufgaben_status_ist_offen(status):
            continue

        eintraege.append({
            "datum": parsed.isoformat(),
            "typ": "Aufgabe",
            "titel": aufgaben_feld(row, "Aufgabe"),
            "objekt": objektordner_fuer_datensatz("Aufgaben", row),
            "mitarbeiter": aufgaben_feld(row, "Verantwortlich"),
            "prioritaet": aufgaben_feld(row, "Priorität"),
            "status": status,
            "bereich": "Arbeitsorganisation PRO",
            "beschreibung": aufgaben_feld(row, "Notiz"),
        })

    for row in DATA.get("Fristen", []):
        datum_text = feldwert("Fristen", row, ["Fällig am", "Termin"])
        parsed = kalender_datum(datum_text)
        if parsed is None:
            continue

        status = feldwert("Fristen", row, ["Status"])
        if norm_key(status) in {"erledigt", "abgeschlossen", "archiviert"}:
            continue

        eintraege.append({
            "datum": parsed.isoformat(),
            "typ": "Frist",
            "titel": feldwert("Fristen", row, ["Titel"]),
            "objekt": objektordner_fuer_datensatz("Fristen", row),
            "mitarbeiter": feldwert("Fristen", row, ["Mieter", "Bereich"]),
            "prioritaet": feldwert("Fristen", row, ["Priorität"]),
            "status": status,
            "bereich": "Fristenmanager PRO",
            "beschreibung": feldwert("Fristen", row, ["Notiz"]),
        })

    eintraege.sort(key=lambda item: (item["datum"], item["typ"], item["titel"]))
    return eintraege


def ics_text(eintraege: list[dict[str, str]]) -> str:
    """Erzeugt eine standardisierte iCalendar-Datei für Outlook, Google und Android."""
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//ImmoVerwaltung//Kalender 4.5//DE",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
    ]

    now_stamp = datetime.now().strftime("%Y%m%dT%H%M%S")

    for index, item in enumerate(eintraege, start=1):
        datum = kalender_datum(item.get("datum", ""))
        if datum is None:
            continue

        start_value = datum.strftime("%Y%m%d")
        end_value = (datum + timedelta(days=1)).strftime("%Y%m%d")

        summary = f'{item.get("typ", "Termin")}: {item.get("titel", "")}'.replace("\n", " ")
        description_parts = [
            f'Objekt: {item.get("objekt", "")}',
            f'Mitarbeiter/Bereich: {item.get("mitarbeiter", "")}',
            f'Priorität: {item.get("prioritaet", "")}',
            f'Status: {item.get("status", "")}',
            item.get("beschreibung", ""),
        ]
        description = "\\n".join(part.replace("\n", " ") for part in description_parts if part)

        uid = f"immoverwaltung-{start_value}-{index}@local"

        lines.extend([
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTAMP:{now_stamp}",
            f"DTSTART;VALUE=DATE:{start_value}",
            f"DTEND;VALUE=DATE:{end_value}",
            f"SUMMARY:{summary}",
            f"DESCRIPTION:{description}",
            f"CATEGORIES:{item.get('typ', 'Termin')}",
            "END:VEVENT",
        ])

    lines.append("END:VCALENDAR")
    return "\r\n".join(lines) + "\r\n"


class KalenderPlanungSeite(QWidget):
    """Interner Planungskalender mit Aufgaben, Fristen und ICS-Anbindung."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.entries: list[dict[str, str]] = []
        self.selected_entries: list[dict[str, str]] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(20, 18, 20, 18)
        root.setSpacing(14)

        title = QLabel("Kalender & Planung")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel(
            "Aufgaben und Fristen werden automatisch im Kalender angezeigt. "
            "Der Kalender kann als ICS-Datei nach Outlook, Google Kalender oder Android exportiert werden."
        )
        info.setObjectName("subTitle")
        info.setWordWrap(True)
        root.addWidget(info)

        toolbar = QHBoxLayout()

        self.objekt_filter = QComboBox()
        self.objekt_filter.addItem("Alle Objektordner")
        self.objekt_filter.addItems(alle_objektordner())
        self.objekt_filter.currentTextChanged.connect(self.refresh)

        self.mitarbeiter_filter = QComboBox()
        self.mitarbeiter_filter.addItem("Alle Mitarbeiter")
        self.mitarbeiter_filter.addItems(["Julia", "Franzi", "Robert", "Ralf", "Admin"])
        self.mitarbeiter_filter.currentTextChanged.connect(self.refresh)

        refresh_btn = QPushButton("Aktualisieren")
        refresh_btn.setObjectName("primaryButton")
        refresh_btn.clicked.connect(self.refresh)

        neue_aufgabe = QPushButton("Neue Aufgabe")
        neue_aufgabe.clicked.connect(self.neue_aufgabe)

        neue_frist = QPushButton("Neue Frist")
        neue_frist.clicked.connect(self.neue_frist)

        export_btn = QPushButton("Kalender exportieren (.ics)")
        export_btn.clicked.connect(self.export_ics)

        toolbar.addWidget(QLabel("Objektordner:"))
        toolbar.addWidget(self.objekt_filter)
        toolbar.addWidget(QLabel("Mitarbeiter:"))
        toolbar.addWidget(self.mitarbeiter_filter)
        toolbar.addWidget(refresh_btn)
        toolbar.addWidget(neue_aufgabe)
        toolbar.addWidget(neue_frist)
        toolbar.addWidget(export_btn)
        root.addLayout(toolbar)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        root.addWidget(splitter, 1)

        left = QWidget()
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(0, 0, 8, 0)

        self.calendar = QCalendarWidget()
        self.calendar.setGridVisible(True)
        self.calendar.setNavigationBarVisible(True)
        self.calendar.setMinimumWidth(390)
        self.calendar.selectionChanged.connect(self.refresh_day)
        left_layout.addWidget(self.calendar)

        self.day_summary = QLabel("")
        self.day_summary.setObjectName("metricTitle")
        self.day_summary.setWordWrap(True)
        left_layout.addWidget(self.day_summary)

        splitter.addWidget(left)

        right = QWidget()
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(8, 0, 0, 0)

        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels([
            "Datum", "Typ", "Titel", "Objektordner",
            "Mitarbeiter/Bereich", "Priorität", "Status", "Beschreibung"
        ])
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.cellDoubleClicked.connect(self.open_entry)
        right_layout.addWidget(self.table, 1)

        splitter.addWidget(right)
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)
        splitter.setSizes([420, 900])

        self.refresh()

    def refresh(self) -> None:
        entries = kalender_eintraege()

        selected_objekt = self.objekt_filter.currentText().strip()
        if selected_objekt != "Alle Objektordner":
            entries = [
                item for item in entries
                if norm_key(item.get("objekt", "")) == norm_key(selected_objekt)
            ]

        selected_mitarbeiter = self.mitarbeiter_filter.currentText().strip()
        if selected_mitarbeiter != "Alle Mitarbeiter":
            entries = [
                item for item in entries
                if norm_key(selected_mitarbeiter) in norm_key(item.get("mitarbeiter", ""))
            ]

        self.entries = entries
        self.refresh_day()

    def refresh_day(self) -> None:
        selected_qdate = self.calendar.selectedDate()
        selected_date = date(
            selected_qdate.year(),
            selected_qdate.month(),
            selected_qdate.day(),
        )
        selected_iso = selected_date.isoformat()

        self.selected_entries = [
            item for item in self.entries
            if item.get("datum") == selected_iso
        ]

        self.day_summary.setText(
            f"{selected_date.strftime('%d.%m.%Y')} – "
            f"{len(self.selected_entries)} Termin(e)"
        )

        self.table.setRowCount(len(self.selected_entries))

        for row_index, item in enumerate(self.selected_entries):
            parsed = kalender_datum(item.get("datum", ""))
            display_date = parsed.strftime("%d.%m.%Y") if parsed else item.get("datum", "")

            values = [
                display_date,
                item.get("typ", ""),
                item.get("titel", ""),
                item.get("objekt", ""),
                item.get("mitarbeiter", ""),
                item.get("prioritaet", ""),
                item.get("status", ""),
                item.get("beschreibung", ""),
            ]

            for col_index, value in enumerate(values):
                self.table.setItem(
                    row_index, col_index, QTableWidgetItem(str(value))
                )

        for col_index, width in enumerate([110, 90, 250, 230, 190, 110, 120, 420]):
            self.table.setColumnWidth(col_index, width)

    def neue_aufgabe(self) -> None:
        values = ["" for _ in SCHEMA["Aufgaben"]]
        selected_date = self.calendar.selectedDate()
        date_text = selected_date.toString("dd.MM.yyyy")

        defaults = {
            "Fällig am": date_text,
            "Verantwortlich": aktueller_mitarbeitername(),
            "Status": "Neu",
            "Priorität": "Normal",
        }

        selected_objekt = self.objekt_filter.currentText().strip()
        if selected_objekt != "Alle Objektordner":
            defaults["Objekt"] = selected_objekt
            defaults["Objektordner"] = selected_objekt

        for index, field in enumerate(SCHEMA["Aufgaben"]):
            if field in defaults:
                values[index] = defaults[field]

        dialog = EingabeDialog("Neue Kalender-Aufgabe", SCHEMA["Aufgaben"], values)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            DATA["Aufgaben"].append(dialog.values())
            speichere_tabelle("Aufgaben")
            self.refresh()

    def neue_frist(self) -> None:
        values = ["" for _ in SCHEMA["Fristen"]]
        selected_date = self.calendar.selectedDate()
        date_text = selected_date.toString("dd.MM.yyyy")

        defaults = {
            "Fällig am": date_text,
            "Status": "Neu",
            "Priorität": "Normal",
        }

        selected_objekt = self.objekt_filter.currentText().strip()
        if selected_objekt != "Alle Objektordner":
            defaults["Objekt"] = selected_objekt
            defaults["Objektordner"] = selected_objekt

        for index, field in enumerate(SCHEMA["Fristen"]):
            if field in defaults:
                values[index] = defaults[field]

        dialog = EingabeDialog("Neue Kalender-Frist", SCHEMA["Fristen"], values)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            DATA["Fristen"].append(dialog.values())
            speichere_tabelle("Fristen")
            self.refresh()

    def open_entry(self, row: int, _column: int) -> None:
        if row < 0 or row >= len(self.selected_entries):
            return

        target = self.selected_entries[row].get("bereich", "")
        if target:
            self.nav(target)

    def export_ics(self) -> None:
        if not self.entries:
            QMessageBox.information(self, "Kalender", "Es sind keine Termine zum Exportieren vorhanden.")
            return

        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        default = EXPORT_DIR / (
            "immoverwaltung_kalender_"
            + datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            + ".ics"
        )

        target, _ = QFileDialog.getSaveFileName(
            self,
            "Kalender exportieren",
            str(default),
            "iCalendar-Dateien (*.ics)",
        )
        if not target:
            return

        target_path = Path(target)
        if target_path.suffix.lower() != ".ics":
            target_path = target_path.with_suffix(".ics")

        target_path.write_text(ics_text(self.entries), encoding="utf-8")

        QMessageBox.information(
            self,
            "Kalender",
            "Kalenderdatei wurde erstellt.\n"
            "Sie kann in Outlook, Google Kalender oder einer Android-Kalender-App importiert werden:\n"
            f"{target_path}",
        )



class EinstellungenSeite(QWidget):
    """Menüseite zum Öffnen und Verwalten der Programmeinstellungen."""

    def __init__(self, nav=None):
        super().__init__()
        self.nav = nav

        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 22)
        root.setSpacing(14)

        title = QLabel("Einstellungen")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel(
            "Speicherorte, Darstellung, Vollbild, Tabellenhöhe und Backups konfigurieren."
        )
        info.setObjectName("subTitle")
        info.setWordWrap(True)
        root.addWidget(info)

        panel = QFrame()
        panel.setObjectName("chartPanel")
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(20, 18, 20, 18)
        layout.setSpacing(12)

        details = QLabel(
            "Aktuelle Konfiguration:\n\n"
            f"Datenordner: {DATEN_DIR}\n"
            f"Dokumente: {DOKUMENTE_DIR}\n"
            f"Exporte: {EXPORT_DIR}\n"
            f"Backups: {BACKUP_DIR}\n"
            f"Darstellung: {CONFIG.get('theme', 'hell')}\n"
            f"Start im Vollbild: {CONFIG.get('start_vollbild', 'ja')}"
        )
        details.setWordWrap(True)
        layout.addWidget(details)

        open_dialog = QPushButton("Einstellungen bearbeiten")
        open_dialog.setObjectName("primaryButton")
        open_dialog.clicked.connect(self.open_settings)
        layout.addWidget(open_dialog)

        backup_btn = QPushButton("Backup jetzt erstellen")
        backup_btn.clicked.connect(self.create_backup)
        layout.addWidget(backup_btn)

        open_config = QPushButton("Konfigurationsordner öffnen")
        open_config.clicked.connect(
            lambda: system_datei_oeffnen(str(CONFIG_FILE.parent))
        )
        layout.addWidget(open_config)

        layout.addStretch()
        root.addWidget(panel)
        root.addStretch()

    def open_settings(self) -> None:
        dialog = EinstellungenDialog()
        dialog.exec()

    def create_backup(self) -> None:
        target = erstelle_schnellbackup()
        if target is not None:
            QMessageBox.information(
                self,
                "Backup",
                f"Backup wurde erstellt:\n{target}",
            )



def benutzer_darf_seite_oeffnen(seite: str) -> bool:
    rolle = str(CURRENT_USER.get("rolle", "") or "").strip().lower()
    name = str(CURRENT_USER.get("name", "") or "").strip().lower()

    if rolle == "admin" or name in {"admin", "julia", "franzi"}:
        return True

    allgemein = {
        "Mein Arbeitstag", "Mobiles Mitarbeiterportal", "DBS Field Import", "DBS Field Eingangszentrale", "Kalender & Planung", "Smart Startseite", "Objektgalerie", "Enterprise Objekt 360°", "Dokumenten-Center 2.0", "Workflow-Regeln PRO", "Dashboard", "Verwaltungsleitstand 5.0",
        "Globale Suche", "Dokumente", "Dokumentenautomatisierung PRO", "Akten-Center",
        "Aufgaben", "Fristen", "Arbeitsorganisation PRO",
        "Objektchronik", "Smart Objektakte", "Digitaler Gebäudezwilling",
        "Fristenmanager PRO", "Workflow-Center PRO",
    }
    buchhaltung = {
        "Rechnungen", "Betriebskosten", "Buchhaltung",
        "Kontoauszug-Import", "Mietkonto-Abgleich",
        "Zahlungsabgleich PRO", "Belegscanner PRO",
        "BK-Automatik", "Jahresprüfung PRO",
    }
    technik = {
        "Schäden", "Dienstleister", "Versorger",
        "Übergabeprotokolle", "Brand- und Arbeitsschutz",
        "Schlüssel", "Wohnungen", "Objekte",
    }
    stammdaten = {
        "Mieter", "Mietverträge", "Wohnungsgeberauskunft",
        "Vermieterauskunft", "Objekte", "Wohnungen",
    }

    erlaubte = set(allgemein)
    if name == "franzi":
        erlaubte.update(buchhaltung)
        erlaubte.update(stammdaten)
    elif name == "robert":
        erlaubte.update(technik)
    elif name == "ralf":
        erlaubte.update(technik)
        erlaubte.update(stammdaten)
    elif name == "julia":
        erlaubte.update(stammdaten)
        erlaubte.update(buchhaltung)

    if seite in {
        "Berichte & Export PRO", "System-Center PRO", "Einstellungen", "Objektordner-Prüfung",
        "Mitarbeiter-Login", "Mitarbeiter",
        "Ereignisprotokoll", "Berechtigungen & Protokoll", "Berichte & Export PRO", "System-Center PRO", "Projektmonitor",
    }:
        return False

    return seite in erlaubte


def aktivitaet_protokollieren(
    bereich: str,
    aktion: str,
    beschreibung: str = "",
    status: str = "Erfolgreich",
) -> None:

    felder = SCHEMA.get("Ereignisprotokoll", [])
    if not felder:
        return

    now = datetime.now()
    values = {
        "Datum": now.strftime("%d.%m.%Y"),
        "Uhrzeit": now.strftime("%H:%M:%S"),
        "Bereich": bereich,
        "Aktion": aktion,
        "Beschreibung": beschreibung,
        "Benutzer": aktueller_mitarbeitername()
        or str(CURRENT_USER.get("benutzername", "")),
        "Status": status,
    }

    row = [values.get(feld, "") for feld in felder]
    DATA.setdefault("Ereignisprotokoll", []).append(row)

    try:
        speichere_tabelle("Ereignisprotokoll")
    except (OSError, PermissionError, ValueError):
        pass


def berechtigungsuebersicht() -> list[dict[str, str]]:
    seiten = [
        "Dashboard", "Verwaltungsleitstand 5.0", "Mein Arbeitstag", "Kalender & Planung", "DBS Field Kalenderimport",
        "Objekte", "Wohnungen", "Mieter", "Mietverträge",
        "Rechnungen", "Betriebskosten", "Buchhaltung",
        "Kontoauszug-Import", "Mietkonto-Abgleich",
        "Zahlungsabgleich PRO", "Belegscanner PRO",
        "Arbeitsorganisation PRO", "Objektchronik",
        "Digitaler Gebäudezwilling", "Smart Objektakte", "Fristenmanager PRO",
        "Workflow-Center PRO", "Jahresprüfung PRO",
        "Einstellungen", "Objektordner-Prüfung",
    ]

    users = [
        ("Julia", "Admin"),
        ("Franzi", "Admin"),
        ("Robert", "Mitarbeiter"),
        ("Ralf", "Mitarbeiter"),
        ("Admin", "Admin"),
    ]

    original = dict(CURRENT_USER)
    rows: list[dict[str, str]] = []

    try:
        for name, role in users:
            CURRENT_USER["name"] = name
            CURRENT_USER["rolle"] = role
            for page in seiten:
                rows.append({
                    "name": name,
                    "rolle": role,
                    "seite": page,
                    "zugriff": "Ja" if benutzer_darf_seite_oeffnen(page) else "Nein",
                })
    finally:
        CURRENT_USER.clear()
        CURRENT_USER.update(original)

    return rows


class BerechtigungenProtokollSeite(QWidget):
    def __init__(self, nav):
        super().__init__()
        self.nav = nav

        root = QVBoxLayout(self)
        root.setContentsMargins(22, 20, 22, 20)
        root.setSpacing(14)

        title = QLabel("Berechtigungen & Protokoll")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel(
            "Adminübersicht über Rollen, Rechte und protokollierte Aktivitäten."
        )
        info.setObjectName("subTitle")
        info.setWordWrap(True)
        root.addWidget(info)

        top = QHBoxLayout()
        refresh = QPushButton("Aktualisieren")
        refresh.setObjectName("primaryButton")
        refresh.clicked.connect(self.laden)

        open_log = QPushButton("Ereignisprotokoll öffnen")
        open_log.clicked.connect(lambda: self.nav("Ereignisprotokoll"))

        export_btn = QPushButton("Rechteübersicht exportieren")
        export_btn.clicked.connect(self.exportieren)

        top.addWidget(refresh)
        top.addWidget(open_log)
        top.addWidget(export_btn)
        top.addStretch()
        root.addLayout(top)

        self.tabs = QTabWidget()
        root.addWidget(self.tabs, 1)

        self.rechte_table = QTableWidget()
        self.log_table = QTableWidget()
        self.tabs.addTab(self.rechte_table, "Rollen & Rechte")
        self.tabs.addTab(self.log_table, "Aktivitätsprotokoll")

        self.laden()

    def laden(self) -> None:
        rechte = berechtigungsuebersicht()

        self.rechte_table.setColumnCount(4)
        self.rechte_table.setHorizontalHeaderLabels(
            ["Mitarbeiter", "Rolle", "Bereich", "Zugriff"]
        )
        self.rechte_table.setRowCount(len(rechte))
        self.rechte_table.setAlternatingRowColors(True)

        for row_index, item in enumerate(rechte):
            for col_index, value in enumerate([
                item["name"], item["rolle"], item["seite"], item["zugriff"]
            ]):
                self.rechte_table.setItem(
                    row_index, col_index, QTableWidgetItem(str(value))
                )

        for col_index, width in enumerate([160, 140, 300, 100]):
            self.rechte_table.setColumnWidth(col_index, width)

        headers = SCHEMA.get("Ereignisprotokoll", [])
        rows = DATA.get("Ereignisprotokoll", [])

        self.log_table.setColumnCount(len(headers))
        self.log_table.setHorizontalHeaderLabels(headers)
        self.log_table.setRowCount(len(rows))
        self.log_table.setAlternatingRowColors(True)
        self.log_table.setHorizontalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAlwaysOn
        )

        for row_index, row in enumerate(rows):
            for col_index in range(len(headers)):
                value = row[col_index] if col_index < len(row) else ""
                self.log_table.setItem(
                    row_index, col_index, QTableWidgetItem(str(value))
                )

        for col_index in range(len(headers)):
            self.log_table.setColumnWidth(col_index, 160)

    def exportieren(self) -> None:

        target, _ = QFileDialog.getSaveFileName(
            self,
            "Rechteübersicht exportieren",
            str(
                EXPORT_DIR
                / (
                    "berechtigungsuebersicht_"
                    + datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                    + ".xlsx"
                )
            ),
            "Excel-Dateien (*.xlsx)",
        )
        if not target:
            return

        target_path = Path(target)
        if target_path.suffix.lower() != ".xlsx":
            target_path = target_path.with_suffix(".xlsx")

        workbook = Workbook()
        sheet = workbook.active
        if not isinstance(sheet, Worksheet):
            return

        sheet.title = "Berechtigungen"
        sheet.append(["Mitarbeiter", "Rolle", "Bereich", "Zugriff"])

        for item in berechtigungsuebersicht():
            sheet.append([
                item["name"], item["rolle"],
                item["seite"], item["zugriff"]
            ])

        for col_index, width in enumerate([22, 18, 38, 14], start=1):
            sheet.column_dimensions[
                sheet.cell(1, col_index).column_letter
            ].width = width

        workbook.save(target_path)
        QMessageBox.information(
            self,
            "Export",
            f"Rechteübersicht exportiert:\n{target_path}",
        )



def leitstand_portfolio_rows() -> list[dict[str, str]]:
    """Portfolioübersicht aus vorhandenen Daten und Beziehungen."""
    rows: list[dict[str, str]] = []

    for objekt in alle_objektordner():
        status = smart_objekt_status(objekt)
        health = objekt_gesundheitsindex(objekt)

        rows.append({
            "ampel": str(health["ampel"]),
            "objekt": objekt,
            "index": f'{health["punkte"]} %',
            "bewertung": str(health["bewertung"]),
            "wohnungen": str(status["wohnungen"]),
            "mieter": str(status["mieter"]),
            "leerstand": str(status["freie_wohnungen"]),
            "aufgaben": str(status["offene_aufgaben"]),
            "rechnungen": str(status["offene_rechnungen"]),
            "schaeden": str(status["offene_schaeden"]),
            "jahresmiete": euro(float(status["jahresmiete"])),
            "saldo": euro(float(status["saldo"])),
        })

    rows.sort(key=lambda item: (item["ampel"] != "🔴", item["ampel"] != "🟡", item["objekt"]))
    return rows


def leitstand_benachrichtigungen() -> list[dict[str, str]]:
    """Erzeugt priorisierte Benachrichtigungen aus vorhandenen Modulen."""
    items: list[dict[str, str]] = []

    for row in ueberfaellige_aufgaben_rows():
        items.append({
            "prioritaet": "🔴",
            "typ": "Überfällige Aufgabe",
            "titel": aufgaben_feld(row, "Aufgabe"),
            "objekt": objektordner_fuer_datensatz("Aufgaben", row),
            "faellig": aufgaben_feld(row, "Fällig am"),
            "bereich": "Arbeitsorganisation PRO",
        })

    for item in fristen_pro_status():
        if item.get("bewertung") in {"Überfällig", "Dringend"}:
            items.append({
                "prioritaet": "🔴" if item.get("bewertung") == "Überfällig" else "🟠",
                "typ": item.get("bewertung", "Frist"),
                "titel": item.get("titel", ""),
                "objekt": item.get("objekt", ""),
                "faellig": item.get("faellig", ""),
                "bereich": "Fristenmanager PRO",
            })

    for warning in plausibilitaetspruefungen():
        items.append({
            "prioritaet": warning.get("ampel", "🟡"),
            "typ": "Plausibilitätsprüfung",
            "titel": warning.get("hinweis", ""),
            "objekt": warning.get("objekt", ""),
            "faellig": "",
            "bereich": warning.get("bereich", "Workflow-Center PRO").split("/")[0],
        })

    rank = {"🔴": 0, "🟠": 1, "🟡": 2, "🟢": 3}
    items.sort(key=lambda item: (rank.get(item["prioritaet"], 9), datum_sort_key(item["faellig"])))
    return items


def leitstand_datenqualitaet() -> list[dict[str, str]]:
    """Prüft Datenqualität ausschließlich lesend."""
    result: list[dict[str, str]] = []

    for item in objektordner_pruefbericht():
        result.append({
            "ampel": "🔴" if "Ungültig" in item["problem"] else "🟡",
            "bereich": item["bereich"],
            "problem": item["problem"],
            "objekt": item.get("objektordner", ""),
            "zeile": item["zeile"],
            "aktion": "Objektzuordnung prüfen",
        })

    for title, rows in DATA.items():
        headers = SCHEMA.get(title, [])
        if not headers:
            continue

        for row_index, row in enumerate(rows, start=1):
            missing = []
            for index, field in enumerate(headers[:4]):
                value = row[index] if index < len(row) else ""
                if not str(value or "").strip():
                    missing.append(field)

            if len(missing) >= 2:
                result.append({
                    "ampel": "🟡",
                    "bereich": title,
                    "problem": "Mehrere wichtige Felder leer: " + ", ".join(missing),
                    "objekt": objektordner_fuer_datensatz(title, row),
                    "zeile": str(row_index),
                    "aktion": "Datensatz vervollständigen",
                })

    return result[:500]


def regelassistent_antwort(query: str, objekt: str = "") -> list[dict[str, str]]:
    """
    Regelbasierter Verwaltungsassistent.
    Keine externe KI und keine Datenübertragung.
    """
    search = str(query or "").strip().lower()
    objekt_key = norm_key(objekt)
    results: list[dict[str, str]] = []

    if not search:
        return results

    if any(word in search for word in ["miete", "mietzahlung", "nicht bezahlt", "offene mieten"]):
        for row in DATA.get("Mietkonto-Abgleich", []):
            status = feldwert("Mietkonto-Abgleich", row, ["Status"])
            relation = objektordner_fuer_datensatz("Mietkonto-Abgleich", row)
            if objekt_key and norm_key(relation) != objekt_key:
                continue
            if norm_key(status) in {"offen", "teilzahlung"}:
                results.append({
                    "bereich": "Mietkonto-Abgleich",
                    "objekt": relation,
                    "titel": feldwert("Mietkonto-Abgleich", row, ["Mieter"]),
                    "status": status,
                    "details": " | ".join(str(v) for v in row),
                })

    elif any(word in search for word in ["rechnung", "rechnungen", "offen"]):
        for row in DATA.get("Rechnungen", []):
            relation = objektordner_fuer_datensatz("Rechnungen", row)
            if objekt_key and norm_key(relation) != objekt_key:
                continue
            status = feldwert("Rechnungen", row, ["Status"])
            if norm_key(status) not in {"bezahlt", "erledigt", "abgeschlossen"}:
                results.append({
                    "bereich": "Rechnungen",
                    "objekt": relation,
                    "titel": feldwert("Rechnungen", row, ["Rechnungsnr.", "Dienstleister"]),
                    "status": status,
                    "details": " | ".join(str(v) for v in row),
                })

    elif any(word in search for word in ["schaden", "schäden", "reparatur"]):
        for row in DATA.get("Schäden", []):
            relation = objektordner_fuer_datensatz("Schäden", row)
            if objekt_key and norm_key(relation) != objekt_key:
                continue
            status = feldwert("Schäden", row, ["Status"])
            if norm_key(status) not in {"erledigt", "abgeschlossen", "behoben"}:
                results.append({
                    "bereich": "Schäden",
                    "objekt": relation,
                    "titel": feldwert("Schäden", row, ["Schaden"]),
                    "status": status,
                    "details": " | ".join(str(v) for v in row),
                })

    elif any(word in search for word in ["frist", "termin", "fällig"]):
        for item in fristen_pro_status():
            if objekt_key and norm_key(item.get("objekt", "")) != objekt_key:
                continue
            if item.get("bewertung") in {"Überfällig", "Dringend", "Bald fällig"}:
                results.append({
                    "bereich": "Fristenmanager PRO",
                    "objekt": item.get("objekt", ""),
                    "titel": item.get("titel", ""),
                    "status": item.get("bewertung", ""),
                    "details": f'{item.get("faellig", "")} | {item.get("bereich", "")}',
                })

    elif any(word in search for word in ["aufgabe", "aufgaben", "todo"]):
        for row in offene_aufgaben_rows():
            relation = objektordner_fuer_datensatz("Aufgaben", row)
            if objekt_key and norm_key(relation) != objekt_key:
                continue
            results.append({
                "bereich": "Arbeitsorganisation PRO",
                "objekt": relation,
                "titel": aufgaben_feld(row, "Aufgabe"),
                "status": aufgaben_feld(row, "Status"),
                "details": " | ".join(str(v) for v in row),
            })

    else:
        for item in globale_dashboard_suche(search, objekt):
            results.append({
                "bereich": item.get("bereich", ""),
                "objekt": item.get("objektordner", ""),
                "titel": item.get("felder", ""),
                "status": "Treffer",
                "details": item.get("inhalt", ""),
            })

    return results[:300]


def systemcheck_ergebnis() -> list[dict[str, str]]:
    """Prüft Verzeichnisse, Daten und Bibliotheksfunktionen."""
    checks: list[dict[str, str]] = []

    folders = [
        ("Datenordner", DATEN_DIR),
        ("Dokumente", DOKUMENTE_DIR),
        ("Exporte", EXPORT_DIR),
        ("Akten", AKTEN_DIR),
        ("Backups", BACKUP_DIR),
    ]

    for name, folder_path in folders:
        exists = folder_path.exists()
        writable = os.access(folder_path, os.W_OK) if exists else False
        checks.append({
            "status": "🟢" if exists and writable else "🔴",
            "bereich": name,
            "ergebnis": str(folder_path),
            "hinweis": "Vorhanden und beschreibbar" if exists and writable else "Pfad oder Schreibrechte prüfen",
        })

    missing_files = [
        title for title in SCHEMA
        if not xlsx_pfad(title).exists()
    ]
    checks.append({
        "status": "🟢" if not missing_files else "🟡",
        "bereich": "Excel-Dateien",
        "ergebnis": f"{len(SCHEMA) - len(missing_files)} von {len(SCHEMA)} vorhanden",
        "hinweis": "Vollständig" if not missing_files else "Fehlend: " + ", ".join(missing_files[:10]),
    })

    checks.append({
        "status": "🟢" if PdfReader is not None else "🟡",
        "bereich": "PDF-Erkennung",
        "ergebnis": "pypdf verfügbar" if PdfReader is not None else "pypdf fehlt",
        "hinweis": "Text-PDF-Import möglich" if PdfReader is not None else "Abhängigkeit installieren",
    })

    checks.append({
        "status": "🟢",
        "bereich": "Objektordner",
        "ergebnis": f"{len(alle_objektordner())} gültige Objektordner",
        "hinweis": f"{len(objektordner_pruefbericht())} Datensätze zu prüfen",
    })

    return checks


class VerwaltungsleitstandSeite(QWidget):
    """Zentraler Leitstand für Portfolio, Warnungen und Assistent."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.notifications: list[dict[str, str]] = []
        self.assistant_rows: list[dict[str, str]] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(20, 18, 20, 18)
        root.setSpacing(14)

        title = QLabel("Verwaltungsleitstand 5.0")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel(
            "Portfolio, Benachrichtigungen, Datenqualität und regelbasierter Verwaltungsassistent in einer Zentrale."
        )
        info.setObjectName("subTitle")
        info.setWordWrap(True)
        root.addWidget(info)

        toolbar = QHBoxLayout()
        refresh = QPushButton("Alles aktualisieren")
        refresh.setObjectName("primaryButton")
        refresh.clicked.connect(self.laden)

        report = QPushButton("Berichte & Export")
        report.clicked.connect(lambda: self.nav("Berichte & Export PRO"))

        system = QPushButton("System-Center")
        system.clicked.connect(lambda: self.nav("System-Center PRO"))

        toolbar.addWidget(refresh)
        toolbar.addWidget(report)
        toolbar.addWidget(system)
        toolbar.addStretch()
        root.addLayout(toolbar)

        self.cards = QGridLayout()
        root.addLayout(self.cards)

        self.tabs = QTabWidget()
        root.addWidget(self.tabs, 1)

        self.portfolio_table = QTableWidget()
        self.notification_table = QTableWidget()
        self.quality_table = QTableWidget()

        assistant_page = QWidget()
        assistant_layout = QVBoxLayout(assistant_page)

        assistant_top = QHBoxLayout()
        self.object_filter = QComboBox()
        self.object_filter.addItem("Alle Objektordner")
        self.object_filter.addItems(alle_objektordner())

        self.query = QLineEdit()
        self.query.setPlaceholderText(
            "Beispiel: offene Mieten, offene Rechnungen, Schäden, Fristen oder beliebiges Schlagwort"
        )
        self.query.returnPressed.connect(self.assistant_search)

        search_btn = QPushButton("Abfrage starten")
        search_btn.setObjectName("primaryButton")
        search_btn.clicked.connect(self.assistant_search)

        assistant_top.addWidget(QLabel("Objektordner:"))
        assistant_top.addWidget(self.object_filter)
        assistant_top.addWidget(self.query, 1)
        assistant_top.addWidget(search_btn)
        assistant_layout.addLayout(assistant_top)

        self.assistant_table = QTableWidget()
        self.assistant_table.setAlternatingRowColors(True)
        self.assistant_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.assistant_table.cellDoubleClicked.connect(self.open_assistant_result)
        assistant_layout.addWidget(self.assistant_table, 1)

        self.tabs.addTab(self.portfolio_table, "Portfolio")
        self.tabs.addTab(self.notification_table, "Benachrichtigungen")
        self.tabs.addTab(self.quality_table, "Datenqualität")
        self.tabs.addTab(assistant_page, "Verwaltungsassistent")

        self.portfolio_table.cellDoubleClicked.connect(self.open_portfolio)
        self.notification_table.cellDoubleClicked.connect(self.open_notification)
        self.quality_table.cellDoubleClicked.connect(self.open_quality)

        self.laden()

    @staticmethod
    def _card(title: str, value: str, icon: str) -> QFrame:
        card = QFrame()
        card.setObjectName("metricCard")
        card.setMinimumHeight(100)
        layout = QVBoxLayout(card)
        label = QLabel(f"{icon}  {title}")
        label.setObjectName("metricTitle")
        value_label = QLabel(value)
        value_label.setObjectName("metricValue")
        value_label.setWordWrap(True)
        layout.addWidget(label)
        layout.addWidget(value_label)
        return card

    def laden(self) -> None:
        portfolio = leitstand_portfolio_rows()
        self.notifications = leitstand_benachrichtigungen()
        quality = leitstand_datenqualitaet()

        while self.cards.count():
            item = self.cards.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        critical_objects = sum(1 for item in portfolio if item["ampel"] == "🔴")
        open_tasks = len(offene_aufgaben_rows())
        urgent = sum(1 for item in self.notifications if item["prioritaet"] in {"🔴", "🟠"})
        quality_count = len(quality)

        cards = [
            ("Objekte", str(len(portfolio)), "🏢"),
            ("Kritische Objekte", str(critical_objects), "🔴"),
            ("Offene Aufgaben", str(open_tasks), "🗂"),
            ("Dringende Hinweise", str(urgent), "⏰"),
            ("Datenprüfungen", str(quality_count), "🔎"),
        ]
        for index, values in enumerate(cards):
            self.cards.addWidget(self._card(*values), 0, index)

        self._fill_portfolio(portfolio)
        self._fill_notifications(self.notifications)
        self._fill_quality(quality)

    def _fill_portfolio(self, rows: list[dict[str, str]]) -> None:
        headers = [
            "Ampel", "Objektordner", "Index", "Bewertung",
            "Wohnungen", "Mieter", "Leerstand", "Aufgaben",
            "Rechnungen", "Schäden", "Jahresmiete", "Saldo",
        ]
        self.portfolio_table.setColumnCount(len(headers))
        self.portfolio_table.setHorizontalHeaderLabels(headers)
        self.portfolio_table.setRowCount(len(rows))
        self.portfolio_table.setAlternatingRowColors(True)
        self.portfolio_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)

        keys = [
            "ampel", "objekt", "index", "bewertung",
            "wohnungen", "mieter", "leerstand", "aufgaben",
            "rechnungen", "schaeden", "jahresmiete", "saldo",
        ]

        for row_index, item in enumerate(rows):
            for col_index, key in enumerate(keys):
                self.portfolio_table.setItem(
                    row_index, col_index, QTableWidgetItem(item[key])
                )

        for col_index, width in enumerate([70, 250, 90, 140, 100, 90, 100, 100, 110, 90, 150, 150]):
            self.portfolio_table.setColumnWidth(col_index, width)

    def _fill_notifications(self, rows: list[dict[str, str]]) -> None:
        headers = ["Priorität", "Typ", "Titel", "Objektordner", "Fällig", "Bereich"]
        self.notification_table.setColumnCount(len(headers))
        self.notification_table.setHorizontalHeaderLabels(headers)
        self.notification_table.setRowCount(len(rows))
        self.notification_table.setAlternatingRowColors(True)

        keys = ["prioritaet", "typ", "titel", "objekt", "faellig", "bereich"]
        for row_index, item in enumerate(rows):
            for col_index, key in enumerate(keys):
                self.notification_table.setItem(
                    row_index, col_index, QTableWidgetItem(item[key])
                )

        for col_index, width in enumerate([80, 180, 500, 240, 110, 220]):
            self.notification_table.setColumnWidth(col_index, width)

    def _fill_quality(self, rows: list[dict[str, str]]) -> None:
        headers = ["Ampel", "Bereich", "Problem", "Objektordner", "Zeile", "Aktion"]
        self.quality_table.setColumnCount(len(headers))
        self.quality_table.setHorizontalHeaderLabels(headers)
        self.quality_table.setRowCount(len(rows))
        self.quality_table.setAlternatingRowColors(True)
        self.quality_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)

        keys = ["ampel", "bereich", "problem", "objekt", "zeile", "aktion"]
        for row_index, item in enumerate(rows):
            for col_index, key in enumerate(keys):
                self.quality_table.setItem(
                    row_index, col_index, QTableWidgetItem(item[key])
                )

        for col_index, width in enumerate([70, 190, 500, 240, 70, 260]):
            self.quality_table.setColumnWidth(col_index, width)

    def assistant_search(self) -> None:
        selected = self.object_filter.currentText().strip()
        objekt = "" if selected == "Alle Objektordner" else selected
        self.assistant_rows = regelassistent_antwort(self.query.text(), objekt)

        headers = ["Bereich", "Objektordner", "Titel", "Status", "Details"]
        self.assistant_table.setColumnCount(len(headers))
        self.assistant_table.setHorizontalHeaderLabels(headers)
        self.assistant_table.setRowCount(len(self.assistant_rows))

        keys = ["bereich", "objekt", "titel", "status", "details"]
        for row_index, item in enumerate(self.assistant_rows):
            for col_index, key in enumerate(keys):
                self.assistant_table.setItem(
                    row_index, col_index, QTableWidgetItem(item.get(key, ""))
                )

        for col_index, width in enumerate([210, 240, 300, 130, 720]):
            self.assistant_table.setColumnWidth(col_index, width)

    def open_portfolio(self, row: int, _column: int) -> None:
        item = self.portfolio_table.item(row, 1)
        if item is not None:
            self.nav("Smart Objektakte")

    def open_notification(self, row: int, _column: int) -> None:
        if 0 <= row < len(self.notifications):
            self.nav(self.notifications[row].get("bereich", "Workflow-Center PRO"))

    def open_quality(self, row: int, _column: int) -> None:
        item = self.quality_table.item(row, 1)
        if item is not None:
            self.nav(item.text().strip())

    def open_assistant_result(self, row: int, _column: int) -> None:
        if 0 <= row < len(self.assistant_rows):
            self.nav(self.assistant_rows[row].get("bereich", "Dashboard"))


class BerichteExportProSeite(QWidget):
    """Portfolio- und Monatsberichte ohne Änderung der Datenstruktur."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav

        root = QVBoxLayout(self)
        root.setContentsMargins(22, 20, 22, 20)
        root.setSpacing(14)

        title = QLabel("Berichte & Export PRO")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel(
            "Portfolio-, Aufgaben-, Fristen- und Datenqualitätsberichte aus den vorhandenen Daten erstellen."
        )
        info.setObjectName("subTitle")
        info.setWordWrap(True)
        root.addWidget(info)

        buttons = QGridLayout()
        actions = [
            ("Portfolio-Bericht", self.export_portfolio),
            ("Benachrichtigungsbericht", self.export_notifications),
            ("Datenqualitätsbericht", self.export_quality),
            ("Gesamtbericht", self.export_complete),
        ]
        for index, (label, callback) in enumerate(actions):
            button = QPushButton(label)
            if index == 3:
                button.setObjectName("primaryButton")
            button.clicked.connect(callback)
            buttons.addWidget(button, index // 2, index % 2)
        root.addLayout(buttons)

        self.preview = QTableWidget()
        self.preview.setAlternatingRowColors(True)
        self.preview.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        root.addWidget(self.preview, 1)

        self.show_preview()

    def show_preview(self) -> None:
        rows = leitstand_portfolio_rows()
        headers = ["Ampel", "Objektordner", "Index", "Bewertung", "Jahresmiete", "Saldo"]
        keys = ["ampel", "objekt", "index", "bewertung", "jahresmiete", "saldo"]

        self.preview.setColumnCount(len(headers))
        self.preview.setHorizontalHeaderLabels(headers)
        self.preview.setRowCount(len(rows))

        for row_index, item in enumerate(rows):
            for col_index, key in enumerate(keys):
                self.preview.setItem(
                    row_index, col_index, QTableWidgetItem(item[key])
                )

        for col_index, width in enumerate([70, 280, 100, 160, 170, 170]):
            self.preview.setColumnWidth(col_index, width)

    @staticmethod
    def _save_workbook(workbook: Workbook, suggested_name: str) -> None:
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        target, _ = QFileDialog.getSaveFileName(
            None,
            "Bericht exportieren",
            str(EXPORT_DIR / suggested_name),
            "Excel-Dateien (*.xlsx)",
        )
        if not target:
            return

        target_path = Path(target)
        if target_path.suffix.lower() != ".xlsx":
            target_path = target_path.with_suffix(".xlsx")

        workbook.save(target_path)
        QMessageBox.information(None, "Export", f"Bericht exportiert:\n{target_path}")

    def export_portfolio(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        if not isinstance(sheet, Worksheet):
            return
        sheet.title = "Portfolio"
        headers = [
            "Ampel", "Objektordner", "Index", "Bewertung",
            "Wohnungen", "Mieter", "Leerstand", "Aufgaben",
            "Rechnungen", "Schäden", "Jahresmiete", "Saldo",
        ]
        sheet.append(headers)
        keys = [
            "ampel", "objekt", "index", "bewertung",
            "wohnungen", "mieter", "leerstand", "aufgaben",
            "rechnungen", "schaeden", "jahresmiete", "saldo",
        ]
        for item in leitstand_portfolio_rows():
            sheet.append([item[key] for key in keys])
        self._save_workbook(workbook, "portfolio_bericht.xlsx")

    def export_notifications(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        if not isinstance(sheet, Worksheet):
            return
        sheet.title = "Benachrichtigungen"
        sheet.append(["Priorität", "Typ", "Titel", "Objektordner", "Fällig", "Bereich"])
        for item in leitstand_benachrichtigungen():
            sheet.append([
                item["prioritaet"], item["typ"], item["titel"],
                item["objekt"], item["faellig"], item["bereich"],
            ])
        self._save_workbook(workbook, "benachrichtigungen.xlsx")

    def export_quality(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        if not isinstance(sheet, Worksheet):
            return
        sheet.title = "Datenqualität"
        sheet.append(["Ampel", "Bereich", "Problem", "Objektordner", "Zeile", "Aktion"])
        for item in leitstand_datenqualitaet():
            sheet.append([
                item["ampel"], item["bereich"], item["problem"],
                item["objekt"], item["zeile"], item["aktion"],
            ])
        self._save_workbook(workbook, "datenqualitaet.xlsx")

    def export_complete(self) -> None:
        workbook = Workbook()

        portfolio_sheet = workbook.active
        if not isinstance(portfolio_sheet, Worksheet):
            return
        portfolio_sheet.title = "Portfolio"
        portfolio_sheet.append([
            "Ampel", "Objektordner", "Index", "Bewertung",
            "Wohnungen", "Mieter", "Leerstand", "Aufgaben",
            "Rechnungen", "Schäden", "Jahresmiete", "Saldo",
        ])
        for item in leitstand_portfolio_rows():
            portfolio_sheet.append([
                item["ampel"], item["objekt"], item["index"], item["bewertung"],
                item["wohnungen"], item["mieter"], item["leerstand"], item["aufgaben"],
                item["rechnungen"], item["schaeden"], item["jahresmiete"], item["saldo"],
            ])

        notification_sheet = workbook.create_sheet("Benachrichtigungen")
        notification_sheet.append(["Priorität", "Typ", "Titel", "Objektordner", "Fällig", "Bereich"])
        for item in leitstand_benachrichtigungen():
            notification_sheet.append([
                item["prioritaet"], item["typ"], item["titel"],
                item["objekt"], item["faellig"], item["bereich"],
            ])

        quality_sheet = workbook.create_sheet("Datenqualität")
        quality_sheet.append(["Ampel", "Bereich", "Problem", "Objektordner", "Zeile", "Aktion"])
        for item in leitstand_datenqualitaet():
            quality_sheet.append([
                item["ampel"], item["bereich"], item["problem"],
                item["objekt"], item["zeile"], item["aktion"],
            ])

        self._save_workbook(workbook, "immoverwaltung_gesamtbericht.xlsx")


class SystemCenterProSeite(QWidget):
    """Systemprüfung, Backup und Wartungsübersicht."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav

        root = QVBoxLayout(self)
        root.setContentsMargins(22, 20, 22, 20)
        root.setSpacing(14)

        title = QLabel("System-Center PRO")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel(
            "Systemzustand, Verzeichnisse, Excel-Dateien, PDF-Erkennung, Objektordner und Backup."
        )
        info.setObjectName("subTitle")
        info.setWordWrap(True)
        root.addWidget(info)

        top = QHBoxLayout()
        refresh = QPushButton("System prüfen")
        refresh.setObjectName("primaryButton")
        refresh.clicked.connect(self.laden)

        backup = QPushButton("Backup erstellen")
        backup.clicked.connect(self.backup)

        settings = QPushButton("Einstellungen öffnen")
        settings.clicked.connect(lambda: self.nav("Einstellungen"))

        top.addWidget(refresh)
        top.addWidget(backup)
        top.addWidget(settings)
        top.addStretch()
        root.addLayout(top)

        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Status", "Bereich", "Ergebnis", "Hinweis"])
        self.table.setAlternatingRowColors(True)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        root.addWidget(self.table, 1)

        self.laden()

    def laden(self) -> None:
        rows = systemcheck_ergebnis()
        self.table.setRowCount(len(rows))

        for row_index, item in enumerate(rows):
            values = [
                item["status"], item["bereich"],
                item["ergebnis"], item["hinweis"],
            ]
            for col_index, value in enumerate(values):
                self.table.setItem(
                    row_index, col_index, QTableWidgetItem(str(value))
                )

        for col_index, width in enumerate([80, 220, 550, 450]):
            self.table.setColumnWidth(col_index, width)

    def backup(self) -> None:
        target = erstelle_schnellbackup()
        if target is not None:
            aktivitaet_protokollieren(
                "System-Center",
                "Backup erstellt",
                beschreibung=str(target),
            )
            QMessageBox.information(self, "Backup", f"Backup erstellt:\n{target}")



USER_UI_FILE = DATEN_DIR / "user_ui_state.json"


def user_ui_state_laden() -> dict[str, Any]:
    try:
        if USER_UI_FILE.exists():
            data = json.loads(USER_UI_FILE.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return data
    except (OSError, ValueError, json.JSONDecodeError):
        pass
    return {}


def user_ui_state_speichern(data: dict[str, Any]) -> None:
    try:
        DATEN_DIR.mkdir(parents=True, exist_ok=True)
        USER_UI_FILE.write_text(
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except (OSError, PermissionError, TypeError):
        pass


def benutzer_ui_key() -> str:
    return norm_key(
        aktueller_mitarbeitername()
        or str(CURRENT_USER.get("benutzername", "benutzer"))
    ) or "benutzer"


def benutzer_favoriten() -> list[str]:
    state = user_ui_state_laden()
    values = state.get(benutzer_ui_key(), {}).get("favoriten", [])
    if not isinstance(values, list):
        return []
    return [str(value) for value in values if str(value).strip()]


def benutzer_zuletzt() -> list[str]:
    state = user_ui_state_laden()
    values = state.get(benutzer_ui_key(), {}).get("zuletzt", [])
    if not isinstance(values, list):
        return []
    return [str(value) for value in values if str(value).strip()]


def benutzer_favorit_umschalten(seite: str) -> bool:
    state = user_ui_state_laden()
    key = benutzer_ui_key()
    section = state.setdefault(key, {})
    favorites = section.setdefault("favoriten", [])

    if seite in favorites:
        favorites.remove(seite)
        active = False
    else:
        favorites.insert(0, seite)
        favorites[:] = favorites[:12]
        active = True

    user_ui_state_speichern(state)
    return active


def benutzer_zuletzt_hinzufuegen(seite: str) -> None:
    if not seite:
        return

    state = user_ui_state_laden()
    key = benutzer_ui_key()
    section = state.setdefault(key, {})
    recent = section.setdefault("zuletzt", [])

    if seite in recent:
        recent.remove(seite)

    recent.insert(0, seite)
    recent[:] = recent[:10]
    user_ui_state_speichern(state)


def objektkarten_daten() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for objekt in alle_objektordner():
        status = smart_objekt_status(objekt)
        health = objekt_gesundheitsindex(objekt)

        rows.append({
            "objekt": objekt,
            "ampel": str(health["ampel"]),
            "index": int(health["punkte"]),
            "bewertung": str(health["bewertung"]),
            "wohnungen": int(status["wohnungen"]),
            "mieter": int(status["mieter"]),
            "leerstand": int(status["freie_wohnungen"]),
            "aufgaben": int(status["offene_aufgaben"]),
            "saldo": euro(float(status["saldo"])),
        })

    rows.sort(key=lambda item: (item["index"], item["objekt"]))
    return rows


class ObjektGalerieSeite(QWidget):
    """Moderne Kartenansicht aller Objektordner."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav

        root = QVBoxLayout(self)
        root.setContentsMargins(20, 18, 20, 18)
        root.setSpacing(14)

        title = QLabel("Objektgalerie")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel(
            "Moderne Kartenansicht mit Gesundheitsindex, Vermietung, Aufgaben und Saldo."
        )
        info.setObjectName("subTitle")
        info.setWordWrap(True)
        root.addWidget(info)

        top = QHBoxLayout()

        self.search = QLineEdit()
        self.search.setPlaceholderText("Objektordner suchen ...")
        self.search.textChanged.connect(self.laden)

        self.filter = QComboBox()
        self.filter.addItems(["Alle Objekte", "Kritisch", "Beobachten", "Gesund"])
        self.filter.currentTextChanged.connect(self.laden)

        refresh = QPushButton("Aktualisieren")
        refresh.setObjectName("primaryButton")
        refresh.clicked.connect(self.laden)

        top.addWidget(self.search, 1)
        top.addWidget(self.filter)
        top.addWidget(refresh)
        root.addLayout(top)

        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setFrameShape(QFrame.Shape.NoFrame)
        root.addWidget(self.scroll, 1)

        self.content = QWidget()
        self.grid = QGridLayout(self.content)
        self.grid.setContentsMargins(4, 4, 14, 14)
        self.grid.setSpacing(14)
        self.scroll.setWidget(self.content)

        self.laden()

    def laden(self) -> None:
        while self.grid.count():
            item = self.grid.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        query = self.search.text().strip().lower()
        selected_filter = self.filter.currentText()

        rows = objektkarten_daten()

        if query:
            rows = [
                item for item in rows
                if query in item["objekt"].lower()
            ]

        if selected_filter != "Alle Objekte":
            rows = [
                item for item in rows
                if item["bewertung"] == selected_filter
            ]

        for index, item in enumerate(rows):
            self.grid.addWidget(
                self._card(item),
                index // 3,
                index % 3,
            )

        self.grid.setRowStretch((len(rows) // 3) + 1, 1)

    def _card(self, item: dict[str, Any]) -> QFrame:
        card = QFrame()
        card.setObjectName("metricCard")
        card.setMinimumWidth(280)
        card.setMinimumHeight(240)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(8)

        top = QHBoxLayout()
        icon = QLabel(str(item["ampel"]))
        icon.setStyleSheet("font-size:28px;")

        name = QLabel(str(item["objekt"]))
        name.setObjectName("metricValue")
        name.setWordWrap(True)

        top.addWidget(icon)
        top.addWidget(name, 1)
        layout.addLayout(top)

        status = QLabel(
            f'{item["index"]} % · {item["bewertung"]}'
        )
        status.setObjectName("metricTitle")
        layout.addWidget(status)

        progress = QProgressBar()
        progress.setRange(0, 100)
        progress.setValue(int(item["index"]))
        progress.setTextVisible(True)
        progress.setFormat("%p %")
        progress.setMinimumHeight(22)
        layout.addWidget(progress)

        details = QGridLayout()
        values = [
            ("Wohnungen", item["wohnungen"]),
            ("Mieter", item["mieter"]),
            ("Leerstand", item["leerstand"]),
            ("Aufgaben", item["aufgaben"]),
            ("Saldo", item["saldo"]),
        ]

        for index, (label, value) in enumerate(values):
            label_widget = QLabel(label)
            label_widget.setObjectName("metricTitle")

            value_widget = QLabel(str(value))
            value_widget.setStyleSheet("font-weight:900; font-size:16px;")

            details.addWidget(label_widget, index, 0)
            details.addWidget(value_widget, index, 1)

        layout.addLayout(details)
        layout.addStretch()

        actions = QHBoxLayout()

        open_button = QPushButton("Objekt öffnen")
        open_button.setObjectName("primaryButton")
        open_button.clicked.connect(
            lambda checked=False, objekt=str(item["objekt"]): self.open_objekt(objekt)
        )

        chronik_button = QPushButton("Chronik")
        chronik_button.clicked.connect(
            lambda checked=False, objekt=str(item["objekt"]): self.open_chronik(objekt)
        )

        actions.addWidget(open_button)
        actions.addWidget(chronik_button)
        layout.addLayout(actions)

        return card

    def open_objekt(self, objekt: str) -> None:
        self.nav("Smart Objektakte")

    def open_chronik(self, objekt: str) -> None:
        self.nav("Objektchronik")


class SmartStartseiteSeite(QWidget):
    """Personalisierte moderne Startseite mit Favoriten und Schnellaktionen."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav

        root = QVBoxLayout(self)
        root.setContentsMargins(20, 18, 20, 18)
        root.setSpacing(16)

        greeting = QLabel(
            f"Guten Tag, {aktueller_mitarbeitername() or 'Mitarbeiter'}"
        )
        greeting.setObjectName("pageTitle")
        root.addWidget(greeting)

        subtitle = QLabel(
            "Ihre wichtigsten Bereiche, letzten Aktivitäten und Schnellaktionen auf einen Blick."
        )
        subtitle.setObjectName("subTitle")
        subtitle.setWordWrap(True)
        root.addWidget(subtitle)

        self.cards = QGridLayout()
        self.cards.setSpacing(12)
        root.addLayout(self.cards)

        quick_group = QGroupBox("Schnellaktionen")
        quick_layout = QGridLayout(quick_group)

        actions = [
            ("Neue Aufgabe", "Arbeitsorganisation PRO", "🗂"),
            ("Neue Frist", "Fristenmanager PRO", "⏰"),
            ("Neue Rechnung", "Rechnungen", "🧾"),
            ("Neuer Mieter", "Mieter", "👤"),
            ("Neues Objekt", "Objekte", "🏢"),
            ("Kontoauszug", "Kontoauszug-Import", "🏦"),
            ("Dokument scannen", "Belegscanner PRO", "📥"),
            ("Kalender", "Kalender & Planung", "📅"),
        ]

        for index, (label, target, icon) in enumerate(actions):
            button = QPushButton(f"{icon}\n{label}")
            button.setMinimumHeight(72)
            button.clicked.connect(
                lambda checked=False, page=target: self.nav(page)
            )
            quick_layout.addWidget(button, index // 4, index % 4)

        root.addWidget(quick_group)

        split = QSplitter(Qt.Orientation.Horizontal)
        root.addWidget(split, 1)

        favorites_group = QGroupBox("Favoriten")
        favorites_layout = QVBoxLayout(favorites_group)
        self.favorite_list = QListWidget()
        self.favorite_list.itemDoubleClicked.connect(
            lambda item: self.nav(item.text())
        )
        favorites_layout.addWidget(self.favorite_list)

        recent_group = QGroupBox("Zuletzt verwendet")
        recent_layout = QVBoxLayout(recent_group)
        self.recent_list = QListWidget()
        self.recent_list.itemDoubleClicked.connect(
            lambda item: self.nav(item.text())
        )
        recent_layout.addWidget(self.recent_list)

        split.addWidget(favorites_group)
        split.addWidget(recent_group)
        split.setSizes([520, 520])

        bottom = QHBoxLayout()

        add_favorite = QPushButton("Aktuellen Bereich als Favorit")
        add_favorite.clicked.connect(self.add_current_favorite)

        object_gallery = QPushButton("Objektgalerie öffnen")
        object_gallery.setObjectName("primaryButton")
        object_gallery.clicked.connect(lambda: self.nav("Objektgalerie"))

        bottom.addWidget(add_favorite)
        bottom.addWidget(object_gallery)
        bottom.addStretch()
        root.addLayout(bottom)

        self.laden()

    @staticmethod
    def _card(title: str, value: str, icon: str) -> QFrame:
        card = QFrame()
        card.setObjectName("metricCard")
        layout = QVBoxLayout(card)

        label = QLabel(f"{icon}  {title}")
        label.setObjectName("metricTitle")

        value_label = QLabel(value)
        value_label.setObjectName("metricValue")
        value_label.setWordWrap(True)

        layout.addWidget(label)
        layout.addWidget(value_label)
        return card

    def laden(self) -> None:
        while self.cards.count():
            item = self.cards.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        notifications = leitstand_benachrichtigungen()
        portfolio = leitstand_portfolio_rows()

        values = [
            ("Offene Aufgaben", str(len(offene_aufgaben_rows())), "🗂"),
            (
                "Dringende Hinweise",
                str(sum(1 for item in notifications if item["prioritaet"] in {"🔴", "🟠"})),
                "🔔",
            ),
            (
                "Kritische Objekte",
                str(sum(1 for item in portfolio if item["ampel"] == "🔴")),
                "🏢",
            ),
            ("Heutige Termine", str(len(heutige_mitarbeiter_fristen(aktueller_mitarbeitername()))), "📅"),
        ]

        for index, card in enumerate(values):
            self.cards.addWidget(self._card(*card), 0, index)

        self.favorite_list.clear()
        self.favorite_list.addItems(benutzer_favoriten())

        self.recent_list.clear()
        self.recent_list.addItems(benutzer_zuletzt())

    def add_current_favorite(self) -> None:
        parent = self.window()
        current_title = ""

        if hasattr(parent, "header"):
            current_title = str(parent.header.text()).strip()

        if not current_title:
            return

        active = benutzer_favorit_umschalten(current_title)
        self.laden()

        QMessageBox.information(
            self,
            "Favoriten",
            (
                f"'{current_title}' wurde zu den Favoriten hinzugefügt."
                if active
                else f"'{current_title}' wurde aus den Favoriten entfernt."
            ),
        )



def enterprise_dokument_index(objektordner: str = "") -> list[dict[str, str]]:
    """Indiziert vorhandene Dokumentpfade aus allen Tabellen, ohne Excel zu verändern."""
    objekt_key = norm_key(objektordner)
    result: list[dict[str, str]] = []

    for titel, rows in DATA.items():
        headers = SCHEMA.get(titel, [])
        document_columns = [
            index for index, field in enumerate(headers)
            if any(token in norm_key(field) for token in [
                "pdf", "dateipfad", "ordnerpfad", "foto", "anlage", "dokument"
            ])
        ]

        if not document_columns:
            continue

        for row_index, row in enumerate(rows, start=1):
            relation = objektordner_fuer_datensatz(titel, row)
            if objekt_key and norm_key(relation) != objekt_key:
                continue

            for col_index in document_columns:
                value = str(row[col_index] if col_index < len(row) else "").strip()
                if not value:
                    continue

                path = Path(value)
                if not path.is_absolute():
                    candidates = [
                        APP_DIR / path,
                        DOKUMENTE_DIR / path,
                        AKTEN_DIR / path,
                    ]
                    resolved = next((candidate for candidate in candidates if candidate.exists()), APP_DIR / path)
                else:
                    resolved = path

                result.append({
                    "bereich": titel,
                    "zeile": str(row_index),
                    "objekt": relation,
                    "feld": headers[col_index] if col_index < len(headers) else f"Spalte {col_index + 1}",
                    "titel": path.name or value,
                    "pfad": str(resolved),
                    "status": "Vorhanden" if resolved.exists() else "Pfad prüfen",
                    "typ": resolved.suffix.lower().lstrip(".").upper() or "DATEI",
                })

    result.sort(key=lambda item: (
        item["status"] != "Pfad prüfen",
        item["objekt"],
        item["bereich"],
        item["titel"],
    ))
    return result


def enterprise_workflow_vorschlaege(objektordner: str = "") -> list[dict[str, str]]:
    """Regelbasierte Workflow-Vorschläge aus bestehenden Daten."""
    object_key = norm_key(objektordner)
    result: list[dict[str, str]] = []

    def match(title: str, row: list[Any]) -> bool:
        return not object_key or norm_key(objektordner_fuer_datensatz(title, row)) == object_key

    for row in DATA.get("Rechnungen", []):
        if not match("Rechnungen", row):
            continue
        status = feldwert("Rechnungen", row, ["Status"])
        if norm_key(status) not in {"bezahlt", "erledigt", "abgeschlossen"}:
            result.append({
                "prioritaet": "Hoch",
                "regel": "Offene Rechnung",
                "objekt": objektordner_fuer_datensatz("Rechnungen", row),
                "titel": f'Rechnung {feldwert("Rechnungen", row, ["Rechnungsnr."])} prüfen',
                "bereich": "Rechnungen",
                "aktion": "Aufgabe zur Rechnungsprüfung anlegen",
                "faellig": feldwert("Rechnungen", row, ["Datum"]),
            })

    for row in DATA.get("Schäden", []):
        if not match("Schäden", row):
            continue
        status = feldwert("Schäden", row, ["Status"])
        if norm_key(status) not in {"erledigt", "abgeschlossen", "behoben"}:
            result.append({
                "prioritaet": feldwert("Schäden", row, ["Priorität"]) or "Hoch",
                "regel": "Offener Schaden",
                "objekt": objektordner_fuer_datensatz("Schäden", row),
                "titel": feldwert("Schäden", row, ["Schaden"]) or "Schaden bearbeiten",
                "bereich": "Schäden",
                "aktion": "Technische Wiedervorlage anlegen",
                "faellig": feldwert("Schäden", row, ["Datum"]),
            })

    for item in fristen_pro_status():
        if object_key and norm_key(item.get("objekt", "")) != object_key:
            continue
        if item.get("bewertung") in {"Überfällig", "Dringend", "Bald fällig"}:
            result.append({
                "prioritaet": "Kritisch" if item.get("bewertung") == "Überfällig" else "Hoch",
                "regel": "Fristüberwachung",
                "objekt": item.get("objekt", ""),
                "titel": item.get("titel", ""),
                "bereich": "Fristenmanager PRO",
                "aktion": f'{item.get("bewertung", "")}: Verantwortlichen informieren',
                "faellig": item.get("faellig", ""),
            })

    for warning in plausibilitaetspruefungen(objektordner):
        result.append({
            "prioritaet": "Kritisch" if warning.get("ampel") == "🔴" else "Normal",
            "regel": "Plausibilitätsprüfung",
            "objekt": warning.get("objekt", ""),
            "titel": warning.get("hinweis", ""),
            "bereich": warning.get("bereich", "").split("/")[0],
            "aktion": warning.get("aktion", "Datensatz prüfen"),
            "faellig": "",
        })

    priority_rank = {"Kritisch": 0, "Hoch": 1, "Normal": 2, "Niedrig": 3}
    result.sort(key=lambda item: (
        priority_rank.get(item["prioritaet"], 9),
        datum_sort_key(item["faellig"]),
        item["objekt"],
    ))
    return result





class EnterpriseObjekt360Seite(QWidget):
    """Enterprise-Objektansicht mit robustem, vollständig scrollbarerm Layout."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.current_objekt = ""

        outer = QVBoxLayout(self)
        outer.setContentsMargins(18, 16, 18, 16)
        outer.setSpacing(12)

        title = QLabel("Enterprise Objekt 360°")
        title.setObjectName("pageTitle")
        outer.addWidget(title)

        subtitle = QLabel(
            "Zentrale 360°-Ansicht für Stammdaten, Mieter, Finanzen, Aufgaben, Dokumente und Chronik."
        )
        subtitle.setObjectName("subTitle")
        subtitle.setWordWrap(True)
        outer.addWidget(subtitle)

        toolbar = QHBoxLayout()
        toolbar.setSpacing(10)

        self.objekt_filter = QComboBox()
        self.objekt_filter.setEditable(True)
        self.objekt_filter.setMinimumWidth(320)
        self.objekt_filter.addItems(alle_objektordner())
        self.objekt_filter.currentTextChanged.connect(self.laden)

        refresh = QPushButton("Aktualisieren")
        refresh.setObjectName("primaryButton")
        refresh.clicked.connect(self.laden)

        open_chronik = QPushButton("Chronik öffnen")
        open_chronik.clicked.connect(lambda: self.nav("Objektchronik"))

        open_docs = QPushButton("Dokumenten-Center")
        open_docs.clicked.connect(lambda: self.nav("Dokumenten-Center 2.0"))

        toolbar.addWidget(QLabel("Objektordner:"))
        toolbar.addWidget(self.objekt_filter, 1)
        toolbar.addWidget(refresh)
        toolbar.addWidget(open_chronik)
        toolbar.addWidget(open_docs)
        outer.addLayout(toolbar)

        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setFrameShape(QFrame.Shape.NoFrame)
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        outer.addWidget(self.scroll, 1)

        self.content = QWidget()
        self.content.setMinimumWidth(980)
        self.content_layout = QVBoxLayout(self.content)
        self.content_layout.setContentsMargins(4, 4, 14, 18)
        self.content_layout.setSpacing(14)
        self.scroll.setWidget(self.content)

        self.status_panel = QFrame()
        self.status_panel.setObjectName("chartPanel")
        self.status_panel.setMinimumHeight(115)
        status_layout = QHBoxLayout(self.status_panel)
        status_layout.setContentsMargins(18, 14, 18, 14)
        status_layout.setSpacing(14)

        self.ampel = QLabel("⚪")
        self.ampel.setStyleSheet("font-size:46px;")
        self.ampel.setMinimumWidth(72)

        status_text = QVBoxLayout()
        self.status_title = QLabel("Kein Objekt ausgewählt")
        self.status_title.setObjectName("metricValue")
        self.status_title.setWordWrap(True)

        self.status_info = QLabel("")
        self.status_info.setWordWrap(True)
        self.status_info.setMinimumHeight(42)

        status_text.addWidget(self.status_title)
        status_text.addWidget(self.status_info)

        status_layout.addWidget(self.ampel, 0, Qt.AlignmentFlag.AlignTop)
        status_layout.addLayout(status_text, 1)
        self.content_layout.addWidget(self.status_panel)

        self.cards_widget = QWidget()
        self.cards_grid = QGridLayout(self.cards_widget)
        self.cards_grid.setContentsMargins(0, 0, 0, 0)
        self.cards_grid.setHorizontalSpacing(12)
        self.cards_grid.setVerticalSpacing(12)
        self.content_layout.addWidget(self.cards_widget)

        self.tabs = QTabWidget()
        self.tabs.setMinimumHeight(500)
        self.tabs.setDocumentMode(True)
        self.content_layout.addWidget(self.tabs, 1)

        self.tables = {}
        for title_name in [
            "Stammdaten",
            "Wohnungen",
            "Mieter",
            "Finanzen",
            "Aufgaben & Schäden",
            "Dokumente",
            "Chronik",
        ]:
            table = QTableWidget()
            table.setAlternatingRowColors(True)
            table.setWordWrap(False)
            table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
            table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
            table.setMinimumHeight(410)
            table.horizontalHeader().setStretchLastSection(False)
            self.tables[title_name] = table
            self.tabs.addTab(table, title_name)

        self.laden()

    @staticmethod
    def _metric_card(title: str, value: str, icon: str) -> QFrame:
        card = QFrame()
        card.setObjectName("metricCard")
        card.setMinimumWidth(210)
        card.setMaximumWidth(360)
        card.setMinimumHeight(110)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 12, 14, 12)
        layout.setSpacing(6)

        title_label = QLabel(f"{icon}  {title}")
        title_label.setObjectName("metricTitle")
        title_label.setWordWrap(True)

        value_label = QLabel(value)
        value_label.setObjectName("metricValue")
        value_label.setWordWrap(True)

        layout.addWidget(title_label)
        layout.addWidget(value_label)
        return card

    @staticmethod
    def _fill_table(table: QTableWidget, headers: list[str], rows: list[list[object]]) -> None:
        table.clear()
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setRowCount(len(rows))

        for row_index, row in enumerate(rows):
            for col_index in range(len(headers)):
                value = row[col_index] if col_index < len(row) else ""
                table.setItem(
                    row_index,
                    col_index,
                    QTableWidgetItem("" if value is None else str(value)),
                )

        for col_index in range(len(headers)):
            width = 180
            if col_index == 0:
                width = 220
            table.setColumnWidth(col_index, width)

    def laden(self) -> None:
        objekt = self.objekt_filter.currentText().strip()
        self.current_objekt = objekt

        if not objekt:
            self.ampel.setText("⚪")
            self.status_title.setText("Kein Objekt ausgewählt")
            self.status_info.setText("")
            return

        status = smart_objekt_status(objekt)
        health = objekt_gesundheitsindex(objekt)

        self.ampel.setText(str(health.get("ampel", "⚪")))
        self.status_title.setText(
            f"{objekt} · {health.get('punkte', 0)} % · {health.get('bewertung', '')}"
        )
        hinweise = status.get("hinweise", [])
        self.status_info.setText(
            " · ".join(str(item) for item in hinweise)
            if hinweise
            else "Keine kritischen Hinweise."
        )

        while self.cards_grid.count():
            item = self.cards_grid.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        cards = [
            ("Wohnungen", str(status.get("wohnungen", 0)), "🏠"),
            ("Mieter", str(status.get("mieter", 0)), "👥"),
            ("Leerstand", str(status.get("freie_wohnungen", 0)), "🚪"),
            ("Offene Aufgaben", str(status.get("offene_aufgaben", 0)), "🗂"),
            ("Überfällig", str(status.get("ueberfaellige_aufgaben", 0)), "⏰"),
            ("Offene Rechnungen", str(status.get("offene_rechnungen", 0)), "🧾"),
            ("Offene Schäden", str(status.get("offene_schaeden", 0)), "⚠"),
            ("Dokumente", str(status.get("dokumente", 0)), "📄"),
            ("Monatsmiete", euro(float(status.get("monatsmiete", 0.0))), "💶"),
            ("Jahresmiete", euro(float(status.get("jahresmiete", 0.0))), "📈"),
            ("Ausgaben", euro(float(status.get("ausgaben", 0.0))), "💸"),
            ("Saldo", euro(float(status.get("saldo", 0.0))), "Σ"),
        ]

        # 3 Spalten verhindern Überlagerungen auf kleineren Bildschirmen.
        for index, card in enumerate(cards):
            self.cards_grid.addWidget(
                self._metric_card(*card),
                index // 3,
                index % 3,
            )

        # Stammdaten
        object_rows = [
            row for row in DATA.get("Objekte", [])
            if norm_key(objektordner_fuer_datensatz("Objekte", row)) == norm_key(objekt)
            or norm_key(feldwert("Objekte", row, ["Objektname", "Objekt"])) == norm_key(objekt)
        ]
        self._fill_table(
            self.tables["Stammdaten"],
            SCHEMA.get("Objekte", []),
            object_rows,
        )

        # Wohnungen
        wohnungen = [
            row for row in DATA.get("Wohnungen", [])
            if norm_key(objektordner_fuer_datensatz("Wohnungen", row)) == norm_key(objekt)
        ]
        self._fill_table(
            self.tables["Wohnungen"],
            SCHEMA.get("Wohnungen", []),
            wohnungen,
        )

        # Mieter
        mieter = [
            row for row in DATA.get("Mieter", [])
            if norm_key(objektordner_fuer_datensatz("Mieter", row)) == norm_key(objekt)
        ]
        self._fill_table(
            self.tables["Mieter"],
            SCHEMA.get("Mieter", []),
            mieter,
        )

        # Finanzen als vereinheitlichte Übersicht
        finance_rows = []
        for title_name in ["Rechnungen", "Betriebskosten", "Zahlungen", "HV-Rechnungen"]:
            for row in DATA.get(title_name, []):
                if norm_key(objektordner_fuer_datensatz(title_name, row)) == norm_key(objekt):
                    finance_rows.append([
                        title_name,
                        *[row[i] if i < len(row) else "" for i in range(min(8, len(row)))],
                    ])
        finance_headers = ["Bereich"] + [f"Feld {i}" for i in range(1, 9)]
        self._fill_table(self.tables["Finanzen"], finance_headers, finance_rows)

        # Aufgaben und Schäden
        task_rows = []
        for title_name in ["Aufgaben", "Schäden", "Fristen"]:
            for row in DATA.get(title_name, []):
                if norm_key(objektordner_fuer_datensatz(title_name, row)) == norm_key(objekt):
                    task_rows.append([
                        title_name,
                        *[row[i] if i < len(row) else "" for i in range(min(8, len(row)))],
                    ])
        task_headers = ["Bereich"] + [f"Feld {i}" for i in range(1, 9)]
        self._fill_table(self.tables["Aufgaben & Schäden"], task_headers, task_rows)

        # Dokumente
        document_rows = []
        for title_name, rows in DATA.items():
            headers = SCHEMA.get(title_name, [])
            pdf_indexes = [
                index for index, field in enumerate(headers)
                if any(token in norm_key(field) for token in ["pdf", "datei", "dokument", "pfad"])
            ]
            if not pdf_indexes:
                continue

            for row in rows:
                if norm_key(objektordner_fuer_datensatz(title_name, row)) != norm_key(objekt):
                    continue
                for pdf_index in pdf_indexes:
                    if pdf_index < len(row) and str(row[pdf_index] or "").strip():
                        document_rows.append([
                            title_name,
                            headers[pdf_index],
                            str(row[pdf_index]),
                        ])

        self._fill_table(
            self.tables["Dokumente"],
            ["Bereich", "Feld", "Datei/Pfad"],
            document_rows,
        )

        # Chronik
        chronik = objektchronik_rows(objekt)
        self._fill_table(
            self.tables["Chronik"],
            ["Datum", "Bereich", "Ereignis", "Status", "Objektordner"],
            chronik,
        )

class DokumentenCenter2Seite(QWidget):
    """Dokumentenindex über alle bestehenden Tabellen und PDF-Felder."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.rows: list[dict[str, str]] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(20, 18, 20, 18)
        root.setSpacing(14)

        title = QLabel("Dokumenten-Center 2.0")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        subtitle = QLabel(
            "Zentrale Suche und Statusprüfung für PDF-, Bild- und Dateipfade aus allen bestehenden Tabellen."
        )
        subtitle.setObjectName("subTitle")
        subtitle.setWordWrap(True)
        root.addWidget(subtitle)

        top = QHBoxLayout()
        self.object_filter = QComboBox()
        self.object_filter.addItem("Alle Objektordner")
        self.object_filter.addItems(alle_objektordner())
        self.object_filter.currentTextChanged.connect(self.refresh)

        self.search = QLineEdit()
        self.search.setPlaceholderText("Titel, Bereich, Feld, Dateityp oder Pfad suchen ...")
        self.search.textChanged.connect(self.refresh)

        self.status_filter = QComboBox()
        self.status_filter.addItems(["Alle Status", "Vorhanden", "Pfad prüfen"])
        self.status_filter.currentTextChanged.connect(self.refresh)

        refresh_button = QPushButton("Index aktualisieren")
        refresh_button.setObjectName("primaryButton")
        refresh_button.clicked.connect(self.refresh)

        top.addWidget(QLabel("Objektordner:"))
        top.addWidget(self.object_filter)
        top.addWidget(self.search, 1)
        top.addWidget(self.status_filter)
        top.addWidget(refresh_button)
        root.addLayout(top)

        self.summary = QLabel("")
        self.summary.setObjectName("metricTitle")
        root.addWidget(self.summary)

        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels(
            ["Status", "Typ", "Titel", "Bereich", "Objektordner", "Feld", "Zeile", "Pfad"]
        )
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.cellDoubleClicked.connect(self.open_document)
        root.addWidget(self.table, 1)

        actions = QHBoxLayout()
        open_button = QPushButton("Ausgewählte Datei öffnen")
        open_button.setObjectName("primaryButton")
        open_button.clicked.connect(self.open_selected)
        area_button = QPushButton("Quellbereich öffnen")
        area_button.clicked.connect(self.open_source_area)
        actions.addWidget(open_button)
        actions.addWidget(area_button)
        actions.addStretch()
        root.addLayout(actions)

        self.refresh()

    def refresh(self) -> None:
        selected = self.object_filter.currentText().strip()
        object_name = "" if selected == "Alle Objektordner" else selected
        rows = enterprise_dokument_index(object_name)

        query = self.search.text().strip().lower()
        if query:
            rows = [
                item for item in rows
                if query in " ".join(item.values()).lower()
            ]

        status = self.status_filter.currentText()
        if status != "Alle Status":
            rows = [item for item in rows if item["status"] == status]

        self.rows = rows
        existing = sum(1 for item in rows if item["status"] == "Vorhanden")
        missing = len(rows) - existing
        self.summary.setText(
            f"{len(rows)} Dokumentverweise · {existing} vorhanden · {missing} Pfad/Paket prüfen"
        )

        self.table.setRowCount(len(rows))
        keys = ["status", "typ", "titel", "bereich", "objekt", "feld", "zeile", "pfad"]
        for row_index, item in enumerate(rows):
            for col_index, key in enumerate(keys):
                self.table.setItem(
                    row_index, col_index, QTableWidgetItem(item.get(key, ""))
                )
        for col_index, width in enumerate([110, 80, 280, 190, 240, 210, 70, 700]):
            self.table.setColumnWidth(col_index, width)

    def selected_row(self) -> dict[str, str] | None:
        index = self.table.currentRow()
        if 0 <= index < len(self.rows):
            return self.rows[index]
        return None

    def open_document(self, row: int, _column: int) -> None:
        if 0 <= row < len(self.rows):
            path = Path(self.rows[row]["pfad"])
            if path.exists():
                system_datei_oeffnen(str(path))
            else:
                QMessageBox.warning(self, "Dokument", f"Datei nicht gefunden:\n{path}")

    def open_selected(self) -> None:
        item = self.selected_row()
        if item is not None:
            path = Path(item["pfad"])
            if path.exists():
                system_datei_oeffnen(str(path))
            else:
                QMessageBox.warning(self, "Dokument", f"Datei nicht gefunden:\n{path}")

    def open_source_area(self) -> None:
        item = self.selected_row()
        if item is not None and item["bereich"]:
            self.nav(item["bereich"])


class WorkflowRegelnProSeite(QWidget):
    """Regelbasierte Vorschläge mit optionaler Übernahme als bestehende Aufgabe."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.rows: list[dict[str, str]] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(20, 18, 20, 18)
        root.setSpacing(14)

        title = QLabel("Workflow-Regeln PRO")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        subtitle = QLabel(
            "Prüft offene Rechnungen, Schäden, Fristen und Plausibilitäten und schlägt konkrete Arbeitsschritte vor."
        )
        subtitle.setObjectName("subTitle")
        subtitle.setWordWrap(True)
        root.addWidget(subtitle)

        top = QHBoxLayout()
        self.object_filter = QComboBox()
        self.object_filter.addItem("Alle Objektordner")
        self.object_filter.addItems(alle_objektordner())
        self.object_filter.currentTextChanged.connect(self.refresh)

        self.priority_filter = QComboBox()
        self.priority_filter.addItems(["Alle Prioritäten", "Kritisch", "Hoch", "Normal", "Niedrig"])
        self.priority_filter.currentTextChanged.connect(self.refresh)

        refresh_button = QPushButton("Regeln ausführen")
        refresh_button.setObjectName("primaryButton")
        refresh_button.clicked.connect(self.refresh)

        create_button = QPushButton("Als Aufgabe übernehmen")
        create_button.clicked.connect(self.create_task)

        top.addWidget(QLabel("Objektordner:"))
        top.addWidget(self.object_filter)
        top.addWidget(QLabel("Priorität:"))
        top.addWidget(self.priority_filter)
        top.addWidget(refresh_button)
        top.addWidget(create_button)
        top.addStretch()
        root.addLayout(top)

        self.summary = QLabel("")
        self.summary.setObjectName("metricTitle")
        root.addWidget(self.summary)

        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(
            ["Priorität", "Regel", "Objektordner", "Titel", "Bereich", "Aktion", "Fällig"]
        )
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.cellDoubleClicked.connect(self.open_area)
        root.addWidget(self.table, 1)

        self.refresh()

    def refresh(self) -> None:
        selected = self.object_filter.currentText().strip()
        object_name = "" if selected == "Alle Objektordner" else selected
        rows = enterprise_workflow_vorschlaege(object_name)

        priority = self.priority_filter.currentText()
        if priority != "Alle Prioritäten":
            rows = [item for item in rows if item["prioritaet"] == priority]

        self.rows = rows
        critical = sum(1 for item in rows if item["prioritaet"] == "Kritisch")
        high = sum(1 for item in rows if item["prioritaet"] == "Hoch")
        self.summary.setText(
            f"{len(rows)} Vorschläge · {critical} kritisch · {high} hoch"
        )

        self.table.setRowCount(len(rows))
        keys = ["prioritaet", "regel", "objekt", "titel", "bereich", "aktion", "faellig"]
        for row_index, item in enumerate(rows):
            for col_index, key in enumerate(keys):
                self.table.setItem(
                    row_index, col_index, QTableWidgetItem(item.get(key, ""))
                )
        for col_index, width in enumerate([110, 190, 240, 420, 190, 420, 110]):
            self.table.setColumnWidth(col_index, width)

    def selected_row(self) -> dict[str, str] | None:
        index = self.table.currentRow()
        return self.rows[index] if 0 <= index < len(self.rows) else None

    def create_task(self) -> None:
        item = self.selected_row()
        if item is None:
            QMessageBox.information(self, "Workflow", "Bitte einen Vorschlag auswählen.")
            return

        values = ["" for _ in SCHEMA["Aufgaben"]]
        defaults = {
            "Aufgabe": item["titel"],
            "Bereich": item["bereich"],
            "Objekt": item["objekt"],
            "Objektordner": item["objekt"],
            "Verantwortlich": aktueller_mitarbeitername(),
            "Priorität": item["prioritaet"],
            "Fällig am": item["faellig"],
            "Status": "Neu",
            "Notiz": item["aktion"],
        }
        for index, field in enumerate(SCHEMA["Aufgaben"]):
            if field in defaults:
                values[index] = defaults[field]

        dialog = EingabeDialog("Workflow-Aufgabe übernehmen", SCHEMA["Aufgaben"], values)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            DATA["Aufgaben"].append(dialog.values())
            speichere_tabelle("Aufgaben")
            aktivitaet_protokollieren(
                "Workflow-Regeln PRO",
                "Aufgabe erzeugt",
                beschreibung=item["titel"],
            )
            QMessageBox.information(self, "Workflow", "Aufgabe wurde übernommen.")
            self.refresh()

    def open_area(self, row: int, _column: int) -> None:
        if 0 <= row < len(self.rows):
            target = self.rows[row].get("bereich", "")
            if target:
                self.nav(target)



SEARCH_INDEX_FILE = DATEN_DIR / "enterprise_search_index.json"


def enterprise_search_index_bauen(force: bool = False) -> list[dict[str, str]]:
    """Erzeugt einen lokalen Volltextindex ohne Änderung der Exceldateien."""
    if not force and SEARCH_INDEX_FILE.exists():
        try:
            cached = json.loads(SEARCH_INDEX_FILE.read_text(encoding="utf-8"))
            if isinstance(cached, list):
                return cached
        except (OSError, ValueError, json.JSONDecodeError):
            pass

    index: list[dict[str, str]] = []
    for titel, rows in DATA.items():
        headers = SCHEMA.get(titel, [])
        for row_number, row in enumerate(rows, start=1):
            values = ["" if value is None else str(value) for value in row]
            objekt = objektordner_fuer_datensatz(titel, row)
            index.append({
                "bereich": titel,
                "zeile": str(row_number),
                "objekt": objekt,
                "titel": values[0] if values else titel,
                "text": " | ".join(values),
                "suchtext": " ".join(values).lower(),
                "felder": " | ".join(headers),
            })

    try:
        DATEN_DIR.mkdir(parents=True, exist_ok=True)
        SEARCH_INDEX_FILE.write_text(
            json.dumps(index, ensure_ascii=False),
            encoding="utf-8",
        )
    except (OSError, PermissionError, TypeError):
        pass
    return index


def enterprise_search(query: str, objekt: str = "") -> list[dict[str, str]]:
    words = [word for word in str(query or "").lower().split() if word]
    objekt_key = norm_key(objekt)
    if not words and not objekt_key:
        return []

    results: list[dict[str, str]] = []
    for item in enterprise_search_index_bauen():
        if objekt_key and norm_key(item.get("objekt", "")) != objekt_key:
            continue
        haystack = item.get("suchtext", "")
        if words and not all(word in haystack for word in words):
            continue
        results.append(item)
    return results[:500]


def projektmonitor_daten() -> list[dict[str, str]]:
    checks: list[dict[str, str]] = []
    total = 0
    passed = 0

    for titel in SCHEMA:
        total += 1
        path = xlsx_pfad(titel)
        ok = path.exists()
        passed += int(ok)
        checks.append({
            "status": "🟢" if ok else "🔴",
            "bereich": "Excel",
            "element": titel,
            "ergebnis": "Vorhanden" if ok else "Fehlt",
            "details": str(path),
        })

    total += 1
    invalid = len(objektordner_pruefbericht())
    ok = invalid == 0
    passed += int(ok)
    checks.append({
        "status": "🟢" if ok else "🟡",
        "bereich": "Objektordner",
        "element": "Zuordnungen",
        "ergebnis": f"{invalid} Datensätze prüfen",
        "details": "Objektordner-Prüfung öffnen",
    })

    total += 1
    documents = dokumenten_index_rows()
    missing = sum(1 for item in documents if item.get("status") != "Vorhanden")
    ok = missing == 0
    passed += int(ok)
    checks.append({
        "status": "🟢" if ok else "🟡",
        "bereich": "Dokumente",
        "element": "Dateipfade",
        "ergebnis": f"{missing} fehlende Datei(en)",
        "details": f"{len(documents)} indexierte Dokumente",
    })

    total += 1
    backup_files = list(BACKUP_DIR.glob("*.zip")) if BACKUP_DIR.exists() else []
    ok = bool(backup_files)
    passed += int(ok)
    checks.append({
        "status": "🟢" if ok else "🟡",
        "bereich": "Backup",
        "element": "Sicherungen",
        "ergebnis": f"{len(backup_files)} Backup(s)",
        "details": str(BACKUP_DIR),
    })

    score = round((passed / total) * 100) if total else 100
    checks.insert(0, {
        "status": "🟢" if score >= 90 else "🟡" if score >= 70 else "🔴",
        "bereich": "Gesamtzustand",
        "element": "Projektmonitor",
        "ergebnis": f"{score} %",
        "details": f"{passed} von {total} Prüfungen bestanden",
    })
    return checks


class EnterpriseAssistentSeite(QWidget):
    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.rows: list[dict[str, str]] = []
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 18, 20, 18)
        layout.setSpacing(14)

        title = QLabel("Enterprise Assistant")
        title.setObjectName("pageTitle")
        layout.addWidget(title)
        info = QLabel("Lokale intelligente Suche über alle Tabellen und Objektordner. Keine Datenübertragung nach außen.")
        info.setObjectName("subTitle")
        info.setWordWrap(True)
        layout.addWidget(info)

        top = QHBoxLayout()
        self.objekt = QComboBox()
        self.objekt.addItem("Alle Objektordner")
        self.objekt.addItems(alle_objektordner())
        self.query = QLineEdit()
        self.query.setPlaceholderText("z. B. Müller, offene Rechnung, Heizung Musterstraße ...")
        self.query.returnPressed.connect(self.suchen)
        search = QPushButton("Suchen")
        search.setObjectName("primaryButton")
        search.clicked.connect(self.suchen)
        rebuild = QPushButton("Index neu aufbauen")
        rebuild.clicked.connect(self.index_neu)
        top.addWidget(QLabel("Objektordner:"))
        top.addWidget(self.objekt)
        top.addWidget(self.query, 1)
        top.addWidget(search)
        top.addWidget(rebuild)
        layout.addLayout(top)

        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Bereich", "Zeile", "Objektordner", "Titel", "Felder", "Inhalt"])
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.cellDoubleClicked.connect(self.oeffnen)
        layout.addWidget(self.table, 1)

    def index_neu(self) -> None:
        enterprise_search_index_bauen(True)
        QMessageBox.information(self, "Suchindex", "Der lokale Suchindex wurde neu aufgebaut.")
        self.suchen()

    def suchen(self) -> None:
        selected = self.objekt.currentText().strip()
        objekt = "" if selected == "Alle Objektordner" else selected
        self.rows = enterprise_search(self.query.text(), objekt)
        self.table.setRowCount(len(self.rows))
        keys = ["bereich", "zeile", "objekt", "titel", "felder", "text"]
        for r, item in enumerate(self.rows):
            for c, key in enumerate(keys):
                self.table.setItem(r, c, QTableWidgetItem(item.get(key, "")))
        for c, width in enumerate([190, 70, 230, 260, 360, 760]):
            self.table.setColumnWidth(c, width)

    def oeffnen(self, row: int, _column: int) -> None:
        if 0 <= row < len(self.rows):
            self.nav(self.rows[row].get("bereich", "Dashboard"))


class ProjektmonitorSeite(QWidget):
    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 18, 20, 18)
        layout.setSpacing(14)
        title = QLabel("Projektmonitor")
        title.setObjectName("pageTitle")
        layout.addWidget(title)
        info = QLabel("Überwacht Exceldateien, Objektordner, Dokumentpfade, Backups und den Gesamtzustand des Projekts.")
        info.setObjectName("subTitle")
        info.setWordWrap(True)
        layout.addWidget(info)
        top = QHBoxLayout()
        refresh = QPushButton("Projekt prüfen")
        refresh.setObjectName("primaryButton")
        refresh.clicked.connect(self.laden)
        system = QPushButton("System-Center öffnen")
        system.clicked.connect(lambda: self.nav("System-Center PRO"))
        object_check = QPushButton("Objektordner-Prüfung")
        object_check.clicked.connect(lambda: self.nav("Objektordner-Prüfung"))
        top.addWidget(refresh)
        top.addWidget(system)
        top.addWidget(object_check)
        top.addStretch()
        layout.addLayout(top)
        self.score = QLabel("")
        self.score.setObjectName("metricValue")
        layout.addWidget(self.score)
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Status", "Bereich", "Element", "Ergebnis", "Details"])
        self.table.setAlternatingRowColors(True)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        layout.addWidget(self.table, 1)
        self.laden()

    def laden(self) -> None:
        rows = projektmonitor_daten()
        self.score.setText(f"Projektzustand: {rows[0]['ergebnis'] if rows else '–'}")
        self.table.setRowCount(len(rows))
        keys = ["status", "bereich", "element", "ergebnis", "details"]
        for r, item in enumerate(rows):
            for c, key in enumerate(keys):
                self.table.setItem(r, c, QTableWidgetItem(item.get(key, "")))
        for c, width in enumerate([80, 180, 260, 190, 620]):
            self.table.setColumnWidth(c, width)


class DokumentVorschauSeite(QWidget):
    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.rows: list[dict[str, str]] = []
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 18, 20, 18)
        layout.setSpacing(12)
        title = QLabel("Dokumentenvorschau")
        title.setObjectName("pageTitle")
        layout.addWidget(title)
        split = QSplitter(Qt.Orientation.Horizontal)
        layout.addWidget(split, 1)

        left = QWidget()
        left_layout = QVBoxLayout(left)
        self.search = QLineEdit()
        self.search.setPlaceholderText("Dokument suchen ...")
        self.search.textChanged.connect(self.laden)
        left_layout.addWidget(self.search)
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Bereich", "Objektordner", "Titel", "Datei"])
        self.table.setAlternatingRowColors(True)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.cellClicked.connect(self.vorschau)
        self.table.cellDoubleClicked.connect(self.extern_oeffnen)
        left_layout.addWidget(self.table, 1)
        split.addWidget(left)

        right = QWidget()
        right_layout = QVBoxLayout(right)
        self.preview_title = QLabel("Dokument auswählen")
        self.preview_title.setObjectName("metricTitle")
        right_layout.addWidget(self.preview_title)
        self.preview = QTextBrowser()
        self.preview.setOpenExternalLinks(False)
        right_layout.addWidget(self.preview, 1)
        self.open_button = QPushButton("Extern öffnen")
        self.open_button.setObjectName("primaryButton")
        self.open_button.clicked.connect(self.extern_oeffnen_current)
        right_layout.addWidget(self.open_button)
        split.addWidget(right)
        split.setSizes([700, 700])
        self.laden()

    def laden(self) -> None:
        query = self.search.text().lower().strip()
        rows = dokumenten_index_rows()
        self.rows = [item for item in rows if not query or query in " ".join(item.values()).lower()]
        self.table.setRowCount(len(self.rows))
        for r, item in enumerate(self.rows):
            values = [item.get("bereich", ""), item.get("objekt", ""), item.get("titel", ""), item.get("pfad", "")]
            for c, value in enumerate(values):
                self.table.setItem(r, c, QTableWidgetItem(value))
        for c, width in enumerate([180, 230, 250, 520]):
            self.table.setColumnWidth(c, width)

    def _path(self, row: int) -> Path | None:
        if not (0 <= row < len(self.rows)):
            return None
        raw = self.rows[row].get("pfad", "")
        path = Path(raw)
        if not path.is_absolute():
            candidates = [APP_DIR / path, DOKUMENTE_DIR / path, path]
            path = next((candidate for candidate in candidates if candidate.exists()), candidates[0])
        return path

    def vorschau(self, row: int, _column: int) -> None:
        path = self._path(row)
        if path is None:
            return
        self.preview_title.setText(path.name)
        if not path.exists():
            self.preview.setPlainText(f"Datei nicht gefunden:\n{path}")
            return
        suffix = path.suffix.lower()
        if suffix in {".png", ".jpg", ".jpeg", ".webp", ".bmp"}:
            self.preview.setHtml(f'<div style="text-align:center"><img src="{path.as_uri()}" style="max-width:95%; max-height:700px"></div>')
        elif suffix == ".pdf" and PdfReader is not None:
            try:
                reader = PdfReader(str(path))
                content = "\n\n".join((page.extract_text() or "") for page in reader.pages[:20])
                self.preview.setPlainText(content or "Das PDF enthält keinen direkt lesbaren Text.")
            except (OSError, ValueError, TypeError) as error:
                self.preview.setPlainText(f"PDF-Vorschau nicht möglich:\n{error}")
        elif suffix in {".txt", ".md", ".csv", ".json"}:
            try:
                self.preview.setPlainText(path.read_text(encoding="utf-8", errors="replace")[:200000])
            except OSError as error:
                self.preview.setPlainText(str(error))
        else:
            self.preview.setPlainText(f"Für {suffix or 'diesen Dateityp'} ist keine interne Vorschau verfügbar.\n\n{path}")

    def extern_oeffnen(self, row: int, _column: int) -> None:
        path = self._path(row)
        if path is not None and path.exists():
            system_datei_oeffnen(path)

    def extern_oeffnen_current(self) -> None:
        self.extern_oeffnen(self.table.currentRow(), 0)



def objektbilder_ordner(objektordner: str) -> Path:
    """Liefert einen sicheren, objektbezogenen Ablageordner für Gebäudebilder."""
    safe_name = re.sub(r"[^0-9A-Za-zÄÖÜäöüß._ -]+", "_", str(objektordner).strip())
    safe_name = safe_name.strip(" .") or "Objekt"
    return DOKUMENTE_DIR / "objektbilder" / safe_name


def gebaeudezwilling_fotos(objektordner: str) -> list[Path]:
    """Ermittelt hochgeladene und bereits indexierte Objektbilder."""
    objekt_key = norm_key(objektordner)
    image_suffixes = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"}
    result: list[Path] = []
    seen: set[str] = set()

    upload_dir = objektbilder_ordner(objektordner)
    if upload_dir.exists():
        for path in sorted(upload_dir.iterdir(), key=lambda item: item.name.lower()):
            path_key = str(path.resolve())
            if path.is_file() and path.suffix.lower() in image_suffixes and path_key not in seen:
                seen.add(path_key)
                result.append(path)

    for item in dokumenten_index_rows():
        if norm_key(item.get("objekt", "")) != objekt_key:
            continue

        path = Path(item.get("pfad", ""))
        try:
            path_key = str(path.resolve())
        except OSError:
            path_key = str(path)

        if (
            item.get("status") == "Vorhanden"
            and path.suffix.lower() in image_suffixes
            and path_key not in seen
        ):
            seen.add(path_key)
            result.append(path)

    return result[:40]


def gebaeudezwilling_einheiten(objektordner: str) -> list[dict[str, str]]:
    """Bereitet Wohnungen und zugehörige Mieter für die visuelle Ansicht auf.

    Zuordnungsreihenfolge:
    1. gleicher Wohnungsordner,
    2. gleicher Objektordner und gleiche Wohnfläche.

    Die im Gebäudezwilling angezeigte Fläche stammt immer aus der Tabelle
    ``Mieter`` aus der Spalte ``Wohnfläche``. Ein Mieterdatensatz wird dabei
    höchstens einer Wohnung zugeordnet.
    """
    objekt_key = norm_key(objektordner)
    result: list[dict[str, str]] = []

    tenants = [
        row for row in DATA.get("Mieter", [])
        if norm_key(objektordner_fuer_datensatz("Mieter", row)) == objekt_key
    ]
    verwendete_mieter: set[int] = set()

    def flaechenwert(value: Any) -> float | None:
        text = str(value or "").strip()
        if not text:
            return None
        number = to_float(text, default=float("nan"))
        return number if number == number else None

    for apartment in DATA.get("Wohnungen", []):
        if norm_key(objektordner_fuer_datensatz("Wohnungen", apartment)) != objekt_key:
            continue

        apartment_name = feldwert(
            "Wohnungen",
            apartment,
            ["Wohnung", "Wohnungsordner", "Bezeichnung", "Nummer"],
        )
        apartment_folder = feldwert("Wohnungen", apartment, ["Wohnungsordner"])
        apartment_area = flaechenwert(
            feldwert("Wohnungen", apartment, ["Größe qm", "Wohnfläche", "Fläche"])
        )
        status = feldwert("Wohnungen", apartment, ["Status"])

        tenant_index: int | None = None

        # 1. Bevorzugt über einen ausdrücklich gepflegten Wohnungsordner verbinden.
        apartment_folder_key = norm_key(apartment_folder)
        if apartment_folder_key:
            for index, tenant in enumerate(tenants):
                if index in verwendete_mieter:
                    continue
                tenant_folder = feldwert("Mieter", tenant, ["Wohnungsordner"])
                if norm_key(tenant_folder) == apartment_folder_key:
                    tenant_index = index
                    break

        # 2. Fallback für bestehende Daten: Objektordner + identische Wohnfläche.
        if tenant_index is None and apartment_area is not None:
            for index, tenant in enumerate(tenants):
                if index in verwendete_mieter:
                    continue
                tenant_area = flaechenwert(feldwert("Mieter", tenant, ["Wohnfläche"]))
                if tenant_area is not None and abs(tenant_area - apartment_area) < 0.01:
                    tenant_index = index
                    break

        tenant_name = ""
        tenant_status = ""
        area = ""

        if tenant_index is not None:
            verwendete_mieter.add(tenant_index)
            tenant = tenants[tenant_index]
            tenant_name = feldwert("Mieter", tenant, ["Mieter", "Name"])
            tenant_status = feldwert("Mieter", tenant, ["Mieter-Status", "Status"])
            # Gewünscht: Fläche ausschließlich aus Mieter.xlsx / Wohnfläche.
            area = feldwert("Mieter", tenant, ["Wohnfläche"])

        combined_status = status or tenant_status
        status_key = norm_key(combined_status)

        if any(word in status_key for word in ["frei", "leer", "unvermietet"]):
            ampel = "🟡"
            display_status = combined_status or "Frei"
        elif tenant_name:
            ampel = "🟢"
            display_status = combined_status or "Vermietet"
        else:
            ampel = "⚪"
            display_status = combined_status or "Nicht zugeordnet"

        result.append({
            "wohnung": apartment_name or "Einheit",
            "mieter": tenant_name or "Kein Mieter zugeordnet",
            "status": display_status,
            "ampel": ampel,
            "flaeche": area,
        })

    return result


def gebaeudezwilling_letzte_ereignisse(
    objektordner: str,
    limit: int = 12,
) -> list[list[str]]:
    rows = objektchronik_rows(objektordner)
    rows.sort(key=lambda row: datum_sort_key(row[0]), reverse=True)
    return rows[:limit]


class DigitalerGebaeudezwillingSeite(QWidget):
    """Visuelle Objektübersicht als digitaler Gebäudezwilling."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.current_objekt = ""
        self.photo_paths: list[Path] = []
        self.photo_index = 0

        outer = QVBoxLayout(self)
        outer.setContentsMargins(18, 16, 18, 16)
        outer.setSpacing(12)

        title = QLabel("Digitaler Gebäudezwilling")
        title.setObjectName("pageTitle")
        outer.addWidget(title)

        subtitle = QLabel(
            "Visuelle Objektakte mit Gebäudebild, Einheiten, Zustand, "
            "Finanzen, Aufgaben und letzten Ereignissen."
        )
        subtitle.setObjectName("subTitle")
        subtitle.setWordWrap(True)
        outer.addWidget(subtitle)

        toolbar = QHBoxLayout()
        toolbar.setSpacing(10)

        self.objekt_filter = QComboBox()
        self.objekt_filter.setEditable(True)
        self.objekt_filter.setMinimumWidth(320)
        self.objekt_filter.addItems(alle_objektordner())
        self.objekt_filter.currentTextChanged.connect(self.laden)

        refresh = QPushButton("Aktualisieren")
        refresh.setObjectName("primaryButton")
        refresh.clicked.connect(self.laden)

        open_360 = QPushButton("Objekt 360°")
        open_360.clicked.connect(lambda: self.nav("Enterprise Objekt 360°"))

        open_docs = QPushButton("Dokumente")
        open_docs.clicked.connect(lambda: self.nav("Dokumenten-Center 2.0"))

        open_tasks = QPushButton("Aufgaben")
        open_tasks.clicked.connect(lambda: self.nav("Arbeitsorganisation PRO"))

        toolbar.addWidget(QLabel("Objektordner:"))
        toolbar.addWidget(self.objekt_filter, 1)
        toolbar.addWidget(refresh)
        toolbar.addWidget(open_360)
        toolbar.addWidget(open_docs)
        toolbar.addWidget(open_tasks)
        outer.addLayout(toolbar)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        outer.addWidget(scroll, 1)

        content = QWidget()
        content.setMinimumWidth(1080)
        layout = QVBoxLayout(content)
        layout.setContentsMargins(4, 4, 14, 18)
        layout.setSpacing(14)
        scroll.setWidget(content)

        hero = QFrame()
        hero.setObjectName("chartPanel")
        hero_layout = QHBoxLayout(hero)
        hero_layout.setContentsMargins(18, 16, 18, 16)
        hero_layout.setSpacing(18)

        photo_panel = QFrame()
        photo_panel.setObjectName("metricCard")
        photo_layout = QVBoxLayout(photo_panel)
        photo_layout.setContentsMargins(10, 10, 10, 10)

        self.photo = QLabel("Kein Objektfoto vorhanden")
        self.photo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.photo.setMinimumSize(390, 260)
        self.photo.setMaximumHeight(340)
        self.photo.setStyleSheet(
            "border:1px solid #dce5f0; border-radius:12px; "
            "background:#f5f8fc; color:#64748b;"
        )
        photo_layout.addWidget(self.photo, 1)

        photo_controls = QHBoxLayout()
        previous_photo = QPushButton("‹")
        previous_photo.setMaximumWidth(45)
        previous_photo.clicked.connect(self.vorheriges_foto)

        self.photo_counter = QLabel("0 / 0")
        self.photo_counter.setAlignment(Qt.AlignmentFlag.AlignCenter)

        next_photo = QPushButton("›")
        next_photo.setMaximumWidth(45)
        next_photo.clicked.connect(self.naechstes_foto)

        upload_photo = QPushButton("Objektbilder hochladen")
        upload_photo.clicked.connect(self.objektbilder_hochladen)

        open_photo = QPushButton("Foto öffnen")
        open_photo.clicked.connect(self.foto_oeffnen)

        photo_controls.addWidget(previous_photo)
        photo_controls.addWidget(self.photo_counter, 1)
        photo_controls.addWidget(next_photo)
        photo_controls.addWidget(upload_photo)
        photo_controls.addWidget(open_photo)
        photo_layout.addLayout(photo_controls)

        hero_layout.addWidget(photo_panel, 0)

        overview_panel = QFrame()
        overview_panel.setObjectName("metricCard")
        overview_layout = QVBoxLayout(overview_panel)
        overview_layout.setContentsMargins(18, 16, 18, 16)
        overview_layout.setSpacing(9)

        self.object_name = QLabel("Kein Objekt ausgewählt")
        self.object_name.setObjectName("metricValue")
        self.object_name.setWordWrap(True)
        overview_layout.addWidget(self.object_name)

        self.health_text = QLabel("")
        self.health_text.setObjectName("metricTitle")
        overview_layout.addWidget(self.health_text)

        self.health_bar = QProgressBar()
        self.health_bar.setRange(0, 100)
        self.health_bar.setMinimumHeight(26)
        self.health_bar.setFormat("Gesundheitsindex: %p %")
        overview_layout.addWidget(self.health_bar)

        self.object_details = QLabel("")
        self.object_details.setWordWrap(True)
        self.object_details.setTextInteractionFlags(
            Qt.TextInteractionFlag.TextSelectableByMouse
        )
        overview_layout.addWidget(self.object_details, 1)

        hero_actions = QGridLayout()
        actions = [
            ("Chronik", "Objektchronik"),
            ("Fristen", "Fristenmanager PRO"),
            ("Rechnungen", "Rechnungen"),
            ("Schäden", "Schäden"),
            ("Mieter", "Mieter"),
            ("Kalender", "Kalender & Planung"),
        ]
        for index, (label, target) in enumerate(actions):
            button = QPushButton(label)
            button.clicked.connect(
                lambda checked=False, page=target: self.nav(page)
            )
            hero_actions.addWidget(button, index // 3, index % 3)
        overview_layout.addLayout(hero_actions)

        hero_layout.addWidget(overview_panel, 1)
        layout.addWidget(hero)

        self.cards_widget = QWidget()
        self.cards_grid = QGridLayout(self.cards_widget)
        self.cards_grid.setContentsMargins(0, 0, 0, 0)
        self.cards_grid.setHorizontalSpacing(12)
        self.cards_grid.setVerticalSpacing(12)
        layout.addWidget(self.cards_widget)

        lower_splitter = QSplitter(Qt.Orientation.Horizontal)
        lower_splitter.setChildrenCollapsible(False)
        layout.addWidget(lower_splitter, 1)

        unit_panel = QGroupBox("Einheiten im Gebäude")
        unit_layout = QVBoxLayout(unit_panel)

        self.unit_grid_widget = QWidget()
        self.unit_grid = QGridLayout(self.unit_grid_widget)
        self.unit_grid.setContentsMargins(2, 2, 12, 12)
        self.unit_grid.setSpacing(10)

        unit_scroll = QScrollArea()
        unit_scroll.setWidgetResizable(True)
        unit_scroll.setFrameShape(QFrame.Shape.NoFrame)
        unit_scroll.setWidget(self.unit_grid_widget)
        unit_layout.addWidget(unit_scroll)

        lower_splitter.addWidget(unit_panel)

        timeline_panel = QGroupBox("Letzte Ereignisse")
        timeline_layout = QVBoxLayout(timeline_panel)

        self.timeline = QTableWidget()
        self.timeline.setColumnCount(4)
        self.timeline.setHorizontalHeaderLabels(
            ["Datum", "Bereich", "Ereignis", "Status"]
        )
        self.timeline.setAlternatingRowColors(True)
        self.timeline.setWordWrap(False)
        self.timeline.setHorizontalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAlwaysOn
        )
        self.timeline.setVerticalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAlwaysOn
        )
        self.timeline.cellDoubleClicked.connect(self.ereignis_oeffnen)
        timeline_layout.addWidget(self.timeline)

        lower_splitter.addWidget(timeline_panel)
        lower_splitter.setSizes([600, 650])

        self.laden()

    @staticmethod
    def _metric_card(title: str, value: str, icon: str) -> QFrame:
        card = QFrame()
        card.setObjectName("metricCard")
        card.setMinimumWidth(190)
        card.setMinimumHeight(100)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 12, 14, 12)

        heading = QLabel(f"{icon}  {title}")
        heading.setObjectName("metricTitle")
        heading.setWordWrap(True)

        amount = QLabel(value)
        amount.setObjectName("metricValue")
        amount.setWordWrap(True)

        layout.addWidget(heading)
        layout.addWidget(amount)
        return card

    @staticmethod
    def _unit_card(item: dict[str, str]) -> QFrame:
        card = QFrame()
        card.setObjectName("metricCard")
        card.setMinimumWidth(205)
        card.setMinimumHeight(145)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(13, 11, 13, 11)
        layout.setSpacing(5)

        heading = QLabel(f'{item["ampel"]}  {item["mieter"]}')
        heading.setStyleSheet("font-size:17px; font-weight:900;")
        heading.setWordWrap(True)

        apartment = QLabel(f'Wohnung: {item["wohnung"]}')
        apartment.setWordWrap(True)

        status = QLabel(item["status"])
        status.setObjectName("metricTitle")

        area = QLabel(
            f'Fläche: {item["flaeche"]}'
            if item["flaeche"]
            else "Fläche: nicht hinterlegt"
        )
        area.setObjectName("metricTitle")

        layout.addWidget(heading)
        layout.addWidget(apartment)
        layout.addWidget(status)
        layout.addWidget(area)
        layout.addStretch()
        return card

    def laden(self) -> None:
        objekt = self.objekt_filter.currentText().strip()
        self.current_objekt = objekt

        if not objekt:
            self.object_name.setText("Kein Objekt ausgewählt")
            self.health_bar.setValue(0)
            self.photo.setText("Kein Objektfoto vorhanden")
            return

        status = smart_objekt_status(objekt)
        health = objekt_gesundheitsindex(objekt)

        self.object_name.setText(
            f'{health.get("ampel", "⚪")}  {objekt}'
        )
        self.health_text.setText(
            f'{health.get("bewertung", "")} · '
            f'{len(health.get("gruende", []))} Bewertungshinweis(e)'
        )
        self.health_bar.setValue(int(health.get("punkte", 0)))

        object_rows = [
            row for row in DATA.get("Objekte", [])
            if (
                norm_key(feldwert("Objekte", row, ["Objektname", "Objekt"]))
                == norm_key(objekt)
            )
        ]
        object_row = object_rows[0] if object_rows else []

        details = [
            ("Adresse", feldwert("Objekte", object_row, ["Adresse", "Straße", "Objektname"])),
            ("Ort", feldwert("Objekte", object_row, ["Ort"])),
            ("Vermieter", feldwert("Objekte", object_row, ["Vermieter", "Eigentümer"])),
            ("Vermieter E-Mail", feldwert("Objekte", object_row, ["Vermieter E-Mail-Adresse", "Vermieter E-Mail"])),
            ("Baujahr", feldwert("Objekte", object_row, ["Baujahr"])),
            ("Grundstücksfläche", feldwert("Objekte", object_row, ["Grundstücksfläche"])),
            ("Wohnungen", str(status.get("wohnungen", 0))),
            ("Mieter", str(status.get("mieter", 0))),
            ("Leerstand", str(status.get("freie_wohnungen", 0))),
        ]
        self.object_details.setText(
            "\n".join(
                f"<b>{label}:</b> {value or 'nicht hinterlegt'}"
                for label, value in details
            )
        )

        while self.cards_grid.count():
            item = self.cards_grid.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        cards = [
            ("Monatsmiete", euro(float(status.get("monatsmiete", 0.0))), "💶"),
            ("Jahresmiete", euro(float(status.get("jahresmiete", 0.0))), "📈"),
            ("Saldo", euro(float(status.get("saldo", 0.0))), "Σ"),
            ("Aufgaben", str(status.get("offene_aufgaben", 0)), "🗂"),
            ("Rechnungen", str(status.get("offene_rechnungen", 0)), "🧾"),
            ("Schäden", str(status.get("offene_schaeden", 0)), "⚠"),
            ("Dokumente", str(status.get("dokumente", 0)), "📄"),
            ("Überfällig", str(status.get("ueberfaellige_aufgaben", 0)), "⏰"),
        ]

        for index, card in enumerate(cards):
            self.cards_grid.addWidget(
                self._metric_card(*card),
                index // 4,
                index % 4,
            )

        self.photo_paths = gebaeudezwilling_fotos(objekt)
        self.photo_index = 0
        self.foto_anzeigen()

        while self.unit_grid.count():
            item = self.unit_grid.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        units = gebaeudezwilling_einheiten(objekt)
        for index, unit in enumerate(units):
            self.unit_grid.addWidget(
                self._unit_card(unit),
                index // 2,
                index % 2,
            )
        self.unit_grid.setRowStretch((len(units) // 2) + 1, 1)

        events = gebaeudezwilling_letzte_ereignisse(objekt)
        self.timeline.setRowCount(len(events))

        for row_index, row in enumerate(events):
            for col_index, value in enumerate(row[:4]):
                self.timeline.setItem(
                    row_index,
                    col_index,
                    QTableWidgetItem(str(value)),
                )

        for col_index, width in enumerate([110, 170, 500, 130]):
            self.timeline.setColumnWidth(col_index, width)

    def objektbilder_hochladen(self) -> None:
        objekt = self.current_objekt.strip()
        if not objekt:
            QMessageBox.information(
                self,
                "Objektbilder",
                "Bitte zuerst einen Objektordner auswählen.",
            )
            return

        paths, _ = QFileDialog.getOpenFileNames(
            self,
            "Objektbilder auswählen",
            "",
            "Bilddateien (*.png *.jpg *.jpeg *.webp *.bmp *.gif)",
        )
        if not paths:
            return

        zielordner = objektbilder_ordner(objekt)
        try:
            zielordner.mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            QMessageBox.warning(self, "Objektbilder", f"Ordner konnte nicht erstellt werden:\n{exc}")
            return

        copied = 0
        errors: list[str] = []
        for source_text in paths:
            source = Path(source_text)
            ziel = zielordner / source.name
            counter = 2
            while ziel.exists():
                ziel = zielordner / f"{source.stem}_{counter}{source.suffix}"
                counter += 1
            try:
                shutil.copy2(source, ziel)
                copied += 1
            except OSError as exc:
                errors.append(f"{source.name}: {exc}")

        self.photo_paths = gebaeudezwilling_fotos(objekt)
        self.photo_index = max(0, len(self.photo_paths) - copied)
        self.foto_anzeigen()

        message = f"{copied} Objektbild(er) wurden gespeichert."
        if errors:
            message += "\n\nNicht gespeichert:\n" + "\n".join(errors)
            QMessageBox.warning(self, "Objektbilder", message)
        else:
            QMessageBox.information(self, "Objektbilder", message)

    def foto_anzeigen(self) -> None:
        if not self.photo_paths:
            self.photo.clear()
            self.photo.setText("Kein Objektfoto vorhanden")
            self.photo_counter.setText("0 / 0")
            return

        self.photo_index %= len(self.photo_paths)
        path = self.photo_paths[self.photo_index]
        pixmap = QPixmap(str(path))

        if pixmap.isNull():
            self.photo.setText(f"Bild konnte nicht geladen werden:\n{path.name}")
        else:
            self.photo.setPixmap(
                pixmap.scaled(
                    self.photo.size(),
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation,
                )
            )

        self.photo.setToolTip(str(path))
        self.photo_counter.setText(
            f"{self.photo_index + 1} / {len(self.photo_paths)}"
        )

    def vorheriges_foto(self) -> None:
        if self.photo_paths:
            self.photo_index = (self.photo_index - 1) % len(self.photo_paths)
            self.foto_anzeigen()

    def naechstes_foto(self) -> None:
        if self.photo_paths:
            self.photo_index = (self.photo_index + 1) % len(self.photo_paths)
            self.foto_anzeigen()

    def foto_oeffnen(self) -> None:
        if self.photo_paths:
            system_datei_oeffnen(str(self.photo_paths[self.photo_index]))

    def ereignis_oeffnen(self, row: int, _column: int) -> None:
        item = self.timeline.item(row, 1)
        if item is not None and item.text().strip():
            self.nav(item.text().strip())



def dokument_text_lesen(path: Path) -> str:
    """Liest Text aus unterstützten Dokumenten, ohne externe Dienste."""
    suffix = path.suffix.lower()

    if suffix == ".pdf":
        if PdfReader is None:
            return ""
        try:
            reader = PdfReader(str(path))
            return "\n".join((page.extract_text() or "") for page in reader.pages)
        except (OSError, ValueError, TypeError):
            return ""

    if suffix in {".txt", ".md", ".csv", ".json", ".xml"}:
        try:
            return path.read_text(encoding="utf-8", errors="ignore")
        except OSError:
            return ""

    return ""


def dokument_klassifizieren(path: Path, text_value: str = "") -> dict[str, str]:
    """
    Regelbasierte, lokale Dokumenterkennung.
    Es werden keine Daten übertragen und keine Excelstrukturen verändert.
    """
    combined = f"{path.name} {text_value}".lower()

    supplier_rules = {
        "Stadtwerke": ["stadtwerke", "energieversorgung"],
        "E.ON": ["e.on", "eon"],
        "EnBW": ["enbw"],
        "Westnetz": ["westnetz"],
        "Techem": ["techem"],
        "Ista": ["ista"],
        "Brunata": ["brunata"],
        "Telekom": ["telekom"],
        "Vodafone": ["vodafone"],
        "Versicherung": ["versicherung", "police", "versicherungsnummer"],
        "Schornsteinfeger": ["schornsteinfeger", "feuerstättenschau"],
    }

    category_rules = {
        "Versorgerrechnung": [
            "strom", "gas", "wasser", "energie", "abschlag",
            "verbrauch", "zählerstand", "kwh", "m³", "m3",
        ],
        "Heizkosten": [
            "heizkosten", "wärme", "heizung", "techem", "ista", "brunata",
        ],
        "Versicherung": [
            "versicherung", "police", "prämie", "schadenversicherung",
        ],
        "Mietvertrag": [
            "mietvertrag", "mietbeginn", "kaltmiete", "warmmiete",
        ],
        "Rechnung": [
            "rechnung", "rechnungsnummer", "rechnungsnr", "brutto", "netto",
        ],
        "Kontoauszug": [
            "kontoauszug", "buchungstag", "wertstellung", "iban", "saldo",
        ],
        "Übergabeprotokoll": [
            "übergabeprotokoll", "wohnungsübergabe", "zählerstände",
        ],
        "Schaden": [
            "schaden", "schadennummer", "reparatur", "mangel",
        ],
    }

    supplier = ""
    supplier_score = 0
    for name, tokens in supplier_rules.items():
        score = sum(1 for token in tokens if token in combined)
        if score > supplier_score:
            supplier = name
            supplier_score = score

    category = "Sonstiges Dokument"
    category_score = 0
    for name, tokens in category_rules.items():
        score = sum(1 for token in tokens if token in combined)
        if score > category_score:
            category = name
            category_score = score

    object_matches: list[str] = []
    for objekt in alle_objektordner():
        if norm_key(objekt) and norm_key(objekt) in norm_key(combined):
            object_matches.append(objekt)

    objekt = object_matches[0] if len(object_matches) == 1 else ""

    invoice_number = ""
    invoice_patterns = [
        r"(?:rechnungsnummer|rechnungsnr\.?|rechnung\s*nr\.?)\s*[:#]?\s*([A-Z0-9\-\/]+)",
        r"(?:invoice\s*no\.?)\s*[:#]?\s*([A-Z0-9\-\/]+)",
    ]
    for pattern in invoice_patterns:
        match = re.search(pattern, text_value, re.IGNORECASE)
        if match:
            invoice_number = match.group(1).strip()
            break

    amount = ""
    amount_match = re.search(
        r"(?:gesamtbetrag|rechnungsbetrag|brutto|zu\s*zahlen)\s*[: ]+\s*"
        r"(\d{1,3}(?:\.\d{3})*,\d{2})\s*€?",
        text_value,
        re.IGNORECASE,
    )
    if amount_match:
        amount = amount_match.group(1)

    confidence = min(
        100,
        20
        + category_score * 18
        + supplier_score * 12
        + (20 if objekt else 0)
        + (10 if invoice_number else 0),
    )

    return {
        "datei": path.name,
        "pfad": str(path),
        "kategorie": category,
        "lieferant": supplier,
        "objekt": objekt,
        "rechnungsnummer": invoice_number,
        "betrag": amount,
        "sicherheit": f"{confidence} %",
        "status": "Erkannt" if confidence >= 60 else "Prüfen",
    }


def dokument_dublette_pruefen(item: dict[str, str]) -> str:
    """Prüft Dateiname, Rechnungsnummer und Pfad gegen vorhandene Dokumente."""
    filename_key = norm_key(item.get("datei", ""))
    invoice_key = norm_key(item.get("rechnungsnummer", ""))
    path_key = norm_key(item.get("pfad", ""))

    for existing in dokumenten_index_rows():
        existing_filename = norm_key(existing.get("datei", ""))
        existing_path = norm_key(existing.get("pfad", ""))
        existing_text = norm_key(existing.get("inhalt", ""))

        if path_key and path_key == existing_path:
            return "Identischer Pfad vorhanden"

        if filename_key and filename_key == existing_filename:
            return "Gleicher Dateiname vorhanden"

        if invoice_key and invoice_key in existing_text:
            return "Rechnungsnummer bereits vorhanden"

    return ""


def dokument_automatik_scan(folder: Path) -> list[dict[str, str]]:
    result: list[dict[str, str]] = []

    supported = {
        ".pdf", ".txt", ".md", ".csv", ".json", ".xml",
        ".png", ".jpg", ".jpeg", ".webp", ".bmp",
    }

    if not folder.exists():
        return result

    for path in sorted(folder.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in supported:
            continue

        text_value = dokument_text_lesen(path)
        item = dokument_klassifizieren(path, text_value)
        item["dublette"] = dokument_dublette_pruefen(item)
        result.append(item)

    return result


class DokumentenautomatisierungProSeite(QWidget):
    """Lokale Dokumentklassifizierung mit kontrollierter Übernahme."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.rows: list[dict[str, str]] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(20, 18, 20, 18)
        root.setSpacing(14)

        title = QLabel("Dokumentenautomatisierung PRO")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        info = QLabel(
            "Dokumente lokal klassifizieren, Objektordner vorschlagen, "
            "Lieferanten und Rechnungsdaten erkennen sowie Dubletten prüfen."
        )
        info.setObjectName("subTitle")
        info.setWordWrap(True)
        root.addWidget(info)

        top = QHBoxLayout()

        self.folder = QLineEdit()
        self.folder.setPlaceholderText("Ordner mit zu prüfenden Dokumenten ...")
        self.folder.setText(str(DOKUMENTE_DIR))

        choose = QPushButton("Ordner wählen")
        choose.clicked.connect(self.ordner_waehlen)

        scan = QPushButton("Dokumente prüfen")
        scan.setObjectName("primaryButton")
        scan.clicked.connect(self.scannen)

        open_center = QPushButton("Dokumenten-Center")
        open_center.clicked.connect(lambda: self.nav("Dokumenten-Center 2.0"))

        top.addWidget(self.folder, 1)
        top.addWidget(choose)
        top.addWidget(scan)
        top.addWidget(open_center)
        root.addLayout(top)

        self.summary = QLabel("")
        self.summary.setObjectName("metricTitle")
        root.addWidget(self.summary)

        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "Status", "Datei", "Kategorie", "Lieferant", "Objektordner",
            "Rechnungsnummer", "Betrag", "Sicherheit", "Dublette", "Pfad",
        ])
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setHorizontalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAlwaysOn
        )
        self.table.setVerticalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAlwaysOn
        )
        self.table.cellDoubleClicked.connect(self.datei_oeffnen)
        root.addWidget(self.table, 1)

        actions = QHBoxLayout()

        create_task = QPushButton("Prüfaufgabe erzeugen")
        create_task.clicked.connect(self.pruefaufgabe_erzeugen)

        export_btn = QPushButton("Ergebnis exportieren")
        export_btn.clicked.connect(self.exportieren)

        actions.addWidget(create_task)
        actions.addWidget(export_btn)
        actions.addStretch()
        root.addLayout(actions)

    def ordner_waehlen(self) -> None:
        selected = QFileDialog.getExistingDirectory(
            self,
            "Dokumentenordner auswählen",
            self.folder.text().strip() or str(DOKUMENTE_DIR),
        )
        if selected:
            self.folder.setText(selected)

    def scannen(self) -> None:
        folder = Path(self.folder.text().strip())
        self.rows = dokument_automatik_scan(folder)

        recognized = sum(1 for item in self.rows if item["status"] == "Erkannt")
        duplicates = sum(1 for item in self.rows if item["dublette"])

        self.summary.setText(
            f"{len(self.rows)} Dokument(e) geprüft · "
            f"{recognized} erkannt · {duplicates} mögliche Dublette(n)"
        )

        self.table.setRowCount(len(self.rows))

        keys = [
            "status", "datei", "kategorie", "lieferant", "objekt",
            "rechnungsnummer", "betrag", "sicherheit", "dublette", "pfad",
        ]

        for row_index, item in enumerate(self.rows):
            for col_index, key in enumerate(keys):
                self.table.setItem(
                    row_index,
                    col_index,
                    QTableWidgetItem(str(item.get(key, ""))),
                )

        for col_index, width in enumerate(
            [100, 240, 180, 160, 230, 170, 110, 100, 240, 650]
        ):
            self.table.setColumnWidth(col_index, width)

    def _selected_item(self) -> dict[str, str] | None:
        row = self.table.currentRow()
        if row < 0 or row >= len(self.rows):
            return None
        return self.rows[row]

    def datei_oeffnen(self, row: int, _column: int) -> None:
        if 0 <= row < len(self.rows):
            path = Path(self.rows[row].get("pfad", ""))
            if path.exists():
                system_datei_oeffnen(str(path))

    def pruefaufgabe_erzeugen(self) -> None:
        item = self._selected_item()
        if item is None:
            QMessageBox.information(
                self,
                "Dokumentenautomatisierung",
                "Bitte zuerst ein Dokument auswählen.",
            )
            return

        fields = SCHEMA.get("Aufgaben", [])
        if not fields:
            return

        values = ["" for _ in fields]
        defaults = {
            "Aufgabe": f'Dokument prüfen: {item.get("datei", "")}',
            "Objekt": item.get("objekt", ""),
            "Objektordner": item.get("objekt", ""),
            "Bereich": "Dokumentenprüfung",
            "Priorität": "Normal" if not item.get("dublette") else "Hoch",
            "Status": "Neu",
            "Verantwortlich": aktueller_mitarbeitername(),
            "Notiz": (
                f'Kategorie: {item.get("kategorie", "")}\n'
                f'Lieferant: {item.get("lieferant", "")}\n'
                f'Rechnungsnummer: {item.get("rechnungsnummer", "")}\n'
                f'Dublette: {item.get("dublette", "")}\n'
                f'Pfad: {item.get("pfad", "")}'
            ),
        }

        for index, field in enumerate(fields):
            if field in defaults:
                values[index] = defaults[field]

        dialog = EingabeDialog(
            "Dokumenten-Prüfaufgabe",
            fields,
            values,
        )
        if dialog.exec() == QDialog.DialogCode.Accepted:
            DATA.setdefault("Aufgaben", []).append(dialog.values())
            speichere_tabelle("Aufgaben")
            aktivitaet_protokollieren(
                "Dokumentenautomatisierung",
                "Prüfaufgabe erzeugt",
                beschreibung=item.get("datei", ""),
            )

    def exportieren(self) -> None:
        if not self.rows:
            QMessageBox.information(
                self,
                "Dokumentenautomatisierung",
                "Keine Ergebnisse vorhanden.",
            )
            return

        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        target, _ = QFileDialog.getSaveFileName(
            self,
            "Ergebnis exportieren",
            str(EXPORT_DIR / "dokumentenautomatisierung.xlsx"),
            "Excel-Dateien (*.xlsx)",
        )
        if not target:
            return

        target_path = Path(target)
        if target_path.suffix.lower() != ".xlsx":
            target_path = target_path.with_suffix(".xlsx")

        workbook = Workbook()
        sheet = workbook.active
        if not isinstance(sheet, Worksheet):
            return

        sheet.title = "Dokumentenautomatisierung"
        headers = [
            "Status", "Datei", "Kategorie", "Lieferant", "Objektordner",
            "Rechnungsnummer", "Betrag", "Sicherheit", "Dublette", "Pfad",
        ]
        sheet.append(headers)

        keys = [
            "status", "datei", "kategorie", "lieferant", "objekt",
            "rechnungsnummer", "betrag", "sicherheit", "dublette", "pfad",
        ]

        for item in self.rows:
            sheet.append([item.get(key, "") for key in keys])

        for col_index, width in enumerate(
            [14, 32, 24, 22, 30, 24, 16, 14, 34, 90],
            start=1,
        ):
            sheet.column_dimensions[
                sheet.cell(1, col_index).column_letter
            ].width = width

        workbook.save(target_path)
        QMessageBox.information(
            self,
            "Dokumentenautomatisierung",
            f"Ergebnis exportiert:\n{target_path}",
        )



MOBILE_PORTAL_DIR = APP_DIR / "mobile_portal"
MOBILE_INBOX_DIR = DATEN_DIR / "mobile_zeiterfassung"
MOBILE_ARCHIVE_DIR = MOBILE_INBOX_DIR / "archiv"


def mobile_verzeichnisse_sicherstellen() -> None:
    for folder in [MOBILE_PORTAL_DIR, MOBILE_INBOX_DIR, MOBILE_ARCHIVE_DIR]:
        try:
            folder.mkdir(parents=True, exist_ok=True)
        except OSError:
            pass


def mobile_stundeneintraege_laden() -> list[dict[str, Any]]:
    mobile_verzeichnisse_sicherstellen()
    rows: list[dict[str, Any]] = []

    for path in sorted(MOBILE_INBOX_DIR.glob("*.json")):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, ValueError, json.JSONDecodeError):
            continue

        entries = data if isinstance(data, list) else [data]
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            item = dict(entry)
            item["_datei"] = str(path)
            item["_status"] = str(item.get("status", "Eingereicht"))
            rows.append(item)

    rows.sort(
        key=lambda item: (
            str(item.get("datum", "")),
            str(item.get("mitarbeiter", "")),
            str(item.get("beginn", "")),
        ),
        reverse=True,
    )
    return rows


def mobile_stunden_berechnen(
    beginn: str,
    ende: str,
    pause_minuten: Any,
) -> float:

    try:
        start = datetime.strptime(str(beginn).strip(), "%H:%M")
        end = datetime.strptime(str(ende).strip(), "%H:%M")
        if end < start:
            end += timedelta(days=1)
        pause = max(0, int(float(str(pause_minuten or 0).replace(",", "."))))
        hours = (end - start).total_seconds() / 3600 - pause / 60
        return round(max(0.0, hours), 2)
    except (ValueError, TypeError):
        return 0.0


def mobile_eintrag_signatur(item: dict[str, Any]) -> str:
    raw = "|".join([
        str(item.get("id", "")),
        str(item.get("mitarbeiter", "")),
        str(item.get("datum", "")),
        str(item.get("objektordner", "")),
        str(item.get("beginn", "")),
        str(item.get("ende", "")),
    ])
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def mobile_bereits_importiert(signatur: str) -> bool:
    marker_file = MOBILE_ARCHIVE_DIR / "importiert.json"
    try:
        values = json.loads(marker_file.read_text(encoding="utf-8"))
        return signatur in values if isinstance(values, list) else False
    except (OSError, ValueError, json.JSONDecodeError):
        return False


def mobile_import_markieren(signatur: str) -> None:
    marker_file = MOBILE_ARCHIVE_DIR / "importiert.json"
    try:
        values = json.loads(marker_file.read_text(encoding="utf-8"))
        if not isinstance(values, list):
            values = []
    except (OSError, ValueError, json.JSONDecodeError):
        values = []

    if signatur not in values:
        values.append(signatur)

    try:
        marker_file.write_text(
            json.dumps(values, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except OSError:
        pass


def mobile_stundennachweis_zeile(item: dict[str, Any]) -> list[Any]:
    fields = SCHEMA.get("Stundennachweise", [])
    hours = mobile_stunden_berechnen(
        str(item.get("beginn", "")),
        str(item.get("ende", "")),
        item.get("pause", 0),
    )

    values = {
        "Mitarbeiter": str(item.get("mitarbeiter", "")),
        "Datum": str(item.get("datum", "")),
        "Objekt": str(item.get("objektordner", "")),
        "Objektordner": str(item.get("objektordner", "")),
        "Wohnung": str(item.get("wohnung", "")),
        "Tätigkeit": str(item.get("taetigkeit", "")),
        "Beginn": str(item.get("beginn", "")),
        "Ende": str(item.get("ende", "")),
        "Pause": str(item.get("pause", "")),
        "Stunden": str(hours).replace(".", ","),
        "Notiz": str(item.get("notiz", "")),
        "Status": "Freigegeben",
    }

    row = ["" for _ in fields]
    for index, field in enumerate(fields):
        if field in values:
            row[index] = values[field]
    return row


def mobile_portal_dateien_erstellen() -> None:
    mobile_verzeichnisse_sicherstellen()

    objects = alle_objektordner()
    portal_config = {
        "mitarbeiter": ["Julia", "Franzi", "Robert", "Ralf"],
        "objektordner": objects,
        "taetigkeiten": [
            "Hausmeister",
            "Reparatur",
            "Reinigung",
            "Gartenpflege",
            "Winterdienst",
            "Wohnungsübergabe",
            "Wohnungsabnahme",
            "Besichtigung",
            "Verwaltung",
            "Sonstiges",
        ],
    }

    (MOBILE_PORTAL_DIR / "config.json").write_text(
        json.dumps(portal_config, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    index_html = """<!DOCTYPE html>
<html lang="de">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover">
  <meta name="theme-color" content="#071a35">
  <link rel="manifest" href="manifest.json">
  <link rel="stylesheet" href="style.css">
  <title>DBS Zeiterfassung</title>
</head>
<body>
  <header>
    <h1>DBS Zeiterfassung</h1>
    <p>Mobile Arbeitszeit- und Tätigkeitsmeldung</p>
  </header>
  <main>
    <section class="card">
      <label>Mitarbeiter<select id="mitarbeiter"></select></label>
      <label>Datum<input id="datum" type="date"></label>
      <label>Objektordner<select id="objekt"></select></label>
      <label>Wohnung<input id="wohnung" placeholder="optional"></label>
      <label>Tätigkeit<select id="taetigkeit"></select></label>
      <div class="grid">
        <label>Beginn<input id="beginn" type="time"></label>
        <label>Ende<input id="ende" type="time"></label>
      </div>
      <label>Pause in Minuten<input id="pause" type="number" min="0" value="0"></label>
      <label>Notiz<textarea id="notiz" rows="4"></textarea></label>
      <label>Foto<input id="foto" type="file" accept="image/*" capture="environment"></label>
      <div id="stunden" class="result">0,00 Stunden</div>
      <button id="speichern">Eintrag speichern</button>
      <button id="exportieren" class="secondary">Einträge exportieren</button>
    </section>
    <section class="card">
      <h2>Gespeicherte Einträge</h2>
      <div id="liste"></div>
    </section>
  </main>
  <script src="app.js"></script>
</body>
</html>
"""
    (MOBILE_PORTAL_DIR / "index.html").write_text(index_html, encoding="utf-8")

    style_css = """
:root{font-family:system-ui,-apple-system,Segoe UI,sans-serif;color:#0b1628;background:#eef3f9}
*{box-sizing:border-box}body{margin:0}header{background:#071a35;color:white;padding:22px 18px}
header h1{margin:0 0 6px;font-size:24px}header p{margin:0;color:#b9cdeb}
main{max-width:680px;margin:auto;padding:16px}.card{background:white;border-radius:18px;padding:18px;
box-shadow:0 10px 30px rgba(15,23,42,.08);margin-bottom:16px}
label{display:block;font-weight:700;margin-bottom:12px}input,select,textarea{width:100%;margin-top:6px;
padding:12px;border:1px solid #d8e2ef;border-radius:11px;font:inherit;background:white}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:12px}.result{font-size:23px;font-weight:900;
padding:14px 0;color:#1d4ed8}button{width:100%;padding:13px;border:0;border-radius:12px;
background:#2563eb;color:white;font-weight:900;font-size:16px;margin-top:8px}
button.secondary{background:#e8eef6;color:#0b1628}.entry{padding:12px;border-bottom:1px solid #e5eaf1}
.entry strong{display:block}.muted{color:#64748b;font-size:13px}
@media(max-width:520px){.grid{grid-template-columns:1fr}}
"""
    (MOBILE_PORTAL_DIR / "style.css").write_text(style_css, encoding="utf-8")

    app_js = r"""
let config={mitarbeiter:[],objektordner:[],taetigkeiten:[]};
const $=id=>document.getElementById(id);
const entries=()=>JSON.parse(localStorage.getItem("dbs_zeiten")||"[]");
const saveEntries=v=>localStorage.setItem("dbs_zeiten",JSON.stringify(v));

function fillSelect(id,values,first="Bitte wählen"){
  const el=$(id); el.innerHTML="";
  const empty=document.createElement("option"); empty.textContent=first; empty.value=""; el.appendChild(empty);
  values.forEach(value=>{const o=document.createElement("option");o.value=value;o.textContent=value;el.appendChild(o);});
}
function calc(){
  const b=$("beginn").value,e=$("ende").value,p=parseFloat($("pause").value||0);
  if(!b||!e){$("stunden").textContent="0,00 Stunden";return 0}
  let [bh,bm]=b.split(":").map(Number),[eh,em]=e.split(":").map(Number);
  let mins=(eh*60+em)-(bh*60+bm); if(mins<0)mins+=1440; mins=Math.max(0,mins-p);
  const h=Math.round(mins/60*100)/100;$("stunden").textContent=h.toFixed(2).replace(".",",")+" Stunden";return h;
}
async function imageData(file){
  if(!file)return "";
  return await new Promise(resolve=>{const r=new FileReader();r.onload=()=>resolve(r.result);r.readAsDataURL(file)});
}
async function store(){
  const item={
    id:crypto.randomUUID?crypto.randomUUID():Date.now().toString(),
    mitarbeiter:$("mitarbeiter").value,datum:$("datum").value,objektordner:$("objekt").value,
    wohnung:$("wohnung").value,taetigkeit:$("taetigkeit").value,beginn:$("beginn").value,
    ende:$("ende").value,pause:$("pause").value,stunden:calc(),notiz:$("notiz").value,
    foto:await imageData($("foto").files[0]),status:"Eingereicht",erstellt_am:new Date().toISOString()
  };
  if(!item.mitarbeiter||!item.datum||!item.objektordner||!item.beginn||!item.ende){
    alert("Bitte Mitarbeiter, Datum, Objektordner, Beginn und Ende ausfüllen.");return;
  }
  const values=entries();values.push(item);saveEntries(values);render();alert("Eintrag lokal gespeichert.");
}
function render(){
  const list=$("liste"),values=entries();list.innerHTML="";
  if(!values.length){list.textContent="Noch keine Einträge gespeichert.";return}
  values.slice().reverse().forEach(item=>{
    const d=document.createElement("div");d.className="entry";
    d.innerHTML=`<strong>${item.datum} · ${item.mitarbeiter}</strong>
    <div>${item.objektordner} · ${item.taetigkeit}</div>
    <div class="muted">${item.beginn}–${item.ende} · ${String(item.stunden).replace(".",",")} Std.</div>`;
    list.appendChild(d);
  });
}
function exportData(){
  const values=entries();if(!values.length){alert("Keine Einträge vorhanden.");return}
  const blob=new Blob([JSON.stringify(values,null,2)],{type:"application/json"});
  const a=document.createElement("a");a.href=URL.createObjectURL(blob);
  a.download=`mobile_zeiterfassung_${new Date().toISOString().slice(0,10)}.json`;a.click();
  setTimeout(()=>URL.revokeObjectURL(a.href),1000);
}
fetch("config.json").then(r=>r.json()).then(value=>{
  config=value;fillSelect("mitarbeiter",config.mitarbeiter);fillSelect("objekt",config.objektordner);
  fillSelect("taetigkeit",config.taetigkeiten);$("datum").value=new Date().toISOString().slice(0,10);
}).catch(()=>{});
["beginn","ende","pause"].forEach(id=>$(id).addEventListener("input",calc));
$("speichern").addEventListener("click",store);$("exportieren").addEventListener("click",exportData);
render();
if("serviceWorker" in navigator)navigator.serviceWorker.register("service-worker.js").catch(()=>{});
"""
    (MOBILE_PORTAL_DIR / "app.js").write_text(app_js, encoding="utf-8")

    manifest = {
        "name": "DBS Mobile Zeiterfassung",
        "short_name": "DBS Zeit",
        "start_url": "./index.html",
        "display": "standalone",
        "background_color": "#eef3f9",
        "theme_color": "#071a35",
        "icons": [],
    }
    (MOBILE_PORTAL_DIR / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    service_worker = """
const CACHE="dbs-mobile-v1";
const FILES=["./","index.html","style.css","app.js","config.json","manifest.json"];
self.addEventListener("install",e=>e.waitUntil(caches.open(CACHE).then(c=>c.addAll(FILES))));
self.addEventListener("fetch",e=>e.respondWith(caches.match(e.request).then(r=>r||fetch(e.request))));
"""
    (MOBILE_PORTAL_DIR / "service-worker.js").write_text(
        service_worker,
        encoding="utf-8",
    )





FIELD_IMPORT_DIR = DATEN_DIR / "dbs_field_import"
FIELD_IMPORT_ARCHIVE_DIR = FIELD_IMPORT_DIR / "archiv"


def field_import_verzeichnisse_sicherstellen() -> None:
    for folder in [FIELD_IMPORT_DIR, FIELD_IMPORT_ARCHIVE_DIR]:
        try:
            folder.mkdir(parents=True, exist_ok=True)
        except OSError:
            pass


def field_schadenpaket_lesen(zip_path: Path) -> dict[str, Any]:
    """
    Liest ein DBS-Field-Schadenpaket, ohne Dateien dauerhaft zu verändern.
    Erwartet `schaden.json` sowie optionale Bilddateien.
    """
    if not zip_path.exists() or zip_path.suffix.lower() != ".zip":
        raise ValueError("Ungültiges Schadenpaket.")

    with zipfile.ZipFile(zip_path, "r") as archive:
        names = archive.namelist()
        json_name = next(
            (name for name in names if Path(name).name.lower() == "schaden.json"),
            None,
        )
        if json_name is None:
            raise ValueError("Die Datei schaden.json fehlt im ZIP-Paket.")

        with archive.open(json_name) as source:
            payload = json.loads(source.read().decode("utf-8"))

        if not isinstance(payload, dict):
            raise ValueError("Die Schadenmeldung enthält kein gültiges JSON-Objekt.")

        photo_names = [
            name
            for name in names
            if Path(name).suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}
        ]

    return {
        "zip_path": str(zip_path),
        "id": str(payload.get("id", "")).strip(),
        "typ": str(payload.get("typ", "")).strip(),
        "mitarbeiter": str(payload.get("mitarbeiter", "")).strip(),
        "objektordner": str(payload.get("objektordner", "")).strip(),
        "wohnung": str(payload.get("wohnung", "")).strip(),
        "raum": str(payload.get("raum", "")).strip(),
        "schadensart": str(payload.get("schadensart", "")).strip(),
        "prioritaet": str(payload.get("prioritaet", "")).strip(),
        "beschreibung": str(payload.get("beschreibung", "")).strip(),
        "status": str(payload.get("status", "Gemeldet")).strip() or "Gemeldet",
        "erstellt_am": str(payload.get("erstellt_am", "")).strip(),
        "fotos": photo_names,
        "roh": payload,
    }


def field_schadenpaket_signatur(item: dict[str, Any]) -> str:
    raw = "|".join([
        str(item.get("id", "")),
        str(item.get("objektordner", "")),
        str(item.get("erstellt_am", "")),
        str(item.get("beschreibung", "")),
        str(item.get("zip_path", "")),
    ])
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def field_import_marker_laden() -> list[str]:
    field_import_verzeichnisse_sicherstellen()
    marker = FIELD_IMPORT_ARCHIVE_DIR / "importiert.json"

    try:
        values = json.loads(marker.read_text(encoding="utf-8"))
        return values if isinstance(values, list) else []
    except (OSError, ValueError, json.JSONDecodeError):
        return []


def field_import_marker_speichern(values: list[str]) -> None:
    field_import_verzeichnisse_sicherstellen()
    marker = FIELD_IMPORT_ARCHIVE_DIR / "importiert.json"

    try:
        marker.write_text(
            json.dumps(sorted(set(values)), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except OSError:
        pass


def field_import_bereits_importiert(item: dict[str, Any]) -> bool:
    return field_schadenpaket_signatur(item) in field_import_marker_laden()


def field_schaden_fotoziel(objektordner: str, package_id: str) -> Path:
    """
    Legt einen Unterordner innerhalb des vorhandenen Objektordners an.
    Die Excelstruktur bleibt dabei unverändert.
    """
    object_folder = normalisiere_gueltigen_objektordner(objektordner)
    if not object_folder:
        raise ValueError("Der Objektordner ist ungültig oder nicht eindeutig.")

    base = Path(object_folder)

    if not base.is_absolute():
        base = DOKUMENTE_DIR / base

    target = base / "Schäden" / (package_id or datetime.now().strftime("%Y%m%d_%H%M%S"))
    target.mkdir(parents=True, exist_ok=True)
    return target


def field_schaden_fotos_extrahieren(
    item: dict[str, Any],
    target_dir: Path,
) -> list[str]:
    zip_path = Path(str(item.get("zip_path", "")))
    extracted: list[str] = []

    with zipfile.ZipFile(zip_path, "r") as archive:
        for index, member in enumerate(item.get("fotos", []), start=1):
            suffix = Path(member).suffix.lower() or ".jpg"
            target = target_dir / f"foto_{index:02d}{suffix}"

            with archive.open(member) as source, target.open("wb") as destination:
                shutil.copyfileobj(source, destination)

            extracted.append(str(target))

        metadata_target = target_dir / "schaden.json"
        metadata_target.write_text(
            json.dumps(item.get("roh", {}), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    return extracted


def field_schaden_zeile_erzeugen(
    item: dict[str, Any],
    photo_paths: list[str],
) -> list[Any]:
    fields = SCHEMA.get("Schäden", [])
    row = ["" for _ in fields]

    values = {
        "Objekt": item.get("objektordner", ""),
        "Objektordner": item.get("objektordner", ""),
        "Wohnung": item.get("wohnung", ""),
        "Raum": item.get("raum", ""),
        "Bereich": item.get("raum", ""),
        "Schadensart": item.get("schadensart", ""),
        "Kategorie": item.get("schadensart", ""),
        "Priorität": item.get("prioritaet", ""),
        "Beschreibung": item.get("beschreibung", ""),
        "Schaden": item.get("beschreibung", ""),
        "Status": item.get("status", "Gemeldet"),
        "Mitarbeiter": item.get("mitarbeiter", ""),
        "Verantwortlich": item.get("mitarbeiter", ""),
        "Datum": item.get("erstellt_am", "")[:10],
        "Erstellt am": item.get("erstellt_am", ""),
        "Foto": photo_paths[0] if photo_paths else "",
        "Fotos": " | ".join(photo_paths),
        "Datei": item.get("zip_path", ""),
        "Dokument": item.get("zip_path", ""),
        "Notiz": (
            f'DBS Field ID: {item.get("id", "")}\n'
            f'Raum: {item.get("raum", "")}\n'
            f'Fotos: {len(photo_paths)}'
        ),
    }

    for index, field in enumerate(fields):
        if field in values:
            row[index] = values[field]

    return row


def field_schaden_importieren(item: dict[str, Any]) -> tuple[list[str], str]:
    """
    Importiert eine geprüfte Schadenmeldung.
    Es wird nur in bestehende Tabellen geschrieben; die Excelstruktur bleibt gleich.
    """
    if field_import_bereits_importiert(item):
        raise ValueError("Dieses Schadenpaket wurde bereits importiert.")

    if norm_key(item.get("typ", "")) not in {"schaden", "schadenmeldung"}:
        raise ValueError("Das ZIP-Paket ist keine DBS-Field-Schadenmeldung.")

    if not item.get("objektordner"):
        raise ValueError("Der Objektordner fehlt.")

    if not item.get("beschreibung"):
        raise ValueError("Die Schadensbeschreibung fehlt.")

    target_dir = field_schaden_fotoziel(
        str(item.get("objektordner", "")),
        str(item.get("id", "")),
    )
    photo_paths = field_schaden_fotos_extrahieren(item, target_dir)

    fields = SCHEMA.get("Schäden", [])
    if not fields:
        raise ValueError("Die bestehende Tabelle Schäden wurde nicht gefunden.")

    DATA.setdefault("Schäden", []).append(
        field_schaden_zeile_erzeugen(item, photo_paths)
    )
    speichere_tabelle("Schäden")

    signature = field_schadenpaket_signatur(item)
    markers = field_import_marker_laden()
    markers.append(signature)
    field_import_marker_speichern(markers)

    archive_target = FIELD_IMPORT_ARCHIVE_DIR / Path(
        str(item.get("zip_path", "schaden.zip"))
    ).name

    try:
        shutil.copy2(str(item.get("zip_path", "")), archive_target)
    except OSError:
        pass

    aktivitaet_protokollieren(
        "DBS Field Import",
        "Schadenmeldung importiert",
        beschreibung=(
            f'{item.get("objektordner", "")} · '
            f'{item.get("schadensart", "")} · '
            f'{item.get("prioritaet", "")}'
        ),
        status=item.get("status", "Gemeldet"),
    )

    return photo_paths, str(target_dir)


class DbsFieldImportSeite(QWidget):
    """Importzentrale für DBS-Field-Schadenpakete."""

    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.current_item: dict[str, Any] | None = None

        root = QVBoxLayout(self)
        root.setContentsMargins(20, 18, 20, 18)
        root.setSpacing(14)

        title = QLabel("DBS Field Import")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        subtitle = QLabel(
            "Schadenpakete aus der Android-App prüfen, Fotos anzeigen "
            "und kontrolliert in die bestehende Schadenstabelle übernehmen."
        )
        subtitle.setObjectName("subTitle")
        subtitle.setWordWrap(True)
        root.addWidget(subtitle)

        toolbar = QHBoxLayout()

        self.path_input = QLineEdit()
        self.path_input.setPlaceholderText("DBS-Field-Schadenpaket (*.zip) ...")

        choose = QPushButton("ZIP auswählen")
        choose.clicked.connect(self.zip_waehlen)

        inspect = QPushButton("Prüfen")
        inspect.setObjectName("primaryButton")
        inspect.clicked.connect(self.pruefen)

        open_archive = QPushButton("Importarchiv öffnen")
        open_archive.clicked.connect(
            lambda: system_datei_oeffnen(str(FIELD_IMPORT_ARCHIVE_DIR))
        )

        toolbar.addWidget(self.path_input, 1)
        toolbar.addWidget(choose)
        toolbar.addWidget(inspect)
        toolbar.addWidget(open_archive)
        root.addLayout(toolbar)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setChildrenCollapsible(False)
        root.addWidget(splitter, 1)

        detail_panel = QGroupBox("Schadendaten")
        detail_layout = QFormLayout(detail_panel)

        self.detail_labels: dict[str, QLabel] = {}
        for key, label_text in [
            ("status", "Prüfstatus"),
            ("mitarbeiter", "Mitarbeiter"),
            ("objektordner", "Objektordner"),
            ("wohnung", "Wohnung"),
            ("raum", "Raum"),
            ("schadensart", "Schadensart"),
            ("prioritaet", "Priorität"),
            ("erstellt_am", "Erstellt am"),
            ("beschreibung", "Beschreibung"),
            ("fotos_count", "Fotos"),
        ]:
            label = QLabel("")
            label.setWordWrap(True)
            label.setTextInteractionFlags(
                Qt.TextInteractionFlag.TextSelectableByMouse
            )
            self.detail_labels[key] = label
            detail_layout.addRow(label_text + ":", label)

        splitter.addWidget(detail_panel)

        photo_panel = QGroupBox("Fotovorschau")
        photo_layout = QVBoxLayout(photo_panel)

        self.photo_list = QListWidget()
        self.photo_list.currentRowChanged.connect(self.foto_anzeigen)
        photo_layout.addWidget(self.photo_list)

        self.photo_preview = QLabel("Kein Foto ausgewählt")
        self.photo_preview.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.photo_preview.setMinimumSize(420, 320)
        self.photo_preview.setStyleSheet(
            "border:1px solid #dce5f0; border-radius:12px; "
            "background:#f5f8fc; color:#64748b;"
        )
        photo_layout.addWidget(self.photo_preview, 1)

        splitter.addWidget(photo_panel)
        splitter.setSizes([520, 680])

        actions = QHBoxLayout()

        import_button = QPushButton("Schaden übernehmen")
        import_button.setObjectName("primaryButton")
        import_button.clicked.connect(self.uebernehmen)

        open_damage = QPushButton("Schäden öffnen")
        open_damage.clicked.connect(lambda: self.nav("Schäden"))

        actions.addWidget(import_button)
        actions.addWidget(open_damage)
        actions.addStretch()
        root.addLayout(actions)

    def zip_waehlen(self) -> None:
        field_import_verzeichnisse_sicherstellen()
        path, _ = QFileDialog.getOpenFileName(
            self,
            "DBS-Field-Schadenpaket auswählen",
            str(FIELD_IMPORT_DIR),
            "ZIP-Dateien (*.zip)",
        )
        if path:
            self.path_input.setText(path)
            self.pruefen()

    def pruefen(self) -> None:
        path = Path(self.path_input.text().strip())

        try:
            self.current_item = field_schadenpaket_lesen(path)
        except (OSError, ValueError, KeyError, json.JSONDecodeError) as exc:
            self.current_item = None
            QMessageBox.warning(
                self,
                "DBS Field Import",
                f"Das Paket konnte nicht gelesen werden:\n{exc}",
            )
            return

        item = self.current_item
        valid_object = normalisiere_gueltigen_objektordner(
            item.get("objektordner", "")
        )
        duplicate = field_import_bereits_importiert(item)

        if duplicate:
            status = "Bereits importiert"
        elif not valid_object:
            status = "Objektordner prüfen"
        elif not item.get("beschreibung"):
            status = "Beschreibung fehlt"
        else:
            status = "Importbereit"

        values = dict(item)
        values["status"] = status
        values["fotos_count"] = str(len(item.get("fotos", [])))

        for key, label in self.detail_labels.items():
            label.setText(str(values.get(key, "")))

        self.photo_list.clear()
        for name in item.get("fotos", []):
            self.photo_list.addItem(Path(name).name)

        if self.photo_list.count():
            self.photo_list.setCurrentRow(0)
        else:
            self.photo_preview.clear()
            self.photo_preview.setText("Keine Fotos im Paket")

    def foto_anzeigen(self, row: int) -> None:
        if (
            self.current_item is None
            or row < 0
            or row >= len(self.current_item.get("fotos", []))
        ):
            return

        member = self.current_item["fotos"][row]
        zip_path = Path(str(self.current_item.get("zip_path", "")))

        try:
            with zipfile.ZipFile(zip_path, "r") as archive:
                raw = archive.read(member)
        except (OSError, KeyError, zipfile.BadZipFile):
            self.photo_preview.setText("Foto konnte nicht gelesen werden")
            return

        pixmap = QPixmap()
        pixmap.loadFromData(raw)

        if pixmap.isNull():
            self.photo_preview.setText("Foto konnte nicht angezeigt werden")
            return

        self.photo_preview.setPixmap(
            pixmap.scaled(
                self.photo_preview.size(),
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation,
            )
        )

    def uebernehmen(self) -> None:
        if self.current_item is None:
            QMessageBox.information(
                self,
                "DBS Field Import",
                "Bitte zuerst ein Schadenpaket prüfen.",
            )
            return

        answer = QMessageBox.question(
            self,
            "Schaden übernehmen",
            "Soll die geprüfte Schadenmeldung jetzt übernommen werden?",
        )
        if answer != QMessageBox.StandardButton.Yes:
            return

        try:
            photo_paths, target_dir = field_schaden_importieren(
                self.current_item
            )
        except (OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
            QMessageBox.warning(
                self,
                "DBS Field Import",
                f"Der Import ist fehlgeschlagen:\n{exc}",
            )
            return

        QMessageBox.information(
            self,
            "DBS Field Import",
            "Die Schadenmeldung wurde übernommen.\n\n"
            f"Fotos: {len(photo_paths)}\n"
            f"Zielordner: {target_dir}",
        )
        self.pruefen()



FIELD_INBOX_DIR = FIELD_IMPORT_DIR / "eingang"


def field_eingang_verzeichnisse_sicherstellen() -> None:
    field_import_verzeichnisse_sicherstellen()
    FIELD_INBOX_DIR.mkdir(parents=True, exist_ok=True)


def field_eingangspakete_laden() -> list[dict[str, Any]]:
    field_eingang_verzeichnisse_sicherstellen()
    result: list[dict[str, Any]] = []

    for zip_path in sorted(FIELD_INBOX_DIR.glob("*.zip")):
        try:
            item = field_schadenpaket_lesen(zip_path)
            valid_object = bool(
                normalisiere_gueltigen_objektordner(
                    item.get("objektordner", "")
                )
            )
            duplicate = field_import_bereits_importiert(item)

            if duplicate:
                status = "Bereits importiert"
            elif not valid_object:
                status = "Objektordner prüfen"
            elif not item.get("beschreibung"):
                status = "Beschreibung fehlt"
            else:
                status = "Importbereit"

            item["pruefstatus"] = status
            item["dateiname"] = zip_path.name
            item["fotoanzahl"] = len(item.get("fotos", []))
            result.append(item)
        except (OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
            result.append({
                "zip_path": str(zip_path),
                "dateiname": zip_path.name,
                "pruefstatus": f"Fehler: {exc}",
                "mitarbeiter": "",
                "objektordner": "",
                "wohnung": "",
                "raum": "",
                "schadensart": "",
                "prioritaet": "",
                "beschreibung": "",
                "erstellt_am": "",
                "fotoanzahl": 0,
                "fotos": [],
            })

    return result


def field_eingang_paket_kopieren(source: Path) -> Path:
    field_eingang_verzeichnisse_sicherstellen()
    target = FIELD_INBOX_DIR / source.name

    if target.exists():
        target = FIELD_INBOX_DIR / (
            f"{source.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            f"{source.suffix}"
        )

    shutil.copy2(source, target)
    return target


def field_pruefaufgabe_erzeugen(item: dict[str, Any]) -> bool:
    fields = SCHEMA.get("Aufgaben", [])
    if not fields:
        return False

    row = ["" for _ in fields]
    values = {
        "Aufgabe": f'DBS Field prüfen: {item.get("dateiname", "")}',
        "Titel": f'DBS Field prüfen: {item.get("dateiname", "")}',
        "Objekt": item.get("objektordner", ""),
        "Objektordner": item.get("objektordner", ""),
        "Bereich": "DBS Field Import",
        "Priorität": (
            "Hoch"
            if norm_key(item.get("prioritaet", "")) in {"hoch", "notfall"}
            else "Normal"
        ),
        "Status": "Neu",
        "Verantwortlich": item.get("mitarbeiter", "") or aktueller_mitarbeitername(),
        "Notiz": (
            f'Prüfstatus: {item.get("pruefstatus", "")}\n'
            f'Schadensart: {item.get("schadensart", "")}\n'
            f'Raum: {item.get("raum", "")}\n'
            f'Fotos: {item.get("fotoanzahl", 0)}\n'
            f'Datei: {item.get("zip_path", "")}'
        ),
    }

    for index, field in enumerate(fields):
        if field in values:
            row[index] = values[field]

    DATA.setdefault("Aufgaben", []).append(row)
    speichere_tabelle("Aufgaben")
    aktivitaet_protokollieren(
        "DBS Field Eingang",
        "Prüfaufgabe erzeugt",
        beschreibung=item.get("dateiname", ""),
    )
    return True


class DbsFieldEingangszentraleSeite(QWidget):
    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.rows: list[dict[str, Any]] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(20, 18, 20, 18)
        root.setSpacing(14)

        title = QLabel("DBS Field Eingangszentrale")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        subtitle = QLabel(
            "Mobile Schadenmeldungen sammeln, prüfen und gemeinsam importieren."
        )
        subtitle.setObjectName("subTitle")
        subtitle.setWordWrap(True)
        root.addWidget(subtitle)

        toolbar = QHBoxLayout()

        add_files = QPushButton("ZIP-Pakete hinzufügen")
        add_files.setObjectName("primaryButton")
        add_files.clicked.connect(self.pakete_hinzufuegen)

        refresh = QPushButton("Aktualisieren")
        refresh.clicked.connect(self.laden)

        open_folder = QPushButton("Eingangsordner öffnen")
        open_folder.clicked.connect(
            lambda: system_datei_oeffnen(str(FIELD_INBOX_DIR))
        )

        toolbar.addWidget(add_files)
        toolbar.addWidget(refresh)
        toolbar.addWidget(open_folder)
        toolbar.addStretch()
        root.addLayout(toolbar)

        self.summary = QLabel("")
        self.summary.setObjectName("metricTitle")
        root.addWidget(self.summary)

        self.table = QTableWidget()
        self.table.setColumnCount(11)
        self.table.setHorizontalHeaderLabels([
            "Prüfstatus", "Datei", "Mitarbeiter", "Objektordner",
            "Wohnung", "Raum", "Schadensart", "Priorität",
            "Fotos", "Erstellt", "Beschreibung",
        ])
        self.table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows
        )
        self.table.setSelectionMode(
            QAbstractItemView.SelectionMode.ExtendedSelection
        )
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setHorizontalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAlwaysOn
        )
        self.table.setVerticalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAlwaysOn
        )
        root.addWidget(self.table, 1)

        actions = QHBoxLayout()

        import_selected = QPushButton("Auswahl importieren")
        import_selected.setObjectName("primaryButton")
        import_selected.clicked.connect(self.auswahl_importieren)

        create_tasks = QPushButton("Prüfaufgaben erzeugen")
        create_tasks.clicked.connect(self.pruefaufgaben_erzeugen)

        single_import = QPushButton("Einzelimport öffnen")
        single_import.clicked.connect(lambda: self.nav("DBS Field Import"))

        actions.addWidget(import_selected)
        actions.addWidget(create_tasks)
        actions.addWidget(single_import)
        actions.addStretch()
        root.addLayout(actions)

        self.laden()

    def pakete_hinzufuegen(self) -> None:
        paths, _ = QFileDialog.getOpenFileNames(
            self,
            "DBS-Field-Schadenpakete auswählen",
            str(APP_DIR),
            "ZIP-Dateien (*.zip)",
        )

        copied = 0
        errors: list[str] = []

        for value in paths:
            try:
                field_eingang_paket_kopieren(Path(value))
                copied += 1
            except OSError as exc:
                errors.append(f"{Path(value).name}: {exc}")

        self.laden()
        message = f"{copied} Paket(e) hinzugefügt."
        if errors:
            message += "\n\nFehler:\n" + "\n".join(errors)
        QMessageBox.information(self, "DBS Field Eingang", message)

    def laden(self) -> None:
        self.rows = field_eingangspakete_laden()
        ready = sum(
            1 for item in self.rows
            if item.get("pruefstatus") == "Importbereit"
        )
        duplicates = sum(
            1 for item in self.rows
            if item.get("pruefstatus") == "Bereits importiert"
        )

        self.summary.setText(
            f"{len(self.rows)} Paket(e) · "
            f"{ready} importbereit · "
            f"{duplicates} bereits importiert"
        )

        self.table.setRowCount(len(self.rows))
        keys = [
            "pruefstatus", "dateiname", "mitarbeiter", "objektordner",
            "wohnung", "raum", "schadensart", "prioritaet",
            "fotoanzahl", "erstellt_am", "beschreibung",
        ]

        for row_index, item in enumerate(self.rows):
            for column_index, key in enumerate(keys):
                self.table.setItem(
                    row_index,
                    column_index,
                    QTableWidgetItem(str(item.get(key, ""))),
                )

        widths = [170, 260, 130, 240, 120, 140, 190, 110, 70, 170, 520]
        for index, width in enumerate(widths):
            self.table.setColumnWidth(index, width)

    def _selected_items(self) -> list[dict[str, Any]]:
        rows = sorted({
            index.row()
            for index in self.table.selectionModel().selectedRows()
        })
        return [
            self.rows[index]
            for index in rows
            if 0 <= index < len(self.rows)
        ]

    def auswahl_importieren(self) -> None:
        items = self._selected_items()
        if not items:
            QMessageBox.information(
                self,
                "DBS Field Eingang",
                "Bitte mindestens ein Paket auswählen.",
            )
            return

        imported = 0
        errors: list[str] = []

        for item in items:
            if item.get("pruefstatus") != "Importbereit":
                errors.append(
                    f'{item.get("dateiname", "")}: '
                    f'{item.get("pruefstatus", "")}'
                )
                continue

            try:
                field_schaden_importieren(item)
                imported += 1
            except (OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
                errors.append(f'{item.get("dateiname", "")}: {exc}')

        self.laden()
        message = f"{imported} Schadenmeldung(en) importiert."
        if errors:
            message += "\n\nNicht importiert:\n" + "\n".join(errors)
        QMessageBox.information(self, "DBS Field Eingang", message)

    def pruefaufgaben_erzeugen(self) -> None:
        items = self._selected_items()
        if not items:
            QMessageBox.information(
                self,
                "DBS Field Eingang",
                "Bitte mindestens ein Paket auswählen.",
            )
            return

        created = sum(
            1 for item in items
            if field_pruefaufgabe_erzeugen(item)
        )
        QMessageBox.information(
            self,
            "DBS Field Eingang",
            f"{created} Prüfaufgabe(n) erzeugt.",
        )



def schadenleitstand_rows() -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for index, row in enumerate(DATA.get("Schäden", [])):
        prio = feldwert("Schäden", row, ["Priorität", "Dringlichkeit"]) or "Niedrig"
        status = feldwert("Schäden", row, ["Status"]) or "Gemeldet"
        key = norm_key(prio)
        ampel = "🔴" if key == "notfall" else "🟠" if key == "hoch" else "🟡" if key == "mittel" else "🟢"
        result.append({
            "index": index,
            "ampel": ampel,
            "prioritaet": prio,
            "status": status,
            "objekt": objektordner_fuer_datensatz("Schäden", row),
            "wohnung": feldwert("Schäden", row, ["Wohnung"]),
            "raum": feldwert("Schäden", row, ["Raum", "Bereich"]),
            "art": feldwert("Schäden", row, ["Schadensart", "Kategorie", "Art"]),
            "beschreibung": feldwert("Schäden", row, ["Beschreibung", "Schaden", "Notiz"]),
            "datum": feldwert("Schäden", row, ["Datum", "Erstellt am"]),
            "mitarbeiter": feldwert("Schäden", row, ["Mitarbeiter", "Verantwortlich"]),
            "fotos": feldwert("Schäden", row, ["Fotos", "Foto", "Datei"]),
        })
    return result


def schadenleitstand_status_setzen(index: int, status: str) -> bool:
    rows = DATA.get("Schäden", [])
    fields = SCHEMA.get("Schäden", [])
    if not 0 <= index < len(rows) or "Status" not in fields:
        return False
    col = fields.index("Status")
    while len(rows[index]) <= col:
        rows[index].append("")
    rows[index][col] = status
    speichere_tabelle("Schäden")
    aktivitaet_protokollieren("Schadenleitstand", "Status geändert", status=status)
    return True


class SchadenleitstandProSeite(QWidget):
    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.rows: list[dict[str, Any]] = []
        self.filtered: list[dict[str, Any]] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(20, 18, 20, 18)
        root.setSpacing(12)

        title = QLabel("Schadenleitstand PRO")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        sub = QLabel("Alle Schäden zentral suchen, priorisieren und bearbeiten.")
        sub.setObjectName("subTitle")
        root.addWidget(sub)

        bar = QHBoxLayout()
        self.search = QLineEdit()
        self.search.setPlaceholderText("Objekt, Wohnung, Raum, Schaden oder Mitarbeiter suchen ...")
        self.search.textChanged.connect(self.filtern)
        self.status = QComboBox()
        self.status.addItems(["Alle Status", "Gemeldet", "In Bearbeitung", "Beauftragt", "Erledigt"])
        self.status.currentTextChanged.connect(self.filtern)
        refresh = QPushButton("Aktualisieren")
        refresh.setObjectName("primaryButton")
        refresh.clicked.connect(self.laden)
        bar.addWidget(self.search, 1)
        bar.addWidget(self.status)
        bar.addWidget(refresh)
        root.addLayout(bar)

        self.info = QLabel("")
        self.info.setObjectName("metricTitle")
        root.addWidget(self.info)

        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "Priorität", "Status", "Objektordner", "Wohnung", "Raum",
            "Schadensart", "Beschreibung", "Datum", "Mitarbeiter", "Fotos"
        ])
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table.cellDoubleClicked.connect(self.oeffnen)
        root.addWidget(self.table, 1)

        actions = QHBoxLayout()
        for label, value in [
            ("In Bearbeitung", "In Bearbeitung"),
            ("Beauftragt", "Beauftragt"),
            ("Erledigt", "Erledigt"),
        ]:
            button = QPushButton(label)
            button.clicked.connect(lambda checked=False, s=value: self.status_setzen(s))
            actions.addWidget(button)
        object_button = QPushButton("Objektordner öffnen")
        object_button.clicked.connect(self.objekt_oeffnen)
        actions.addWidget(object_button)
        actions.addStretch()
        root.addLayout(actions)
        self.laden()

    def laden(self) -> None:
        self.rows = schadenleitstand_rows()
        self.filtern()

    def filtern(self) -> None:
        query = norm_key(self.search.text())
        selected = self.status.currentText()
        self.filtered = []
        for item in self.rows:
            haystack = norm_key(" | ".join(str(item.get(k, "")) for k in [
                "objekt", "wohnung", "raum", "art", "beschreibung", "mitarbeiter"
            ]))
            if query and query not in haystack:
                continue
            if selected != "Alle Status" and norm_key(item["status"]) != norm_key(selected):
                continue
            self.filtered.append(item)

        open_count = sum(1 for item in self.rows if norm_key(item["status"]) not in {"erledigt", "geschlossen"})
        urgent = sum(1 for item in self.rows if norm_key(item["prioritaet"]) in {"hoch", "notfall"} and norm_key(item["status"]) != "erledigt")
        self.info.setText(f"{len(self.rows)} Schäden · {open_count} offen · {urgent} dringend")
        self.table.setRowCount(len(self.filtered))

        for r, item in enumerate(self.filtered):
            values = [
                f'{item["ampel"]} {item["prioritaet"]}', item["status"], item["objekt"],
                item["wohnung"], item["raum"], item["art"], item["beschreibung"],
                item["datum"], item["mitarbeiter"], item["fotos"]
            ]
            for c, value in enumerate(values):
                self.table.setItem(r, c, QTableWidgetItem(str(value)))
        for c, width in enumerate([120, 150, 240, 120, 130, 180, 520, 120, 140, 420]):
            self.table.setColumnWidth(c, width)

    def _item(self) -> dict[str, Any] | None:
        row = self.table.currentRow()
        return self.filtered[row] if 0 <= row < len(self.filtered) else None

    def status_setzen(self, status: str) -> None:
        item = self._item()
        if item is None:
            QMessageBox.information(self, "Schadenleitstand", "Bitte einen Schaden auswählen.")
            return
        if schadenleitstand_status_setzen(int(item["index"]), status):
            self.laden()

    def objekt_oeffnen(self) -> None:
        item = self._item()
        if item is None:
            return
        folder = normalisiere_gueltigen_objektordner(item["objekt"])
        if not folder:
            QMessageBox.warning(self, "Objektordner", "Objektordner nicht gefunden.")
            return
        path = Path(folder)
        if not path.is_absolute():
            path = DOKUMENTE_DIR / path
        system_datei_oeffnen(str(path))

    def oeffnen(self, row: int, column: int) -> None:
        if 0 <= row < len(self.filtered) and column == 9:
            first = str(self.filtered[row].get("fotos", "")).split("|")[0].strip()
            if first:
                system_datei_oeffnen(first)
                return
        self.objekt_oeffnen()



def dbs_field_json_lesen(path: Path) -> dict[str, Any]:
    if not path.exists() or path.suffix.lower() != ".json":
        raise ValueError("Bitte eine gültige JSON-Datei auswählen.")

    try:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"JSON-Datei konnte nicht gelesen werden: {exc}") from exc

    # Manche Mailprogramme verpacken den eigentlichen Inhalt als JSON-String.
    if isinstance(payload, str):
        try:
            payload = json.loads(payload)
        except json.JSONDecodeError as exc:
            raise ValueError("Die Datei enthält nur Text, aber kein DBS-Field-JSON.") from exc

    if not isinstance(payload, dict):
        raise ValueError("Die Datei enthält kein gültiges JSON-Objekt.")

    payload["_datei"] = str(path)
    return payload


def dbs_field_typ_normalisieren(value: Any) -> str:
    compact = norm_key(value)
    aliases = {
        "stundennachweisemehrtag": "stundennachweise_mehrtag",
        "stundennachweismehrtag": "stundennachweise_mehrtag",
        "mehrtagstundennachweise": "stundennachweise_mehrtag",
        "zeiterfassung": "stundennachweise_mehrtag",
        "timesheets": "stundennachweise_mehrtag",
        "terminplaner": "terminplaner",
        "termine": "terminplaner",
        "appointments": "terminplaner",
        "kalender": "terminplaner",
    }
    return aliases.get(compact, compact)


def dbs_field_liste_finden(payload: dict[str, Any], typ: str) -> tuple[list[Any], str]:
    keys = (
        ["eintraege", "einträge", "entries", "daten", "stundennachweise", "timesheets"]
        if typ == "stundennachweise_mehrtag"
        else ["termine", "appointments", "entries", "daten", "kalender"]
    )
    for key in keys:
        value = payload.get(key)
        if isinstance(value, list):
            return value, key
        if isinstance(value, str) and value.strip():
            try:
                decoded = json.loads(value)
            except json.JSONDecodeError:
                continue
            if isinstance(decoded, list):
                return decoded, key
    return [], ""


def dbs_field_objektordner_pruefen(value: str) -> tuple[str, str]:
    raw = str(value or "").strip()
    if not raw:
        return "", "Objektordner fehlt"
    folder = normalisiere_gueltigen_objektordner(raw)
    if folder:
        return folder, ""
    # Außendienstwerte dürfen importiert werden, auch wenn der Ordner lokal
    # noch nicht angelegt oder anders benannt ist. So gehen keine Daten verloren.
    return raw, f"Objektordner nicht lokal zugeordnet: {raw}"


def dbs_field_signatur(entry: dict[str, Any], typ: str) -> str:
    if typ == "stundennachweise_mehrtag":
        parts = [
            entry.get("id", ""), entry.get("mitarbeiter", ""),
            entry.get("datum", ""), entry.get("objektordner", ""),
            entry.get("taetigkeit", entry.get("tätigkeit", "")),
            entry.get("beginn", ""), entry.get("ende", ""),
            entry.get("pause", ""), entry.get("stunden", ""),
        ]
    else:
        parts = [
            entry.get("id", ""), entry.get("titel", ""),
            entry.get("datum", ""), entry.get("beginn", ""),
            entry.get("ende", ""), entry.get("objektordner", ""),
            entry.get("ort", ""),
        ]
    return "|".join(str(value).strip().casefold() for value in parts)


def dbs_field_bestehende_signaturen(target: str, typ: str) -> set[str]:
    fields = SCHEMA.get(target, [])
    signatures: set[str] = set()

    def field_value(row: list[Any], names: list[str]) -> Any:
        for name in names:
            if name in fields:
                index = fields.index(name)
                if index < len(row):
                    return row[index]
        return ""

    for row in DATA.get(target, []):
        if not isinstance(row, list):
            continue
        if typ == "stundennachweise_mehrtag":
            entry = {
                "id": field_value(row, ["ID"]),
                "mitarbeiter": field_value(row, ["Mitarbeiter"]),
                "datum": field_value(row, ["Datum"]),
                "objektordner": field_value(row, ["Objektordner", "Objekt"]),
                "taetigkeit": field_value(row, ["Tätigkeit", "Taetigkeit"]),
                "beginn": field_value(row, ["Beginn"]),
                "ende": field_value(row, ["Ende"]),
                "pause": field_value(row, ["Pause"]),
                "stunden": field_value(row, ["Stunden"]),
            }
        else:
            entry = {
                "id": field_value(row, ["ID"]),
                "titel": field_value(row, ["Titel", "Termin"]),
                "datum": field_value(row, ["Datum"]),
                "beginn": field_value(row, ["Beginn"]),
                "ende": field_value(row, ["Ende"]),
                "objektordner": field_value(row, ["Objektordner", "Objekt"]),
                "ort": field_value(row, ["Ort"]),
            }
        signatures.add(dbs_field_signatur(entry, typ))
    return signatures


def dbs_field_import_pruefen(payload: dict[str, Any]) -> dict[str, Any]:
    typ_raw = payload.get("typ", payload.get("type", payload.get("art", "")))
    typ = dbs_field_typ_normalisieren(typ_raw)

    if typ == "stundennachweise_mehrtag":
        target = "Stundennachweise" if "Stundennachweise" in SCHEMA else "Zeiterfassung"
    elif typ == "terminplaner":
        target = "Termine" if "Termine" in SCHEMA else "Kalender"
    else:
        raise ValueError(
            f"Unbekannter DBS-Field-Dateityp: {typ_raw!r}. "
            "Erwartet werden Stundennachweise oder Terminplaner."
        )

    entries, source_key = dbs_field_liste_finden(payload, typ)
    if not source_key:
        available = ", ".join(sorted(str(key) for key in payload.keys() if not str(key).startswith("_")))
        raise ValueError(
            "Keine Datensatzliste gefunden. Vorhandene JSON-Felder: "
            f"{available or 'keine'}."
        )
    if not SCHEMA.get(target):
        raise ValueError(f"Die Zieltabelle „{target}“ wurde nicht gefunden.")

    existing = dbs_field_bestehende_signaturen(target, typ)
    current: set[str] = set()
    valid: list[dict[str, Any]] = []
    duplicates: list[dict[str, Any]] = []
    invalid: list[tuple[dict[str, Any], str]] = []
    warnings: list[tuple[dict[str, Any], str]] = []

    for raw in entries:
        if not isinstance(raw, dict):
            invalid.append(({"wert": raw}, "Datensatz ist kein JSON-Objekt"))
            continue
        entry = dict(raw)
        if typ == "stundennachweise_mehrtag":
            entry["taetigkeit"] = entry.get("taetigkeit", entry.get("tätigkeit", entry.get("activity", "")))
            entry["mitarbeiter"] = entry.get("mitarbeiter", entry.get("employee", ""))
            entry["datum"] = entry.get("datum", entry.get("date", ""))
            entry["beginn"] = entry.get("beginn", entry.get("start", ""))
            entry["ende"] = entry.get("ende", entry.get("end", ""))
            entry["objektordner"] = entry.get("objektordner", entry.get("objekt", entry.get("property", "")))
            if not str(entry.get("datum", "")).strip():
                invalid.append((entry, "Datum fehlt")); continue
            if not str(entry.get("beginn", "")).strip() or not str(entry.get("ende", "")).strip():
                invalid.append((entry, "Beginn oder Ende fehlt")); continue
            folder, warning = dbs_field_objektordner_pruefen(entry.get("objektordner", ""))
            entry["objektordner"] = folder
            if warning:
                warnings.append((entry, warning))
        else:
            entry["id"] = entry.get("id", entry.get("termin_id", entry.get("appointment_id", "")))
            entry["titel"] = entry.get("titel", entry.get("title", entry.get("termin", entry.get("subject", ""))))
            entry["datum"] = entry.get("datum", entry.get("date", entry.get("start_date", "")))
            entry["beginn"] = entry.get("beginn", entry.get("start", entry.get("start_time", "")))
            entry["ende"] = entry.get("ende", entry.get("end", entry.get("end_time", "")))
            entry["ort"] = entry.get("ort", entry.get("location", entry.get("adresse", "")))
            entry["notiz"] = entry.get(
                "notiz",
                entry.get("note", entry.get("beschreibung", entry.get("description", ""))),
            )
            entry["status"] = entry.get("status", entry.get("state", "Geplant"))
            entry["objektordner"] = entry.get(
                "objektordner",
                entry.get("objekt", entry.get("property", entry.get("property_folder", ""))),
            )
            if not str(entry.get("titel", "")).strip() or not str(entry.get("datum", "")).strip():
                invalid.append((entry, "Titel oder Datum fehlt")); continue
            if str(entry.get("objektordner", "")).strip():
                folder, warning = dbs_field_objektordner_pruefen(entry["objektordner"])
                entry["objektordner"] = folder
                if warning:
                    warnings.append((entry, warning))

        signature = dbs_field_signatur(entry, typ)
        if signature in existing or signature in current:
            duplicates.append(entry)
            continue
        current.add(signature)
        valid.append(entry)

    return {
        "typ": typ, "typ_raw": str(typ_raw), "source_key": source_key,
        "target": target, "valid": valid, "duplicates": duplicates,
        "invalid": invalid, "warnings": warnings, "total": len(entries),
        "reported_total": payload.get("anzahl", payload.get("count", "")),
    }


def dbs_field_zeiteintrag_zeile(
    entry: dict[str, Any],
    fields: list[str],
) -> list[Any]:
    objekt = str(entry.get("objektordner", "")).strip()
    values = {
        "Mitarbeiter": entry.get("mitarbeiter", ""),
        "Datum": entry.get("datum", ""),
        "Objekt": objekt,
        "Objektordner": objekt,
        "Tätigkeit": entry.get("taetigkeit", ""),
        "Taetigkeit": entry.get("taetigkeit", ""),
        "Beginn": entry.get("beginn", ""),
        "Ende": entry.get("ende", ""),
        "Pause": entry.get("pause", ""),
        "Stunden": entry.get("stunden", ""),
        "Notiz": entry.get("notiz", ""),
        "Status": entry.get("status", "Eingereicht"),
        "ID": entry.get("id", ""),
    }
    return [values.get(field, "") for field in fields]


def dbs_field_termin_zeile(
    entry: dict[str, Any],
    fields: list[str],
) -> list[Any]:
    objekt = str(entry.get("objektordner", "")).strip()
    values = {
        "Titel": entry.get("titel", ""),
        "Termin": entry.get("titel", ""),
        "Datum": entry.get("datum", ""),
        "Beginn": entry.get("beginn", ""),
        "Ende": entry.get("ende", ""),
        "Objekt": objekt,
        "Objektordner": objekt,
        "Ort": entry.get("ort", ""),
        "Notiz": entry.get("notiz", ""),
        "Status": entry.get("status", "Geplant"),
        "ID": entry.get("id", ""),
    }
    return [values.get(field, "") for field in fields]


def dbs_field_importprotokoll_speichern(
    payload: dict[str, Any],
    result: dict[str, Any],
    backup: Path | None,
) -> Path:
    log_dir = EXPORT_DIR / "dbs_field_importprotokolle"
    log_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    log_path = log_dir / f"dbs_field_import_{timestamp}.json"

    log_payload = {
        "zeitpunkt": datetime.now().isoformat(timespec="seconds"),
        "quelldatei": payload.get("_datei", ""),
        "typ": result["typ"],
        "zieltabelle": result["target"],
        "gesamt": result["total"],
        "importiert": len(result["valid"]),
        "dubletten": len(result["duplicates"]),
        "ungueltig": len(result["invalid"]),
        "backup": str(backup) if backup else "",
        "ungueltige_eintraege": [
            {"eintrag": entry, "grund": reason}
            for entry, reason in result["invalid"]
        ],
    }
    log_path.write_text(
        json.dumps(log_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return log_path


def dbs_field_json_importieren(
    payload: dict[str, Any],
) -> tuple[str, int, int, int, Path | None, Path]:
    result = dbs_field_import_pruefen(payload)
    target = result["target"]
    fields = SCHEMA[target]

    backup = erstelle_schnellbackup()

    if result["typ"] == "stundennachweise_mehrtag":
        rows = [
            dbs_field_zeiteintrag_zeile(item, fields)
            for item in result["valid"]
        ]
    else:
        rows = [
            dbs_field_termin_zeile(item, fields)
            for item in result["valid"]
        ]

    if rows:
        DATA.setdefault(target, []).extend(rows)
        speichere_tabelle(target)

    log_path = dbs_field_importprotokoll_speichern(
        payload,
        result,
        backup,
    )

    aktivitaet_protokollieren(
        "DBS Field Datenimport",
        "Sicherer Import abgeschlossen",
        beschreibung=(
            f"{len(rows)} importiert, "
            f"{len(result['duplicates'])} Dubletten, "
            f"{len(result['invalid'])} ungültig"
        ),
        status="Importiert",
    )

    return (
        target,
        len(rows),
        len(result["duplicates"]),
        len(result["invalid"]),
        backup,
        log_path,
    )


class DbsFieldDatenimportSeite(QWidget):
    def __init__(self, nav):
        super().__init__()
        self.nav = nav
        self.payload: dict[str, Any] | None = None
        self.check_result: dict[str, Any] | None = None

        root = QVBoxLayout(self)
        root.setContentsMargins(20, 18, 20, 18)
        root.setSpacing(14)

        title = QLabel("DBS Field Datenimport 7.4")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        subtitle = QLabel(
            "Importiert Stundennachweise und Kalendertermine mit "
            "Objektordnerprüfung, Dublettenfilter, Backup und Importprotokoll."
        )
        subtitle.setObjectName("subTitle")
        subtitle.setWordWrap(True)
        root.addWidget(subtitle)

        bar = QHBoxLayout()
        self.path_input = QLineEdit()
        self.path_input.setPlaceholderText(
            "JSON-Datei aus DBS Field auswählen ..."
        )

        choose = QPushButton("JSON auswählen")
        choose.setObjectName("primaryButton")
        choose.clicked.connect(self.datei_waehlen)

        check = QPushButton("Datei prüfen")
        check.clicked.connect(self.pruefen)

        bar.addWidget(self.path_input, 1)
        bar.addWidget(choose)
        bar.addWidget(check)
        root.addLayout(bar)

        summary = QHBoxLayout()
        self.total_label = QLabel("Gesamt: 0")
        self.valid_label = QLabel("Importierbar: 0")
        self.duplicate_label = QLabel("Dubletten: 0")
        self.invalid_label = QLabel("Ungültig: 0")

        for label in [
            self.total_label,
            self.valid_label,
            self.duplicate_label,
            self.invalid_label,
        ]:
            label.setStyleSheet(
                "background:#f5f8fc; border:1px solid #dce5f0; "
                "border-radius:10px; padding:10px; font-weight:600;"
            )
            summary.addWidget(label)

        root.addLayout(summary)

        self.info = QLabel("Noch keine Datei geprüft.")
        self.info.setWordWrap(True)
        self.info.setTextInteractionFlags(
            Qt.TextInteractionFlag.TextSelectableByMouse
        )
        root.addWidget(self.info)

        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setHorizontalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAlwaysOn
        )
        root.addWidget(self.table, 1)

        actions = QHBoxLayout()

        import_button = QPushButton(
            "Backup erstellen und gültige Daten importieren"
        )
        import_button.setObjectName("primaryButton")
        import_button.clicked.connect(self.importieren)

        protocol_button = QPushButton("Importprotokolle öffnen")
        protocol_button.clicked.connect(self.protokolle_oeffnen)

        actions.addWidget(import_button)
        actions.addWidget(protocol_button)
        actions.addStretch()
        root.addLayout(actions)

    def datei_waehlen(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "DBS-Field-JSON auswählen",
            str(APP_DIR),
            "JSON-Dateien (*.json)",
        )
        if path:
            self.path_input.setText(path)
            self.pruefen()

    def pruefen(self) -> None:
        try:
            self.payload = dbs_field_json_lesen(
                Path(self.path_input.text().strip())
            )
            self.check_result = dbs_field_import_pruefen(self.payload)
        except ValueError as exc:
            self.payload = None
            self.check_result = None
            QMessageBox.warning(
                self,
                "DBS Field Datenimport",
                str(exc),
            )
            return

        result = self.check_result
        self.total_label.setText(f"Gesamt: {result['total']}")
        self.valid_label.setText(
            f"Importierbar: {len(result['valid'])}"
        )
        self.duplicate_label.setText(
            f"Dubletten: {len(result['duplicates'])}"
        )
        self.invalid_label.setText(
            f"Ungültig: {len(result['invalid'])}"
        )

        label = (
            "Mehrtag-Stundennachweise"
            if result["typ"] == "stundennachweise_mehrtag"
            else "Terminplaner"
        )
        self.info.setText(
            f"<b>Dateityp:</b> {label}<br>"
            f"<b>Erkannter Typ:</b> {result['typ']} (Original: {result['typ_raw']})<br>"
            f"<b>Datenfeld:</b> {result['source_key']}<br>"
            f"<b>Zieltabelle:</b> {result['target']}<br>"
            f"<b>Datei:</b> {self.payload.get('_datei', '')}<br>"
            "Grün = importierbar, Gelb = Dublette, Rot = ungültig."
        )

        if result["typ"] == "stundennachweise_mehrtag":
            headers = [
                "Prüfung", "Mitarbeiter", "Datum", "Objektordner",
                "Tätigkeit", "Beginn", "Ende", "Pause", "Stunden",
                "Notiz",
            ]
            keys = [
                "mitarbeiter", "datum", "objektordner", "taetigkeit",
                "beginn", "ende", "pause", "stunden", "notiz",
            ]
        else:
            headers = [
                "Prüfung", "Titel", "Datum", "Beginn", "Ende",
                "Objektordner / Ort", "Notiz", "Status",
            ]
            keys = [
                "titel", "datum", "beginn", "ende",
                "objektordner", "notiz", "status",
            ]

        rows: list[tuple[str, dict[str, Any]]] = []
        rows.extend(("Importierbar", item) for item in result["valid"])
        rows.extend(("Dublette", item) for item in result["duplicates"])
        rows.extend(
            (f"Ungültig: {reason}", item)
            for item, reason in result["invalid"]
        )

        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(rows))

        for row_index, (status, item) in enumerate(rows):
            self.table.setItem(
                row_index,
                0,
                QTableWidgetItem(status),
            )
            for column_index, key in enumerate(keys, start=1):
                self.table.setItem(
                    row_index,
                    column_index,
                    QTableWidgetItem(str(item.get(key, ""))),
                )

        self.table.setColumnWidth(0, 250)
        for column in range(1, len(headers)):
            self.table.setColumnWidth(column, 165)

    def importieren(self) -> None:
        if self.payload is None or self.check_result is None:
            QMessageBox.information(
                self,
                "DBS Field Datenimport",
                "Bitte zuerst eine JSON-Datei prüfen.",
            )
            return

        import_count = len(self.check_result["valid"])
        if import_count == 0:
            QMessageBox.information(
                self,
                "DBS Field Datenimport",
                "Es wurden keine neuen gültigen Einträge gefunden.",
            )
            return

        answer = QMessageBox.question(
            self,
            "DBS Field Datenimport",
            (
                f"{import_count} neue Einträge werden importiert.\n"
                "Vorher wird automatisch ein vollständiges Backup erstellt.\n\n"
                "Import jetzt starten?"
            ),
        )
        if answer != QMessageBox.StandardButton.Yes:
            return

        try:
            (
                target,
                imported,
                duplicates,
                invalid,
                backup,
                log_path,
            ) = dbs_field_json_importieren(self.payload)
        except (OSError, ValueError, KeyError) as exc:
            QMessageBox.warning(
                self,
                "DBS Field Datenimport",
                f"Import fehlgeschlagen:\n{exc}",
            )
            return

        backup_text = str(backup) if backup else "Backup nicht erstellt"
        QMessageBox.information(
            self,
            "DBS Field Datenimport",
            (
                f"Zieltabelle: {target}\n"
                f"Importiert: {imported}\n"
                f"Dubletten übersprungen: {duplicates}\n"
                f"Ungültige Einträge: {invalid}\n\n"
                f"Backup: {backup_text}\n"
                f"Protokoll: {log_path}"
            ),
        )
        self.pruefen()

    def protokolle_oeffnen(self) -> None:
        path = EXPORT_DIR / "dbs_field_importprotokolle"
        path.mkdir(parents=True, exist_ok=True)
        system_datei_oeffnen(str(path))

class DbsFieldKalenderimportSeite(DbsFieldDatenimportSeite):
    """Eigene Importseite ausschließlich für Termine aus DBS Field."""

    def __init__(self, nav):
        super().__init__(nav)

        title_labels = self.findChildren(QLabel, "pageTitle")
        if title_labels:
            title_labels[0].setText("DBS Field Kalenderimport 7.4")

        subtitle_labels = self.findChildren(QLabel, "subTitle")
        if subtitle_labels:
            subtitle_labels[0].setText(
                "Importiert Termine aus DBS Field direkt in die Tabelle "
                "„Termine“. "
                "Stundennachweis-Dateien werden auf dieser Seite abgewiesen."
            )

        self.path_input.setPlaceholderText(
            "Termin-JSON aus DBS Field auswählen ..."
        )

    def pruefen(self) -> None:
        super().pruefen()

        if self.check_result is None:
            return

        if self.check_result.get("typ") != "terminplaner":
            self.payload = None
            self.check_result = None
            self.table.setRowCount(0)
            self.total_label.setText("Gesamt: 0")
            self.valid_label.setText("Importierbar: 0")
            self.duplicate_label.setText("Dubletten: 0")
            self.invalid_label.setText("Ungültig: 0")
            self.info.setText(
                "Die ausgewählte Datei enthält keine Kalendertermine."
            )
            QMessageBox.warning(
                self,
                "DBS Field Kalenderimport",
                "Diese Seite akzeptiert ausschließlich Termin-JSON-Dateien "
                "aus dem DBS-Field-Terminplaner."
            )
            return

        self.info.setText(
            self.info.text()
            + "<br><b>Kalenderimport:</b> Datei wurde als Terminplaner erkannt."
        )


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"{APP_NAME} · {APP_VERSION}")
        if APP_ICON_FILE.exists():
            self.setWindowIcon(QIcon(str(APP_ICON_FILE)))
        elif LOGO_FILE.exists():
            self.setWindowIcon(QIcon(str(LOGO_FILE)))
        self.resize(1580, 960)
        self.nav_buttons: dict[str, QPushButton] = {}
        self.sidebar_collapsed = False
        self.page_animation: QPropertyAnimation | None = None

        splitter = QSplitter()
        self.setCentralWidget(splitter)

        self.sidebar = QFrame()
        self.sidebar.setObjectName("sidebar")
        self.sidebar.setMinimumWidth(292)
        self.sidebar.setMaximumWidth(330)

        side = QVBoxLayout(self.sidebar)
        side.setContentsMargins(0, 0, 0, 0)
        side.setSpacing(0)

        logo_panel = QFrame()
        logo_panel.setObjectName("logoPanel")
        logo_layout = QVBoxLayout(logo_panel)
        logo_layout.setContentsMargins(14, 12, 14, 8)

        logo_row = QHBoxLayout()
        logo = QLabel()
        logo.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        if LOGO_FILE.exists():
            pix = QPixmap(str(LOGO_FILE))
            logo.setPixmap(
                pix.scaled(
                    185,
                    82,
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation,
                )
            )
        else:
            logo.setText("DBS")
            logo.setObjectName("pageTitle")

        self.sidebar_toggle = QPushButton("☰")
        self.sidebar_toggle.setObjectName("sidebarToggle")
        self.sidebar_toggle.setToolTip("Navigation ein-/ausklappen")
        self.sidebar_toggle.clicked.connect(self.toggle_sidebar)

        logo_row.addWidget(logo, 1)
        logo_row.addWidget(self.sidebar_toggle)
        logo_layout.addLayout(logo_row)

        self.sidebar_heading = QLabel("IMMOVERWALTUNG")
        self.sidebar_heading.setObjectName("sidebarTitle")
        logo_layout.addWidget(self.sidebar_heading)

        brand_line = QLabel("DBS · Kirchstraße 3 · 04703 Leisnig")
        brand_line.setObjectName("brandText")
        logo_layout.addWidget(brand_line)
        side.addWidget(logo_panel)

        nav_scroll = QScrollArea()
        nav_scroll.setWidgetResizable(True)
        nav_scroll.setFrameShape(QFrame.Shape.NoFrame)
        nav_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        nav_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        nav_inner = QWidget()
        self.nav_layout = QVBoxLayout(nav_inner)
        self.nav_layout.setContentsMargins(0, 10, 0, 12)
        self.nav_layout.setSpacing(4)

        content = QFrame()
        content.setObjectName("content")
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.setSpacing(0)

        topbar = QFrame()
        topbar.setObjectName("topbar")
        topbar.setFixedHeight(62)
        top_layout = QHBoxLayout(topbar)
        top_layout.setContentsMargins(18, 8, 18, 8)
        top_layout.setSpacing(10)

        self.header = QLabel("Immobilienverwaltung")
        self.header.setObjectName("header")
        top_layout.addWidget(self.header)

        top_layout.addStretch()

        self.top_search = QLineEdit()
        self.top_search.setObjectName("topSearch")
        self.top_search.setPlaceholderText("Suchen oder Strg+K drücken …")
        self.top_search.returnPressed.connect(self.open_global_search)
        top_layout.addWidget(self.top_search)

        search_button = QPushButton("⌕")
        search_button.setObjectName("topAction")
        search_button.setToolTip("Globale Suche")
        search_button.clicked.connect(self.open_global_search)
        top_layout.addWidget(search_button)

        notification_count = len(leitstand_benachrichtigungen())
        self.notification_button = QPushButton(f"🔔 {notification_count}")
        self.notification_button.setObjectName("notificationButton")
        self.notification_button.setToolTip("Benachrichtigungscenter öffnen")
        self.notification_button.clicked.connect(
            lambda: self.navigate("Verwaltungsleitstand 5.0")
        )
        top_layout.addWidget(self.notification_button)

        theme_button = QPushButton("◐")
        theme_button.setObjectName("topAction")
        theme_button.setToolTip("Darstellung in Einstellungen ändern")
        theme_button.clicked.connect(lambda: self.navigate("Einstellungen"))
        top_layout.addWidget(theme_button)

        self.user_label = QLabel(
            f"{CURRENT_USER.get('name', '')} · {CURRENT_USER.get('rolle', '')}"
        )
        self.user_label.setObjectName("userPill")
        top_layout.addWidget(self.user_label)

        content_layout.addWidget(topbar)

        self.stack = QStackedWidget()
        content_layout.addWidget(self.stack, 1)

        splitter.addWidget(self.sidebar)
        splitter.addWidget(content)
        splitter.setSizes([305, 1275])
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)

        self.titles = []
        self.add_page("Dashboard", Dashboard(self.navigate))
        self.add_page("Smart Startseite", SmartStartseiteSeite(self.navigate))
        self.add_page("Objektgalerie", ObjektGalerieSeite(self.navigate))
        self.add_page("Enterprise Objekt 360°", EnterpriseObjekt360Seite(self.navigate))
        self.add_page("Digitaler Gebäudezwilling", DigitalerGebaeudezwillingSeite(self.navigate))
        self.add_page("Enterprise Assistant", EnterpriseAssistentSeite(self.navigate))
        self.add_page("Projektmonitor", ProjektmonitorSeite(self.navigate))
        self.add_page("Dokumentenvorschau", DokumentVorschauSeite(self.navigate))
        self.add_page("Dokumenten-Center 2.0", DokumentenCenter2Seite(self.navigate))
        self.add_page("Dokumentenautomatisierung PRO", DokumentenautomatisierungProSeite(self.navigate))
        self.add_page("Workflow-Regeln PRO", WorkflowRegelnProSeite(self.navigate))
        self.add_page("Verwaltungsleitstand 5.0", VerwaltungsleitstandSeite(self.navigate))
        self.add_page("Berichte & Export PRO", BerichteExportProSeite(self.navigate))
        self.add_page("System-Center PRO", SystemCenterProSeite(self.navigate))
        self.add_page("Verwaltungs-Cockpit", VerwaltungsCockpitSeite(self.navigate))
        self.add_page("Objekt-Cockpit", ObjektCockpitSeite(self.navigate))
        self.add_page("Arbeitsorganisation PRO", ArbeitsorganisationSeite(self.navigate))
        self.add_page("Objektchronik", ObjektchronikSeite(self.navigate))
        self.add_page("Smart Objektakte", SmartObjektakteSeite(self.navigate))
        self.add_page("Fristenmanager PRO", FristenmanagerProSeite(self.navigate))
        self.add_page("Workflow-Center PRO", WorkflowCenterSeite(self.navigate))
        self.add_page("Mein Arbeitstag", MitarbeiterStartcenterSeite(self.navigate))
        self.add_page("DBS Field Import", DbsFieldImportSeite(self.navigate))
        self.add_page("DBS Field Eingangszentrale", DbsFieldEingangszentraleSeite(self.navigate))
        self.add_page("DBS Field Datenimport", DbsFieldDatenimportSeite(self.navigate))
        self.add_page("DBS Field Kalenderimport", DbsFieldKalenderimportSeite(self.navigate))
        self.add_page("Schadenleitstand PRO", SchadenleitstandProSeite(self.navigate))
        self.add_page("Kalender & Planung", KalenderPlanungSeite(self.navigate))
        self.add_page("Jahresprüfung PRO", JahrespruefungProSeite(self.navigate))
        self.add_page("Einstellungen", EinstellungenSeite(self.navigate))
        self.add_page("Objektordner-Prüfung", ObjektordnerPruefungSeite(self.navigate))
        self.add_page("Berechtigungen & Protokoll", BerechtigungenProtokollSeite(self.navigate))
        self.add_page("Globale Suche", GlobaleSucheSeite(self.navigate))
        self.add_page("Akten-Center", AktenCenterSeite())
        self.add_page("Analyse-Center", BeziehungsAnalyseSeite())
        self.add_page("Zahlenanalyse", ZahlenAnalyseSeite())
        self.add_page("Buchhaltung", BuchhaltungSeite())
        self.add_page("Kontoauszug-Import", KontoauszugImportSeite(self.navigate))
        self.add_page("Mietkonto-Abgleich", MietkontoAbgleichSeite(self.navigate))
        self.add_page("Zahlungsabgleich PRO", ZahlungsabgleichProSeite(self.navigate))
        self.add_page("Belegscanner PRO", BelegscannerSeite(self.navigate))
        self.add_page("BK-Automatik", BKAutomatikSeite())

        for title in SCHEMA.keys():
            self.add_page(title, TabellenSeite(title))

        # Version 11.0a: aufgeräumte Navigation. Technische Altseiten bleiben im
        # Programm registriert, damit bestehende interne Verknüpfungen weiterhin
        # funktionieren. Aus der Seitenleiste werden nur redundante Einstiege entfernt.
        # Der produktive DBS-Field-App-Import ist ausdrücklich geschützt und vollständig sichtbar.
        nav_groups = [
            ("START", [
                "Smart Startseite", "Verwaltungsleitstand 5.0",
                "Mein Arbeitstag", "Kalender & Planung", "Termine",
            ]),
            ("OBJEKTE & VERMIETUNG", [
                "Objekte", "Wohnungen", "Mieter", "Mietverträge",
                "Enterprise Objekt 360°", "Digitaler Gebäudezwilling", "Objektchronik",
                "Wohnungsgeberauskunft", "Übergabeprotokolle",
                "Schlüssel", "Schäden",
            ]),
            ("FINANZEN", [
                "Zahlungen", "Zahlungsprüfung", "Mahnwesen",
                "Rechnungen", "HV-Rechnungen", "Betriebskosten",
                "BK-Automatik", "Buchhaltung", "Kontoauszug-Import",
                "Mietkonto-Abgleich", "Zahlungsabgleich PRO",
                "Belegscanner PRO", "Grundsteuer", "Versicherungen",
            ]),
            ("DOKUMENTE & ORGANISATION", [
                "Dokumenten-Center 2.0", "Dokumente", "Akten-Center",
                "Arbeitsorganisation PRO", "Aufgaben", "Fristen",
                "Jahresprüfung PRO", "Globale Suche",
            ]),
            ("PERSONAL", [
                "Mitarbeiter", "Stundennachweise", "Mitarbeiter-Login",
                "Berechtigungen & Protokoll",
            ]),
            ("APP-IMPORT · PRODUKTIV", [
                "DBS Field Import", "DBS Field Eingangszentrale",
                "DBS Field Datenimport", "DBS Field Kalenderimport",
            ]),
            ("WEITERE VERWALTUNG", [
                "Dienstleister", "Versorger", "Wichtige Verträge",
                "Vermieterauskunft", "Brand- und Arbeitsschutz",
                "E-Mail", "Aktenstruktur", "Ereignisprotokoll",
            ]),
            ("AUSWERTUNG & SYSTEM", [
                "Berichte & Export PRO", "Analyse-Center", "Zahlenanalyse",
                "System-Center PRO", "Objektordner-Prüfung", "Einstellungen",
            ]),
        ]

        used_titles: set[str] = set()

        for group_title, group_items in nav_groups:
            visible_items = [item for item in group_items if item in self.titles]
            if not visible_items:
                continue

            group_label = QLabel(group_title)
            group_label.setStyleSheet(
                "color:#7ea1cc; font-size:10px; font-weight:900;"
                "padding:12px 20px 4px 20px; letter-spacing:1px;"
            )
            self.nav_layout.addWidget(group_label)

            for title in visible_items:
                self._add_nav_button(title)
                used_titles.add(title)

        # Absichtlich kein automatischer Block „Weitere Bereiche“ mehr:
        # dadurch tauchen alte, doppelte Cockpits und Übergangsseiten nicht erneut auf.

        self.nav_layout.addStretch()

        footer = QLabel("●  System bereit")
        footer.setStyleSheet(
            "color:#51d88a; padding:12px 22px; font-weight:800;"
        )
        self.nav_layout.addWidget(footer)

        nav_scroll.setWidget(nav_inner)
        side.addWidget(nav_scroll, 1)

        self.search_shortcut = QShortcut(QKeySequence("Ctrl+K"), self)
        self.search_shortcut.activated.connect(self.focus_top_search)

        self.command_shortcut = QShortcut(QKeySequence("Ctrl+Shift+P"), self)
        self.command_shortcut.activated.connect(self.open_command_palette)

        initial_page = (
            "Mein Arbeitstag"
            if aktuelle_mitarbeiterrolle().lower() != "admin"
            and "Mein Arbeitstag" in self.titles
            else "Smart Startseite"
        )
        self.navigate(initial_page)

    def _add_nav_button(self, title: str) -> None:
        button = QPushButton(self._nav_label(title))
        button.setObjectName("navButton")
        button.setMinimumHeight(42)
        button.setMinimumWidth(250)
        button.setToolTip(title)
        button.clicked.connect(
            lambda checked=False, target=title: self.navigate(target)
        )
        self.nav_buttons[title] = button
        self.nav_layout.addWidget(button)

    @staticmethod
    def _nav_label(title: str) -> str:
        icons = {
            "Smart Startseite": "✨",
            "Objektgalerie": "🗃",
            "Enterprise Objekt 360°": "🌐",
            "Digitaler Gebäudezwilling": "🏙",
            "Enterprise Assistant": "🤖",
            "Projektmonitor": "📡",
            "Dokumentenvorschau": "👁",
            "Dokumenten-Center 2.0": "📚",
            "Dokumentenautomatisierung PRO": "🧠",
            "Workflow-Regeln PRO": "⚙",
            "Dashboard": "⌂",
            "Verwaltungsleitstand 5.0": "🚀",
            "Berichte & Export PRO": "📊",
            "System-Center PRO": "🛠",
            "Objekte": "▦",
            "Wohnungen": "⌂",
            "Mieter": "👥",
            "Mietverträge": "▤",
            "Zahlungen": "▭",
            "Betriebskosten": "▧",
            "Rechnungen": "▤",
            "Dokumente": "□",
            "Wohnungsgeberauskunft": "▣",
            "Übergabeprotokolle": "▣",
            "Analyse-Center": "▥",
            "Zahlenanalyse": "◔",
            "Buchhaltung": "€",
            "BK-Automatik": "▣",
            "Arbeitsorganisation PRO": "🗂",
            "Objektchronik": "🕒",
            "Smart Objektakte": "🏢",
            "Fristenmanager PRO": "⏰",
            "Workflow-Center PRO": "⚡",
            "Mein Arbeitstag": "📌",
 ##           "Mobiles Mitarbeiterportal": "📱",
            "DBS Field Import": "📦",
            "DBS Field Datenimport": "📥",
            "DBS Field Kalenderimport": "📅",
            "DBS Field Eingangszentrale": "📥",
            "Kalender & Planung": "📅",
            "Termine": "📆",
            "Objektordner-Prüfung": "🔎",
 ##            "Berechtigungen & Protokoll": "🔐",
            "Jahresprüfung PRO": "✅",
            "Einstellungen": "⚙",
            "Globale Suche": "⌕",
            "Akten-Center": "🗄",
            "Kontoauszug-Import": "🏦",
            "Mietkonto-Abgleich": "🧮",
            "Zahlungsabgleich PRO": "🚦",
            "Belegscanner PRO": "📥",
        }
        return f"{icons.get(title, '•')}   {title}"

    def add_page(self, title, widget):
        self.titles.append(title)
        self.stack.addWidget(widget)

    def aktualisiere_navigation(self, active_title: str) -> None:
        for title, button in self.nav_buttons.items():
            button.setProperty(
                "active",
                "true" if title == active_title else "false",
            )
            button.style().unpolish(button)
            button.style().polish(button)
            button.update()

    def toggle_sidebar(self) -> None:
        self.sidebar_collapsed = not self.sidebar_collapsed

        if self.sidebar_collapsed:
            self.sidebar.setMinimumWidth(78)
            self.sidebar.setMaximumWidth(78)
            self.sidebar_heading.hide()

            for title, button in self.nav_buttons.items():
                label = self._nav_label(title)
                icon = label.split("   ", 1)[0]
                button.setText(icon)
                button.setMinimumWidth(54)
                button.setMaximumWidth(54)
        else:
            self.sidebar.setMinimumWidth(292)
            self.sidebar.setMaximumWidth(330)
            self.sidebar_heading.show()

            for title, button in self.nav_buttons.items():
                button.setText(self._nav_label(title))
                button.setMinimumWidth(250)
                button.setMaximumWidth(16777215)

    def focus_top_search(self) -> None:
        self.top_search.setFocus()
        self.top_search.selectAll()

    def open_global_search(self) -> None:
        query = self.top_search.text().strip()
        dialog = DashboardSucheDialog(self)
        if query:
            dialog.search.setText(query)
        dialog.exec()
        self.top_search.clear()

    def open_command_palette(self) -> None:
        dialog = QDialog(self)
        dialog.setWindowTitle("Schnellaktionen")
        dialog.resize(620, 520)

        layout = QVBoxLayout(dialog)
        title = QLabel("Schnellaktionen")
        title.setObjectName("pageTitle")
        layout.addWidget(title)

        search = QLineEdit()
        search.setPlaceholderText("Bereich oder Aktion suchen …")
        layout.addWidget(search)

        list_widget = QListWidget()
        commands = [
            "Smart Startseite",
            "Objektgalerie",
            "Enterprise Objekt 360°",
            "Dokumenten-Center 2.0",
            "Workflow-Regeln PRO",
            "Dashboard",
            "Verwaltungsleitstand 5.0",
            "Mein Arbeitstag",
            "Kalender & Planung",
            "Mobiles Mitarbeiterportal",
            "DBS Field Import",
            "DBS Field Eingangszentrale",
            "DBS Field Kalenderimport",
            "Neue Aufgabe",
            "Neue Frist",
            "Neue Rechnung",
            "Neuer Mieter",
            "Neues Objekt",
            "Belegscanner PRO",
            "Dokumentenautomatisierung PRO",
            "Kontoauszug-Import",
            "Backup erstellen",
            "Einstellungen",
        ]
        list_widget.addItems(commands)
        layout.addWidget(list_widget, 1)

        def filter_commands(value: str) -> None:
            search_key = value.lower().strip()
            for index in range(list_widget.count()):
                item = list_widget.item(index)
                item.setHidden(search_key not in item.text().lower())

        def execute_command() -> None:
            item = list_widget.currentItem()
            if item is None:
                return

            command = item.text()
            dialog.accept()

            navigation_map = {
                "Neue Aufgabe": "Arbeitsorganisation PRO",
                "Neue Frist": "Fristenmanager PRO",
                "Neue Rechnung": "Rechnungen",
                "Neuer Mieter": "Mieter",
                "Neues Objekt": "Objekte",
                "Backup erstellen": "System-Center PRO",
            }
            self.navigate(navigation_map.get(command, command))

        search.textChanged.connect(filter_commands)
        search.returnPressed.connect(execute_command)
        list_widget.itemDoubleClicked.connect(lambda _item: execute_command())

        search.setFocus()
        dialog.exec()

    def animate_current_page(self) -> None:
        page = self.stack.currentWidget()
        if page is None:
            return

        effect = QGraphicsOpacityEffect(page)
        page.setGraphicsEffect(effect)

        animation = QPropertyAnimation(effect, b"opacity", self)
        animation.setDuration(170)
        animation.setStartValue(0.35)
        animation.setEndValue(1.0)
        animation.setEasingCurve(QEasingCurve.Type.OutCubic)

        def cleanup() -> None:
            page.setGraphicsEffect(None)

        animation.finished.connect(cleanup)
        self.page_animation = animation
        animation.start()

    def navigate(self, title):
        mapping = {
            "Akten": "Aktenstruktur",
            "Offene Rechnungen": "Rechnungen",
            "Offene Forderungen": "Mahnwesen",
            # Kompatibilitätsweiterleitungen für ausgeblendete, redundante Einstiege.
            "Dashboard": "Smart Startseite",
            "Objektgalerie": "Enterprise Objekt 360°",
            "Smart Objektakte": "Enterprise Objekt 360°",
            "Enterprise Assistant": "Verwaltungsleitstand 5.0",
            "Projektmonitor": "Verwaltungsleitstand 5.0",
            "Dokumentenvorschau": "Dokumenten-Center 2.0",
            "Dokumentenautomatisierung PRO": "Dokumenten-Center 2.0",
            "Workflow-Regeln PRO": "Arbeitsorganisation PRO",
            "Workflow-Center PRO": "Arbeitsorganisation PRO",
            "Fristenmanager PRO": "Fristen",
            "Verwaltungs-Cockpit": "Verwaltungsleitstand 5.0",
            "Objekt-Cockpit": "Enterprise Objekt 360°",
        }
        target = mapping.get(title, title)

        if not benutzer_darf_seite_oeffnen(target):
            aktivitaet_protokollieren(
                "Navigation",
                "Zugriff verweigert",
                beschreibung=f"Kein Zugriff auf {target}",
                status="Abgelehnt",
            )
            QMessageBox.warning(
                self,
                "Keine Berechtigung",
                f"Für den Bereich '{target}' besteht keine Berechtigung.",
            )
            return

        if target not in self.titles:
            return

        aktivitaet_protokollieren(
            "Navigation",
            "Bereich geöffnet",
            beschreibung=target,
        )

        self.stack.setCurrentIndex(self.titles.index(target))
        self.stack.update()
        self.header.setText(target or "")
        benutzer_zuletzt_hinzufuegen(target)
        self.aktualisiere_navigation(target)
        self.notification_button.setText(
            f"🔔 {len(leitstand_benachrichtigungen())}"
        )
        self.animate_current_page()

def main():
    app = QApplication(sys.argv)
    app.setApplicationDisplayName(APP_NAME)
    if APP_ICON_FILE.exists():
        app.setWindowIcon(QIcon(str(APP_ICON_FILE)))
    elif LOGO_FILE.exists():
        app.setWindowIcon(QIcon(str(LOGO_FILE)))
    app.setStyleSheet(baue_stylesheet())
    lade_alle()

    login = LoginDialog()
    if login.exec() != QDialog.DialogCode.Accepted:
        sys.exit(0)

    if str(CONFIG.get("auto_backup_start", "nein")).lower() == "ja":
        erstelle_schnellbackup()

    win = MainWindow()
    if str(CONFIG.get("start_vollbild", "ja")).lower() == "ja":
        win.showMaximized()
    else:
        win.show()

    sys.exit(app.exec())

if __name__=="__main__": main()
